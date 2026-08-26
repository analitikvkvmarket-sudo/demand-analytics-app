from __future__ import annotations

import base64
import hashlib
import html
import calendar as pycalendar
import io
import json
import math
import os
import re
import tempfile
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from plotly.subplots import make_subplots
import psycopg
import streamlit as st
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
BUILD_ID = "75.11.17-WRITEOFF-CATEGORY-SKU-POINT-DRILLDOWN"


def resolve_app_file(filename: str, *name_fragments: str) -> Path:
    """Find a bundled workbook next to app.py, including renamed legacy files."""
    search_dirs = []
    for folder in (APP_DIR, Path.cwd(), APP_DIR.parent):
        try:
            resolved = folder.resolve()
        except OSError:
            resolved = folder
        if resolved not in search_dirs:
            search_dirs.append(resolved)

    aliases = {
        "entities.xlsx": ("таблица сущности.xlsx",),
        "combo_matrix.xlsx": ("2.3 Матрица КОМБО.xlsx",),
        "analyst_logic.xlsx": ("логика аналитика.xlsx",),
    }
    exact_names = (filename, *aliases.get(filename, ()))
    for folder in search_dirs:
        for exact_name in exact_names:
            candidate = folder / exact_name
            if candidate.exists():
                return candidate

    fragments = tuple(fragment.casefold() for fragment in name_fragments if fragment)
    if fragments:
        for folder in search_dirs:
            if not folder.exists() or not folder.is_dir():
                continue
            try:
                workbooks = sorted(folder.glob("*.xlsx"))
            except OSError:
                continue
            for candidate in workbooks:
                stem = candidate.stem.casefold()
                if any(fragment in stem for fragment in fragments):
                    return candidate

    # Last-resort content check for the entity dictionary. This also works if
    # a ZIP extractor damaged a Cyrillic filename but left the .xlsx intact.
    if filename == "entities.xlsx":
        for folder in search_dirs:
            if not folder.exists() or not folder.is_dir():
                continue
            try:
                workbooks = sorted(folder.glob("*.xlsx"))
            except OSError:
                continue
            for candidate in workbooks:
                try:
                    preview = pd.read_excel(candidate, header=None, nrows=20)
                except Exception:
                    continue
                for _, row in preview.iterrows():
                    values = {str(value).strip().casefold() for value in row if pd.notna(value)}
                    if "код" in values and "название блюда" in values:
                        return candidate

    return APP_DIR / filename


ENTITY_FILE = resolve_app_file("entities.xlsx", "сущност", "entit")
ANALYST_LOGIC_FILE = resolve_app_file("analyst_logic.xlsx", "логика", "аналит", "analyst")
COMBO_MATRIX_FILE = resolve_app_file("combo_matrix.xlsx", "матрица", "комбо", "combo")
AUTO_UNIT_FILE = resolve_app_file("auto_unit_points_vm.xlsx", "авто юнит", "auto_unit", "точки вм")
MATRIX_APPS_SCRIPT_URL = os.getenv(
    "MATRIX_APPS_SCRIPT_URL",
    "https://script.google.com/macros/s/AKfycbxzpVJGvkHTn8YogBOLO4PtvdQqmkoMx-chNCZ2ijmMIRc_-kcd2WUQ283PNyo5hCo/exec",
).strip()
MATRIX_APPS_SCRIPT_KEY = os.getenv(
    "MATRIX_APPS_SCRIPT_KEY",
    "VK_MATRIX_2026_8f31c5a7d942",
).strip()
GOOGLE_MATRIX_REFRESH_SECONDS = 15 * 60
GOOGLE_MATRIX_CACHE_FILE = Path(tempfile.gettempdir()) / "vkusnomarket_combo_plan_cache_v2.xlsx"
MATRIX_PLAN_SHEETS = [
    "План 1-я неделя",
    "План 2-я неделя",
    "План 3-я неделя",
    "План 4-я неделя",
]
MATRIX_ENTITY_SHEET = "Справочник + атрибуты"

st.set_page_config(
    page_title="Аналитика спроса",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)
load_dotenv(APP_DIR / ".env", override=True)

REMEMBERED_PG_FILE = APP_DIR / ".remembered_pg.json"
REMEMBERED_PG_DAYS = 30

MONTH_NAMES_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

def _coerce_api_cell(value: object) -> object:
    """Convert Apps Script display values to simple Excel-friendly Python values."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text == "":
        return None
    # Keep dates and labels as text; parsers below already understand Russian date strings.
    # Convert plain numeric cells so plan quantities remain numeric in the reconstructed XLSX.
    numeric = text.replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", numeric):
        try:
            number = float(numeric)
            return int(number) if number.is_integer() else number
        except ValueError:
            pass
    return text


def _build_matrix_xlsx_from_apps_script(
    sheet_payloads: dict[str, list[list[object]]],
    sheet_order: list[str] | tuple[str, ...] | None = None,
) -> bytes:
    """Rebuild a minimal XLSX from the exact sheets returned by Apps Script."""
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ordered_sheets = list(sheet_order or sheet_payloads.keys())
    for sheet_name in ordered_sheets:
        if sheet_name not in sheet_payloads:
            continue
        values = sheet_payloads.get(sheet_name, [])
        worksheet = workbook.create_sheet(title=sheet_name)
        for row in values:
            worksheet.append([_coerce_api_cell(value) for value in row])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


@st.cache_data(ttl=GOOGLE_MATRIX_REFRESH_SECONDS, show_spinner=False)
def _fetch_apps_script_matrix_snapshot(
    api_url: str,
    api_key: str,
) -> tuple[bytes, str, str, str]:
    """Read ONLY current plan sheets from Apps Script every 15 minutes.

    The SKU/entity reference is intentionally fetched by a separate function.
    Failure of «Справочник + атрибуты» must never make the plan loader fall back
    to an old bundled workbook.
    """
    checked_at = datetime.now().isoformat(timespec="seconds")
    if not api_url or not api_key:
        return b"", "", checked_at, "не указан URL или ключ Apps Script"

    try:
        import requests

        session = requests.Session()
        meta_response = session.get(
            api_url,
            params={"key": api_key, "action": "meta"},
            timeout=60,
            allow_redirects=True,
        )
        meta_response.raise_for_status()
        meta = meta_response.json()
        if not meta.get("ok"):
            raise RuntimeError(meta.get("error") or "Apps Script вернул ошибку meta")

        available_sheets = set(meta.get("sheets") or [])
        missing = [name for name in MATRIX_PLAN_SHEETS if name not in available_sheets]
        if missing:
            raise RuntimeError("в Apps Script не найдены листы планов: " + ", ".join(missing))

        sheet_payloads: dict[str, list[list[object]]] = {}
        for sheet_name in MATRIX_PLAN_SHEETS:
            sheet_response = session.get(
                api_url,
                params={"key": api_key, "action": "sheet", "name": sheet_name},
                timeout=90,
                allow_redirects=True,
            )
            sheet_response.raise_for_status()
            payload = sheet_response.json()
            if not payload.get("ok"):
                raise RuntimeError(
                    f"{sheet_name}: {payload.get('error') or 'Apps Script вернул ошибку'}"
                )
            values = payload.get("values")
            if not isinstance(values, list):
                raise RuntimeError(f"{sheet_name}: Apps Script не вернул массив values")
            sheet_payloads[sheet_name] = values

        content = _build_matrix_xlsx_from_apps_script(
            sheet_payloads,
            MATRIX_PLAN_SHEETS,
        )
        if content[:2] != b"PK":
            raise RuntimeError("не удалось собрать XLSX планов из ответа Apps Script")

        updated_at = str(meta.get("updatedAt") or "").replace("T", " ").replace("Z", " UTC")
        source = "Apps Script · 2.3 Матрица КОМБО · планы 1–4 недель"
        if updated_at:
            source += f" · обновлена {updated_at}"
        return content, source, checked_at, ""
    except Exception as error:
        return b"", "", checked_at, f"Apps Script (планы): {error}"


@st.cache_data(ttl=GOOGLE_MATRIX_REFRESH_SECONDS, show_spinner=False)
def _fetch_apps_script_entity_reference(
    api_url: str,
    api_key: str,
) -> tuple[pd.DataFrame, str, str, str]:
    """Read «Справочник + атрибуты» independently from the plan snapshot.

    If the Apps Script deployment does not expose this sheet, only the entity
    reference falls back. Plan dates remain live and are not affected.
    """
    checked_at = datetime.now().isoformat(timespec="seconds")
    empty = pd.DataFrame()
    if not api_url or not api_key:
        return empty, "", checked_at, "не указан URL или ключ Apps Script"

    try:
        import requests

        session = requests.Session()
        sheet_response = session.get(
            api_url,
            params={"key": api_key, "action": "sheet", "name": MATRIX_ENTITY_SHEET},
            timeout=90,
            allow_redirects=True,
        )
        sheet_response.raise_for_status()
        payload = sheet_response.json()
        if not payload.get("ok"):
            raise RuntimeError(
                f"{MATRIX_ENTITY_SHEET}: {payload.get('error') or 'Apps Script вернул ошибку'}"
            )
        values = payload.get("values")
        if not isinstance(values, list):
            raise RuntimeError(f"{MATRIX_ENTITY_SHEET}: Apps Script не вернул массив values")

        raw = pd.DataFrame(values)
        frame = _parse_entity_reference_raw(raw)
        if frame.empty:
            raise RuntimeError(f"{MATRIX_ENTITY_SHEET}: справочник пуст")
        return (
            frame,
            f"Apps Script · 2.3 Матрица КОМБО · лист «{MATRIX_ENTITY_SHEET}»",
            checked_at,
            "",
        )
    except Exception as error:
        return empty, "", checked_at, f"Apps Script (справочник): {error}"


def get_current_combo_matrix_snapshot() -> tuple[bytes, str, str, str]:
    """Return current PLAN snapshot; entity-reference failures are isolated from this path."""
    google_bytes, google_source, checked_at, google_error = _fetch_apps_script_matrix_snapshot(
        MATRIX_APPS_SCRIPT_URL,
        MATRIX_APPS_SCRIPT_KEY,
    )
    if google_bytes:
        try:
            GOOGLE_MATRIX_CACHE_FILE.write_bytes(google_bytes)
        except OSError:
            pass
        return google_bytes, google_source, checked_at, ""

    if GOOGLE_MATRIX_CACHE_FILE.exists():
        try:
            cached_bytes = GOOGLE_MATRIX_CACHE_FILE.read_bytes()
            if cached_bytes[:2] == b"PK":
                return (
                    cached_bytes,
                    "Резерв · последняя Apps Script-копия",
                    checked_at,
                    google_error,
                )
        except OSError:
            pass

    if COMBO_MATRIX_FILE.exists():
        try:
            bundled_bytes = COMBO_MATRIX_FILE.read_bytes()
            return (
                bundled_bytes,
                "Резерв · combo_matrix.xlsx из сборки",
                checked_at,
                google_error,
            )
        except OSError as error:
            google_error = f"{google_error}; локальная матрица: {error}".strip("; ")

    return b"", "Матрица недоступна", checked_at, google_error

def _combo_matrix_signature(matrix_bytes: bytes, source: str) -> str:
    digest = hashlib.sha256(matrix_bytes).hexdigest() if matrix_bytes else "empty"
    return f"{digest}|{source}"


@st.fragment(run_every="15m")
def _matrix_auto_refresh_watcher() -> None:
    """Refresh the Apps Script matrix in the background while the app session is open."""
    if "analysis" not in st.session_state:
        return
    matrix_bytes, source, _, _ = get_current_combo_matrix_snapshot()
    signature = _combo_matrix_signature(matrix_bytes, source)
    state_key = "combo_matrix_signature_v761"
    previous = st.session_state.get(state_key)
    if previous is None:
        st.session_state[state_key] = signature
        return
    if previous != signature:
        st.session_state[state_key] = signature
        st.rerun()



def _dpapi_encrypt_text(value: str) -> str:
    """Encrypt a secret for the current Windows user using DPAPI."""
    if os.name != "nt":
        raise RuntimeError("Запоминание пароля доступно в локальной Windows-сборке приложения.")

    import ctypes
    from ctypes import wintypes

    class DATA_BLOB(ctypes.Structure):
        _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]

    raw = value.encode("utf-8")
    raw_buffer = ctypes.create_string_buffer(raw)
    input_blob = DATA_BLOB(len(raw), ctypes.cast(raw_buffer, ctypes.POINTER(ctypes.c_byte)))
    output_blob = DATA_BLOB()

    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    if not crypt32.CryptProtectData(
        ctypes.byref(input_blob),
        ctypes.c_wchar_p("DemandAnalytics PostgreSQL"),
        None,
        None,
        None,
        0,
        ctypes.byref(output_blob),
    ):
        raise ctypes.WinError()
    try:
        encrypted = ctypes.string_at(output_blob.pbData, output_blob.cbData)
    finally:
        kernel32.LocalFree(output_blob.pbData)
    return base64.b64encode(encrypted).decode("ascii")


def _dpapi_decrypt_text(value: str) -> str:
    """Decrypt a DPAPI-protected secret for the current Windows user."""
    if os.name != "nt":
        raise RuntimeError("Запоминание пароля доступно в локальной Windows-сборке приложения.")

    import ctypes
    from ctypes import wintypes

    class DATA_BLOB(ctypes.Structure):
        _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]

    encrypted = base64.b64decode(value.encode("ascii"))
    encrypted_buffer = ctypes.create_string_buffer(encrypted)
    input_blob = DATA_BLOB(
        len(encrypted), ctypes.cast(encrypted_buffer, ctypes.POINTER(ctypes.c_byte))
    )
    output_blob = DATA_BLOB()

    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    if not crypt32.CryptUnprotectData(
        ctypes.byref(input_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(output_blob),
    ):
        raise ctypes.WinError()
    try:
        raw = ctypes.string_at(output_blob.pbData, output_blob.cbData)
    finally:
        kernel32.LocalFree(output_blob.pbData)
    return raw.decode("utf-8")


def load_remembered_pg_credentials() -> dict[str, str]:
    """Load a 30-day local PostgreSQL login saved for the current Windows user."""
    if os.name != "nt" or not REMEMBERED_PG_FILE.exists():
        return {}
    try:
        payload = json.loads(REMEMBERED_PG_FILE.read_text(encoding="utf-8"))
        expires_at = datetime.fromisoformat(str(payload.get("expires_at", "")))
        if datetime.now() >= expires_at:
            REMEMBERED_PG_FILE.unlink(missing_ok=True)
            return {}
        password = _dpapi_decrypt_text(str(payload.get("password_dpapi", "")))
        if not password:
            return {}
        return {
            "PGHOST": str(payload.get("host", "")),
            "PGPORT": str(payload.get("port", "6432")),
            "PGDATABASE": str(payload.get("database", "")),
            "PGUSER": str(payload.get("user", "")),
            "PGPASSWORD": password,
            "expires_at": expires_at.isoformat(timespec="seconds"),
        }
    except Exception:
        # A copied file from another Windows account/computer cannot be decrypted.
        try:
            REMEMBERED_PG_FILE.unlink(missing_ok=True)
        except OSError:
            pass
        return {}


def save_remembered_pg_credentials(
    host: str, port: int, database: str, user: str, password: str
) -> datetime:
    """Remember PostgreSQL credentials for 30 days without storing the password in plain text."""
    if not password:
        raise ValueError("Пароль пустой — сохранять нечего.")
    expires_at = datetime.now() + timedelta(days=REMEMBERED_PG_DAYS)
    payload = {
        "version": 1,
        "host": host.strip(),
        "port": int(port),
        "database": database.strip(),
        "user": user.strip(),
        "password_dpapi": _dpapi_encrypt_text(password),
        "expires_at": expires_at.isoformat(timespec="seconds"),
    }
    temp_file = REMEMBERED_PG_FILE.with_suffix(".tmp")
    temp_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_file.replace(REMEMBERED_PG_FILE)
    return expires_at


def forget_remembered_pg_credentials() -> None:
    try:
        REMEMBERED_PG_FILE.unlink(missing_ok=True)
    except OSError:
        pass


remembered_pg = load_remembered_pg_credentials()
for env_name in ("PGHOST", "PGPORT", "PGDATABASE", "PGUSER", "PGPASSWORD"):
    if not os.getenv(env_name) and remembered_pg.get(env_name):
        os.environ[env_name] = remembered_pg[env_name]


def normalize_sku(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    try:
        return str(int(float(text)))
    except (ValueError, TypeError):
        return None

def numeric_series(series: pd.Series) -> pd.Series:
    """Return a plain float64 Series even when source values are object/nullable."""
    return pd.to_numeric(series, errors="coerce").astype("float64")


def safe_ratio(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    """Numeric division that stays numeric on pandas 2.3+/Python 3.14."""
    num = numeric_series(numerator)
    den = numeric_series(denominator).mask(lambda values: values.eq(0))
    return num.div(den)


STYLER_SAFE_MAX_CELLS = 240_000


def styler_safe_preview(frame: pd.DataFrame, preferred_rows: int | None = None) -> tuple[pd.DataFrame, int, bool]:
    """Limit only the on-screen styled preview so Pandas Styler never exceeds its cell cap.

    Calculations and downloads continue to use the complete DataFrame.
    """
    column_count = max(int(frame.shape[1]), 1)
    max_rows = max(STYLER_SAFE_MAX_CELLS // column_count, 1)
    if preferred_rows is not None:
        max_rows = min(max_rows, max(int(preferred_rows), 1))
    truncated = len(frame) > max_rows
    return frame.head(max_rows).copy(), max_rows, truncated


def _parse_entity_reference_raw(raw: pd.DataFrame) -> pd.DataFrame:
    """Normalize SKU/category/entity mapping from the matrix reference sheet."""
    if raw is None or raw.empty:
        raise ValueError("Лист справочника пуст.")

    header_row = None
    for index, row in raw.head(30).iterrows():
        lowered = {str(value).strip().casefold() for value in row if pd.notna(value)}
        if "код" in lowered and "название блюда" in lowered:
            header_row = int(index)
            break
    if header_row is None:
        raise ValueError("В справочнике не найдена строка с колонками «код» и «Название блюда».")

    header_values = [
        str(value).strip() if pd.notna(value) else ""
        for value in raw.iloc[header_row].tolist()
    ]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header_values
    lookup: dict[str, str] = {}
    for column in data.columns:
        key = str(column).strip().casefold()
        if key and key not in lookup:
            lookup[key] = column

    canonical = {
        "код": "код",
        "название блюда": "Название блюда",
        "категория": "Категория",
        "атрибут 1": "Атрибут 1",
        "атрибут 2": "Атрибут 2",
        "атрибут 3": "Атрибут 3",
    }
    missing = [name for key, name in canonical.items() if key not in lookup]
    if missing:
        raise ValueError(f"В справочнике отсутствуют колонки: {', '.join(missing)}")

    result = pd.DataFrame(
        {name: data[lookup[key]] for key, name in canonical.items()}
    )
    result["sku"] = result["код"].map(normalize_sku)
    result = result[result["sku"].notna()].drop_duplicates("sku", keep="last")
    result = result.rename(
        columns={
            "Название блюда": "entity_product_name",
            "Категория": "category",
            "Атрибут 1": "attribute_1",
            "Атрибут 2": "attribute_2",
            "Атрибут 3": "attribute_3",
        }
    )
    for column in ["entity_product_name", "category", "attribute_1", "attribute_2", "attribute_3"]:
        result[column] = result[column].fillna("").astype(str).str.strip()
    result["category"] = result["category"].replace("", "Не сопоставлено")
    attributes = result[["attribute_1", "attribute_2", "attribute_3"]]
    result["entity"] = attributes.apply(
        lambda row: " • ".join(value for value in row if value) or "Не задана",
        axis=1,
    )
    return result[
        ["sku", "entity_product_name", "category", "attribute_1", "attribute_2", "attribute_3", "entity"]
    ].reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_entities(path: str, modified_at: float) -> pd.DataFrame:
    """Legacy fallback: read the old standalone entity workbook."""
    raw = pd.read_excel(path, header=None)
    return _parse_entity_reference_raw(raw)


@st.cache_data(show_spinner=False)
def load_entities_from_matrix_bytes(matrix_bytes: bytes) -> pd.DataFrame:
    """Read SKU/category/entity mapping from «Справочник + атрибуты» in 2.3 Matrix COMBO."""
    if not matrix_bytes or matrix_bytes[:2] != b"PK":
        raise ValueError("Матрица 2.3 не содержит корректный XLSX.")
    try:
        raw = pd.read_excel(
            io.BytesIO(matrix_bytes),
            sheet_name=MATRIX_ENTITY_SHEET,
            header=None,
        )
    except ValueError as error:
        raise ValueError(
            f"В матрице 2.3 не найден лист «{MATRIX_ENTITY_SHEET}»."
        ) from error
    return _parse_entity_reference_raw(raw)


def _entity_reference_signature(frame: pd.DataFrame) -> str:
    columns = [
        "sku", "entity_product_name", "category",
        "attribute_1", "attribute_2", "attribute_3", "entity",
    ]
    normalized = frame.reindex(columns=columns).fillna("").astype(str)
    normalized = normalized.sort_values("sku", kind="stable").reset_index(drop=True)
    return hashlib.sha256(normalized.to_csv(index=False).encode("utf-8")).hexdigest()


def get_current_entity_reference() -> tuple[pd.DataFrame, str, str, str, str]:
    """Load SKU/category/entity mapping independently from the live plan snapshot."""
    live_frame, live_source, checked_at, live_error = _fetch_apps_script_entity_reference(
        MATRIX_APPS_SCRIPT_URL,
        MATRIX_APPS_SCRIPT_KEY,
    )
    errors: list[str] = []
    if live_error:
        errors.append(live_error)

    if live_frame is not None and not live_frame.empty:
        return (
            live_frame,
            live_source,
            checked_at,
            "",
            _entity_reference_signature(live_frame),
        )

    # Reference fallback is independent: it may be old while PLAN sheets stay live.
    # This branch must never replace the current plan snapshot.
    if COMBO_MATRIX_FILE.exists():
        try:
            local_bytes = COMBO_MATRIX_FILE.read_bytes()
            frame = load_entities_from_matrix_bytes(local_bytes)
            return (
                frame,
                f"Резерв справочника · {COMBO_MATRIX_FILE.name} · лист «{MATRIX_ENTITY_SHEET}»",
                checked_at,
                "; ".join(errors),
                _entity_reference_signature(frame),
            )
        except Exception as error:
            errors.append(f"локальная матрица справочника: {error}")

    if ENTITY_FILE.exists():
        try:
            frame = load_entities(str(ENTITY_FILE), ENTITY_FILE.stat().st_mtime)
            return (
                frame,
                f"Аварийный резерв справочника · {ENTITY_FILE.name}",
                checked_at,
                "; ".join(errors),
                _entity_reference_signature(frame),
            )
        except Exception as error:
            errors.append(f"старый справочник: {error}")

    raise RuntimeError("; ".join(errors) or "Справочник SKU/сущностей недоступен.")


def connection_settings() -> dict[str, object]:
    required = ["PGHOST", "PGDATABASE", "PGUSER", "PGPASSWORD"]
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        raise RuntimeError(f"Не заполнены настройки: {', '.join(missing)}")
    return {
        "host": os.environ["PGHOST"],
        "port": int(os.getenv("PGPORT", "6432")),
        "dbname": os.environ["PGDATABASE"],
        "user": os.environ["PGUSER"],
        "password": os.environ["PGPASSWORD"],
        "sslmode": os.getenv("PGSSLMODE", "require"),
        "connect_timeout": 15,
    }


@st.cache_data(ttl=900, show_spinner="Загружаю продажи из PostgreSQL…")
def load_sales(date_from: date, date_to_exclusive: date, points: tuple[int, ...]) -> pd.DataFrame:
    query = """
        SELECT
            business_date,
            sale_datetime,
            shop_number,
            COALESCE(
                NULLIF(TRIM(erp_code), ''),
                NULLIF(TRIM(product_code), ''),
                NULLIF(TRIM(barcode), ''),
                NULLIF(TRIM(product_hash), ''),
                'БЕЗ_SKU'
            ) AS sku,
            MAX(product_name) AS product_name,
            SUM(net_quantity)::numeric AS sold_quantity,
            SUM(net_line_amount)::numeric AS revenue
        FROM dwh.v_sales_item
        WHERE business_date >= %(date_from)s
          AND business_date < %(date_to)s
          AND shop_number = ANY(%(points)s)
        GROUP BY business_date, sale_datetime, shop_number,
                 COALESCE(
                     NULLIF(TRIM(erp_code), ''),
                     NULLIF(TRIM(product_code), ''),
                     NULLIF(TRIM(barcode), ''),
                     NULLIF(TRIM(product_hash), ''),
                     'БЕЗ_SKU'
                 )
        ORDER BY business_date, sale_datetime, shop_number,
                 COALESCE(
                     NULLIF(TRIM(erp_code), ''),
                     NULLIF(TRIM(product_code), ''),
                     NULLIF(TRIM(barcode), ''),
                     NULLIF(TRIM(product_hash), ''),
                     'БЕЗ_SKU'
                 )
    """
    with psycopg.connect(**connection_settings()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                query,
                {"date_from": date_from, "date_to": date_to_exclusive, "points": list(points)},
            )
            records = cursor.fetchall()
            columns = [description.name for description in cursor.description]
    frame = pd.DataFrame(records, columns=columns)
    if frame.empty:
        return frame
    frame["sku"] = frame["sku"].map(normalize_sku)
    frame["sold_quantity"] = pd.to_numeric(frame["sold_quantity"], errors="coerce").fillna(0.0)
    frame["revenue"] = pd.to_numeric(frame["revenue"], errors="coerce").fillna(0.0)
    return frame


@st.cache_data(ttl=300, show_spinner="Ищу магазины с продажами…")
def load_available_shops(date_from: date, date_to_exclusive: date) -> pd.DataFrame:
    query = """
        SELECT
            shop_number,
            COUNT(DISTINCT source_purchase_id) AS receipts,
            SUM(net_quantity)::numeric AS sold_quantity,
            MIN(business_date) AS first_sale_date,
            MAX(business_date) AS last_sale_date
        FROM dwh.v_sales_item
        WHERE business_date >= %(date_from)s
          AND business_date < %(date_to)s
        GROUP BY shop_number
        ORDER BY shop_number
    """
    with psycopg.connect(**connection_settings()) as connection:
        return pd.read_sql_query(
            query,
            connection,
            params={"date_from": date_from, "date_to": date_to_exclusive},
        )


REQUIRED_POINT_SHOPS = {25: "Т25"}


def ensure_required_shops(frame: pd.DataFrame) -> pd.DataFrame:
    """Добавляет обязательные точки, даже если в периоде по ним нет продаж."""
    result = frame.copy()
    existing = set(
        pd.to_numeric(result.get("shop_number", pd.Series(dtype=float)), errors="coerce")
        .dropna()
        .astype(int)
        .tolist()
    )
    missing_rows = []
    for shop_number in REQUIRED_POINT_SHOPS:
        if shop_number not in existing:
            missing_rows.append(
                {
                    "shop_number": shop_number,
                    "receipts": 0,
                    "sold_quantity": 0.0,
                    "first_sale_date": pd.NaT,
                    "last_sale_date": pd.NaT,
                }
            )
    if missing_rows:
        result = pd.concat([result, pd.DataFrame(missing_rows)], ignore_index=True)
    if "shop_number" in result.columns:
        result = result.sort_values("shop_number", kind="stable").reset_index(drop=True)
    return result


@st.cache_data(ttl=900, show_spinner="Загружаю историю для прогноза…")
def load_forecast_history(date_from: date, date_to_exclusive: date, points: tuple[int, ...]) -> pd.DataFrame:
    query = """
        SELECT
            business_date,
            sale_datetime,
            shop_number,
            COALESCE(
                NULLIF(TRIM(erp_code), ''),
                NULLIF(TRIM(product_code), ''),
                NULLIF(TRIM(barcode), ''),
                NULLIF(TRIM(product_hash), ''),
                'БЕЗ_SKU'
            ) AS sku,
            MAX(product_name) AS product_name,
            SUM(net_quantity)::numeric AS sold_quantity,
            SUM(net_line_amount)::numeric AS revenue
        FROM dwh.v_sales_item
        WHERE business_date >= %(date_from)s
          AND business_date < %(date_to)s
          AND shop_number = ANY(%(points)s)
        GROUP BY business_date, sale_datetime, shop_number,
                 COALESCE(
                     NULLIF(TRIM(erp_code), ''),
                     NULLIF(TRIM(product_code), ''),
                     NULLIF(TRIM(barcode), ''),
                     NULLIF(TRIM(product_hash), ''),
                     'БЕЗ_SKU'
                 )
    """
    with psycopg.connect(**connection_settings()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                query,
                {"date_from": date_from, "date_to": date_to_exclusive, "points": list(points)},
            )
            records = cursor.fetchall()
            columns = [description.name for description in cursor.description]
    result = pd.DataFrame(records, columns=columns)
    if result.empty:
        return result
    result["sku"] = result["sku"].map(normalize_sku)
    result["business_date"] = pd.to_datetime(result["business_date"]).dt.date
    result["sale_datetime"] = pd.to_datetime(result["sale_datetime"], errors="coerce")
    result["sold_quantity"] = pd.to_numeric(result["sold_quantity"], errors="coerce").fillna(0.0)
    result["revenue"] = pd.to_numeric(result["revenue"], errors="coerce").fillna(0.0)
    return result


def normalize_matrix_category(value: object) -> str:
    text = str(value or "").strip().lower()
    mappings = {
        "завтрак": "Завтраки",
        "завтраки": "Завтраки",
        "салат": "Салаты",
        "салаты": "Салаты",
        "второе": "Вторые блюда",
        "вторые блюда": "Вторые блюда",
        "суп": "Супы",
        "супы": "Супы",
        "сэндвич": "Сэндвичи",
        "сэндвичи": "Сэндвичи",
        "десерт": "Десерты",
        "десерты": "Десерты",
        "напиток": "Напитки",
        "напитки": "Напитки",
        "япония": "Япония",
        "хлеб": "Хлеб",
        "премиум": "Салаты",
    }
    return mappings.get(text, str(value or "").strip())


def product_lifecycle_days(category: object) -> int:
    """Возвращает срок продажи партии, отсчитываемый со дня после отгрузки."""
    normalized = normalize_matrix_category(category)
    if normalized == "Вторые блюда":
        return 5
    if normalized == "Япония":
        return 2
    if normalized == "Напитки":
        return 7
    return 3


def product_green_days(category: object) -> int:
    """Возвращает число основных (зелёных) дней внутри срока продажи."""
    normalized = normalize_matrix_category(category)
    if normalized == "Вторые блюда":
        return 3
    if normalized == "Япония":
        return 1
    if normalized == "Напитки":
        return 4
    return 2


def forecast_coverage_days(category: object) -> int:
    """Возвращает множитель среднего SKU, используемый только в прогнозе плана."""
    normalized = normalize_matrix_category(category)
    if normalized == "Япония":
        return 1
    if normalized == "Вторые блюда":
        return 3
    if normalized == "Напитки":
        return 4
    return 2


def parse_excel_date(value: object) -> date | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.date()
    if hasattr(value, "date"):
        try:
            return value.date()
        except (TypeError, ValueError):
            pass
    if isinstance(value, (int, float)):
        try:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")).date()
        except (TypeError, ValueError, OverflowError):
            return None
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    return None if pd.isna(parsed) else parsed.date()


@st.cache_data(show_spinner="Читаю матрицу меню…")
def parse_menu_matrix(file_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Читает меню из XLSX и устойчиво пропускает пустые/служебные листы.

    Поддерживаются рабочие файлы, где строка с датой находится не обязательно
    в первом столбце, а строка заголовков может идти через несколько строк после неё.
    """
    source = io.BytesIO(file_bytes)
    excel = pd.ExcelFile(source)
    blocks: list[pd.DataFrame] = []

    def row_values(frame: pd.DataFrame, row_index: int) -> list[object]:
        if row_index < 0 or row_index >= len(frame):
            return []
        return frame.iloc[row_index].tolist()

    def normalized_text(value: object) -> str:
        return str(value or "").strip().replace("T", "Т").replace("t", "т")

    def header_info(values: list[object]) -> tuple[int | None, int | None, int | None, int | None, bool]:
        header = [normalized_text(value) if pd.notna(value) else "" for value in values]
        code_col = next(
            (
                i for i, value in enumerate(header)
                if value.casefold().replace(" ", "") in {"код", "код№", "sku", "ску"}
            ),
            None,
        )
        name_col = next(
            (
                i for i, value in enumerate(header)
                if "название блюда" in value.casefold()
                or "название товара" in value.casefold()
                or "наименование" == value.casefold()
            ),
            None,
        )
        category_col = next(
            (i for i, value in enumerate(header) if "категор" in value.casefold()),
            None,
        )
        price_col = next(
            (i for i, value in enumerate(header) if value.casefold() == "цена"),
            None,
        )
        has_point_columns = any(
            re.fullmatch(r"Т\s*0*1", value.upper().replace(" ", ""))
            for value in header if value
        )
        return code_col, name_col, category_col, price_col, has_point_columns

    # В рабочих файлах меню блоки могут находиться как на листах «План ...»,
    # так и на обычном «Лист1». На одном листе может быть несколько дат.
    for sheet_name in excel.sheet_names:
        try:
            raw = pd.read_excel(excel, sheet_name=sheet_name, header=None)
        except Exception:
            # Один повреждённый/служебный лист не должен ломать весь файл.
            continue

        # Ключевой фикс: пустой Excel-лист может иметь 0 столбцов.
        if raw is None or raw.empty or raw.shape[1] == 0:
            continue

        # Ищем строку даты по ВСЕЙ строке, а не только по первому столбцу.
        date_rows: list[int] = []
        for row_index in range(len(raw)):
            values = row_values(raw, row_index)
            texts = [normalized_text(value).casefold() for value in values if pd.notna(value)]
            is_service_row = any("участок комплектации" in text for text in texts)
            if is_service_row:
                continue
            has_date_label = any(text.startswith("дата") for text in texts)
            has_plan_label = any("план на день кухня" in text for text in texts)
            parsed_date = next(
                (parsed for value in values if (parsed := parse_excel_date(value)) is not None),
                None,
            )
            if parsed_date is not None and (has_date_label or has_plan_label):
                date_rows.append(row_index)

        for position, date_row in enumerate(date_rows):
            date_values = row_values(raw, date_row)
            target_date = next(
                (parsed for value in date_values if (parsed := parse_excel_date(value)) is not None),
                None,
            )
            if target_date is None:
                continue

            # Заголовок обычно следующей строкой, но допускаем до 5 служебных строк.
            header_row: int | None = None
            header_meta: tuple[int | None, int | None, int | None, int | None, bool] | None = None
            search_end = min(len(raw), date_row + 7)
            for candidate in range(date_row + 1, search_end):
                meta = header_info(row_values(raw, candidate))
                code_col, name_col, category_col, _, has_point_columns = meta
                if code_col is not None and name_col is not None and category_col is not None and has_point_columns:
                    header_row = candidate
                    header_meta = meta
                    break
            if header_row is None or header_meta is None:
                continue

            code_col, name_col, category_col, price_col, _ = header_meta
            assert code_col is not None and name_col is not None and category_col is not None

            # Блок заканчивается перед следующей датой или перед служебным
            # разделом «Участок комплектации», если он встречается раньше.
            next_date_row = date_rows[position + 1] if position + 1 < len(date_rows) else len(raw)
            end_row = next_date_row
            for candidate in range(header_row + 1, next_date_row):
                label = " ".join(
                    normalized_text(value).casefold()
                    for value in row_values(raw, candidate)
                    if pd.notna(value)
                )
                if "участок комплектации" in label:
                    end_row = candidate
                    break

            data_rows = raw.iloc[header_row + 1 : end_row].copy()
            if data_rows.empty:
                continue

            block = pd.DataFrame(index=data_rows.index)
            block["sku"] = data_rows.iloc[:, code_col]
            block["price"] = (
                data_rows.iloc[:, price_col]
                if price_col is not None and price_col < data_rows.shape[1]
                else pd.NA
            )
            block["matrix_category"] = data_rows.iloc[:, category_col]
            block["product_name"] = data_rows.iloc[:, name_col]
            block["excel_row"] = block.index + 1
            block["sku"] = block["sku"].map(normalize_sku)
            block = block[block["sku"].notna() & block["product_name"].notna()].copy()
            if block.empty:
                continue

            block["price"] = pd.to_numeric(block["price"], errors="coerce")
            block["matrix_category"] = block["matrix_category"].map(normalize_matrix_category)
            block["target_date"] = target_date
            block["sheet"] = sheet_name
            block["header_excel_row"] = header_row + 1

            day_label = WEEKDAY_RU.get(target_date.weekday(), "")
            block["day_label"] = day_label
            blocks.append(block)

    menu = pd.concat(blocks, ignore_index=True) if blocks else pd.DataFrame()
    if not menu.empty:
        menu = menu.drop_duplicates(
            subset=["target_date", "sheet", "sku", "excel_row"], keep="first"
        ).reset_index(drop=True)
    capacity = pd.DataFrame()
    if "Ёмкость точек" in excel.sheet_names:
        try:
            capacity_raw = pd.read_excel(excel, sheet_name="Ёмкость точек", header=2)
        except Exception:
            capacity_raw = pd.DataFrame()
        if not capacity_raw.empty and len(capacity_raw.columns) >= 3:
            capacity_raw.columns = [str(column).strip() for column in capacity_raw.columns]
            capacity = capacity_raw.copy()
            capacity = capacity.rename(
                columns={capacity.columns[0]: "point_number", capacity.columns[1]: "point_name"}
            )
            capacity["point_number"] = pd.to_numeric(capacity["point_number"], errors="coerce")
            capacity = capacity[capacity["point_number"].notna()].copy()
            capacity["point_number"] = capacity["point_number"].astype(int)
    return menu, capacity


@st.cache_data(show_spinner="Читаю планы аналитика за 1–4 недели…")
def parse_analyst_plan_history(file_bytes: bytes) -> pd.DataFrame:
    """Читает только основные блоки «План на день кухня» с листов 1–4 недели."""
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    rows: list[dict[str, object]] = []
    plan_sheets = [
        name for name in workbook.sheetnames
        if name.lower().startswith("план ") and "недел" in name.lower()
    ]

    for sheet_name in plan_sheets:
        sheet = workbook[sheet_name]
        main_blocks: list[tuple[int, date]] = []
        for excel_row in range(1, sheet.max_row + 1):
            row_values = [sheet.cell(excel_row, column).value for column in range(1, min(sheet.max_column, 12) + 1)]
            row_label = " ".join(str(value or "").strip().lower() for value in row_values)
            if "план на день кухня" not in row_label:
                continue
            target_date = next(
                (parsed for value in row_values if (parsed := parse_excel_date(value)) is not None),
                None,
            )
            if target_date is not None:
                main_blocks.append((excel_row, target_date))

        for block_index, (date_row, target_date) in enumerate(main_blocks):
            header_row = date_row + 1
            headers = {
                str(sheet.cell(header_row, column).value or "").strip(): column
                for column in range(1, sheet.max_column + 1)
            }
            code_column = headers.get("Код", headers.get("Код №"))
            name_column = headers.get("Название блюда")
            category_column = headers.get("Категория", headers.get("категория"))
            price_column = headers.get("Цена")
            if not code_column or not name_column or not category_column:
                continue

            # Основной блок иногда заканчивается строкой «Участок комплектации»,
            # где слово «Дата» отсутствует в колонке A. Поэтому ищем границу
            # по содержимому всей строки, а не только по первой ячейке.
            next_main_row = (
                main_blocks[block_index + 1][0]
                if block_index + 1 < len(main_blocks)
                else sheet.max_row + 1
            )
            end_row = next_main_row - 1
            for candidate_row in range(header_row + 1, next_main_row):
                candidate_label = " ".join(
                    str(sheet.cell(candidate_row, column).value or "").strip().lower()
                    for column in range(1, min(sheet.max_column, 12) + 1)
                )
                if "участок комплектации" in candidate_label:
                    end_row = candidate_row - 1
                    break

            normalized_headers = {
                str(sheet.cell(header_row, column).value or "").replace(" ", "").strip(): column
                for column in range(1, sheet.max_column + 1)
            }
            for excel_row in range(header_row + 1, end_row + 1):
                sku = normalize_sku(sheet.cell(excel_row, code_column).value)
                product_name = sheet.cell(excel_row, name_column).value
                if sku is None or not product_name:
                    continue
                category = normalize_matrix_category(sheet.cell(excel_row, category_column).value)
                unit_price = pd.to_numeric(
                    sheet.cell(excel_row, price_column).value if price_column else None,
                    errors="coerce",
                )
                for point_number in range(1, 30):
                    if point_number == 11:
                        continue
                    point_column = normalized_headers.get(f"Т{point_number}")
                    if point_column is None:
                        continue
                    plan = pd.to_numeric(sheet.cell(excel_row, point_column).value, errors="coerce")
                    if pd.isna(plan):
                        # В матрице пустая ячейка точки означает нулевой план.
                        plan = 0.0
                    rows.append(
                        {
                            "plan_date": target_date,
                            "plan_sheet": sheet_name,
                            "point_number": point_number,
                            "sku": sku,
                            "product_name": str(product_name).strip(),
                            "matrix_category": category,
                            "unit_price": 0.0 if pd.isna(unit_price) else max(0.0, float(unit_price)),
                            "analyst_plan": max(0.0, float(plan)),
                        }
                    )
    if not rows:
        return pd.DataFrame()

    result = pd.DataFrame(rows)
    # Если один и тот же SKU на дату случайно продублирован в матрице,
    # используем последнее значение в основном блоке и не удваиваем отгрузку.
    return result.drop_duplicates(
        ["plan_date", "point_number", "sku"], keep="last"
    ).sort_values(["plan_date", "point_number", "matrix_category", "product_name"]).reset_index(drop=True)


@st.cache_data(show_spinner="Читаю матрицу 2.3 для окна свежести…")
def parse_freshness_plan(file_bytes: bytes) -> pd.DataFrame:
    """Читает планы из матрицы 2.3; fallback оставлен для совместимости формата."""
    try:
        historical = parse_analyst_plan_history(file_bytes)
    except Exception:
        historical = pd.DataFrame()
    if not historical.empty:
        return historical

    menu, _ = parse_menu_matrix(file_bytes)
    if menu.empty:
        return pd.DataFrame()
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    rows: list[dict[str, object]] = []
    for _, item in menu.iterrows():
        sheet = workbook[str(item["sheet"])]
        header_row = int(item["header_excel_row"])
        headers = {
            str(sheet.cell(header_row, column).value or "").strip(): column
            for column in range(1, sheet.max_column + 1)
        }
        for point_number in range(1, 30):
            if point_number == 11:
                continue
            point_column = headers.get(f"Т{point_number}")
            if point_column is None:
                continue
            plan_value = pd.to_numeric(
                sheet.cell(int(item["excel_row"]), point_column).value,
                errors="coerce",
            )
            rows.append(
                {
                    "plan_date": item["target_date"],
                    "point_number": point_number,
                    "sku": item["sku"],
                    "product_name": item["product_name"],
                    "matrix_category": item["matrix_category"],
                    "unit_price": (
                        0.0 if pd.isna(item["price"]) else max(0.0, float(item["price"]))
                    ),
                    "analyst_plan": 0.0 if pd.isna(plan_value) else max(0.0, float(plan_value)),
                }
            )
    return pd.DataFrame(rows).drop_duplicates(
        ["plan_date", "point_number", "sku"], keep="last"
    )



WEEKDAY_RU = {
    0: "Понедельник",
    1: "Вторник",
    2: "Среда",
    3: "Четверг",
    4: "Пятница",
    5: "Суббота",
    6: "Воскресенье",
}


def normalize_point_label(value: object) -> str | None:
    """Нормализует Т7 / T7 / 7 в формат Т7; Т11 исключается бизнес-правилом."""
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().upper().replace("T", "Т")
    if text.startswith("Т"):
        text = text[1:].strip()
    try:
        number = int(float(text))
    except (TypeError, ValueError):
        return None
    if number < 1 or number > 29 or number == 11:
        return None
    return f"Т{number}"


def _normalized_column_name(value: object) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е")
    for char in ["\n", "\r", "\t", ".", ",", ":", ";", "-", "_", "(", ")"]:
        text = text.replace(char, " ")
    return " ".join(text.split())


def _pick_stock_column(columns: list[object], aliases: tuple[str, ...]) -> object | None:
    normalized = {_normalized_column_name(column): column for column in columns}
    alias_norms = {_normalized_column_name(alias) for alias in aliases}
    for alias in alias_norms:
        if alias in normalized:
            return normalized[alias]
    # Fallback: достаточно, чтобы все слова алиаса присутствовали в названии колонки.
    for alias in alias_norms:
        words = set(alias.split())
        for normalized_name, original in normalized.items():
            if words and words.issubset(set(normalized_name.split())):
                return original
    return None


@st.cache_data(show_spinner=False)
def parse_stock_balances(file_bytes: bytes) -> pd.DataFrame:
    """Читает необязательный Excel со снимками остатков по дата + SKU + точка/магазин."""
    if not file_bytes:
        return pd.DataFrame()
    workbook = pd.ExcelFile(io.BytesIO(file_bytes))
    frames: list[pd.DataFrame] = []
    errors: list[str] = []
    for sheet_name in workbook.sheet_names:
        try:
            preview = pd.read_excel(workbook, sheet_name=sheet_name, header=None, nrows=20)
        except Exception as error:
            errors.append(f"{sheet_name}: {error}")
            continue
        header_row = None
        for row_index, row in preview.iterrows():
            labels = [_normalized_column_name(value) for value in row.tolist() if pd.notna(value)]
            has_date = any(label in {"дата", "дата остатка", "дата снимка", "snapshot date"} for label in labels)
            has_sku = any(label in {"sku", "ску", "код", "код sku", "код ску"} for label in labels)
            has_stock = any("остат" in label or label in {"stock", "balance", "quantity"} for label in labels)
            has_location = any(
                label in {"точка", "тт", "point", "магазин", "номер магазина", "shop number", "shop_number"}
                for label in labels
            )
            if has_date and has_sku and has_stock and has_location:
                header_row = int(row_index)
                break
        if header_row is None:
            continue
        frame = pd.read_excel(workbook, sheet_name=sheet_name, header=header_row)
        frame.columns = [str(column).strip() for column in frame.columns]
        date_col = _pick_stock_column(frame.columns.tolist(), ("Дата", "Дата остатка", "Дата снимка", "snapshot_date"))
        sku_col = _pick_stock_column(frame.columns.tolist(), ("SKU", "СКУ", "Код", "Код SKU"))
        stock_col = _pick_stock_column(
            frame.columns.tolist(),
            ("Остаток, шт.", "Остаток шт", "Остаток", "Количество остатка", "stock", "balance"),
        )
        point_col = _pick_stock_column(frame.columns.tolist(), ("Точка", "ТТ", "point"))
        shop_col = _pick_stock_column(
            frame.columns.tolist(), ("Магазин", "Номер магазина", "shop_number", "shop number")
        )
        product_col = _pick_stock_column(
            frame.columns.tolist(), ("Название товара", "Название блюда", "Наименование", "product_name")
        )
        if date_col is None or sku_col is None or stock_col is None or (point_col is None and shop_col is None):
            continue
        result = pd.DataFrame()
        result["snapshot_date"] = pd.to_datetime(frame[date_col], errors="coerce").dt.date
        result["sku"] = frame[sku_col].map(normalize_sku)
        result["actual_stock"] = pd.to_numeric(frame[stock_col], errors="coerce")
        result["point"] = frame[point_col].map(normalize_point_label) if point_col is not None else None
        result["shop_number"] = (
            pd.to_numeric(frame[shop_col], errors="coerce") if shop_col is not None else pd.NA
        )
        result["product_name"] = (
            frame[product_col].map(lambda value: str(value).strip() if pd.notna(value) else "")
            if product_col is not None else ""
        )
        result["source_sheet"] = str(sheet_name)
        result = result[
            result["snapshot_date"].notna()
            & result["sku"].notna()
            & result["actual_stock"].notna()
            & (result["actual_stock"] >= 0)
        ].copy()
        if not result.empty:
            frames.append(result)
    if not frames:
        extra = f" Ошибки чтения: {'; '.join(errors[:3])}" if errors else ""
        raise ValueError(
            "Не найдены подходящие данные. Нужны колонки: Дата, SKU, Остаток и Точка или Магазин."
            + extra
        )
    result = pd.concat(frames, ignore_index=True)
    result["shop_number"] = pd.to_numeric(result["shop_number"], errors="coerce")
    return result


def resolve_stock_balance_points(
    stock_balances: pd.DataFrame,
    point_mapping: dict[int, str],
) -> pd.DataFrame:
    """Заполняет точку по номеру магазина и наоборот, не трогая Т11."""
    if stock_balances is None or stock_balances.empty:
        return pd.DataFrame()
    result = stock_balances.copy()
    shop_to_point = {
        int(shop): normalize_point_label(label)
        for shop, label in point_mapping.items()
        if normalize_point_label(label) is not None
    }
    point_to_shop = {point: shop for shop, point in shop_to_point.items()}
    if "point" not in result.columns:
        result["point"] = None
    if "shop_number" not in result.columns:
        result["shop_number"] = pd.NA
    result["point"] = result["point"].map(normalize_point_label)
    shop_numeric = pd.to_numeric(result["shop_number"], errors="coerce")
    missing_point = result["point"].isna() & shop_numeric.notna()
    result.loc[missing_point, "point"] = shop_numeric[missing_point].astype(int).map(shop_to_point)
    missing_shop = shop_numeric.isna() & result["point"].notna()
    result.loc[missing_shop, "shop_number"] = result.loc[missing_shop, "point"].map(point_to_shop)
    result["shop_number"] = pd.to_numeric(result["shop_number"], errors="coerce")
    result = result[result["point"].notna() & result["point"].ne("Т11")].copy()
    result["weekday"] = result["snapshot_date"].map(
        lambda value: WEEKDAY_RU.get(value.weekday(), "") if pd.notna(value) else ""
    )
    # Один SKU в одной точке на одну дату должен иметь один итоговый остаток. Если файл
    # содержит несколько складских строк, суммируем их.
    result = (
        result.groupby(["snapshot_date", "weekday", "point", "sku"], as_index=False, dropna=False)
        .agg(
            actual_stock=("actual_stock", "sum"),
            shop_number=("shop_number", "first"),
            product_name=("product_name", lambda values: next((str(v).strip() for v in values if str(v).strip()), "")),
        )
    )
    return result.sort_values(["snapshot_date", "point", "sku"], kind="stable")


@st.cache_data(show_spinner=False)
def stock_balance_template_bytes() -> bytes:
    """Небольшой шаблон для ручной выгрузки фактических остатков."""
    sample = pd.DataFrame(columns=["Дата", "Точка", "SKU", "Название товара", "Остаток, шт."])
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sample.to_excel(writer, index=False, sheet_name="Остатки")
        pd.DataFrame(
            {
                "Как заполнять": [
                    "Одна строка = один SKU в одной точке на конкретную дату снимка.",
                    "Дата — обязательна. День недели приложение определит само.",
                    "Точка — Т1…Т29 без Т11. Вместо точки можно использовать колонку «Магазин».",
                    "SKU — код товара. Остаток, шт. — фактическое количество на момент снимка.",
                ]
            }
        ).to_excel(writer, index=False, sheet_name="Инструкция")
        sheet = writer.book["Остатки"]
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        widths = {"A": 14, "B": 10, "C": 14, "D": 38, "E": 16}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for cell in sheet["A"][1:]:
            cell.number_format = "DD.MM.YYYY"
    return buffer.getvalue()


def build_calculated_stock_snapshot(
    plan_rows: pd.DataFrame,
    sales: pd.DataFrame,
    entities: pd.DataFrame,
    point_to_shop: dict[str, int],
    point_labels: list[str] | tuple[str, ...],
    snapshot_date: date,
) -> pd.DataFrame:
    """Расчётный живой остаток по точке/SKU на конкретную дату снимка."""
    if plan_rows is None or plan_rows.empty or not point_labels:
        return pd.DataFrame()
    results: list[pd.DataFrame] = []
    max_lifecycle_days = 7
    display_start = snapshot_date - timedelta(days=max_lifecycle_days)
    for point_label in point_labels:
        normalized_point = normalize_point_label(point_label)
        if normalized_point is None or normalized_point not in point_to_shop:
            continue
        point_number = int(normalized_point[1:])
        shop_number = int(point_to_shop[normalized_point])
        point_plans = plan_rows[
            (pd.to_numeric(plan_rows["point_number"], errors="coerce") == point_number)
            & pd.to_datetime(plan_rows["plan_date"], errors="coerce").dt.date.between(
                display_start, snapshot_date, inclusive="both"
            )
        ].copy()
        point_sales = sales[
            pd.to_numeric(sales.get("shop_number"), errors="coerce") == shop_number
        ].copy() if sales is not None and not sales.empty else pd.DataFrame()
        detail = build_sales_time_period(
            point_plans,
            point_sales,
            entities,
            display_start,
            snapshot_date,
            as_of_date=snapshot_date,
        )
        if detail.empty:
            continue
        detail["Расчётный живой остаток, шт."] = (
            numeric_series(detail["Расчётный остаток"]).fillna(0)
            - numeric_series(detail["Списания"]).fillna(0)
        ).clip(lower=0)
        summary = (
            detail.groupby(["SKU", "Название товара", "Категория"], as_index=False, dropna=False)
            .agg(**{"Расчётный живой остаток, шт.": ("Расчётный живой остаток, шт.", "sum")})
        )
        summary.insert(0, "Точка", normalized_point)
        summary.insert(1, "Магазин", shop_number)
        results.append(summary)
    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


def merge_actual_and_calculated_stock(
    actual_snapshot: pd.DataFrame,
    calculated_snapshot: pd.DataFrame,
    selected_points: list[str] | tuple[str, ...],
) -> pd.DataFrame:
    """Сводит факт и расчёт; факт относится ко всей точке/SKU, а не к отдельной партии."""
    actual = actual_snapshot.copy() if actual_snapshot is not None else pd.DataFrame()
    calculated = calculated_snapshot.copy() if calculated_snapshot is not None else pd.DataFrame()
    if actual.empty and calculated.empty:
        return pd.DataFrame()
    if not actual.empty:
        actual = actual.rename(
            columns={
                "point": "Точка",
                "shop_number": "Магазин",
                "sku": "SKU",
                "actual_stock": "Факт. остаток, шт.",
                "product_name": "Название из файла",
            }
        )
        actual["Точка"] = actual["Точка"].map(normalize_point_label)
        actual = actual[["Точка", "Магазин", "SKU", "Название из файла", "Факт. остаток, шт."]]
    else:
        actual = pd.DataFrame(columns=["Точка", "Магазин", "SKU", "Название из файла", "Факт. остаток, шт."])
    if calculated.empty:
        calculated = pd.DataFrame(
            columns=["Точка", "Магазин", "SKU", "Название товара", "Категория", "Расчётный живой остаток, шт."]
        )
    comparison = calculated.merge(
        actual,
        on=["Точка", "SKU"],
        how="outer",
        suffixes=("", "_факт"),
    )
    if "Магазин_факт" in comparison.columns:
        comparison["Магазин"] = comparison["Магазин"].fillna(comparison["Магазин_факт"])
        comparison = comparison.drop(columns="Магазин_факт")
    comparison["Название товара"] = comparison.get("Название товара", pd.Series(index=comparison.index, dtype=object)).fillna(
        comparison.get("Название из файла", "")
    )
    comparison["Расчётный живой остаток, шт."] = pd.to_numeric(
        comparison.get("Расчётный живой остаток, шт."), errors="coerce"
    ).fillna(0.0)
    comparison["Факт. остаток, шт."] = pd.to_numeric(
        comparison.get("Факт. остаток, шт."), errors="coerce"
    )
    comparison["Отклонение, шт."] = (
        comparison["Факт. остаток, шт."] - comparison["Расчётный живой остаток, шт."]
    )
    def status(row: pd.Series) -> str:
        if pd.isna(row["Факт. остаток, шт."]):
            return "Нет фактического остатка"
        delta = float(row["Отклонение, шт."])
        calc = float(row["Расчётный живой остаток, шт."])
        if abs(delta) < 0.5:
            return "Совпадает"
        material = abs(delta) >= 2 or abs(delta) / max(calc, 1.0) >= 0.25
        direction = "Факт выше расчёта" if delta > 0 else "Факт ниже расчёта"
        return f"{direction} · заметно" if material else direction
    comparison["Статус сверки"] = comparison.apply(status, axis=1)
    allowed_points = {normalize_point_label(value) for value in selected_points}
    comparison = comparison[comparison["Точка"].isin(allowed_points)].copy()
    point_sort = pd.to_numeric(comparison["Точка"].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce")
    comparison["_point_sort"] = point_sort
    return comparison.sort_values(["_point_sort", "SKU"], kind="stable").drop(columns="_point_sort")


def style_stock_comparison(row: pd.Series) -> list[str]:
    styles = [""] * len(row)
    status = str(row.get("Статус сверки", ""))
    target_columns = [column for column in ["Факт. остаток, шт.", "Отклонение, шт.", "Статус сверки"] if column in row.index]
    if status == "Совпадает":
        style = "background-color: #D9EAD3; color: #274E13"
    elif "заметно" in status:
        style = "background-color: #F4CCCC; color: #990000; font-weight: 700"
    elif status.startswith("Нет"):
        style = "background-color: #F2F2F2; color: #777777"
    else:
        style = "background-color: #FFF2CC; color: #7F6000"
    for column in target_columns:
        styles[row.index.get_loc(column)] = style
    return styles


def allocate_integer_budget(weights: pd.Series, budget: int, caps: pd.Series | None = None) -> list[int]:
    count = len(weights)
    if count == 0:
        return []
    budget = max(count, int(round(budget)))
    clean = pd.to_numeric(weights, errors="coerce").fillna(0).clip(lower=0)
    if clean.sum() <= 0:
        clean = pd.Series([1.0] * count, index=weights.index)
    shares = clean / clean.sum()
    remainder_budget = budget - count
    raw_extra = shares * remainder_budget
    extras = raw_extra.apply(lambda value: int(value // 1))
    missing = remainder_budget - int(extras.sum())
    if missing > 0:
        order = (raw_extra - extras).sort_values(ascending=False).index[:missing]
        extras.loc[order] += 1
    allocations = (extras + 1).astype(int)
    if caps is None:
        return allocations.tolist()
    clean_caps = pd.to_numeric(caps, errors="coerce").fillna(budget).clip(lower=1).round().astype(int)
    allocations = pd.Series(
        [min(int(value), int(clean_caps.iloc[position])) for position, value in enumerate(allocations)],
        index=weights.index,
        dtype=int,
    )
    while int(allocations.sum()) < budget:
        eligible = allocations.index[allocations < clean_caps]
        if len(eligible) == 0:
            # Исторические пределы являются жёсткими. Нераспределённый остаток
            # относится к ассортименту категории, которого нет в текущем меню,
            # и не должен принудительно увеличивать присутствующие SKU.
            break
        ideal = shares * budget
        candidate = (ideal.loc[eligible] - allocations.loc[eligible]).idxmax()
        allocations.loc[candidate] += 1
    return allocations.tolist()


def calculate_learned_forecast(
    menu: pd.DataFrame,
    history: pd.DataFrame,
    entities: pd.DataFrame,
    target_date: date,
    point_mapping: dict[int, str],
    analyst_plans: pd.DataFrame,
) -> pd.DataFrame:
    menu_enriched = menu.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    menu_enriched["category"] = menu_enriched["category"].fillna(menu_enriched["matrix_category"])
    menu_enriched["entity"] = menu_enriched["entity"].fillna("Не сопоставлено")

    history_enriched = history.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    history_enriched = history_enriched[history_enriched["category"].notna()].copy()
    history_enriched["sold_quantity"] = history_enriched["sold_quantity"].clip(lower=0)
    history_enriched["business_date"] = pd.to_datetime(history_enriched["business_date"]).dt.date

    plans = analyst_plans.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    plans["category"] = plans["category"].fillna(plans["matrix_category"])
    plans_prior = plans[plans["plan_date"] < target_date].copy()
    plans_target = plans[plans["plan_date"] == target_date].copy()

    results: list[dict[str, object]] = []
    for shop_number, point_name in point_mapping.items():
        point_number = int(str(point_name).lstrip("Тт"))
        point_history = history_enriched[history_enriched["shop_number"] == shop_number].copy()
        point_plans = plans_prior[plans_prior["point_number"] == point_number].copy()
        point_target_plans = plans_target[plans_target["point_number"] == point_number].copy()
        for category, category_menu in menu_enriched.groupby("category", dropna=False):
            category_daily = (
                point_history[
                    (point_history["category"] == category)
                    & (point_history["business_date"].map(lambda value: value.weekday()) == target_date.weekday())
                    & (point_history["business_date"] < target_date)
                ]
                .groupby("business_date", as_index=False)["sold_quantity"].sum()
                .sort_values("business_date")
            )
            if category_daily.empty:
                sql_demand = 0.0
            else:
                category_daily["weeks_ago"] = category_daily["business_date"].map(
                    lambda value: max(0, (target_date - value).days // 7)
                )
                category_daily["weight"] = 0.85 ** category_daily["weeks_ago"]
                weighted = float(
                    (category_daily["sold_quantity"] * category_daily["weight"]).sum()
                    / category_daily["weight"].sum()
                )
                sql_demand = max(weighted, float(category_daily["sold_quantity"].median()))

            plan_category = (
                point_plans[point_plans["category"] == category]
                .groupby("plan_date", as_index=False)["analyst_plan"].sum()
                .sort_values("plan_date")
            )
            calibration_values: list[float] = []
            for _, plan_row in plan_category.iterrows():
                actual = point_history[
                    (point_history["category"] == category)
                    & (point_history["business_date"] == plan_row["plan_date"])
                ]["sold_quantity"].sum()
                if actual > 0:
                    calibration_values.append(float(plan_row["analyst_plan"] / actual))
            calibration = float(pd.Series(calibration_values).median()) if calibration_values else 1.0
            calibration = min(2.0, max(0.5, calibration))
            prior_plan_level = float(plan_category["analyst_plan"].median()) if not plan_category.empty else 0.0
            category_plan_rows = point_plans[point_plans["category"] == category]
            target_category_rows = point_target_plans[point_target_plans["category"] == category]
            has_target_category = not target_category_rows.empty
            target_category_level = float(target_category_rows["analyst_plan"].sum())

            # Сначала оцениваем долю сущностей текущего меню в полном плане
            # аналитика по категории. SKU используется только как резерв, когда
            # сущность не определена: главным является профиль спроса категории.
            current_skus = set(category_menu["sku"].dropna().tolist())
            current_entities = set(category_menu["entity"].dropna().tolist()) - {"Не сопоставлено"}
            coverage_values: list[float] = []
            covered_plan_values: list[float] = []
            coverage_method = "Нет истории — вся категория"
            for plan_date, date_group in category_plan_rows.groupby("plan_date"):
                total_plan = float(date_group["analyst_plan"].sum())
                if total_plan <= 0:
                    continue
                if current_entities:
                    covered_value = float(
                        date_group[date_group["entity"].isin(current_entities)]["analyst_plan"].sum()
                    )
                    if covered_value > 0:
                        coverage_method = "Доля сущностей текущего меню в прошлых планах"
                else:
                    covered_value = float(
                        date_group[date_group["sku"].isin(current_skus)]["analyst_plan"].sum()
                    )
                    if covered_value > 0:
                        coverage_method = "Резервная доля текущих SKU в прошлых планах"
                if covered_value > 0:
                    coverage_values.append(min(1.0, covered_value / total_plan))
                    covered_plan_values.append(covered_value)

            menu_coverage = float(pd.Series(coverage_values).median()) if coverage_values else 1.0
            covered_prior_level = (
                float(pd.Series(covered_plan_values).median()) if covered_plan_values else 0.0
            )
            sales_category_level = max(sql_demand * calibration, prior_plan_level)
            if has_target_category and target_category_level > 0:
                # План аналитика на выбранную дату задаёт основной масштаб категории,
                # а продажи дают ограниченную корректировку в пределах ±20%.
                bounded_sales_level = min(
                    target_category_level * 1.20,
                    max(target_category_level * 0.80, sales_category_level),
                )
                full_category_budget = max(
                    len(category_menu),
                    int(round(0.85 * target_category_level + 0.15 * bounded_sales_level)),
                )
                category_budget = full_category_budget
                category_anchor = "85% план аналитика на дату + 15% продажи (корректировка ограничена ±20%)"
            elif has_target_category:
                # Ноль в плане аналитика — это не отсутствие информации. Не
                # заменяем его большим автоматическим прогнозом: оставляем минимум
                # и передаём ячейки категории на ручное решение знаком #.
                full_category_budget = len(category_menu)
                category_budget = len(category_menu)
                category_anchor = "В плане аналитика на дату стоит 0 — ручная проверка"
            else:
                full_category_budget = max(
                    len(category_menu), int(round(sales_category_level))
                )
                # Без плана на дату используем историческое покрытие, но не даём
                # ему снизить бюджет ниже типичного плана текущих сущностей.
                category_budget = max(
                    len(category_menu),
                    int(round(max(full_category_budget * menu_coverage, covered_prior_level))),
                )
                category_anchor = "История категории и покрытие текущими сущностями"

            item_weights: list[float] = []
            item_caps: list[float] = []
            weight_methods: list[str] = []
            review_reasons: list[str] = []
            analyst_entity_shares: list[float] = []
            sql_entity_shares: list[float] = []
            same_day_history = point_history[
                (point_history["business_date"].map(lambda value: value.weekday()) == target_date.weekday())
                & (point_history["business_date"] < target_date)
            ]
            category_per_dish = category_plan_rows.groupby("plan_date").apply(
                lambda group: group["analyst_plan"].sum() / max(1, group["sku"].nunique()),
                include_groups=False,
            ) if not category_plan_rows.empty else pd.Series(dtype=float)
            category_typical = float(category_per_dish.median()) if not category_per_dish.empty else max(
                1.0, category_budget / max(1, len(category_menu))
            )
            category_sales_total = float(
                same_day_history[same_day_history["category"] == category]["sold_quantity"].sum()
            )
            entity_counts_in_menu = category_menu["entity"].value_counts().to_dict()
            for _, item in category_menu.iterrows():
                exact_plans = category_plan_rows[category_plan_rows["sku"] == item["sku"]]["analyst_plan"]
                exact_target = target_category_rows[
                    target_category_rows["sku"] == item["sku"]
                ]["analyst_plan"]
                exact_reference = pd.concat([exact_plans, exact_target], ignore_index=True)
                entity_plan_rows = category_plan_rows[category_plan_rows["entity"] == item["entity"]]
                entity_plans = entity_plan_rows["analyst_plan"]
                target_entity_rows = target_category_rows[
                    target_category_rows["entity"] == item["entity"]
                ]
                sku_daily = same_day_history[same_day_history["sku"] == item["sku"]].groupby(
                    "business_date"
                )["sold_quantity"].sum()
                entity_history = same_day_history[
                    (same_day_history["category"] == category)
                    & (same_day_history["entity"] == item["entity"])
                ]
                entity_daily = entity_history.groupby("business_date")["sold_quantity"].sum()
                entity_sku_count = max(1, entity_history["sku"].nunique())
                entity_plan_by_date = entity_plan_rows.groupby("plan_date")["analyst_plan"].sum()
                entity_plan_share_values: list[float] = []
                for plan_date, entity_value in entity_plan_by_date.items():
                    category_value = float(
                        category_plan_rows[category_plan_rows["plan_date"] == plan_date]["analyst_plan"].sum()
                    )
                    if category_value > 0:
                        entity_plan_share_values.append(float(entity_value) / category_value)
                analyst_entity_share = (
                    float(pd.Series(entity_plan_share_values).median())
                    if entity_plan_share_values else 0.0
                )
                target_entity_share = (
                    float(target_entity_rows["analyst_plan"].sum()) / target_category_level
                    if target_category_level > 0 else 0.0
                )
                if target_entity_share > 0:
                    analyst_entity_share = (
                        0.80 * target_entity_share + 0.20 * analyst_entity_share
                        if analyst_entity_share > 0 else target_entity_share
                    )
                sql_entity_share = (
                    float(entity_history["sold_quantity"].sum()) / category_sales_total
                    if category_sales_total > 0 else analyst_entity_share
                )
                entity_share = max(0.0, 0.75 * analyst_entity_share + 0.25 * sql_entity_share)
                menu_entity_count = max(1, int(entity_counts_in_menu.get(item["entity"], 1)))
                entity_based_typical = category_budget * entity_share / menu_entity_count

                combined_entity_plans = pd.concat(
                    [entity_plans, target_entity_rows["analyst_plan"]], ignore_index=True
                )
                if not combined_entity_plans.empty:
                    sku_adjustment = (
                        float(exact_reference.median())
                        if not exact_reference.empty else entity_based_typical
                    )
                    analyst_typical = max(1.0, 0.85 * entity_based_typical + 0.15 * sku_adjustment)
                    sql_typical = (
                        float(entity_daily.median()) / entity_sku_count
                        if not entity_daily.empty else analyst_typical
                    )
                    target_entity_total = float(target_entity_rows["analyst_plan"].sum())
                    entity_historical_max = float(entity_plan_by_date.max()) if not entity_plan_by_date.empty else 0.0
                    entity_cap_per_item = max(target_entity_total, entity_historical_max) / menu_entity_count + 2
                    exact_cap = float(exact_reference.max()) + 2 if not exact_reference.empty else entity_cap_per_item
                    cap = max(1.0, min(entity_cap_per_item, exact_cap))
                    method = "План категории → преобладание сущности → распределение по SKU"
                else:
                    analyst_typical = category_typical
                    sql_typical = category_typical
                    cap = max(category_typical + 2, category_typical * 1.5)
                    method = "Нет истории сущности — типичная доля блюда в категории"
                reasons: list[str] = []
                if len(category_daily) < 2:
                    reasons.append("меньше двух аналогичных дней продаж категории")
                if not coverage_values and target_category_level <= 0:
                    reasons.append("нет истории доли текущего ассортимента в категории")
                if has_target_category and target_category_level <= 0:
                    reasons.append("в плане аналитика на выбранную дату категория имеет нулевой план")
                if target_entity_rows.empty and len(entity_plan_by_date) < 2:
                    reasons.append("недостаточно решений аналитика по сущности внутри категории")
                if sku_daily.empty and entity_daily.empty:
                    reasons.append("нет продаж SKU и сущности в аналогичный день недели")
                score = max(1.0, 0.85 * analyst_typical + 0.15 * sql_typical)
                item_weights.append(score)
                item_caps.append(max(1.0, cap))
                weight_methods.append(method)
                review_reasons.append("; ".join(reasons))
                analyst_entity_shares.append(analyst_entity_share)
                sql_entity_shares.append(sql_entity_share)

            allocations = allocate_integer_budget(
                pd.Series(item_weights), category_budget, pd.Series(item_caps)
            )
            allocated_total = int(sum(allocations))
            unallocated_demand = max(0, int(full_category_budget - allocated_total))
            for (_, item), recommended, method, weight, cap, review_reason, analyst_share, sql_share in zip(
                category_menu.iterrows(), allocations, weight_methods, item_weights, item_caps,
                review_reasons, analyst_entity_shares, sql_entity_shares,
            ):
                results.append(
                    {
                        "Дата плана": target_date,
                        "Точка": point_name,
                        "Номер точки": point_number,
                        "Номер магазина": int(shop_number),
                        "Категория": category,
                        "Бюджет категории": category_budget,
                        "Основа бюджета категории": category_anchor,
                        "План аналитика на выбранную дату": round(target_category_level, 1),
                        "Полный прогноз категории": full_category_budget,
                        "Покрытие категории текущим меню": round(menu_coverage, 3),
                        "Метод покрытия": coverage_method,
                        "Нераспределённый спрос категории": unallocated_demand,
                        "Спрос категории SQL": round(sql_demand, 1),
                        "Коэффициент аналитика": round(calibration, 3),
                        "Медиана прошлых планов категории": round(prior_plan_level, 1),
                        "Сущность": item["entity"],
                        "Доля сущности по планам аналитика": round(analyst_share, 3),
                        "Доля сущности по продажам": round(sql_share, 3),
                        "SKU": item["sku"],
                        "Название товара": item["product_name"],
                        "Вес распределения": round(float(weight), 2),
                        "Исторический верхний ориентир": round(float(cap), 1),
                        "Метод распределения": method,
                        "Рекомендованный план": int(recommended),
                        "Статус прогноза": "Требует ручной проверки" if review_reason else "Рассчитано",
                        "Причина ручной проверки": review_reason,
                        "Лист": item["sheet"],
                        "Строка Excel": int(item["excel_row"]),
                        "Строка заголовков": int(item["header_excel_row"]),
                    }
                )
    return pd.DataFrame(results).sort_values(["Номер точки", "Категория", "Сущность", "SKU"])


SEVEN_DAY_POINTS = {1, 4, 10, 22, 26, 27}


def calculate_sku_daily_forecast(
    menu: pd.DataFrame,
    history: pd.DataFrame,
    entities: pd.DataFrame,
    target_date: date,
    point_mapping: dict[int, str],
    lookback_weeks: int = 8,
) -> pd.DataFrame:
    """Прогнозирует каждую ячейку Т по среднему SKU за выбранный исторический период."""
    menu_enriched = menu.merge(
        entities[["sku", "category", "entity"]], on="sku", how="left"
    )
    menu_enriched["category"] = menu_enriched["category"].fillna(
        menu_enriched["matrix_category"]
    ).map(normalize_matrix_category)
    menu_enriched["entity"] = menu_enriched["entity"].fillna("Не сопоставлено")

    sold = history.copy()
    sold["business_date"] = pd.to_datetime(sold["business_date"], errors="coerce").dt.date
    sold["sku"] = sold["sku"].map(normalize_sku)
    sold["sold_quantity"] = pd.to_numeric(
        sold["sold_quantity"], errors="coerce"
    ).fillna(0.0)
    sold = sold[
        sold["business_date"].notna()
        & (sold["business_date"] < target_date)
        & sold["sku"].notna()
    ].copy()

    results: list[dict[str, object]] = []
    for shop_number, point_name in point_mapping.items():
        point_number = int(str(point_name).lstrip("Тт"))
        point_type = "7 дней" if point_number in SEVEN_DAY_POINTS else "5 дней"
        point_history = sold[
            pd.to_numeric(sold["shop_number"], errors="coerce") == int(shop_number)
        ]
        for _, item in menu_enriched.iterrows():
            sku_history = point_history[point_history["sku"] == item["sku"]]
            daily_sales = (
                sku_history.groupby("business_date", as_index=False)["sold_quantity"].sum()
                if not sku_history.empty
                else pd.DataFrame(columns=["business_date", "sold_quantity"])
            )
            daily_sales = daily_sales[daily_sales["sold_quantity"] > 0]
            sale_days = int(daily_sales["business_date"].nunique())
            sold_total = float(daily_sales["sold_quantity"].sum()) if sale_days else 0.0
            used_default_average = sale_days == 0
            low_sales_day_count = sale_days in (1, 2)
            average_per_sale_day = sold_total / sale_days if sale_days else 1.0
            lifecycle_days = product_lifecycle_days(item["category"])
            coverage_days = forecast_coverage_days(item["category"])
            average_rule = (
                "продаж не было, принято базовое среднее 1 шт./день"
                if used_default_average
                else "среднее рассчитано по фактическим дням продаж"
            )
            coverage_rule = (
                f"{average_rule}; загрузка по категории: "
                f"среднее SKU × {coverage_days} дн."
            )

            calculated_need = (
                float(average_per_sale_day) * coverage_days
                if pd.notna(average_per_sale_day)
                else pd.NA
            )
            recommended = (
                int(math.ceil(float(calculated_need)))
                if pd.notna(calculated_need) and float(calculated_need) > 0
                else pd.NA
            )
            results.append(
                {
                    "Дата плана": target_date,
                    "День недели": WEEKDAY_RU.get(target_date.weekday(), ""),
                    "Точка": point_name,
                    "Номер точки": point_number,
                    "Номер магазина": int(shop_number),
                    "Тип точки": point_type,
                    "Категория": item["category"],
                    "Сущность": item["entity"],
                    "SKU": item["sku"],
                    "Название товара": item["product_name"],
                    "Период среднего, недель": int(lookback_weeks),
                    "Продано за выбранный период, шт.": round(sold_total, 3),
                    "Дней с продажами": sale_days,
                    "Среднее SKU за день продажи": (
                        round(float(average_per_sale_day), 2)
                        if pd.notna(average_per_sale_day)
                        else pd.NA
                    ),
                    "Жизненный цикл, дней": lifecycle_days,
                    "Дней покрытия поставкой": coverage_days,
                    "Идеальный жизненный цикл для плана, дней": coverage_days,
                    "Правило расчёта": coverage_rule,
                    "Расчётная потребность": (
                        round(float(calculated_need), 2)
                        if pd.notna(calculated_need)
                        else pd.NA
                    ),
                    "Рекомендованный план": recommended,
                    "Статус прогноза": (
                        "Базовое значение 1 шт./день"
                        if used_default_average
                        else (
                            f"Мало данных: {sale_days} дн. продаж"
                            if low_sales_day_count
                            else "Рассчитано"
                        )
                    ),
                    "Причина": (
                        "Продаж SKU на точке не было: среднее принято равным 1 шт./день"
                        if used_default_average
                        else (
                            f"За выбранные {int(lookback_weeks)} нед. продажи были только в {sale_days} дн.; "
                            "прогноз рассчитан, но выборка недостаточна для уверенного решения"
                            if low_sales_day_count
                            else ""
                        )
                    ),
                    "Лист": item["sheet"],
                    "Строка Excel": int(item["excel_row"]),
                    "Строка заголовков": int(item["header_excel_row"]),
                }
            )
    return pd.DataFrame(results).sort_values(
        ["Номер точки", "Категория", "Сущность", "SKU"]
    )


def calculate_entity_forecast(
    menu: pd.DataFrame,
    history: pd.DataFrame,
    entities: pd.DataFrame,
    target_date: date,
    point_mapping: dict[int, str],
    capacity: pd.DataFrame,
) -> pd.DataFrame:
    menu_enriched = menu.merge(
        entities[["sku", "category", "entity"]], on="sku", how="left"
    )
    menu_enriched["category"] = menu_enriched["category"].fillna(menu_enriched["matrix_category"])
    menu_enriched["entity"] = menu_enriched["entity"].fillna("Не сопоставлено")

    history_enriched = history.merge(
        entities[["sku", "category", "entity"]], on="sku", how="left"
    )
    history_enriched = history_enriched[
        history_enriched["category"].notna() & history_enriched["entity"].notna()
    ].copy()
    history_enriched = history_enriched[
        pd.to_datetime(history_enriched["business_date"]).dt.weekday == target_date.weekday()
    ].copy()
    history_enriched["sold_quantity"] = history_enriched["sold_quantity"].clip(lower=0)
    history_enriched["weeks_ago"] = history_enriched["business_date"].map(
        lambda value: max(0, (target_date - value).days // 7)
    )
    history_enriched["weight"] = 0.85 ** history_enriched["weeks_ago"]

    entity_daily = (
        history_enriched.groupby(
            ["shop_number", "business_date", "category", "entity"], as_index=False
        )["sold_quantity"].sum()
    )
    entity_daily["weeks_ago"] = entity_daily["business_date"].map(
        lambda value: max(0, (target_date - value).days // 7)
    )
    entity_daily["weight"] = 0.85 ** entity_daily["weeks_ago"]

    category_daily = (
        history_enriched.groupby(["shop_number", "business_date", "category"], as_index=False)["sold_quantity"].sum()
    )
    category_daily["weeks_ago"] = category_daily["business_date"].map(
        lambda value: max(0, (target_date - value).days // 7)
    )
    category_daily["weight"] = 0.85 ** category_daily["weeks_ago"]

    global_entity = history_enriched.groupby(["category", "entity"], as_index=False)["sold_quantity"].sum()
    global_entity["category_total"] = global_entity.groupby("category")["sold_quantity"].transform("sum")
    global_entity["entity_share"] = global_entity["sold_quantity"] / global_entity["category_total"].replace(0, pd.NA)
    global_share = global_entity.set_index(["category", "entity"])["entity_share"].to_dict()

    capacity_lookup: dict[tuple[int, str], float] = {}
    if not capacity.empty:
        for _, row in capacity.iterrows():
            point_number = int(row["point_number"])
            for column in capacity.columns[2:]:
                value = pd.to_numeric(row[column], errors="coerce")
                if pd.notna(value):
                    capacity_lookup[(point_number, normalize_matrix_category(column))] = float(value)

    rows: list[dict[str, object]] = []
    for shop_number, point_name in point_mapping.items():
        point_number = int(str(point_name).lstrip("Тт"))
        for (category, entity), menu_group in menu_enriched.groupby(["category", "entity"], dropna=False):
            category_history = category_daily[
                (category_daily["shop_number"] == shop_number)
                & (category_daily["category"] == category)
            ]
            if not category_history.empty and category_history["weight"].sum() > 0:
                category_forecast = (
                    category_history["sold_quantity"] * category_history["weight"]
                ).sum() / category_history["weight"].sum()
            else:
                category_forecast = 0.0
            category_history = category_history.sort_values("business_date")
            category_mode = 0.0
            category_growth_factor = 1.0
            if not category_history.empty:
                rounded_category = category_history["sold_quantity"].round().clip(lower=0)
                if not rounded_category.mode().empty:
                    category_mode = float(rounded_category.mode().iloc[0])
                if len(category_history) >= 3:
                    category_recent = float(category_history.tail(2)["sold_quantity"].mean())
                    category_previous = float(category_history.iloc[:-2]["sold_quantity"].mean())
                    if category_previous > 0 and category_recent > category_previous:
                        category_growth_factor = min(2.0, category_recent / category_previous)
            # На снижении не уменьшаем базу; при росте категории применяем только положительный коэффициент.
            category_forecast = max(category_forecast, category_mode)
            same_category_entities = menu_enriched[menu_enriched["category"] == category]["entity"].nunique()
            share = global_share.get((category, entity))
            if pd.isna(share) or share is None:
                share = 1 / max(1, same_category_entities)
            category_entity_forecast = category_forecast * float(share)

            exact = entity_daily[
                (entity_daily["shop_number"] == shop_number)
                & (entity_daily["category"] == category)
                & (entity_daily["entity"] == entity)
            ]
            observations = exact["business_date"].nunique()
            if not exact.empty and exact["weight"].sum() > 0:
                own_entity_forecast = (exact["sold_quantity"] * exact["weight"]).sum() / exact["weight"].sum()
                # Категория задаёт общий спрос точки, сущность уточняет распределение внутри категории.
                entity_forecast = 0.70 * own_entity_forecast + 0.30 * category_entity_forecast
                method = "Спрос категории + история сущности на точке"
                confidence = "Высокая" if observations >= 4 else "Средняя" if observations >= 2 else "Низкая"
            else:
                entity_forecast = category_entity_forecast
                observations = 0
                method = "Доля сущности в категории"
                confidence = "Низкая"

            sku_history = history_enriched[
                (history_enriched["shop_number"] == shop_number)
                & (history_enriched["category"] == category)
                & (history_enriched["entity"] == entity)
                & (history_enriched["sku"].isin(menu_group["sku"]))
            ].groupby("sku")["sold_quantity"].sum()
            sku_history_total = sku_history.sum()
            equal_share = 1 / max(1, len(menu_group))
            capacity_value = capacity_lookup.get((point_number, category))
            for _, item in menu_group.iterrows():
                sku_share = float(sku_history.get(item["sku"], 0) / sku_history_total) if sku_history_total > 0 else equal_share
                raw_forecast = max(0.0, entity_forecast * sku_share)
                trend_source = history_enriched[
                    (history_enriched["shop_number"] == shop_number)
                    & (history_enriched["sku"] == item["sku"])
                ].groupby("business_date", as_index=False)["sold_quantity"].sum().sort_values("business_date")
                sku_mode = 0
                if not trend_source.empty:
                    rounded_sku = trend_source["sold_quantity"].round().clip(lower=0)
                    if not rounded_sku.mode().empty:
                        sku_mode = int(rounded_sku.mode().iloc[0])
                base_quantity = max(1, int(round(raw_forecast)), sku_mode)
                recommended = max(1, int(round(base_quantity * category_growth_factor)))
                trend_adjustment = recommended - base_quantity
                trend_label = "Рост категории" if category_growth_factor > 1 else "Без снижения"
                trend_basis = "Категория точки"
                limited = False
                if capacity_value is not None:
                    if capacity_value <= 0:
                        # Нулевая ёмкость не должна обнулять обязательную поставку блюда.
                        recommended = 1
                        limited = True
                    elif recommended > capacity_value:
                        recommended = max(1, int(capacity_value))
                        limited = True
                rows.append(
                    {
                        "Дата плана": target_date,
                        "Точка": point_name,
                        "Номер точки": point_number,
                        "Номер магазина": int(shop_number),
                        "Категория": category,
                        "Сущность": entity,
                        "SKU": item["sku"],
                        "Название товара": item["product_name"],
                        "Лист": item["sheet"],
                        "Строка Excel": int(item["excel_row"]),
                        "Строка заголовков": int(item["header_excel_row"]),
                        "Историческое среднее": round(entity_forecast * sku_share, 1),
                        "Спрос категории на точке": round(category_forecast, 1),
                        "Преобладающий уровень категории": round(category_mode, 1),
                        "Коэффициент роста категории": round(category_growth_factor, 3),
                        "Преобладающее количество SKU": sku_mode,
                        "Доля сущности в категории": round(float(share), 4),
                        "Рекомендованный план": recommended,
                        "Аналогичных дней": int(observations),
                        "Уверенность": confidence,
                        "Метод": method,
                        "Тренд": trend_label,
                        "Основа тренда": trend_basis,
                        "Корректировка тренда": trend_adjustment,
                        "Ёмкость категории": capacity_value,
                        "Ограничено ёмкостью": "Да" if limited else "Нет",
                    }
                )
    return pd.DataFrame(rows).sort_values(["Номер точки", "Категория", "Сущность", "SKU"])


def export_forecast_excel(forecast: pd.DataFrame, menu: pd.DataFrame, lookback_weeks: int) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        forecast.to_excel(writer, sheet_name="Рекомендации", index=False)
        menu.to_excel(writer, sheet_name="Меню без изменений", index=False)
        methodology = pd.DataFrame(
            {
                "Параметр": ["Исторический период", "Сравнение", "Основной метод", "Резервный метод", "Ограничение"],
                "Значение": [
                    f"Последние {lookback_weeks} недель до даты плана",
                    "Тот же день недели",
                    "70% истории сущности + 30% спроса категории на точке",
                    "Прогноз категории × историческая доля сущности",
                    "Ёмкость категории на точке",
                ],
            }
        )
        methodology.to_excel(writer, sheet_name="Методика", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cells in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 45)
                sheet.column_dimensions[cells[0].column_letter].width = width
    return buffer.getvalue()


def _is_forecast_fact_header(value: object) -> bool:
    """True for matrix columns marked with Cyrillic 'Ф' (fact columns)."""
    text = str(value or "").strip().upper().replace(" ", "")
    return text.startswith("Ф")


def _delete_columns_preserving_merges(sheet, columns: list[int]) -> None:
    """Delete worksheet columns while remapping merged ranges and widths."""
    delete_columns = sorted({int(column) for column in columns if 1 <= int(column) <= sheet.max_column})
    if not delete_columns:
        return

    delete_set = set(delete_columns)
    old_max_column = sheet.max_column
    column_map: dict[int, int | None] = {}
    shift = 0
    for old_column in range(1, old_max_column + 1):
        if old_column in delete_set:
            column_map[old_column] = None
            shift += 1
        else:
            column_map[old_column] = old_column - shift

    merged_specs: list[dict[str, object]] = []
    for merged_range in list(sheet.merged_cells.ranges):
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
        merged_specs.append(
            {
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
                "value": top_left.value,
                "style": copy(top_left._style),
                "number_format": top_left.number_format,
                "alignment": copy(top_left.alignment),
                "border": copy(top_left.border),
                "fill": copy(top_left.fill),
                "font": copy(top_left.font),
                "protection": copy(top_left.protection),
                "comment": copy(top_left.comment),
            }
        )

    # Clearing the merge registry in one step is much faster than thousands of
    # unmerge_cells() calls on the menu sheets. MergedCell placeholders are then
    # removed and only the few ranges that still span multiple cells are restored.
    sheet.merged_cells.ranges.clear()
    for coordinate, cell in list(sheet._cells.items()):
        if cell.__class__.__name__ == "MergedCell":
            del sheet._cells[coordinate]

    dimension_specs: dict[int, dict[str, object]] = {}
    for old_column in range(1, old_max_column + 1):
        old_dimension = sheet.column_dimensions.get(get_column_letter(old_column))
        if old_dimension is None:
            continue
        dimension_specs[old_column] = {
            "width": old_dimension.width,
            "hidden": old_dimension.hidden,
            "bestFit": old_dimension.bestFit,
            "outlineLevel": old_dimension.outlineLevel,
            "collapsed": old_dimension.collapsed,
        }

    for column in reversed(delete_columns):
        sheet.delete_cols(column, 1)

    for old_column, new_column in column_map.items():
        if new_column is None or old_column not in dimension_specs:
            continue
        spec = dimension_specs[old_column]
        new_dimension = sheet.column_dimensions[get_column_letter(new_column)]
        new_dimension.width = spec["width"]
        new_dimension.hidden = spec["hidden"]
        new_dimension.bestFit = spec["bestFit"]
        new_dimension.outlineLevel = spec["outlineLevel"]
        new_dimension.collapsed = spec["collapsed"]

    transformed_merges: list[tuple[int, int, int, int, dict[str, object]]] = []
    for spec in merged_specs:
        kept_columns = [
            old_column
            for old_column in range(int(spec["min_col"]), int(spec["max_col"]) + 1)
            if column_map.get(old_column) is not None
        ]
        if not kept_columns:
            continue
        new_min_col = int(column_map[kept_columns[0]])
        new_max_col = int(column_map[kept_columns[-1]])
        min_row = int(spec["min_row"])
        max_row = int(spec["max_row"])
        transformed_merges.append((min_row, max_row, new_min_col, new_max_col, spec))

    # Restore top-left content/styles before creating any merged cells. This avoids
    # attempting to write into a MergedCell when row insertions caused adjacent
    # source merges to touch after the fact columns were removed.
    for min_row, max_row, new_min_col, new_max_col, spec in transformed_merges:
        target_cell = sheet.cell(min_row, new_min_col)
        target_cell.value = spec["value"]
        target_cell._style = copy(spec["style"])
        target_cell.number_format = spec["number_format"]
        target_cell.alignment = copy(spec["alignment"])
        target_cell.border = copy(spec["border"])
        target_cell.fill = copy(spec["fill"])
        target_cell.font = copy(spec["font"])
        target_cell.protection = copy(spec["protection"])
        target_cell.comment = copy(spec["comment"])

    occupied_merge_cells: set[tuple[int, int]] = set()
    seen_ranges: set[tuple[int, int, int, int]] = set()
    for min_row, max_row, new_min_col, new_max_col, _ in sorted(
        transformed_merges, key=lambda item: (item[0], item[2], item[1], item[3])
    ):
        if min_row == max_row and new_min_col == new_max_col:
            continue
        range_key = (min_row, max_row, new_min_col, new_max_col)
        if range_key in seen_ranges:
            continue
        cells = {
            (row, column)
            for row in range(min_row, max_row + 1)
            for column in range(new_min_col, new_max_col + 1)
        }
        if cells & occupied_merge_cells:
            continue
        sheet.merge_cells(
            start_row=min_row,
            start_column=new_min_col,
            end_row=max_row,
            end_column=new_max_col,
        )
        seen_ranges.add(range_key)
        occupied_merge_cells.update(cells)



def _forecast_export_date_label(forecast: pd.DataFrame) -> str:
    """Human-readable range of the plan dates that are actually exported."""
    if forecast is None or forecast.empty or "Дата плана" not in forecast.columns:
        return "выбранные даты"
    dates = sorted(set(
        pd.to_datetime(forecast["Дата плана"], errors="coerce").dropna().dt.date.tolist()
    ))
    if not dates:
        return "выбранные даты"
    if len(dates) == 1:
        return dates[0].strftime("%d.%m.%Y")
    return f"{dates[0]:%d.%m.%Y}–{dates[-1]:%d.%m.%Y}"


def _find_selected_plan_blocks(sheet, selected_dates: set[date]) -> list[tuple[int, int]]:
    """Return only matrix row ranges that belong to the selected plan dates."""
    date_rows: list[tuple[int, date]] = []
    max_scan_column = min(sheet.max_column, 12)
    for row_number in range(1, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, max_scan_column + 1)]
        texts = [str(value or "").strip().casefold() for value in values if value is not None]
        if any("участок комплектации" in text for text in texts):
            continue
        has_date_label = any(text.startswith("дата") for text in texts)
        has_plan_label = any("план на день кухня" in text for text in texts)
        parsed_date = next(
            (parsed for value in values if (parsed := parse_excel_date(value)) is not None),
            None,
        )
        if parsed_date is not None and (has_date_label or has_plan_label):
            date_rows.append((row_number, parsed_date))

    selected_ranges: list[tuple[int, int]] = []
    for position, (date_row, block_date) in enumerate(date_rows):
        if block_date not in selected_dates:
            continue
        next_date_row = (
            date_rows[position + 1][0]
            if position + 1 < len(date_rows)
            else sheet.max_row + 1
        )
        end_row = next_date_row - 1
        for candidate_row in range(date_row + 1, next_date_row):
            candidate_values = [
                sheet.cell(candidate_row, column).value
                for column in range(1, max_scan_column + 1)
            ]
            candidate_label = " ".join(
                str(value or "").strip().casefold()
                for value in candidate_values
                if value is not None
            )
            if "участок комплектации" in candidate_label:
                end_row = candidate_row - 1
                break
        if end_row >= date_row:
            selected_ranges.append((date_row, end_row))
    return selected_ranges


def _copy_selected_plan_sheet(source_sheet, target_sheet, row_ranges: list[tuple[int, int]]) -> None:
    """Copy selected plan blocks to a clean worksheet while keeping useful formatting."""
    from openpyxl.utils.cell import range_boundaries

    row_map: dict[int, int] = {}
    next_target_row = 1
    for range_index, (start_row, end_row) in enumerate(row_ranges):
        if range_index:
            next_target_row += 1
        for source_row_number in range(start_row, end_row + 1):
            target_row_number = next_target_row
            row_map[source_row_number] = target_row_number
            source_dimension = source_sheet.row_dimensions[source_row_number]
            target_dimension = target_sheet.row_dimensions[target_row_number]
            target_dimension.height = source_dimension.height
            target_dimension.hidden = source_dimension.hidden
            target_dimension.outlineLevel = source_dimension.outlineLevel

            for column in range(1, source_sheet.max_column + 1):
                source_cell = source_sheet.cell(source_row_number, column)
                target_cell = target_sheet.cell(target_row_number, column)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                target_cell.number_format = source_cell.number_format
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.font = copy(source_cell.font)
                target_cell.protection = copy(source_cell.protection)
                if source_cell.comment is not None:
                    target_cell.comment = copy(source_cell.comment)
                if source_cell.hyperlink is not None:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
            next_target_row += 1

    for column_letter, source_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel

    for merged_range in list(source_sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row not in row_map or max_row not in row_map:
            continue
        if row_map[max_row] - row_map[min_row] != max_row - min_row:
            continue
        target_sheet.merge_cells(
            start_row=row_map[min_row],
            start_column=min_col,
            end_row=row_map[max_row],
            end_column=max_col,
        )

    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
    target_sheet.sheet_properties.tabColor = copy(source_sheet.sheet_properties.tabColor)


def _restrict_forecast_workbook_to_selected_dates(workbook, forecast: pd.DataFrame):
    """Build a download workbook containing only the selected menu dates, never the full matrix."""
    from openpyxl import Workbook

    if forecast is None or forecast.empty:
        return workbook

    normalized = forecast.copy()
    normalized["_export_date"] = pd.to_datetime(
        normalized["Дата плана"], errors="coerce"
    ).dt.date
    selected_by_sheet: dict[str, set[date]] = {}
    for sheet_name, group in normalized.dropna(subset=["_export_date"]).groupby("Лист"):
        selected_by_sheet[str(sheet_name)] = set(group["_export_date"].tolist())

    if not selected_by_sheet:
        return workbook

    export_workbook = Workbook()
    export_workbook.remove(export_workbook.active)
    copied_any = False
    for sheet_name in workbook.sheetnames:
        if sheet_name not in selected_by_sheet:
            continue
        source_sheet = workbook[sheet_name]
        row_ranges = _find_selected_plan_blocks(source_sheet, selected_by_sheet[sheet_name])
        if not row_ranges:
            continue
        target_sheet = export_workbook.create_sheet(title=sheet_name)
        _copy_selected_plan_sheet(source_sheet, target_sheet, row_ranges)
        copied_any = True

    if copied_any:
        return export_workbook
    return workbook


def fill_forecast_into_matrix(
    file_bytes: bytes,
    forecast: pd.DataFrame,
    history_from: date,
    target_date: date,
    hide_average_values: bool = False,
) -> bytes:
    workbook = load_workbook(io.BytesIO(file_bytes))

    # Remember all matrix columns marked with 'Ф' before rows are inserted.
    # They are removed only from the downloaded copy; the uploaded source file is untouched.
    fact_columns_by_sheet: dict[str, list[int]] = {}
    for (sheet_name, header_row), _ in forecast.groupby(["Лист", "Строка заголовков"], sort=False):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        fact_columns_by_sheet.setdefault(sheet_name, [])
        for column in range(1, sheet.max_column + 1):
            if _is_forecast_fact_header(sheet.cell(int(header_row), column).value):
                fact_columns_by_sheet[sheet_name].append(column)

    for (sheet_name, header_row), group in forecast.groupby(["Лист", "Строка заголовков"]):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = {
            str(sheet.cell(row=int(header_row), column=column).value or "").strip(): column
            for column in range(1, sheet.max_column + 1)
        }
        menu_rows = sorted({int(value) for value in group["Строка Excel"]})
        # Т11 должна быть пустой. Т4 участвует в расчёте; Ф-столбцы не изменяются.
        for point_number in (11,):
            target_column = headers.get(f"Т{point_number}")
            if target_column is not None:
                for excel_row in menu_rows:
                    sheet.cell(row=excel_row, column=target_column).value = None
        for _, record in group.iterrows():
            point_number = int(record["Номер точки"])
            if point_number == 11 or not 1 <= point_number <= 29:
                continue
            forecast_header = f"Т{point_number}"
            forecast_column = headers.get(forecast_header)
            if forecast_column is None:
                continue
            target_cell = sheet.cell(row=int(record["Строка Excel"]), column=forecast_column)
            if pd.notna(record.get("Рекомендованный план")):
                target_cell.value = int(record["Рекомендованный план"])
                sale_days = int(record.get("Дней с продажами", 0) or 0)
                lookback_weeks = int(record.get("Период среднего, недель", 0) or 0)
                period_label = f"{lookback_weeks} нед." if lookback_weeks > 0 else "выбранный период"
                low_sales_day_count = sale_days in (1, 2)
                warning_line = (
                    "\n⚠ Мало данных: итог подсвечен светло-красным; "
                    "не рекомендуется опираться на него как на устойчивое среднее."
                    if low_sales_day_count
                    else ""
                )
                target_cell.comment = Comment(
                    (
                        f"Среднее SKU за день продажи: "
                        f"{float(record['Среднее SKU за день продажи']):.2f}\n"
                        f"Дней с продажами за период ({period_label}): {sale_days}\n"
                        f"Основание среднего: {record['Статус прогноза']}\n"
                        f"Идеальный цикл поставки: {int(record['Дней покрытия поставкой'])} дн.\n"
                        f"Срок категории: {int(record['Жизненный цикл, дней'])} дн.\n"
                        f"Расчёт: {float(record['Расчётная потребность']):.2f} → "
                        f"{int(record['Рекомендованный план'])}"
                        f"{warning_line}"
                    ),
                    "Прогноз плана",
                )
                if low_sales_day_count:
                    target_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            else:
                target_cell.value = None
                target_cell.comment = Comment(
                    "Нет продаж этого SKU на выбранной точке за выбранный период среднего.",
                    "Прогноз плана",
                )

        plan_column = headers.get("ПЛАН")
        point_columns = [
            headers.get(f"Т{point_number}")
            for point_number in range(1, 30)
            if point_number != 11
        ]
        point_columns = [column for column in point_columns if column is not None]
        if plan_column is not None:
            for excel_row in menu_rows:
                total = sum(
                    float(sheet.cell(row=excel_row, column=column).value or 0)
                    for column in point_columns
                    if isinstance(sheet.cell(row=excel_row, column=column).value, (int, float))
                )
                sheet.cell(row=excel_row, column=plan_column).value = int(round(total))

    # Одна компактная служебная строка размещается непосредственно под каждым SKU.
    # Средние показываются по точкам, а цикл загрузки — только в колонке категории.
    # Вставляем снизу вверх, чтобы сохранённые номера исходных строк не смещались.
    for sheet_name, sheet_group in forecast.groupby("Лист", sort=False):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for excel_row, row_group in sorted(
            sheet_group.groupby("Строка Excel"), key=lambda item: int(item[0]), reverse=True
        ):
            source_row = int(excel_row)
            header_row = int(row_group["Строка заголовков"].iloc[0])
            headers = {
                str(sheet.cell(header_row, column).value or "").strip(): column
                for column in range(1, sheet.max_column + 1)
            }
            sheet.insert_rows(source_row + 1, amount=1)
            for new_row in (source_row + 1,):
                sheet.row_dimensions[new_row].height = sheet.row_dimensions[source_row].height
                for column in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(source_row, column)
                    target_cell = sheet.cell(new_row, column)
                    if source_cell.has_style:
                        target_cell._style = copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.border = copy(source_cell.border)
                    target_cell.value = None
                    target_cell.comment = None

            average_row = source_row + 1
            sheet.cell(average_row, 1).value = "СР"
            category_column = headers.get("категория", headers.get("Категория", 4))
            name_column = headers.get("Название блюда", 5)
            sheet.cell(average_row, name_column).value = "Среднее SKU за день продажи"
            category_name = str(row_group["Категория"].iloc[0])
            coverage_days = int(row_group["Дней покрытия поставкой"].iloc[0])
            sheet.cell(average_row, category_column).value = (
                f"Цикл загрузки: {coverage_days} дн."
            )
            for column in range(1, sheet.max_column + 1):
                sheet.cell(average_row, column).fill = PatternFill("solid", fgColor="DDEBF7")
            sheet.cell(average_row, name_column).font = Font(bold=True, color="1F4E78")
            sheet.cell(average_row, category_column).font = Font(bold=True, color="7F6000")
            sheet.cell(average_row, category_column).comment = Comment(
                f"Категория: {category_name}. Множитель прогноза: {coverage_days} дня.",
                "Прогноз плана",
            )

            average_total = 0.0
            if not hide_average_values:
                for _, point_row in row_group.iterrows():
                    point_number = int(point_row["Номер точки"])
                    point_column = headers.get(f"Т{point_number}")
                    if point_column is None:
                        continue
                    average_value = point_row["Среднее SKU за день продажи"]
                    if pd.notna(average_value):
                        sheet.cell(average_row, point_column).value = round(float(average_value), 2)
                        average_total += float(average_value)
                plan_column = headers.get("ПЛАН")
                if plan_column is not None:
                    sheet.cell(average_row, plan_column).value = round(average_total, 2)

    # Compact the downloaded menu by removing all 'Ф' columns (Ф1...Ф29, ФСПК, Ф.Стол, Ф.Цена, etc.).
    # Blue 'СР' rows stay in place; only their numeric averages can be hidden by the option above.
    for sheet_name, columns in fact_columns_by_sheet.items():
        if sheet_name in workbook.sheetnames:
            _delete_columns_preserving_merges(workbook[sheet_name], columns)

    # Critical export rule: the downloaded plan must contain ONLY the dates selected
    # in «Прогноз плана». The source matrix can contain four weeks and many dates,
    # but none of the unselected blocks are carried into the downloaded workbook.
    workbook = _restrict_forecast_workbook_to_selected_dates(workbook, forecast)

    calculation_name = "Расчёт по меню"
    if calculation_name in workbook.sheetnames:
        del workbook[calculation_name]
    calculation_sheet = workbook.create_sheet(calculation_name)
    calculation_rows: list[dict[str, object]] = []
    calculation_keys = [
        "Дата плана", "День недели", "Лист", "Строка Excel", "Категория", "Сущность",
        "SKU", "Название товара", "Жизненный цикл, дней", "Период среднего, недель",
    ]
    for _, group in forecast.groupby(calculation_keys, dropna=False, sort=False):
        first = group.iloc[0]
        output_row = {key: first[key] for key in calculation_keys}
        total_plan = 0
        for _, point_row in group.sort_values("Номер точки").iterrows():
            point = str(point_row["Точка"])
            output_row[f"Среднее {point}"] = point_row["Среднее SKU за день продажи"]
            output_row[f"Дней продаж {point}"] = point_row["Дней с продажами"]
            output_row[f"Идеальный цикл {point}"] = point_row["Дней покрытия поставкой"]
            output_row[f"План {point}"] = point_row["Рекомендованный план"]
            if pd.notna(point_row["Рекомендованный план"]):
                total_plan += int(point_row["Рекомендованный план"])
        output_row["ВСЕГО"] = total_plan
        calculation_rows.append(output_row)
    calculation_frame = pd.DataFrame(calculation_rows)
    calculation_headers = calculation_frame.columns.tolist()
    calculation_sheet.append(calculation_headers)
    for row in calculation_frame.itertuples(index=False, name=None):
        calculation_sheet.append([
            None if value is pd.NA or (not isinstance(value, str) and pd.isna(value)) else value
            for value in row
        ])
    # Подсвечиваем итоговый план светло-красным, если среднее основано только
    # на 1–2 днях продаж за выбранный пользователем период. Сам расчёт при этом не меняется.
    calculation_header_map = {
        str(cell.value): cell.column for cell in calculation_sheet[1] if cell.value is not None
    }
    for row_number in range(2, calculation_sheet.max_row + 1):
        for point_number in range(1, 30):
            if point_number == 11:
                continue
            days_col = calculation_header_map.get(f"Дней продаж Т{point_number}")
            plan_col = calculation_header_map.get(f"План Т{point_number}")
            if days_col is None or plan_col is None:
                continue
            days_value = calculation_sheet.cell(row=row_number, column=days_col).value
            if days_value in (1, 2):
                calculation_sheet.cell(row=row_number, column=plan_col).fill = PatternFill(
                    "solid", fgColor="F4CCCC"
                )

    calculation_sheet.freeze_panes = "A2"
    calculation_sheet.auto_filter.ref = calculation_sheet.dimensions
    calculation_sheet.sheet_properties.tabColor = "70AD47"
    for cells in calculation_sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 32)
        calculation_sheet.column_dimensions[cells[0].column_letter].width = width

    detail_name = "Обоснование прогноза"
    if detail_name in workbook.sheetnames:
        del workbook[detail_name]
    detail_sheet = workbook.create_sheet(detail_name)
    preferred_export_columns = [
        "Дата плана",
        "День недели",
        "Точка",
        "Тип точки",
        "Номер магазина",
        "Категория",
        "Сущность",
        "SKU",
        "Название товара",
        "Период среднего, недель",
        "Продано за выбранный период, шт.",
        "Дней с продажами",
        "Среднее SKU за день продажи",
        "Жизненный цикл, дней",
        "Дней покрытия поставкой",
        "Идеальный жизненный цикл для плана, дней",
        "Правило расчёта",
        "Расчётная потребность",
        "Рекомендованный план",
        "Статус прогноза",
        "Причина",
        "Бюджет категории",
        "Основа бюджета категории",
        "План аналитика на выбранную дату",
        "Полный прогноз категории",
        "Покрытие категории текущим меню",
        "Метод покрытия",
        "Нераспределённый спрос категории",
        "Спрос категории SQL",
        "Коэффициент аналитика",
        "Медиана прошлых планов категории",
        "Сущность",
        "Доля сущности по планам аналитика",
        "Доля сущности по продажам",
        "SKU",
        "Название товара",
        "Вес распределения",
        "Исторический верхний ориентир",
        "Метод распределения",
        "Историческое среднее",
        "Спрос категории на точке",
        "Преобладающий уровень категории",
        "Коэффициент роста категории",
        "Преобладающее количество SKU",
        "Доля сущности в категории",
        "Рекомендованный план",
        "Статус прогноза",
        "Причина ручной проверки",
        "Аналогичных дней",
        "Уверенность",
        "Метод",
        "Тренд",
        "Основа тренда",
        "Корректировка тренда",
        "Ёмкость категории",
        "Ограничено ёмкостью",
    ]
    export_columns = list(dict.fromkeys(
        column for column in preferred_export_columns if column in forecast.columns
    ))
    detail_sheet.append(export_columns)
    for row in forecast[export_columns].itertuples(index=False, name=None):
        detail_sheet.append([
            None if value is pd.NA or (not isinstance(value, str) and pd.isna(value)) else value
            for value in row
        ])
    detail_header_map = {
        str(cell.value): cell.column for cell in detail_sheet[1] if cell.value is not None
    }
    days_col = detail_header_map.get("Дней с продажами")
    plan_col = detail_header_map.get("Рекомендованный план")
    if days_col is not None and plan_col is not None:
        for row_number in range(2, detail_sheet.max_row + 1):
            if detail_sheet.cell(row=row_number, column=days_col).value in (1, 2):
                detail_sheet.cell(row=row_number, column=plan_col).fill = PatternFill(
                    "solid", fgColor="F4CCCC"
                )

    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    detail_sheet.sheet_properties.tabColor = "5B9BD5"
    for cells in detail_sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 45)
        detail_sheet.column_dimensions[cells[0].column_letter].width = width
    metadata_column = len(export_columns) + 2
    detail_sheet.cell(row=1, column=metadata_column).value = "Период выгрузки плана"
    detail_sheet.cell(row=2, column=metadata_column).value = _forecast_export_date_label(forecast)
    selected_lookback = int(pd.to_numeric(forecast.get("Период среднего, недель"), errors="coerce").dropna().iloc[0]) if "Период среднего, недель" in forecast.columns and not pd.to_numeric(forecast.get("Период среднего, недель"), errors="coerce").dropna().empty else 0
    detail_sheet.cell(row=3, column=metadata_column).value = "Период для расчёта среднего"
    detail_sheet.cell(row=4, column=metadata_column).value = (
        f"{selected_lookback} нед. ({selected_lookback * 7} дн.) перед каждой датой плана"
        if selected_lookback > 0 else "Выбран пользователем"
    )
    detail_sheet.cell(row=5, column=metadata_column).value = "Общий диапазон загруженной истории"
    detail_sheet.cell(row=6, column=metadata_column).value = (
        f"{history_from:%d.%m.%Y}–{(target_date - timedelta(days=1)):%d.%m.%Y}"
    )
    detail_sheet.cell(row=7, column=metadata_column).value = "Легенда"
    detail_sheet.cell(row=8, column=metadata_column).value = (
        "Светло-красный план = продажи SKU на точке были только в 1–2 дня "
        "за выбранный период среднего; значение требует осторожной оценки."
    )
    detail_sheet.cell(row=8, column=metadata_column).fill = PatternFill(
        "solid", fgColor="F4CCCC"
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_sales_time_menu(
    plan_rows: pd.DataFrame,
    sales: pd.DataFrame,
    entities: pd.DataFrame,
    shipment_date: date,
) -> pd.DataFrame:
    """Сопоставляет партию из плана с продажами по дням её срока годности."""
    if plan_rows.empty:
        return pd.DataFrame()
    menu = plan_rows.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    menu["category"] = menu["category"].fillna(menu["matrix_category"])
    menu["category"] = menu["category"].map(normalize_matrix_category)
    menu["entity"] = menu["entity"].fillna("Не сопоставлено")
    menu["analyst_plan"] = pd.to_numeric(menu["analyst_plan"], errors="coerce").fillna(0).clip(lower=0)

    sold = sales.copy()
    if sold.empty:
        sold = pd.DataFrame(columns=["business_date", "sku", "sold_quantity"])
    sold["business_date"] = pd.to_datetime(sold["business_date"], errors="coerce").dt.date
    sold["sku"] = sold["sku"].map(normalize_sku)
    sold["sold_quantity"] = pd.to_numeric(sold["sold_quantity"], errors="coerce").fillna(0).clip(lower=0)
    sold["day_after_shipment"] = sold["business_date"].map(
        lambda value: (value - shipment_date).days if pd.notna(value) else None
    )
    daily = (
        sold[sold["day_after_shipment"].between(1, 7, inclusive="both")]
        .groupby(["sku", "day_after_shipment"], as_index=False)["sold_quantity"].sum()
    )
    daily_map = {
        (row["sku"], int(row["day_after_shipment"])): float(row["sold_quantity"])
        for _, row in daily.iterrows()
    }

    rows: list[dict[str, object]] = []
    for _, item in menu.iterrows():
        shelf_days = product_lifecycle_days(item["category"])
        green_days = product_green_days(item["category"])
        day_sales = [daily_map.get((item["sku"], day_number), 0.0) for day_number in range(1, 8)]
        green_sales = sum(day_sales[:green_days])
        grey_sales = sum(day_sales[green_days:shelf_days])
        rows.append(
            {
                "Категория": item["category"],
                "Сущность": item["entity"],
                "SKU": item["sku"],
                "Название товара": item["product_name"],
                "Отгружено по плану": int(round(float(item["analyst_plan"]))),
                "Срок годности, дней": shelf_days,
                "Зелёный период, дней": green_days,
                "Продано в зелёный период": round(green_sales, 3),
                "Продано в серый период": round(grey_sales, 3),
                "Продано за срок": round(sum(day_sales[:shelf_days]), 3),
                "Расчётный остаток": round(
                    max(0.0, float(item["analyst_plan"]) - sum(day_sales[:shelf_days])), 3
                ),
                "Последний день срока": shipment_date + timedelta(days=shelf_days),
                "Скидка в последний день": "−40%",
                **{f"День {day_number}": round(day_sales[day_number - 1], 3) for day_number in range(1, 8)},
            }
        )
    return pd.DataFrame(rows).sort_values(["Категория", "Сущность", "Название товара"])


def build_sales_time_period(
    allocation_plan_rows: pd.DataFrame,
    sales: pd.DataFrame,
    entities: pd.DataFrame,
    display_start: date,
    display_end: date,
    as_of_date: date | None = None,
) -> pd.DataFrame:
    """Распределяет продажи между партиями SKU по FIFO и списывает только завершённые партии."""
    as_of_date = as_of_date or date.today()
    if allocation_plan_rows.empty:
        return pd.DataFrame()
    batches = allocation_plan_rows.merge(
        entities[["sku", "category", "entity"]], on="sku", how="left"
    )
    batches["category"] = batches["category"].fillna(batches["matrix_category"]).map(
        normalize_matrix_category
    )
    batches["entity"] = batches["entity"].fillna("Не сопоставлено")
    batches["analyst_plan"] = pd.to_numeric(
        batches["analyst_plan"], errors="coerce"
    ).fillna(0).clip(lower=0)
    batches["unit_price"] = pd.to_numeric(
        batches.get("unit_price", 0), errors="coerce"
    ).fillna(0).clip(lower=0)
    batches = batches.reset_index(drop=True)
    batches["batch_id"] = batches.index
    batches["shelf_days"] = batches["category"].map(product_lifecycle_days)
    batches["expiry_date"] = batches.apply(
        lambda row: row["plan_date"] + timedelta(days=int(row["shelf_days"])), axis=1
    )
    remaining = batches.set_index("batch_id")["analyst_plan"].astype(float).to_dict()
    allocated: dict[tuple[int, int], float] = {}
    allocated_revenue: dict[tuple[int, int], float] = {}
    allocated_elapsed_hours: dict[int, float] = {}
    allocated_speed_quantity: dict[int, float] = {}

    sold = sales.copy()
    if sold.empty:
        sold = pd.DataFrame(
            columns=["business_date", "sale_datetime", "sku", "sold_quantity", "revenue"]
        )
    sold["business_date"] = pd.to_datetime(sold["business_date"], errors="coerce").dt.date
    sold["sale_datetime"] = pd.to_datetime(sold.get("sale_datetime"), errors="coerce")
    sold["sku"] = sold["sku"].map(normalize_sku)
    sold["sold_quantity"] = pd.to_numeric(
        sold["sold_quantity"], errors="coerce"
    ).fillna(0).clip(lower=0)
    sold["revenue"] = pd.to_numeric(
        sold.get("revenue", 0), errors="coerce"
    ).fillna(0).clip(lower=0)
    # Будущие календарные даты не участвуют в фактических продажах партии.
    sold = sold[
        sold["business_date"].notna() & (sold["business_date"] <= as_of_date)
    ].copy()
    sold_events = (
        sold.dropna(subset=["sale_datetime"])
        .groupby(["business_date", "sale_datetime", "sku"], as_index=False)
        .agg(sold_quantity=("sold_quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values(["sale_datetime", "sku"])
    )
    for _, sale_row in sold_events.iterrows():
        sale_date = sale_row["business_date"]
        sale_datetime = sale_row["sale_datetime"]
        sale_left = float(sale_row["sold_quantity"])
        revenue_per_unit = (
            float(sale_row["revenue"]) / sale_left if sale_left > 0 else 0.0
        )
        eligible = batches[
            (batches["sku"] == sale_row["sku"])
            & (batches["plan_date"] < sale_date)
            & (batches["expiry_date"] >= sale_date)
        ].sort_values("plan_date")
        for _, batch in eligible.iterrows():
            batch_id = int(batch["batch_id"])
            if sale_left <= 0:
                break
            available = float(remaining.get(batch_id, 0.0))
            if available <= 0:
                continue
            quantity = min(available, sale_left)
            day_number = (sale_date - batch["plan_date"]).days
            allocated[(batch_id, day_number)] = allocated.get((batch_id, day_number), 0.0) + quantity
            allocated_revenue[(batch_id, day_number)] = (
                allocated_revenue.get((batch_id, day_number), 0.0)
                + quantity * revenue_per_unit
            )
            freshness_start = pd.Timestamp(batch["plan_date"] + timedelta(days=1))
            elapsed_hours = max(
                0.0, (pd.Timestamp(sale_datetime) - freshness_start).total_seconds() / 3600
            )
            allocated_elapsed_hours[batch_id] = (
                allocated_elapsed_hours.get(batch_id, 0.0) + quantity * elapsed_hours
            )
            allocated_speed_quantity[batch_id] = (
                allocated_speed_quantity.get(batch_id, 0.0) + quantity
            )
            remaining[batch_id] = available - quantity
            sale_left -= quantity

    output_batches = batches[
        batches["plan_date"].between(display_start, display_end, inclusive="both")
    ]
    rows: list[dict[str, object]] = []
    for _, item in output_batches.iterrows():
        batch_id = int(item["batch_id"])
        shelf_days = int(item["shelf_days"])
        green_days = product_green_days(item["category"])
        day_sales = [allocated.get((batch_id, day_number), 0.0) for day_number in range(1, 8)]
        day_revenue = [
            allocated_revenue.get((batch_id, day_number), 0.0)
            for day_number in range(1, 8)
        ]
        remaining_quantity = max(0.0, remaining.get(batch_id, 0.0))
        expiry_date = item["expiry_date"]
        # Остаток становится списанием только ПОСЛЕ окончания последнего дня срока.
        # Если дата отгрузки ещё не наступила или партия всё ещё в жизненном цикле,
        # списание равно нулю и остаток остаётся живым.
        batch_started = item["plan_date"] <= as_of_date
        batch_expired = batch_started and as_of_date > expiry_date
        writeoff_quantity = remaining_quantity if batch_expired else 0.0
        if not batch_started:
            batch_status = "Плановая дата не наступила"
        elif not batch_expired:
            batch_status = "В жизненном цикле" if remaining_quantity > 0 else "Реализована"
        else:
            batch_status = (
                "Срок завершён · списание" if remaining_quantity > 0
                else "Срок завершён · реализована"
            )
        speed_quantity = allocated_speed_quantity.get(batch_id, 0.0)
        elapsed_hours_weight = allocated_elapsed_hours.get(batch_id, 0.0)
        depletion_hours = elapsed_hours_weight / speed_quantity if speed_quantity > 0 else pd.NA
        rows.append(
            {
                "Дата отгрузки": item["plan_date"],
                "Категория": item["category"],
                "Сущность": item["entity"],
                "SKU": item["sku"],
                "Название товара": item["product_name"],
                "Отгружено по плану": int(round(float(item["analyst_plan"]))),
                "Срок годности, дней": shelf_days,
                "Зелёный период, дней": green_days,
                "Продано в зелёный период": round(sum(day_sales[:green_days]), 3),
                "Продано в серый период": round(sum(day_sales[green_days:shelf_days]), 3),
                "Продано за срок": round(sum(day_sales[:shelf_days]), 3),
                "Расчётный остаток": round(remaining_quantity, 3),
                "Списания": round(writeoff_quantity, 3),
                "Статус партии": batch_status,
                "Выручка SKU, ₽": round(sum(day_revenue[:shelf_days]), 2),
                "Убыток от списания, ₽": round(
                    writeoff_quantity * float(item.get("unit_price", 0.0)), 2
                ),
                "Скорость выбытия, ч": (
                    round(float(depletion_hours), 1) if pd.notna(depletion_hours) else pd.NA
                ),
                "Вес часов выбытия": elapsed_hours_weight,
                "Количество для скорости": speed_quantity,
                "Последний день срока": item["expiry_date"],
                "Скидка в последний день": "−40%",
                **{f"День {day_number}": round(day_sales[day_number - 1], 3) for day_number in range(1, 8)},
            }
        )
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(
        ["Дата отгрузки", "Категория", "Сущность", "Название товара"]
    )



FRESHNESS_POINT_DAY_COLUMNS = [f"День {day_number}" for day_number in range(1, 8)]


def prepare_freshness_point_menu_view(
    point_rows: pd.DataFrame,
    as_of_date: date | None = None,
) -> pd.DataFrame:
    """Готовит меню точки с фактом продаж по каждому дню окна свежести."""
    as_of_date = as_of_date or date.today()
    if point_rows is None or point_rows.empty:
        return pd.DataFrame()

    work = point_rows.copy()
    work["Отгружено по плану"] = pd.to_numeric(
        work.get("Отгружено по плану"), errors="coerce"
    ).fillna(0.0)
    # Меню точки — только позиции, которым в матрице реально задан план > 0.
    work = work[work["Отгружено по плану"] > 0].copy()
    if work.empty:
        return pd.DataFrame()

    for column in FRESHNESS_POINT_DAY_COLUMNS:
        work[column] = pd.to_numeric(work.get(column), errors="coerce").fillna(0.0)

    work["Продано в свежесть, шт."] = pd.to_numeric(
        work.get("Продано в зелёный период"), errors="coerce"
    ).fillna(0.0)
    work["Продано за срок, шт."] = pd.to_numeric(
        work.get("Продано за срок"), errors="coerce"
    ).fillna(0.0)
    work["Остаток, шт."] = pd.to_numeric(
        work.get("Расчётный остаток"), errors="coerce"
    ).fillna(0.0)
    work["Дата меню"] = pd.to_datetime(work.get("Дата отгрузки"), errors="coerce")

    # Текстовый экран: 0 остаётся нулём, будущие дни и дни за пределами срока — «—».
    display = work.copy()
    for column in FRESHNESS_POINT_DAY_COLUMNS:
        display[column] = display[column].map(
            lambda value: f"{float(value):,.0f}".replace(",", " ")
        ).astype(object)

    for row_index, row in display.iterrows():
        category = row.get("Категория", "")
        shelf_days = product_lifecycle_days(category)
        menu_date_ts = pd.to_datetime(row.get("Дата меню"), errors="coerce")
        for day_number, column in enumerate(FRESHNESS_POINT_DAY_COLUMNS, start=1):
            if day_number > shelf_days:
                display.at[row_index, column] = "—"
                continue
            if pd.notna(menu_date_ts):
                fact_date = menu_date_ts.date() + timedelta(days=day_number)
                if fact_date > as_of_date:
                    display.at[row_index, column] = "—"

    display = display.rename(
        columns={
            "Отгружено по плану": "План, шт.",
            "Срок годности, дней": "Окно, дней",
        }
    )
    columns = [
        "Точка", "Дата меню", "Категория", "Сущность", "SKU", "Название товара",
        "План, шт.", "Окно, дней", *FRESHNESS_POINT_DAY_COLUMNS,
        "Продано в свежесть, шт.", "Продано за срок, шт.", "Остаток, шт.", "Статус партии",
    ]
    columns = [column for column in columns if column in display.columns]
    display = display[columns].copy()
    display["_point_sort"] = pd.to_numeric(
        display.get("Точка", pd.Series(dtype=str)).astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    display["_category_sort"] = display.get("Категория", pd.Series(dtype=str)).map(
        {
            "Завтраки": 0, "Салаты": 1, "Супы": 2, "Вторые блюда": 3,
            "Сэндвичи": 4, "Япония": 5, "Десерты": 6, "Напитки": 7, "Хлеб": 8,
        }
    ).fillna(99)
    return (
        display.sort_values(
            ["_point_sort", "Дата меню", "_category_sort", "Категория", "Название товара", "SKU"],
            kind="stable",
        )
        .drop(columns=["_point_sort", "_category_sort"])
        .reset_index(drop=True)
    )


def build_freshness_point_menu_excel(
    point_rows: pd.DataFrame,
    period_start: date,
    period_end: date,
) -> bytes:
    """Excel: отдельный лист на каждую точку, план + продажи по дням свежести."""
    from openpyxl import Workbook

    view = prepare_freshness_point_menu_view(point_rows, as_of_date=date.today())
    if view.empty:
        raise ValueError("Нет строк меню с планом больше нуля для выбранных точек и периода.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E3D")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    green_positive_fill = PatternFill("solid", fgColor="B6D7A8")
    grey_fill = PatternFill("solid", fgColor="E7E6E6")
    outside_fill = PatternFill("solid", fgColor="F2F2F2")
    imported_font = Font(color="008000")
    title_font = Font(bold=True, size=12)

    ordered_points = sorted(
        view["Точка"].dropna().astype(str).unique(),
        key=lambda label: int(re.search(r"\d+", label).group()) if re.search(r"\d+", label) else 999,
    )
    for point_label in ordered_points:
        point_view = view[view["Точка"].astype(str).eq(point_label)].copy()
        sheet = workbook.create_sheet(title=str(point_label)[:31])
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = f"Окно свежести · {point_label}"
        sheet["A1"].font = title_font
        sheet["A2"] = f"Период формирования меню: {period_start:%d.%m.%Y}–{period_end:%d.%m.%Y}"
        sheet["A3"] = (
            "День 1 начинается на следующий день после даты меню. "
            "Зелёный = основной период свежести; серый = завершающий срок."
        )
        header_row = 5
        export_columns = [column for column in point_view.columns if column != "Точка"]
        for col_idx, column in enumerate(export_columns, start=1):
            cell = sheet.cell(header_row, col_idx, column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for out_row, (_, row) in enumerate(point_view.iterrows(), start=header_row + 1):
            category = row.get("Категория", "")
            green_days = product_green_days(category)
            shelf_days = product_lifecycle_days(category)
            for col_idx, column in enumerate(export_columns, start=1):
                value = row.get(column)
                cell = sheet.cell(out_row, col_idx)
                if column == "Дата меню":
                    parsed = pd.to_datetime(value, errors="coerce")
                    cell.value = parsed.to_pydatetime() if pd.notna(parsed) else None
                    cell.number_format = "DD.MM.YYYY"
                elif column in FRESHNESS_POINT_DAY_COLUMNS and value == "—":
                    cell.value = "—"
                else:
                    # Экранные строки Day N возвращаем в число для удобной выгрузки.
                    if column in FRESHNESS_POINT_DAY_COLUMNS:
                        numeric_value = pd.to_numeric(str(value).replace(" ", ""), errors="coerce")
                        cell.value = 0 if pd.isna(numeric_value) else float(numeric_value)
                        cell.number_format = "0"
                    else:
                        cell.value = value
                cell.font = imported_font
                if column in FRESHNESS_POINT_DAY_COLUMNS:
                    day_number = int(column.split()[-1])
                    if day_number <= green_days:
                        numeric = pd.to_numeric(str(value).replace(" ", ""), errors="coerce")
                        cell.fill = green_positive_fill if pd.notna(numeric) and float(numeric) > 0 else green_fill
                    elif day_number <= shelf_days:
                        cell.fill = grey_fill
                    else:
                        cell.fill = outside_fill
                cell.alignment = Alignment(
                    horizontal="left" if column in {"Категория", "Сущность", "Название товара", "Статус партии"} else "center",
                    vertical="center",
                )

        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(export_columns))}{header_row + len(point_view)}"
        widths = {
            "Дата меню": 13, "Категория": 18, "Сущность": 24, "SKU": 12,
            "Название товара": 38, "План, шт.": 12, "Окно, дней": 12,
            "Продано в свежесть, шт.": 20, "Продано за срок, шт.": 18,
            "Остаток, шт.": 14, "Статус партии": 26,
        }
        for col_idx, column in enumerate(export_columns, start=1):
            width = 10 if column in FRESHNESS_POINT_DAY_COLUMNS else widths.get(column, 14)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.row_dimensions[header_row].height = 32

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def freshness_stage_for_sale(category: object, loading_date: object, sale_date: object) -> str:
    """Возвращает цветовой этап свежести для продажи, привязанной к партии."""
    loading_ts = pd.to_datetime(loading_date, errors="coerce")
    sale_ts = pd.to_datetime(sale_date, errors="coerce")
    if pd.isna(loading_ts) or pd.isna(sale_ts):
        return "Нет партии"
    day_number = (sale_ts.date() - loading_ts.date()).days
    if day_number <= 0:
        return "День отгрузки"
    green_days = product_green_days(str(category))
    shelf_days = product_lifecycle_days(str(category))
    if day_number <= green_days:
        return "Основной период"
    if day_number <= shelf_days:
        return "Завершающий период"
    return "После срока"


def attach_loading_dates_to_sales(
    sales_detail: pd.DataFrame,
    plan_rows: pd.DataFrame,
    as_of_date: date | None = None,
) -> pd.DataFrame:
    """
    Привязывает продажи к плановым партиям FIFO тем же способом, что окно свежести.

    Помимо даты загрузки сохраняет объём плана партии и её итоговый остаток/списание.
    Остаток становится списанием только после полного окончания жизненного цикла партии.
    При переходе одной продажи через границу партий строка делится пропорционально,
    поэтому итоговые количество и выручка сохраняются.
    """
    as_of_date = as_of_date or date.today()
    batch_columns = {
        "loading_date": pd.NaT,
        "freshness_stage": "Нет партии",
        "plan_quantity": pd.NA,
        "batch_expiry_date": pd.NaT,
        "batch_sold_total": pd.NA,
        "batch_live_remaining": pd.NA,
        "batch_writeoff_quantity": pd.NA,
        "batch_status": "Нет партии",
    }
    if sales_detail.empty:
        result = sales_detail.copy()
        for column, default in batch_columns.items():
            result[column] = default
        return result

    source = sales_detail.copy()
    source["sku"] = source["sku"].map(normalize_sku)
    source["business_date"] = pd.to_datetime(source["business_date"], errors="coerce").dt.date
    source["sale_datetime"] = pd.to_datetime(source["sale_datetime"], errors="coerce")
    source["sales"] = pd.to_numeric(source["sales"], errors="coerce").fillna(0.0).clip(lower=0)
    source["revenue"] = pd.to_numeric(source["revenue"], errors="coerce").fillna(0.0).clip(lower=0)
    source["_point_number"] = pd.to_numeric(
        source["point"].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce"
    )

    plans = plan_rows.copy() if plan_rows is not None else pd.DataFrame()
    if plans.empty:
        for column, default in batch_columns.items():
            source[column] = default
        return source.drop(columns=["_point_number"], errors="ignore")

    plans["sku"] = plans["sku"].map(normalize_sku)
    plans["plan_date"] = pd.to_datetime(plans["plan_date"], errors="coerce").dt.date
    plans["point_number"] = pd.to_numeric(plans["point_number"], errors="coerce")
    plans["analyst_plan"] = pd.to_numeric(
        plans["analyst_plan"], errors="coerce"
    ).fillna(0.0).clip(lower=0)
    plans["matrix_category"] = plans["matrix_category"].map(normalize_matrix_category)
    plans = plans[
        plans["plan_date"].notna()
        & plans["point_number"].notna()
        & plans["sku"].notna()
        & (plans["analyst_plan"] > 0)
    ].copy()
    plans["point_number"] = plans["point_number"].astype(int)
    plans = plans.sort_values(["point_number", "sku", "plan_date"]).reset_index(drop=True)

    batch_map: dict[tuple[int, str], list[dict[str, object]]] = {}
    batch_lookup: dict[tuple[int, str, date], dict[str, object]] = {}
    for _, batch in plans.iterrows():
        category = normalize_matrix_category(batch.get("matrix_category"))
        plan_date = batch["plan_date"]
        point_number = int(batch["point_number"])
        sku = str(batch["sku"])
        plan_quantity = float(batch["analyst_plan"])
        batch_key = (point_number, sku, plan_date)
        batch_record = {
            "batch_key": batch_key,
            "plan_date": plan_date,
            "category": category,
            "expiry_date": plan_date + timedelta(days=product_lifecycle_days(category)),
            "plan_quantity": plan_quantity,
            "remaining": plan_quantity,
            "sold_allocated": 0.0,
        }
        batch_map.setdefault((point_number, sku), []).append(batch_record)
        batch_lookup[batch_key] = batch_record

    allocated_rows: list[dict[str, object]] = []
    source = source.sort_values(["sale_datetime", "point", "sku"], kind="stable")
    for _, sale in source.iterrows():
        sale_record = sale.to_dict()
        point_number = sale.get("_point_number")
        sale_date = sale.get("business_date")
        quantity_left = float(sale.get("sales", 0.0))
        revenue_total = float(sale.get("revenue", 0.0))
        revenue_per_unit = revenue_total / quantity_left if quantity_left > 0 else 0.0
        batches_for_sku = (
            batch_map.get((int(point_number), str(sale.get("sku"))), [])
            if pd.notna(point_number) else []
        )

        for batch in batches_for_sku:
            if quantity_left <= 0:
                break
            if sale_date is None or pd.isna(sale_date):
                continue
            if not (batch["plan_date"] < sale_date <= batch["expiry_date"]):
                continue
            available = float(batch["remaining"] or 0.0)
            if available <= 0:
                continue
            allocated_quantity = min(available, quantity_left)
            output = sale_record.copy()
            output["sales"] = allocated_quantity
            output["revenue"] = allocated_quantity * revenue_per_unit
            output["loading_date"] = batch["plan_date"]
            output["freshness_stage"] = freshness_stage_for_sale(
                sale.get("category", batch["category"]), batch["plan_date"], sale_date
            )
            output["_batch_key"] = batch["batch_key"]
            allocated_rows.append(output)
            batch["remaining"] = available - allocated_quantity
            batch["sold_allocated"] = float(batch["sold_allocated"]) + allocated_quantity
            quantity_left -= allocated_quantity

        if quantity_left > 1e-9 or float(sale.get("sales", 0.0)) <= 0:
            output = sale_record.copy()
            output["sales"] = max(0.0, quantity_left)
            output["revenue"] = max(0.0, quantity_left) * revenue_per_unit
            output["loading_date"] = pd.NaT
            output["freshness_stage"] = "Нет партии"
            output["_batch_key"] = None
            allocated_rows.append(output)

    batch_metrics: dict[tuple[int, str, date], dict[str, object]] = {}
    for batch_key, batch in batch_lookup.items():
        plan_date = batch["plan_date"]
        expiry_date = batch["expiry_date"]
        remaining = max(0.0, float(batch["remaining"] or 0.0))
        started = plan_date <= as_of_date
        expired = started and as_of_date > expiry_date
        writeoff = remaining if expired else 0.0
        live_remaining = remaining if not expired else 0.0
        if not started:
            status = "Плановая дата не наступила"
        elif not expired:
            status = "В жизненном цикле" if remaining > 0 else "Реализована"
        else:
            status = (
                "Срок завершён · списание" if remaining > 0
                else "Срок завершён · реализована"
            )
        batch_metrics[batch_key] = {
            "plan_quantity": float(batch["plan_quantity"]),
            "batch_expiry_date": expiry_date,
            "batch_sold_total": float(batch["sold_allocated"]),
            "batch_live_remaining": live_remaining,
            "batch_writeoff_quantity": writeoff,
            "batch_status": status,
        }

    result = pd.DataFrame(allocated_rows)
    if result.empty:
        result = source.copy()
        result["loading_date"] = pd.NaT
        result["freshness_stage"] = "Нет партии"
        result["_batch_key"] = None

    for column in [
        "plan_quantity",
        "batch_expiry_date",
        "batch_sold_total",
        "batch_live_remaining",
        "batch_writeoff_quantity",
        "batch_status",
    ]:
        result[column] = result["_batch_key"].map(
            lambda key: batch_metrics.get(key, {}).get(column, pd.NA) if key is not None else pd.NA
        )
    result["batch_status"] = result["batch_status"].fillna("Нет партии")
    result["batch_expiry_date"] = pd.to_datetime(result["batch_expiry_date"], errors="coerce")
    result = result.drop(columns=["_point_number", "_batch_key"], errors="ignore")
    return result.sort_values(["business_date", "sale_datetime", "point", "sku"], kind="stable")


def build_loading_batch_summary(
    plan_rows: pd.DataFrame,
    sales_with_loading: pd.DataFrame,
    point_labels: list[str] | tuple[str, ...] | set[str] | None = None,
    sku: str | None = None,
    category: str | None = None,
    display_start: date | None = None,
    display_end: date | None = None,
    as_of_date: date | None = None,
) -> pd.DataFrame:
    """Строит поштучную сводку плановых партий для детализации загрузки."""
    as_of_date = as_of_date or date.today()
    plans = plan_rows.copy() if plan_rows is not None else pd.DataFrame()
    if plans.empty:
        return pd.DataFrame()

    plans["sku"] = plans["sku"].map(normalize_sku)
    plans["plan_date"] = pd.to_datetime(plans["plan_date"], errors="coerce").dt.date
    plans["point_number"] = pd.to_numeric(plans["point_number"], errors="coerce")
    plans["analyst_plan"] = pd.to_numeric(plans["analyst_plan"], errors="coerce").fillna(0).clip(lower=0)
    plans["matrix_category"] = plans["matrix_category"].map(normalize_matrix_category)
    plans = plans[
        plans["plan_date"].notna()
        & plans["point_number"].notna()
        & plans["sku"].notna()
        & (plans["analyst_plan"] > 0)
    ].copy()
    if plans.empty:
        return pd.DataFrame()
    plans["point_number"] = plans["point_number"].astype(int)
    plans["point"] = plans["point_number"].map(lambda value: f"Т{int(value)}")

    if point_labels is not None:
        normalized_points = {str(value).strip() for value in point_labels}
        plans = plans[plans["point"].isin(normalized_points)]
    if sku is not None:
        normalized_sku = normalize_sku(sku)
        plans = plans[plans["sku"].eq(normalized_sku)]
    if category is not None:
        normalized_category = normalize_matrix_category(category)
        plans = plans[plans["matrix_category"].eq(normalized_category)]
    if display_start is not None:
        plans = plans[plans["plan_date"] >= display_start]
    if display_end is not None:
        plans = plans[plans["plan_date"] <= display_end]
    if plans.empty:
        return pd.DataFrame()

    sold = sales_with_loading.copy() if sales_with_loading is not None else pd.DataFrame()
    if not sold.empty:
        sold["sku"] = sold["sku"].map(normalize_sku)
        sold["loading_date"] = pd.to_datetime(sold.get("loading_date"), errors="coerce").dt.date
        sold["sales"] = pd.to_numeric(sold.get("sales", 0), errors="coerce").fillna(0).clip(lower=0)
        sold["point"] = sold["point"].astype(str).str.strip()
        sold = sold[sold["loading_date"].notna()].copy()
        sold_by_batch = (
            sold.groupby(["point", "sku", "loading_date"], as_index=False)["sales"].sum()
            .rename(columns={"sales": "sold_from_batch"})
        )
    else:
        sold_by_batch = pd.DataFrame(columns=["point", "sku", "loading_date", "sold_from_batch"])

    result = plans.merge(
        sold_by_batch,
        left_on=["point", "sku", "plan_date"],
        right_on=["point", "sku", "loading_date"],
        how="left",
    )
    result["sold_from_batch"] = pd.to_numeric(result["sold_from_batch"], errors="coerce").fillna(0).clip(lower=0)
    result["remaining_quantity"] = (
        result["analyst_plan"] - result["sold_from_batch"]
    ).clip(lower=0)
    result["expiry_date"] = result.apply(
        lambda row: row["plan_date"] + timedelta(days=product_lifecycle_days(row["matrix_category"])),
        axis=1,
    )
    result["batch_started"] = result["plan_date"].map(lambda value: value <= as_of_date)
    result["batch_expired"] = result.apply(
        lambda row: bool(row["batch_started"] and as_of_date > row["expiry_date"]), axis=1
    )
    result["writeoff_quantity"] = result.apply(
        lambda row: float(row["remaining_quantity"]) if row["batch_expired"] else 0.0, axis=1
    )
    result["live_remaining"] = result.apply(
        lambda row: 0.0 if row["batch_expired"] else float(row["remaining_quantity"]), axis=1
    )

    def status_for_batch(row: pd.Series) -> str:
        if not row["batch_started"]:
            return "Плановая дата не наступила"
        if not row["batch_expired"]:
            return "В жизненном цикле" if row["remaining_quantity"] > 0 else "Реализована"
        return (
            "Срок завершён · списание" if row["remaining_quantity"] > 0
            else "Срок завершён · реализована"
        )

    result["batch_status"] = result.apply(status_for_batch, axis=1)
    return result.sort_values(["plan_date", "point_number", "matrix_category", "product_name", "sku"])

def style_loading_date_rows(row: pd.Series) -> list[str]:
    """Красит только дату загрузки в гамме окна свежести."""
    styles = [""] * len(row)
    if "Дата загрузки" not in row.index:
        return styles
    stage = str(row.get("Статус свежести", ""))
    position = row.index.get_loc("Дата загрузки")
    if stage == "Основной период":
        styles[position] = "background-color: #D9EAD3; color: #274E13; font-weight: 700"
    elif stage == "Завершающий период":
        styles[position] = "background-color: #D9D9D9; color: #333333; font-weight: 700"
    elif stage == "После срока":
        styles[position] = "background-color: #F4CCCC; color: #990000; font-weight: 700"
    elif stage == "Нет партии":
        styles[position] = "background-color: #F2F2F2; color: #777777"
    return styles


def add_freshness_bands_to_depletion_chart(
    figure: go.Figure,
    loading_dates: list[date],
    category: str,
) -> None:
    """Добавляет на почасовой график фоновые интервалы жизненного цикла партий."""
    green_days = product_green_days(category)
    shelf_days = product_lifecycle_days(category)
    for loading_date in sorted(set(loading_dates)):
        green_start = pd.Timestamp(loading_date + timedelta(days=1))
        green_end = pd.Timestamp(loading_date + timedelta(days=green_days + 1))
        shelf_end = pd.Timestamp(loading_date + timedelta(days=shelf_days + 1))
        figure.add_vrect(
            x0=green_start, x1=green_end, fillcolor="#D9EAD3", opacity=0.30,
            line_width=0, layer="below",
        )
        if green_days < shelf_days:
            figure.add_vrect(
                x0=green_end, x1=shelf_end, fillcolor="#D9D9D9", opacity=0.35,
                line_width=0, layer="below",
            )
        if normalize_matrix_category(category) == "Япония":
            # Третий календарный день после загрузки — уже день списания.
            figure.add_vrect(
                x0=shelf_end,
                x1=shelf_end + pd.Timedelta(days=1),
                fillcolor="#F4CCCC",
                opacity=0.30,
                line_width=0,
                layer="below",
            )
        figure.add_vline(
            x=green_start, line_width=1, line_dash="dot", line_color="#6B8E23"
        )
        figure.add_annotation(
            x=green_start, y=1.03, xref="x", yref="paper",
            text=f"Загрузка {loading_date:%d.%m}", showarrow=False,
            font=dict(size=10, color="#444444"), xanchor="left",
        )


def calculate_product_abc(sales: pd.DataFrame, entities: pd.DataFrame) -> pd.DataFrame:
    """Рассчитывает ABC по количеству и выручке отдельно внутри каждой категории."""
    if sales.empty:
        return pd.DataFrame()
    source = sales.copy()
    source["sku"] = source["sku"].map(normalize_sku)
    source["sold_quantity"] = pd.to_numeric(
        source["sold_quantity"], errors="coerce"
    ).fillna(0.0)
    source["revenue"] = pd.to_numeric(source["revenue"], errors="coerce").fillna(0.0)
    source = source.merge(
        entities[["sku", "category"]], on="sku", how="left", validate="many_to_one"
    )
    source["category"] = source["category"].fillna("Не сопоставлено")
    summary = (
        source.groupby(["category", "sku"], as_index=False, dropna=False)
        .agg(
            product_name=("product_name", "last"),
            sold_quantity=("sold_quantity", "sum"),
            revenue=("revenue", "sum"),
        )
    )
    summary["sold_quantity"] = summary["sold_quantity"].clip(lower=0)
    summary["revenue"] = summary["revenue"].clip(lower=0)
    summary = summary[(summary["sold_quantity"] > 0) | (summary["revenue"] > 0)].copy()
    if summary.empty:
        return summary

    quantity_rank = summary.sort_values(
        ["category", "sold_quantity", "sku"], ascending=[True, False, True]
    ).copy()
    quantity_rank["quantity_total_category"] = quantity_rank.groupby("category")[
        "sold_quantity"
    ].transform("sum")
    quantity_rank["quantity_share_category"] = quantity_rank["sold_quantity"].div(
        quantity_rank["quantity_total_category"].replace(0, pd.NA)
    )
    quantity_rank["quantity_cumulative_share"] = quantity_rank.groupby("category")[
        "sold_quantity"
    ].cumsum().div(quantity_rank["quantity_total_category"].replace(0, pd.NA))
    quantity_rank["abc_quantity"] = quantity_rank["quantity_cumulative_share"].map(
        lambda value: "A" if value <= 0.80 else ("B" if value <= 0.95 else "C")
    )
    quantity_rank["quantity_rank"] = quantity_rank.groupby("category").cumcount() + 1

    revenue_rank = summary.sort_values(
        ["category", "revenue", "sku"], ascending=[True, False, True]
    ).copy()
    revenue_rank["revenue_total_category"] = revenue_rank.groupby("category")["revenue"].transform(
        "sum"
    )
    revenue_rank["revenue_share_category"] = revenue_rank["revenue"].div(
        revenue_rank["revenue_total_category"].replace(0, pd.NA)
    )
    revenue_rank["revenue_cumulative_share"] = revenue_rank.groupby("category")[
        "revenue"
    ].cumsum().div(revenue_rank["revenue_total_category"].replace(0, pd.NA))
    revenue_rank["abc_revenue"] = revenue_rank["revenue_cumulative_share"].map(
        lambda value: "A" if value <= 0.80 else ("B" if value <= 0.95 else "C")
    )
    revenue_rank["revenue_rank"] = revenue_rank.groupby("category").cumcount() + 1

    result = quantity_rank.merge(
        revenue_rank[
            [
                "category", "sku", "revenue_share_category", "revenue_cumulative_share",
                "abc_revenue", "revenue_rank",
            ]
        ],
        on=["category", "sku"],
        how="left",
        validate="one_to_one",
    )
    result["abc_matrix"] = result["abc_quantity"] + result["abc_revenue"]
    return result.sort_values(
        ["category", "quantity_rank", "revenue_rank"], ascending=[True, True, True]
    )




def _normalize_product_name_for_classification(value: object) -> str:
    """Нормализация названия товара для осторожной классификации retail-SKU."""
    text = str(value or "").casefold().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    return " ".join(text.split())


def suggest_unmapped_retail_category(product_name: object) -> tuple[str, int, str]:
    """Предлагает только две retail-категории; сомнительные товары остаются на проверку."""
    name = _normalize_product_name_for_classification(product_name)
    if not name:
        return "Не определено", 0, "Пустое название товара"

    # Если в названии явно присутствует готовое блюдо/выпечка, одно слово «шоколад»
    # не должно автоматически превращать позицию в шоколадку.
    prepared_food_terms = (
        "пирог", "пирожн", "торт", "маффин", "кекс", "круассан", "пончик", "мусс",
        "сырник", "каша", "блин", "панкейк", "запеканк", "десерт",
        "вафли с куриц", "вафли с томлен", "вафли с томлён", "венская вафля",
        "вафли сырные", "печень кур", "печень по ",
    )

    soda_brands = (
        "coca cola", "coca-cola", "кока кола", "pepsi", "пепси", "sprite",
        "спрайт", "fanta", "фанта", "evervess", "эвервесс", "schweppes",
        "швепс", "mirinda", "миринда", "7up", "7 up", "добрый cola", "добрый кола",
    )
    soda_high = ("газирован", "газ напит", "газ напиток", "содовая", "soda", "тоник")
    soda_medium = ("лимонад", "дюшес", "тархун", "байкал")
    energy_brands = ("red bull", "ред бул", "adrenaline", "адреналин", "burn", "берн", "энергетик")

    snack_brands = (
        "snickers", "сникерс", "twix", "твикс", "mars", "марс", "bounty", "баунти",
        "kitkat", "kit kat", "киткат", "кит кат", "milky way", "милки вей", "nuts",
        "натс", "picnic", "пикник", "lion", "kinder", "киндер", "ritter sport",
        "альпен гольд", "alpen gold", "lays", "лейс", "pringles", "принглс",
    )
    snack_high = (
        "шоколад", "шоколадка", "батончик", "чипс", "сухарик", "сухари", "снек",
        "попкорн", "крекер",
    )
    snack_medium = ("печенье", "арахис", "семеч", "орех")

    if "негазирован" in name:
        return "Не определено", 10, "Явно указано: негазированный напиток"

    for keyword in soda_brands:
        if keyword.replace("-", " ") in name:
            return "Газированные напитки", 98, f"Бренд/маркер: {keyword}"
    for keyword in soda_high:
        if keyword in name:
            return "Газированные напитки", 92, f"Явный маркер: {keyword}"
    for keyword in soda_medium:
        if keyword in name:
            return "Газированные напитки", 82, f"Тип напитка: {keyword}"
    for keyword in energy_brands:
        if keyword in name:
            return "Газированные напитки", 68, f"Энергетический напиток: {keyword}; проверить"

    if any(term in name for term in prepared_food_terms):
        return "Не определено", 25, "Похоже на готовое блюдо/десерт; не относим к retail-снекам автоматически"

    # Вафли считаем retail-снеком только если есть признак фасованного товара
    # (масса/упаковка) или слово «батончик». Это отсеивает блюда вроде «вафли с курицей».
    if "вафл" in name:
        packaged_waffle = bool(
            re.search(r"\b\d+(?:[.,]\d+)?\s*(?:г|гр|g)\b", name)
            or "батончик" in name
            or "яшкино" in name
        )
        if packaged_waffle:
            return "Шоколад и снеки", 90, "Фасованные вафли/вафельный батончик"
        return "Не определено", 45, "Вафли без явного признака фасованного retail-товара; проверить вручную"

    for keyword in snack_brands:
        if keyword in name:
            return "Шоколад и снеки", 98, f"Бренд/маркер: {keyword}"
    for keyword in snack_high:
        if keyword in name:
            return "Шоколад и снеки", 90, f"Явный маркер: {keyword}"
    for keyword in snack_medium:
        if keyword in name:
            return "Шоколад и снеки", 70, f"Вероятный снек: {keyword}; проверить"

    return "Не определено", 0, "Нет надёжных маркеров в названии"


def build_unmapped_sku_classification(source: pd.DataFrame) -> pd.DataFrame:
    """Агрегирует не сопоставленные SKU и добавляет осторожное предложение категории."""
    if source is None or source.empty or "category" not in source.columns:
        return pd.DataFrame()
    work = source[source["category"].astype(str).eq("Не сопоставлено")].copy()
    if work.empty:
        return pd.DataFrame()
    work["business_date"] = pd.to_datetime(work.get("business_date"), errors="coerce")
    work["sales"] = pd.to_numeric(work.get("sales"), errors="coerce").fillna(0.0)
    work["revenue"] = pd.to_numeric(work.get("revenue"), errors="coerce").fillna(0.0)
    grouped = (
        work.groupby(["sku", "product_name"], as_index=False, dropna=False)
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_points=("point", "nunique"),
            first_sale=("business_date", "min"),
            last_sale=("business_date", "max"),
            sale_rows=("sales", "size"),
        )
    )
    suggestions = grouped["product_name"].map(suggest_unmapped_retail_category)
    grouped["suggested_category"] = suggestions.map(lambda item: item[0])
    grouped["confidence"] = suggestions.map(lambda item: item[1]).astype(int)
    grouped["reason"] = suggestions.map(lambda item: item[2])
    grouped["review_priority"] = grouped.apply(
        lambda row: (
            "Высокий" if row["suggested_category"] != "Не определено" and row["confidence"] >= 90
            else "Средний" if row["suggested_category"] != "Не определено"
            else "Не определено"
        ),
        axis=1,
    )
    grouped = grouped.sort_values(
        ["review_priority", "amount", "quantity"],
        ascending=[True, False, False],
        kind="stable",
    ).reset_index(drop=True)
    return grouped



def build_unmapped_monthly_sales(
    source: pd.DataFrame,
    classification: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Помесячная детализация не сопоставленных SKU для ручной классификации."""
    if source is None or source.empty or "category" not in source.columns:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    work = source[source["category"].astype(str).eq("Не сопоставлено")].copy()
    if work.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    work["business_date"] = pd.to_datetime(work.get("business_date"), errors="coerce")
    work = work[work["business_date"].notna()].copy()
    if work.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    work["sales"] = pd.to_numeric(work.get("sales"), errors="coerce").fillna(0.0)
    work["revenue"] = pd.to_numeric(work.get("revenue"), errors="coerce").fillna(0.0)
    work["month_start"] = work["business_date"].dt.to_period("M").dt.to_timestamp()

    class_cols = [
        "sku", "product_name", "suggested_category", "confidence", "reason", "review_priority"
    ]
    class_lookup = classification[[c for c in class_cols if c in classification.columns]].copy()
    if class_lookup.empty:
        suggestions = work["product_name"].map(suggest_unmapped_retail_category)
        work["suggested_category"] = suggestions.map(lambda item: item[0])
        work["confidence"] = suggestions.map(lambda item: item[1]).astype(int)
        work["reason"] = suggestions.map(lambda item: item[2])
        work["review_priority"] = ""
    else:
        class_lookup = class_lookup.drop_duplicates(["sku", "product_name"], keep="first")
        work = work.merge(class_lookup, on=["sku", "product_name"], how="left")
        missing = work["suggested_category"].isna()
        if missing.any():
            suggestions = work.loc[missing, "product_name"].map(suggest_unmapped_retail_category)
            work.loc[missing, "suggested_category"] = suggestions.map(lambda item: item[0])
            work.loc[missing, "confidence"] = suggestions.map(lambda item: item[1])
            work.loc[missing, "reason"] = suggestions.map(lambda item: item[2])

    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }

    detail = (
        work.groupby(
            ["month_start", "sku", "product_name", "suggested_category", "confidence"],
            as_index=False,
            dropna=False,
        )
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_points=("point", "nunique"),
            first_sale=("business_date", "min"),
            last_sale=("business_date", "max"),
        )
        .sort_values(["month_start", "suggested_category", "amount"], ascending=[True, True, False], kind="stable")
        .reset_index(drop=True)
    )
    detail["month_label"] = detail["month_start"].map(
        lambda value: f"{month_names.get(value.month, value.strftime('%m'))} {value.year}"
    )

    candidates = detail[detail["suggested_category"].ne("Не определено")].copy()
    if candidates.empty:
        return detail, pd.DataFrame(), pd.DataFrame()

    candidate_work = work[work["suggested_category"].ne("Не определено")].copy()
    candidate_work["month_label"] = candidate_work["month_start"].map(
        lambda value: f"{month_names.get(value.month, value.strftime('%m'))} {value.year}"
    )
    summary = (
        candidate_work.groupby(["month_start", "month_label", "suggested_category"], as_index=False)
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_sku=("sku", "nunique"),
            active_points=("point", "nunique"),
        )
        .sort_values(["suggested_category", "month_start"], kind="stable")
        .reset_index(drop=True)
    )
    summary["quantity_mom_pct"] = summary.groupby("suggested_category")["quantity"].pct_change(fill_method=None) * 100
    summary["amount_mom_pct"] = summary.groupby("suggested_category")["amount"].pct_change(fill_method=None) * 100

    wanted = ["Газированные напитки", "Шоколад и снеки"]
    all_months = sorted(candidates["month_start"].dropna().unique().tolist())
    wide_rows = []
    for month_value in all_months:
        month_ts = pd.Timestamp(month_value)
        row = {
            "month_start": month_ts,
            "month_label": f"{month_names.get(month_ts.month, month_ts.strftime('%m'))} {month_ts.year}",
        }
        for category in wanted:
            subset = summary[
                summary["month_start"].eq(month_ts) & summary["suggested_category"].eq(category)
            ]
            row[f"{category} · quantity"] = float(subset["quantity"].sum()) if not subset.empty else 0.0
            row[f"{category} · amount"] = float(subset["amount"].sum()) if not subset.empty else 0.0
        wide_rows.append(row)
    wide = pd.DataFrame(wide_rows)
    return detail, summary, wide


def build_category_point_monthly_report(source: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Готовит второй управленческий срез: категория товаров -> точки -> месяцы."""
    if source is None or source.empty:
        return (pd.DataFrame(),) * 5

    required = {"business_date", "point", "sku", "product_name", "category", "sales", "revenue"}
    if not required.issubset(source.columns):
        return (pd.DataFrame(),) * 5

    work = source.copy()
    work["business_date"] = pd.to_datetime(work["business_date"], errors="coerce")
    work = work[work["business_date"].notna()].copy()
    if work.empty:
        return (pd.DataFrame(),) * 5

    work["point"] = work["point"].astype(str).str.strip()
    work = work[work["point"].ne("Т11")].copy()
    work["sales"] = pd.to_numeric(work["sales"], errors="coerce").fillna(0.0)
    work["revenue"] = pd.to_numeric(work["revenue"], errors="coerce").fillna(0.0)
    work["category_report"] = work["category"].fillna("Не сопоставлено").astype(str)
    work["classification_source"] = "Справочник"
    work["classification_confidence"] = pd.NA

    # Для второго отчёта используем ту же осторожную автоклассификацию retail-SKU,
    # чтобы газировка и шоколад/снеки были отдельны от «Не сопоставлено».
    unmapped_mask = work["category_report"].eq("Не сопоставлено")
    if unmapped_mask.any():
        suggestions = work.loc[unmapped_mask, "product_name"].map(suggest_unmapped_retail_category)
        suggested_category = suggestions.map(lambda item: item[0])
        suggested_confidence = suggestions.map(lambda item: item[1])
        accepted = suggested_category.ne("Не определено")
        accepted_index = suggested_category.index[accepted]
        if len(accepted_index):
            work.loc[accepted_index, "category_report"] = suggested_category.loc[accepted_index]
            work.loc[accepted_index, "classification_source"] = "Автоклассификация retail"
            work.loc[accepted_index, "classification_confidence"] = suggested_confidence.loc[accepted_index].astype(float)

    # В отчёте напитки должны быть визуально разделены.
    work["category_report"] = work["category_report"].replace({
        "Напитки": "Напитки (негазированные)",
    })
    work["month_start"] = work["business_date"].dt.to_period("M").dt.to_timestamp()
    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }
    work["month_label"] = work["month_start"].map(
        lambda value: f"{month_names.get(value.month, value.strftime('%m'))} {value.year}"
    )

    group_cols = ["month_start", "month_label", "point", "category_report"]
    if "shop_number" in work.columns:
        group_cols.insert(3, "shop_number")
    monthly = (
        work.groupby(group_cols, as_index=False, dropna=False)
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_sku=("sku", "nunique"),
            active_days=("business_date", lambda s: pd.Series(s).dt.normalize().nunique()),
        )
    )
    monthly["quantity"] = pd.to_numeric(monthly["quantity"], errors="coerce").fillna(0.0)
    monthly["amount"] = pd.to_numeric(monthly["amount"], errors="coerce").fillna(0.0)
    monthly = monthly.sort_values(["point", "category_report", "month_start"], kind="stable").reset_index(drop=True)
    monthly["quantity_mom_pct"] = monthly.groupby(["point", "category_report"])["quantity"].pct_change(fill_method=None) * 100
    monthly["amount_mom_pct"] = monthly.groupby(["point", "category_report"])["amount"].pct_change(fill_method=None) * 100

    period_group = ["point", "category_report"]
    if "shop_number" in work.columns:
        period_group.insert(1, "shop_number")
    point_category = (
        work.groupby(period_group, as_index=False, dropna=False)
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_sku=("sku", "nunique"),
            active_months=("month_start", "nunique"),
        )
    )
    point_totals = point_category.groupby("point", as_index=False).agg(
        point_quantity=("quantity", "sum"),
        point_amount=("amount", "sum"),
    )
    point_category = point_category.merge(point_totals, on="point", how="left")
    point_category["quantity_share_pct"] = point_category["quantity"].div(point_category["point_quantity"].replace(0, pd.NA)).mul(100)
    point_category["amount_share_pct"] = point_category["amount"].div(point_category["point_amount"].replace(0, pd.NA)).mul(100)
    point_category = point_category.sort_values(["point", "amount"], ascending=[True, False], kind="stable").reset_index(drop=True)

    category_point = point_category.sort_values(["category_report", "amount"], ascending=[True, False], kind="stable").reset_index(drop=True)

    point_month_group = ["month_start", "month_label", "point"]
    if "shop_number" in work.columns:
        point_month_group.append("shop_number")
    point_month = (
        work.groupby(point_month_group, as_index=False, dropna=False)
        .agg(
            quantity=("sales", "sum"),
            amount=("revenue", "sum"),
            active_sku=("sku", "nunique"),
            active_categories=("category_report", "nunique"),
        )
        .sort_values(["month_start", "amount"], ascending=[True, False], kind="stable")
        .reset_index(drop=True)
    )

    retail_audit = work[work["classification_source"].eq("Автоклассификация retail")].copy()
    if retail_audit.empty:
        retail_sku = pd.DataFrame()
    else:
        retail_sku = (
            retail_audit.groupby(
                ["sku", "product_name", "category_report", "classification_confidence"],
                as_index=False,
                dropna=False,
            )
            .agg(
                quantity=("sales", "sum"),
                amount=("revenue", "sum"),
                active_points=("point", "nunique"),
            )
            .sort_values(["category_report", "amount"], ascending=[True, False], kind="stable")
            .reset_index(drop=True)
        )

    return monthly, point_category, category_point, point_month, retail_sku


def category_point_report_bytes(source: pd.DataFrame, period_value: tuple[date, date]) -> bytes:
    """Excel-выгрузка для PDF №2: продажи категорий по точкам."""
    monthly, point_category, category_point, point_month, retail_sku = build_category_point_monthly_report(source)
    buffer = io.BytesIO()

    def rename_common(frame: pd.DataFrame) -> pd.DataFrame:
        return frame.rename(columns={
            "month_label": "Месяц",
            "month_start": "Начало месяца",
            "point": "Точка",
            "shop_number": "Магазин",
            "category_report": "Категория",
            "quantity": "Количество, шт.",
            "amount": "Сумма, ₽",
            "active_sku": "Активных SKU",
            "active_days": "Дней с продажами",
            "active_months": "Активных месяцев",
            "quantity_mom_pct": "Количество к пред. месяцу, %",
            "amount_mom_pct": "Сумма к пред. месяцу, %",
            "point_quantity": "Количество точки, шт.",
            "point_amount": "Сумма точки, ₽",
            "quantity_share_pct": "Доля категории в количестве точки, %",
            "amount_share_pct": "Доля категории в выручке точки, %",
            "active_categories": "Активных категорий",
            "product_name": "Название из PostgreSQL",
            "classification_confidence": "Уверенность, %",
            "active_points": "Активных точек",
            "sku": "SKU",
        })

    monthly_export = rename_common(monthly.copy())
    point_category_export = rename_common(point_category.copy())
    category_point_export = rename_common(category_point.copy())
    point_month_export = rename_common(point_month.copy())
    retail_export = rename_common(retail_sku.copy())

    info = pd.DataFrame([
        {"Параметр": "Период", "Значение": f"{period_value[0]:%d.%m.%Y} - {period_value[1]:%d.%m.%Y}"},
        {"Параметр": "Источник", "Значение": "PostgreSQL · dwh.v_sales_item / данные приложения"},
        {"Параметр": "Разрез", "Значение": "Категория -> точка -> календарный месяц"},
        {"Параметр": "Т11", "Значение": "Исключена"},
        {"Параметр": "Напитки", "Значение": "Напитки (негазированные) и Газированные напитки разделены"},
        {"Параметр": "Retail-категории", "Значение": "Газированные напитки и Шоколад и снеки выделяются из Не сопоставлено по текущей автоклассификации; лист Retail SKU содержит аудит"},
    ])

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        info.to_excel(writer, sheet_name="Описание", index=False)
        monthly_export.to_excel(writer, sheet_name="Месяц точка категория", index=False)
        point_category_export.to_excel(writer, sheet_name="Итого по точкам", index=False)
        category_point_export.to_excel(writer, sheet_name="Точки по категориям", index=False)
        point_month_export.to_excel(writer, sheet_name="Итого точка месяц", index=False)
        if not retail_export.empty:
            retail_export.to_excel(writer, sheet_name="Retail SKU", index=False)

        header_fill = PatternFill("solid", fgColor="17365D")
        header_font = Font(color="FFFFFF", bold=True)
        alt_fill = PatternFill("solid", fgColor="F2F6FA")
        positive_fill = PatternFill("solid", fgColor="E2F0D9")
        negative_fill = PatternFill("solid", fgColor="FCE4D6")
        note_fill = PatternFill("solid", fgColor="FFF2CC")

        for sheet in writer.book.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet.freeze_panes = "A2"
            if sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if sheet.title != "Описание":
                sheet.auto_filter.ref = sheet.dimensions
            for row_idx in range(2, sheet.max_row + 1):
                if row_idx % 2 == 0:
                    for cell in sheet[row_idx]:
                        cell.fill = alt_fill
            for column_cells in sheet.columns:
                values = [str(cell.value or "") for cell in column_cells[: min(len(column_cells), 500)]]
                max_len = min(max([len(v) for v in values] + [8]) + 2, 38)
                sheet.column_dimensions[column_cells[0].column_letter].width = max_len
            headers = {cell.column: str(cell.value or "") for cell in sheet[1]}
            for col_idx, header in headers.items():
                for row_idx in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row_idx, col_idx)
                    if "Сумма" in header and "%" not in header:
                        cell.number_format = '#,##0.00'
                    elif "Количество" in header and "%" not in header:
                        cell.number_format = '#,##0'
                    elif "%" in header:
                        try:
                            value = float(cell.value)
                        except (TypeError, ValueError):
                            continue
                        cell.value = value / 100
                        cell.number_format = '+0.0%;[Red]-0.0%;0.0%'
                        if "пред. месяцу" in header:
                            cell.fill = positive_fill if value > 0 else negative_fill if value < 0 else alt_fill
            if sheet.title == "Описание":
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row_idx, 1).font = Font(bold=True, color="595959")
                    if "Retail" in str(sheet.cell(row_idx, 1).value or "") or "Напитки" in str(sheet.cell(row_idx, 1).value or ""):
                        for cell in sheet[row_idx]:
                            cell.fill = note_fill

    return buffer.getvalue()

def unmapped_classification_report_bytes(
    classification: pd.DataFrame,
    period_value: tuple[date, date],
    source: pd.DataFrame | None = None,
) -> bytes:
    """Excel для проверки кандидатов перед изменением справочника сущностей."""
    buffer = io.BytesIO()
    data = classification.copy()
    candidates = data[data["suggested_category"].ne("Не определено")].copy() if not data.empty else data.copy()
    monthly_detail, monthly_summary, monthly_wide = build_unmapped_monthly_sales(source, data)

    rename_map = {
        "sku": "SKU",
        "product_name": "Название из PostgreSQL",
        "suggested_category": "Предлагаемая категория",
        "confidence": "Уверенность, %",
        "reason": "Почему предложено",
        "review_priority": "Приоритет проверки",
        "quantity": "Продано, шт.",
        "amount": "Выручка, ₽",
        "active_points": "Активных точек",
        "first_sale": "Первая продажа",
        "last_sale": "Последняя продажа",
        "sale_rows": "Строк продаж",
    }
    columns = [
        "sku", "product_name", "suggested_category", "confidence", "reason",
        "review_priority", "quantity", "amount", "active_points", "first_sale", "last_sale", "sale_rows",
    ]
    candidates_export = candidates[[c for c in columns if c in candidates.columns]].rename(columns=rename_map)
    all_export = data[[c for c in columns if c in data.columns]].rename(columns=rename_map)
    # Колонку оставляем пустой специально: пользователь может подтвердить или исправить категорию.
    if not candidates_export.empty:
        candidates_export.insert(3, "Подтверждённая категория", "")
    if not all_export.empty:
        all_export.insert(3, "Подтверждённая категория", "")

    summary = pd.DataFrame(
        [
            {"Показатель": "Период", "Значение": f"{period_value[0]:%d.%m.%Y}–{period_value[1]:%d.%m.%Y}"},
            {"Показатель": "Не сопоставленных SKU", "Значение": int(data["sku"].nunique()) if not data.empty else 0},
            {"Показатель": "Кандидатов в две retail-категории", "Значение": int(len(candidates))},
            {"Показатель": "Газированные напитки · кандидатов", "Значение": int((candidates.get("suggested_category", pd.Series(dtype=str)) == "Газированные напитки").sum())},
            {"Показатель": "Шоколад и снеки · кандидатов", "Значение": int((candidates.get("suggested_category", pd.Series(dtype=str)) == "Шоколад и снеки").sum())},
            {"Показатель": "Выручка кандидатов, ₽", "Значение": float(candidates.get("amount", pd.Series(dtype=float)).sum()) if not candidates.empty else 0.0},
        ]
    )
    instruction = pd.DataFrame(
        {
            "Инструкция": [
                "Отчёт построен только по SKU, которых нет в листе «Справочник + атрибуты» матрицы 2.3.",
                "Предложенная категория формируется по названию товара из PostgreSQL и не применяется автоматически.",
                "Заполните «Подтверждённая категория», если хотите зафиксировать решение вручную; либо просто пришлите файл на проверку.",
                "Низкая/средняя уверенность означает, что название неоднозначно и требует ручного просмотра.",
                "Лист «Продажи по месяцам» показывает каждый SKU отдельно по календарному месяцу; «Сводка по месяцам» готова для переноса в PDF-отчёт.",
                "После подтверждения SKU добавьте его в лист «Справочник + атрибуты» матрицы 2.3 — приложение подхватит категорию и сущность автоматически.",
            ]
        }
    )

    detail_export = monthly_detail.rename(
        columns={
            "month_label": "Месяц",
            "sku": "SKU",
            "product_name": "Название из PostgreSQL",
            "suggested_category": "Предлагаемая категория",
            "confidence": "Уверенность, %",
            "quantity": "Продано, шт.",
            "amount": "Выручка, ₽",
            "active_points": "Активных точек",
            "first_sale": "Первая продажа месяца",
            "last_sale": "Последняя продажа месяца",
        }
    )
    detail_cols = [
        "Месяц", "SKU", "Название из PostgreSQL", "Предлагаемая категория", "Уверенность, %",
        "Продано, шт.", "Выручка, ₽", "Активных точек", "Первая продажа месяца", "Последняя продажа месяца",
    ]
    detail_export = detail_export[[c for c in detail_cols if c in detail_export.columns]]

    summary_export = monthly_summary.rename(
        columns={
            "month_label": "Месяц",
            "suggested_category": "Категория",
            "quantity": "Кол-во, шт.",
            "amount": "Сумма, ₽",
            "quantity_mom_pct": "Кол-во к пред. месяцу, %",
            "amount_mom_pct": "Сумма к пред. месяцу, %",
            "active_sku": "Активных SKU",
            "active_points": "Активных точек",
        }
    )
    summary_cols = [
        "Месяц", "Категория", "Кол-во, шт.", "Сумма, ₽",
        "Кол-во к пред. месяцу, %", "Сумма к пред. месяцу, %", "Активных SKU", "Активных точек",
    ]
    summary_export = summary_export[[c for c in summary_cols if c in summary_export.columns]]

    wide_export = monthly_wide.rename(
        columns={
            "month_label": "Месяц",
            "Газированные напитки · quantity": "Газированные напитки, шт.",
            "Газированные напитки · amount": "Газированные напитки, ₽",
            "Шоколад и снеки · quantity": "Шоколад и снеки, шт.",
            "Шоколад и снеки · amount": "Шоколад и снеки, ₽",
        }
    )
    wide_cols = [
        "Месяц", "Газированные напитки, шт.", "Газированные напитки, ₽",
        "Шоколад и снеки, шт.", "Шоколад и снеки, ₽",
    ]
    wide_export = wide_export[[c for c in wide_cols if c in wide_export.columns]]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Сводка", index=False)
        wide_export.to_excel(writer, sheet_name="Сводка по месяцам", index=False)
        summary_export.to_excel(writer, sheet_name="Динамика категорий", index=False)
        detail_export.to_excel(writer, sheet_name="Продажи по месяцам", index=False)
        candidates_export.to_excel(writer, sheet_name="Кандидаты", index=False)
        all_export.to_excel(writer, sheet_name="Все не сопоставлено", index=False)
        instruction.to_excel(writer, sheet_name="Инструкция", index=False)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        high_fill = PatternFill("solid", fgColor="E2F0D9")
        medium_fill = PatternFill("solid", fgColor="FFF2CC")
        unknown_fill = PatternFill("solid", fgColor="F2F2F2")
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
            if sheet.max_column > 0:
                sheet.auto_filter.ref = sheet.dimensions
            for column_cells in sheet.columns:
                values = [len(str(cell.value or "")) for cell in column_cells]
                width = min(max(values + [8]) + 2, 48)
                sheet.column_dimensions[column_cells[0].column_letter].width = width
            if sheet.title in {"Кандидаты", "Все не сопоставлено", "Продажи по месяцам", "Динамика категорий", "Сводка по месяцам"} and sheet.max_row > 1:
                headers = {str(cell.value): cell.column for cell in sheet[1]}
                priority_col = headers.get("Приоритет проверки")
                amount_col = headers.get("Выручка, ₽") or headers.get("Сумма, ₽")
                qty_col = headers.get("Продано, шт.") or headers.get("Кол-во, шт.")
                confidence_col = headers.get("Уверенность, %")
                first_col = headers.get("Первая продажа") or headers.get("Первая продажа месяца")
                last_col = headers.get("Последняя продажа") or headers.get("Последняя продажа месяца")
                percent_cols = [
                    headers.get("Кол-во к пред. месяцу, %"),
                    headers.get("Сумма к пред. месяцу, %"),
                ]
                ruble_cols = [
                    headers.get("Газированные напитки, ₽"),
                    headers.get("Шоколад и снеки, ₽"),
                ]
                unit_cols = [
                    headers.get("Газированные напитки, шт."),
                    headers.get("Шоколад и снеки, шт."),
                ]
                for row_index in range(2, sheet.max_row + 1):
                    if priority_col:
                        priority = str(sheet.cell(row_index, priority_col).value or "")
                        fill = high_fill if priority == "Высокий" else medium_fill if priority == "Средний" else unknown_fill
                        sheet.cell(row_index, priority_col).fill = fill
                    if amount_col:
                        sheet.cell(row_index, amount_col).number_format = '#,##0.00 [$₽-ru-RU]'
                    if qty_col:
                        sheet.cell(row_index, qty_col).number_format = '#,##0.00'
                    if confidence_col:
                        sheet.cell(row_index, confidence_col).number_format = '0"%"'
                    for col_index in percent_cols:
                        if col_index:
                            sheet.cell(row_index, col_index).number_format = '0.0"%"'
                    for col_index in ruble_cols:
                        if col_index:
                            sheet.cell(row_index, col_index).number_format = '#,##0.00 [$₽-ru-RU]'
                    for col_index in unit_cols:
                        if col_index:
                            sheet.cell(row_index, col_index).number_format = '#,##0.00'
                    if first_col:
                        sheet.cell(row_index, first_col).number_format = 'dd.mm.yyyy'
                    if last_col:
                        sheet.cell(row_index, last_col).number_format = 'dd.mm.yyyy'
    return buffer.getvalue()

def prepare_analysis(
    sales: pd.DataFrame, entities: pd.DataFrame, point_mapping: dict[int, str]
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    sales = sales.copy()
    sales["point"] = sales["shop_number"].astype(int).map(point_mapping)
    sales = sales[sales["point"].notna()].copy()
    merged = sales.merge(entities, on="sku", how="left", validate="many_to_one")
    merged["category"] = merged["category"].fillna("Не сопоставлено")
    merged["entity"] = merged["entity"].fillna("Не сопоставлено")
    # Для найденного SKU название, категория и сущность считаются справочными данными
    # листа «Справочник + атрибуты». PostgreSQL остаётся источником факта продажи.
    if "entity_product_name" in merged.columns:
        sql_names = merged["product_name"].fillna("").astype(str).str.strip()
        matrix_names = merged["entity_product_name"].fillna("").astype(str).str.strip()
        merged["product_name"] = matrix_names.where(matrix_names.ne(""), sql_names)

    daily_detail = (
        merged.groupby(
            ["business_date", "sale_datetime", "point", "shop_number", "sku", "product_name", "category", "entity"],
            dropna=False,
            as_index=False,
        )
        .agg(sales=("sold_quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values(
            ["business_date", "sale_datetime", "point", "category", "entity", "sales"],
            ascending=[True, True, True, True, True, False],
        )
    )

    sku_point = (
        merged.groupby(["point", "shop_number", "sku", "product_name", "category", "entity"], dropna=False, as_index=False)
        .agg(sales=("sold_quantity", "sum"), revenue=("revenue", "sum"))
    )
    point_totals = sku_point.groupby("point", as_index=False).agg(point_sales=("sales", "sum"))
    category_profile = (
        sku_point.groupby(["point", "category"], as_index=False)
        .agg(category_sales=("sales", "sum"), category_revenue=("revenue", "sum"), active_sku=("sku", "nunique"))
        .merge(point_totals, on="point", how="left")
    )
    category_profile["category_share_point"] = category_profile["category_sales"].div(
        category_profile["point_sales"].replace(0, pd.NA)
    )

    entity_profile = (
        sku_point.groupby(["point", "category", "entity"], as_index=False)
        .agg(entity_sales=("sales", "sum"), entity_revenue=("revenue", "sum"), active_sku=("sku", "nunique"))
        .merge(point_totals, on="point", how="left")
    )
    entity_profile["entity_share_point"] = entity_profile["entity_sales"].div(
        entity_profile["point_sales"].replace(0, pd.NA)
    )
    entity_profile["entity_rank_category"] = entity_profile.groupby(["point", "category"])["entity_sales"].rank(
        method="first", ascending=False
    ).astype(int)
    return sku_point, category_profile, entity_profile, daily_detail



def plan_check_minimum(category: object) -> int:
    """Business minimum for one SKU in one point when validating a ready plan."""
    normalized = normalize_matrix_category(category)
    if normalized == "Вторые блюда":
        return 3
    if normalized == "Напитки":
        return 5
    if normalized == "Япония":
        return 1
    return 2


PLAN_CHECK_LIGHT_VEGETABLE_DEFAULT = (
    "овощ", "винегрет", "свекл", "морков", "капуст", "брокколи", "огур",
    "томат", "помидор", "зелень", "редис", "баклаж", "кабач", "тыкв",
)


def _plan_check_trend(recent_value: float, previous_value: float) -> float:
    recent = max(0.0, float(recent_value or 0.0))
    previous = max(0.0, float(previous_value or 0.0))
    if previous <= 0:
        return 1.0 if recent > 0 else 0.0
    return (recent - previous) / previous


def _auto_unit_number(value: object) -> float | None:
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return None if pd.isna(number) else float(number)


@st.cache_data(show_spinner=False)
def parse_auto_unit_points(file_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Read point parameters and daily economics from «Авто Юнит точки ВМ»."""
    if not file_bytes:
        return pd.DataFrame(), pd.DataFrame()
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    profile_rows: list[dict[str, object]] = []
    daily_rows: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        if not re.fullmatch(r"\d+", str(sheet_name).strip()):
            continue
        fallback_point = int(str(sheet_name).strip())
        if not 1 <= fallback_point <= 29:
            continue
        sheet = workbook[sheet_name]
        point_number_value = _auto_unit_number(sheet["B1"].value)
        point_number = int(point_number_value) if point_number_value is not None else fallback_point
        if not 1 <= point_number <= 29:
            continue

        point_type = str(sheet["F1"].value or "").strip()
        active_text = str(sheet["H1"].value or "").strip()
        work_days: float | None = None
        work_hours = ""
        audience_start_row: int | None = None
        audience_label_col: int | None = None

        scan_max_row = min(sheet.max_row, 120)
        scan_max_col = min(sheet.max_column, 20)
        for row_number in range(1, scan_max_row + 1):
            for column_number in range(1, scan_max_col + 1):
                value = str(sheet.cell(row_number, column_number).value or "").strip().casefold()
                if "топ-3 целевой аудитории" in value:
                    audience_start_row = row_number + 2
                    audience_label_col = column_number
                if "график работы" in value:
                    days_value = _auto_unit_number(sheet.cell(row_number + 1, column_number + 1).value)
                    if days_value is not None:
                        work_days = days_value
                    work_hours = str(sheet.cell(row_number + 1, column_number + 2).value or "").strip()

        female_share = 0.0
        male_share = 0.0
        audience_parts: list[str] = []
        if audience_start_row is not None and audience_label_col is not None:
            for row_number in range(audience_start_row, audience_start_row + 3):
                sex = str(sheet.cell(row_number, audience_label_col).value or "").strip()
                age = str(sheet.cell(row_number, audience_label_col + 1).value or "").strip()
                share = _auto_unit_number(sheet.cell(row_number, audience_label_col + 2).value) or 0.0
                sex_low = sex.casefold()
                if "жен" in sex_low:
                    female_share += share
                elif sex_low:
                    # В рабочем файле встречаются опечатки «мудчины»/«муччины».
                    male_share += share
                if sex or age or share:
                    audience_parts.append(
                        f"{sex or '—'} {age or ''} {share * 100:.0f}%".strip()
                    )

        female_dominant = female_share > male_share and female_share > 0
        profile_rows.append(
            {
                "point_number": point_number,
                "point": f"Т{point_number}",
                "point_type": point_type,
                "active": active_text.casefold() in {"да", "yes", "true", "1"},
                "work_days_per_week": int(round(work_days)) if work_days is not None else pd.NA,
                "work_hours": work_hours,
                "female_top3_share": round(female_share, 4),
                "male_top3_share": round(male_share, 4),
                "female_dominant": bool(female_dominant),
                "audience_top3": " · ".join(audience_parts),
            }
        )

        header_row: int | None = None
        header_map: dict[str, int] = {}
        for row_number in range(1, min(sheet.max_row, 12) + 1):
            row_headers = {
                str(sheet.cell(row_number, column_number).value or "").strip(): column_number
                for column_number in range(1, min(sheet.max_column, 15) + 1)
                if sheet.cell(row_number, column_number).value is not None
            }
            if "Дата" in row_headers and ("$ Итог" in row_headers or "Ср.-й чек" in row_headers):
                header_row = row_number
                header_map = row_headers
                break
        if header_row is None:
            continue

        date_col = header_map.get("Дата", 1)
        revenue_col = header_map.get("$ Итог")
        checks_col = header_map.get("Чеки Итог")
        average_check_col = header_map.get("Ср.-й чек")
        writeoff_col = header_map.get("Списания")
        for row_number in range(header_row + 1, sheet.max_row + 1):
            raw_date = sheet.cell(row_number, date_col).value
            parsed_date = parse_excel_date(raw_date)
            if parsed_date is None:
                continue
            revenue = _auto_unit_number(sheet.cell(row_number, revenue_col).value) if revenue_col else None
            checks = _auto_unit_number(sheet.cell(row_number, checks_col).value) if checks_col else None
            avg_check = _auto_unit_number(sheet.cell(row_number, average_check_col).value) if average_check_col else None
            writeoff = _auto_unit_number(sheet.cell(row_number, writeoff_col).value) if writeoff_col else None
            if revenue is None and checks is None and avg_check is None and writeoff is None:
                continue
            daily_rows.append(
                {
                    "point_number": point_number,
                    "business_date": parsed_date,
                    "revenue": max(0.0, revenue or 0.0),
                    "checks": max(0.0, checks or 0.0),
                    "average_check": max(0.0, avg_check or 0.0),
                    "writeoff_rub": max(0.0, writeoff or 0.0),
                }
            )

    profiles = pd.DataFrame(profile_rows)
    if not profiles.empty:
        profiles = profiles.drop_duplicates("point_number", keep="last").sort_values("point_number").reset_index(drop=True)
    daily = pd.DataFrame(daily_rows)
    if not daily.empty:
        daily["business_date"] = pd.to_datetime(daily["business_date"], errors="coerce").dt.date
        daily = daily[daily["business_date"].notna()].sort_values(["point_number", "business_date"]).reset_index(drop=True)
    return profiles, daily


def build_auto_unit_period_metrics(
    daily: pd.DataFrame,
    analysis_date: date,
    window_days: int = 14,
) -> dict[int, dict[str, object]]:
    """Compare recent and previous equal periods for average check and daily revenue."""
    if daily is None or daily.empty:
        return {}
    source = daily.copy()
    source["business_date"] = pd.to_datetime(source["business_date"], errors="coerce").dt.date
    source["point_number"] = pd.to_numeric(source["point_number"], errors="coerce")
    source = source[source["business_date"].notna() & source["point_number"].notna()].copy()
    source["point_number"] = source["point_number"].astype(int)
    result: dict[int, dict[str, object]] = {}

    def summarize(period: pd.DataFrame) -> dict[str, float]:
        if period.empty:
            return {"revenue": 0.0, "checks": 0.0, "avg_check": 0.0, "revenue_day": 0.0, "days": 0.0}
        revenue = float(pd.to_numeric(period["revenue"], errors="coerce").fillna(0).clip(lower=0).sum())
        checks = float(pd.to_numeric(period["checks"], errors="coerce").fillna(0).clip(lower=0).sum())
        trading = period[(pd.to_numeric(period["revenue"], errors="coerce").fillna(0) > 0) | (pd.to_numeric(period["checks"], errors="coerce").fillna(0) > 0)]
        days = int(trading["business_date"].nunique())
        if checks > 0:
            avg_check = revenue / checks
        else:
            avg_check = float(pd.to_numeric(period["average_check"], errors="coerce").dropna().mean() or 0.0)
        return {
            "revenue": revenue,
            "checks": checks,
            "avg_check": avg_check,
            "revenue_day": revenue / max(1, days),
            "days": float(days),
        }

    for point_number, group in source.groupby("point_number"):
        eligible = group[group["business_date"] <= analysis_date].copy()
        if eligible.empty:
            continue
        point_as_of = min(analysis_date, max(eligible["business_date"]))
        recent_start = point_as_of - timedelta(days=window_days - 1)
        previous_end = recent_start - timedelta(days=1)
        previous_start = previous_end - timedelta(days=window_days - 1)
        recent = summarize(eligible[eligible["business_date"].between(recent_start, point_as_of, inclusive="both")])
        previous = summarize(eligible[eligible["business_date"].between(previous_start, previous_end, inclusive="both")])
        result[int(point_number)] = {
            "as_of": point_as_of,
            "recent_start": recent_start,
            "previous_start": previous_start,
            "previous_end": previous_end,
            "avg_check_recent": recent["avg_check"],
            "avg_check_previous": previous["avg_check"],
            "avg_check_trend": _plan_check_trend(recent["avg_check"], previous["avg_check"]),
            "revenue_day_recent": recent["revenue_day"],
            "revenue_day_previous": previous["revenue_day"],
            "revenue_day_trend": _plan_check_trend(recent["revenue_day"], previous["revenue_day"]),
            "recent_days": int(recent["days"]),
            "previous_days": int(previous["days"]),
        }
    return result


def _plan_check_light_vegetable(
    category: object,
    entity: object,
    product_name: object,
    keywords: tuple[str, ...] | list[str],
) -> bool:
    text = " ".join(
        [
            str(normalize_matrix_category(category) or ""),
            str(entity or ""),
            str(product_name or ""),
        ]
    ).casefold().replace("ё", "е")
    for keyword in keywords:
        normalized_keyword = str(keyword or "").strip().casefold().replace("ё", "е")
        if normalized_keyword and normalized_keyword in text:
            return True
    return False


def build_ready_plan_check(
    matrix_plans: pd.DataFrame,
    sales: pd.DataFrame,
    entities: pd.DataFrame,
    target_date: date,
    point_mapping: dict[int, str],
    trend_threshold: float = 0.10,
    writeoff_threshold: float = 0.20,
    minimum_sale_days: int = 3,
    selected_points: tuple[str, ...] | list[str] | None = None,
    selected_categories: tuple[str, ...] | list[str] | None = None,
    auto_unit_profiles: pd.DataFrame | None = None,
    auto_unit_daily: pd.DataFrame | None = None,
    avg_check_growth_target: float = 0.05,
    revenue_growth_target: float = 0.05,
    light_vegetable_keywords: tuple[str, ...] | list[str] = PLAN_CHECK_LIGHT_VEGETABLE_DEFAULT,
) -> pd.DataFrame:
    """Validate an already-filled matrix plan and only adjust the existing values."""
    if matrix_plans is None or matrix_plans.empty:
        return pd.DataFrame()

    plans = matrix_plans.copy()
    plans["plan_date"] = pd.to_datetime(plans["plan_date"], errors="coerce").dt.date
    plans["point_number"] = pd.to_numeric(plans["point_number"], errors="coerce")
    plans["analyst_plan"] = pd.to_numeric(plans["analyst_plan"], errors="coerce").fillna(0.0).clip(lower=0)
    plans["unit_price"] = pd.to_numeric(plans.get("unit_price", 0), errors="coerce").fillna(0.0).clip(lower=0)
    plans["sku"] = plans["sku"].map(normalize_sku)
    plans = plans[plans["plan_date"].notna() & plans["point_number"].notna() & plans["sku"].notna()].copy()
    plans["point_number"] = plans["point_number"].astype(int)
    plans = plans[plans["point_number"].ne(11)].copy()

    plans = plans.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    plans["category"] = plans["category"].fillna(plans["matrix_category"]).map(normalize_matrix_category)
    plans["entity"] = plans["entity"].fillna("Не сопоставлено")

    target = plans[plans["plan_date"].eq(target_date)].copy()
    if target.empty:
        return pd.DataFrame()

    point_to_shop = {
        int(str(label).lstrip("Тт")): int(shop)
        for shop, label in point_mapping.items()
        if str(label).strip().startswith("Т")
        and str(label).strip().lstrip("Тт").isdigit()
        and int(str(label).strip().lstrip("Тт")) != 11
    }
    target["point"] = target["point_number"].map(lambda number: f"Т{int(number)}")
    target["shop_number"] = target["point_number"].map(point_to_shop)
    target["shop_number"] = target["shop_number"].fillna(target["point_number"]).astype(int)

    if selected_points:
        allowed_points = {normalize_point_label(value) for value in selected_points}
        target = target[target["point"].isin(allowed_points)].copy()
    if selected_categories:
        allowed_categories = {normalize_matrix_category(value) for value in selected_categories}
        target = target[target["category"].isin(allowed_categories)].copy()
    if target.empty:
        return pd.DataFrame()

    analysis_date = min(date.today(), target_date - timedelta(days=1))
    two_month_start = (pd.Timestamp(target_date) - pd.DateOffset(months=2)).date()
    trend_start = analysis_date - timedelta(days=13)
    previous_end = analysis_date - timedelta(days=7)
    recent_start = analysis_date - timedelta(days=6)

    sold = sales.copy() if sales is not None else pd.DataFrame()
    if sold.empty:
        sold = pd.DataFrame(columns=["business_date", "shop_number", "sku", "sold_quantity"])
    sold["business_date"] = pd.to_datetime(sold.get("business_date"), errors="coerce").dt.date
    sold["shop_number"] = pd.to_numeric(sold.get("shop_number"), errors="coerce")
    sold["sku"] = sold.get("sku", pd.Series(index=sold.index, dtype=object)).map(normalize_sku)
    sold["sold_quantity"] = pd.to_numeric(sold.get("sold_quantity"), errors="coerce").fillna(0.0).clip(lower=0)
    sold = sold[
        sold["business_date"].notna()
        & sold["shop_number"].notna()
        & sold["sku"].notna()
        & sold["business_date"].between(two_month_start - timedelta(days=7), analysis_date, inclusive="both")
    ].copy()
    sold["shop_number"] = sold["shop_number"].astype(int)

    matrix_sku_category = plans[["sku", "category"]].dropna().drop_duplicates("sku", keep="last").set_index("sku")["category"].to_dict()
    matrix_sku_entity = plans[["sku", "entity"]].dropna().drop_duplicates("sku", keep="last").set_index("sku")["entity"].to_dict()
    sold = sold.merge(entities[["sku", "category", "entity"]], on="sku", how="left")
    sold["category"] = sold["category"].fillna(sold["sku"].map(matrix_sku_category)).map(normalize_matrix_category)
    sold["entity"] = sold["entity"].fillna(sold["sku"].map(matrix_sku_entity)).fillna("Не сопоставлено")

    sku_window = sold[sold["business_date"].between(two_month_start, analysis_date, inclusive="both")].copy()
    sku_daily = (
        sku_window.groupby(["shop_number", "sku", "business_date"], as_index=False)["sold_quantity"].sum()
        if not sku_window.empty
        else pd.DataFrame(columns=["shop_number", "sku", "business_date", "sold_quantity"])
    )
    sku_daily = sku_daily[sku_daily["sold_quantity"] > 0].copy()
    sku_stats: dict[tuple[int, str], dict[str, float]] = {}
    if not sku_daily.empty:
        for (shop_number, sku), group in sku_daily.groupby(["shop_number", "sku"]):
            sku_stats[(int(shop_number), str(sku))] = {
                "sale_days": float(group["business_date"].nunique()),
                "sold_total": float(group["sold_quantity"].sum()),
                "avg_sale_day": float(group["sold_quantity"].sum() / max(1, group["business_date"].nunique())),
            }

    trend_sales = sold[sold["business_date"].between(trend_start, analysis_date, inclusive="both")].copy()

    def build_trend_lookup(group_columns: list[str]) -> dict[tuple, dict[str, float]]:
        lookup: dict[tuple, dict[str, float]] = {}
        if trend_sales.empty:
            return lookup
        grouped = trend_sales.groupby([*group_columns, "business_date"], as_index=False)["sold_quantity"].sum()
        for keys, group in grouped.groupby(group_columns):
            if not isinstance(keys, tuple):
                keys = (keys,)
            previous_sum = float(group.loc[group["business_date"].between(trend_start, previous_end, inclusive="both"), "sold_quantity"].sum())
            recent_sum = float(group.loc[group["business_date"].between(recent_start, analysis_date, inclusive="both"), "sold_quantity"].sum())
            total_14 = previous_sum + recent_sum
            lookup[tuple(keys)] = {
                "previous_avg": previous_sum / 7.0,
                "recent_avg": recent_sum / 7.0,
                "avg_14": total_14 / 14.0,
                "trend": _plan_check_trend(recent_sum / 7.0, previous_sum / 7.0),
            }
        return lookup

    sku_trends = build_trend_lookup(["shop_number", "sku"])
    category_trends = build_trend_lookup(["shop_number", "category"])
    entity_trends = build_trend_lookup(["shop_number", "category", "entity"])

    freshness_lookup: dict[tuple[int, str], dict[str, float | int | str]] = {}
    point_numbers = sorted(target["point_number"].unique().tolist())
    freshness_start = two_month_start - timedelta(days=7)
    historical_plans = plans[plans["plan_date"].between(freshness_start, analysis_date, inclusive="both")].copy()
    for point_number in point_numbers:
        shop_number = int(point_to_shop.get(int(point_number), int(point_number)))
        point_plans = historical_plans[historical_plans["point_number"].eq(int(point_number))].copy()
        if point_plans.empty:
            continue
        point_sales = sold[sold["shop_number"].eq(shop_number)].copy()
        try:
            freshness = build_sales_time_period(point_plans, point_sales, entities, two_month_start, analysis_date, as_of_date=analysis_date)
        except Exception:
            freshness = pd.DataFrame()
        if freshness.empty:
            continue
        freshness["Дата отгрузки"] = pd.to_datetime(freshness["Дата отгрузки"], errors="coerce").dt.date
        for sku, group in freshness.groupby("SKU", dropna=False):
            sku_key = normalize_sku(sku)
            if sku_key is None:
                continue
            completed_mask = group["Статус партии"].astype(str).str.startswith("Срок завершён") | group["Статус партии"].astype(str).eq("Реализована")
            completed = group[completed_mask].copy()
            expired = group[group["Статус партии"].astype(str).str.startswith("Срок завершён")].copy()
            green_avg = float(pd.to_numeric(completed.get("Продано в зелёный период"), errors="coerce").fillna(0).mean()) if not completed.empty else 0.0
            grey_avg = float(pd.to_numeric(completed.get("Продано в серый период"), errors="coerce").fillna(0).mean()) if not completed.empty else 0.0
            expired_plan = float(pd.to_numeric(expired.get("Отгружено по плану"), errors="coerce").fillna(0).sum()) if not expired.empty else 0.0
            expired_writeoff = float(pd.to_numeric(expired.get("Списания"), errors="coerce").fillna(0).sum()) if not expired.empty else 0.0
            writeoff_rate = expired_writeoff / expired_plan if expired_plan > 0 else 0.0
            green_remaining = 0
            live_green_sales = 0.0
            live = group[group["Статус партии"].astype(str).eq("В жизненном цикле")].sort_values("Дата отгрузки")
            if not live.empty:
                latest = live.iloc[-1]
                shipment_date = latest["Дата отгрузки"]
                green_days = int(pd.to_numeric(pd.Series([latest.get("Зелёный период, дней")]), errors="coerce").fillna(0).iloc[0])
                if pd.notna(shipment_date):
                    elapsed = max(0, (analysis_date - shipment_date).days)
                    green_remaining = max(0, green_days - elapsed)
                live_green_sales = float(pd.to_numeric(pd.Series([latest.get("Продано в зелёный период")]), errors="coerce").fillna(0).iloc[0])
            freshness_lookup[(int(point_number), sku_key)] = {
                "completed_batches": int(len(completed)),
                "green_avg": green_avg,
                "grey_avg": grey_avg,
                "writeoff_rate": writeoff_rate,
                "writeoff_qty": expired_writeoff,
                "green_remaining": int(green_remaining),
                "live_green_sales": live_green_sales,
            }

    profile_lookup: dict[int, dict[str, object]] = {}
    if auto_unit_profiles is not None and not auto_unit_profiles.empty:
        for _, profile in auto_unit_profiles.iterrows():
            point_value = pd.to_numeric(pd.Series([profile.get("point_number")]), errors="coerce").iloc[0]
            if pd.notna(point_value):
                profile_lookup[int(point_value)] = profile.to_dict()
    economics_lookup = build_auto_unit_period_metrics(auto_unit_daily, analysis_date, window_days=14)
    point_price_median = (
        target[target["unit_price"] > 0].groupby("point_number")["unit_price"].median().to_dict()
        if "unit_price" in target.columns else {}
    )

    rows: list[dict[str, object]] = []
    for _, item in target.iterrows():
        point_number = int(item["point_number"])
        shop_number = int(item["shop_number"])
        sku = str(item["sku"])
        category = normalize_matrix_category(item["category"])
        entity = str(item.get("entity") or "Не сопоставлено")
        product_name = str(item.get("product_name") or "")
        unit_price = float(pd.to_numeric(pd.Series([item.get("unit_price", 0)]), errors="coerce").fillna(0).iloc[0])
        current_plan = int(round(float(item["analyst_plan"])))
        nominal_minimum = plan_check_minimum(category)
        nominal_coverage_days = forecast_coverage_days(category)

        profile = profile_lookup.get(point_number, {})
        work_days_value = pd.to_numeric(pd.Series([profile.get("work_days_per_week")]), errors="coerce").iloc[0]
        work_days = int(work_days_value) if pd.notna(work_days_value) else 0
        five_day_thursday = work_days == 5 and target_date.weekday() == 3
        minimum_applies = not five_day_thursday
        effective_minimum = nominal_minimum if minimum_applies else 0
        effective_coverage_days = 1 if five_day_thursday else nominal_coverage_days
        female_dominant = bool(profile.get("female_dominant", False))
        female_share = float(profile.get("female_top3_share", 0.0) or 0.0)
        male_share = float(profile.get("male_top3_share", 0.0) or 0.0)
        light_vegetable = _plan_check_light_vegetable(category, entity, product_name, light_vegetable_keywords)

        stats = sku_stats.get((shop_number, sku), {})
        sale_days = int(stats.get("sale_days", 0))
        avg_sale_day = float(stats.get("avg_sale_day", 0.0))
        sku_orientation = (
            max(effective_minimum, int(math.ceil(avg_sale_day * effective_coverage_days)))
            if avg_sale_day > 0 else effective_minimum
        )

        sku_trend = sku_trends.get((shop_number, sku), {"previous_avg": 0.0, "recent_avg": 0.0, "avg_14": 0.0, "trend": 0.0})
        cat_trend = category_trends.get((shop_number, category), {"previous_avg": 0.0, "recent_avg": 0.0, "avg_14": 0.0, "trend": 0.0})
        ent_trend = entity_trends.get((shop_number, category, entity), {"previous_avg": 0.0, "recent_avg": 0.0, "avg_14": 0.0, "trend": 0.0})
        fresh = freshness_lookup.get((point_number, sku), {
            "completed_batches": 0, "green_avg": 0.0, "grey_avg": 0.0,
            "writeoff_rate": 0.0, "writeoff_qty": 0.0, "green_remaining": 0,
            "live_green_sales": 0.0,
        })
        economics = economics_lookup.get(point_number, {})
        avg_check_trend = float(economics.get("avg_check_trend", 0.0) or 0.0)
        revenue_day_trend = float(economics.get("revenue_day_trend", 0.0) or 0.0)
        avg_check_below_goal = bool(economics) and avg_check_trend < float(avg_check_growth_target)
        revenue_below_goal = bool(economics) and revenue_day_trend < float(revenue_growth_target)
        price_median = float(point_price_median.get(point_number, 0.0) or 0.0)
        high_value_item = unit_price > 0 and price_median > 0 and unit_price >= price_median

        # Важно: обновление идёт ОТ уже заполненного значения категорийщика.
        recommended = current_plan
        reasons: list[str] = []
        checks: list[str] = []
        if minimum_applies and current_plan < nominal_minimum:
            recommended = nominal_minimum
            reasons.append(f"минимум категории {nominal_minimum} шт.")
        elif five_day_thursday:
            checks.append("Авто Юнит: 5-дневная точка, четверг = загрузка 1 дня; нижний минимум категории отключён")

        sparse_sku = sale_days < int(minimum_sale_days)
        if sparse_sku:
            checks.append(f"SKU: только {sale_days} дн. продаж — опора на категорию/сущность")

        if int(fresh["green_remaining"]) == 1 and float(fresh["live_green_sales"]) > 0:
            recommended += 1
            reasons.append("+1: живая партия продаётся в зелёный период, остался 1 зелёный день")

        if float(ent_trend["trend"]) >= float(trend_threshold):
            if float(cat_trend["trend"]) > -float(trend_threshold):
                recommended += 1
                reasons.append(f"+1: сущность растёт {float(ent_trend['trend']) * 100:+.1f}% за 7д к предыдущим 7д")
            else:
                checks.append("рост сущности не подтверждён категорией")

        # Авто Юнит: женская аудитория -> поддержка лёгких овощных блюд, но только без сильного встречного риска.
        if female_dominant and light_vegetable:
            if (
                float(fresh["writeoff_rate"]) < float(writeoff_threshold)
                and float(ent_trend["trend"]) > -float(trend_threshold)
                and (avg_sale_day > 0 or float(ent_trend["recent_avg"]) > 0)
            ):
                recommended += 1
                reasons.append(
                    f"+1: Авто Юнит — в ТОП-3 аудитории преобладают женщины ({female_share * 100:.0f}% против {male_share * 100:.0f}%), лёгкое овощное блюдо"
                )
            else:
                checks.append("женская аудитория поддерживает лёгкое блюдо, но падение/списания не дают автоматически увеличить")

        if (
            not sparse_sku
            and float(sku_trend["trend"]) <= -float(trend_threshold)
            and float(fresh["writeoff_rate"]) >= float(writeoff_threshold)
        ):
            recommended -= 1
            reasons.append(f"−1: SKU падает {float(sku_trend['trend']) * 100:+.1f}% и списание {float(fresh['writeoff_rate']) * 100:.1f}%")

        green_avg = float(fresh["green_avg"])
        grey_avg = float(fresh["grey_avg"])
        if int(fresh["completed_batches"]) > 0 and grey_avg > green_avg and green_avg > 0:
            if five_day_thursday:
                green_anchor_raw = green_avg / max(1, product_green_days(category))
            else:
                green_anchor_raw = green_avg
            green_anchor = max(effective_minimum, int(math.ceil(green_anchor_raw)))
            if recommended > green_anchor:
                recommended = green_anchor
                reasons.append(f"ориентир по зелёным продажам: серый период преобладает ({grey_avg:.1f} > {green_avg:.1f})")

        if float(ent_trend["trend"]) <= -float(trend_threshold):
            recommended -= 2
            reasons.append(f"−2: сущность падает {float(ent_trend['trend']) * 100:+.1f}% за 7д к предыдущим 7д")

        recommended = max(effective_minimum, int(round(recommended)))
        delta = int(recommended - current_plan)
        if delta > 0:
            action = f"Добавить +{delta}"
        elif delta < 0:
            action = f"Убавить {delta}"
        else:
            action = "Оставить"

        if current_plan >= sku_orientation + 2:
            checks.append(f"текущий план выше SKU-ориентира {sku_orientation} шт.")
        elif current_plan + 2 <= sku_orientation:
            checks.append(f"текущий план ниже SKU-ориентира {sku_orientation} шт.")

        economic_recommendation = "Нет данных Авто Юнит за сравнимые периоды"
        if economics:
            avg_check_recent = float(economics.get("avg_check_recent", 0.0) or 0.0)
            avg_check_previous = float(economics.get("avg_check_previous", 0.0) or 0.0)
            revenue_recent = float(economics.get("revenue_day_recent", 0.0) or 0.0)
            revenue_previous = float(economics.get("revenue_day_previous", 0.0) or 0.0)
            gaps: list[str] = []
            if avg_check_below_goal:
                gaps.append(f"ср. чек {avg_check_trend * 100:+.1f}% при цели ≥ {avg_check_growth_target * 100:+.1f}%")
            if revenue_below_goal:
                gaps.append(f"выручка/раб.день {revenue_day_trend * 100:+.1f}% при цели ≥ {revenue_growth_target * 100:+.1f}%")
            if gaps:
                if high_value_item and max(float(sku_trend["trend"]), float(ent_trend["trend"]), float(cat_trend["trend"])) >= 0 and float(fresh["writeoff_rate"]) < float(writeoff_threshold):
                    economic_recommendation = "Поддерживать позицию: цена не ниже медианы меню, спрос не падает; может помогать росту среднего чека/дохода"
                else:
                    economic_recommendation = "Цель среднего чека/дохода не достигнута: увеличить стоит прежде всего растущие позиции с ценой не ниже медианы и без высокого списания"
                checks.append("Авто Юнит: " + "; ".join(gaps))
            else:
                economic_recommendation = "Средний чек и выручка/раб.день соответствуют заданной цели роста"
        else:
            avg_check_recent = avg_check_previous = revenue_recent = revenue_previous = 0.0

        strong_conflict = (
            abs(delta) >= 2
            or (
                float(sku_trend["trend"]) <= -float(trend_threshold)
                and float(fresh["writeoff_rate"]) >= float(writeoff_threshold)
                and float(ent_trend["trend"]) <= -float(trend_threshold)
            )
        )
        if strong_conflict:
            agreement = "Красное · полное несогласие"
        elif delta != 0 or sparse_sku or avg_check_below_goal or revenue_below_goal or (female_dominant and light_vegetable):
            agreement = "Жёлтое · сомнение / рекомендация"
        else:
            agreement = "Зелёное · согласие"

        if not reasons:
            reasons.append("сигналов для изменения количества нет")

        rows.append(
            {
                "Дата плана": target_date,
                "День недели": WEEKDAY_RU.get(target_date.weekday(), ""),
                "Точка": item["point"],
                "Номер магазина": shop_number,
                "Категория": category,
                "Сущность": entity,
                "SKU": sku,
                "Название товара": product_name,
                "Цена": round(unit_price, 2),
                "Текущий план": current_plan,
                "Рекомендованный план": recommended,
                "Изменение": delta,
                "Действие": action,
                "Согласие системы": agreement,
                "Минимум номинальный": nominal_minimum,
                "Минимум действует": "Нет · четверг 5-дневной точки" if five_day_thursday else "Да",
                "Дней покрытия проверки": effective_coverage_days,
                "СР SKU / день продаж · 2 мес": round(avg_sale_day, 2),
                "Дней продаж SKU · 2 мес": sale_days,
                "Ориентир SKU × цикл": sku_orientation,
                "Категория СР/день · 14д": round(float(cat_trend["avg_14"]), 2),
                "Категория предыдущие 7д": round(float(cat_trend["previous_avg"]), 2),
                "Категория последние 7д": round(float(cat_trend["recent_avg"]), 2),
                "Тренд категории, %": round(float(cat_trend["trend"]) * 100, 1),
                "Сущность предыдущие 7д": round(float(ent_trend["previous_avg"]), 2),
                "Сущность последние 7д": round(float(ent_trend["recent_avg"]), 2),
                "Тренд сущности, %": round(float(ent_trend["trend"]) * 100, 1),
                "SKU предыдущие 7д": round(float(sku_trend["previous_avg"]), 2),
                "SKU последние 7д": round(float(sku_trend["recent_avg"]), 2),
                "Тренд SKU, %": round(float(sku_trend["trend"]) * 100, 1),
                "Зелёные продажи / партия": round(green_avg, 2),
                "Серые продажи / партия": round(grey_avg, 2),
                "Завершённых партий": int(fresh["completed_batches"]),
                "Списание, %": round(float(fresh["writeoff_rate"]) * 100, 1),
                "Списание, шт.": round(float(fresh["writeoff_qty"]), 2),
                "Осталось зелёных дней": int(fresh["green_remaining"]),
                "Продано в зелёном у живой партии": round(float(fresh["live_green_sales"]), 2),
                "Авто Юнит · дней/нед": work_days if work_days else pd.NA,
                "Авто Юнит · тип точки": str(profile.get("point_type", "")),
                "Женская доля ТОП-3, %": round(female_share * 100, 1),
                "Мужская доля ТОП-3, %": round(male_share * 100, 1),
                "Женская аудитория преобладает": "Да" if female_dominant else "Нет",
                "Лёгкое овощное блюдо": "Да" if light_vegetable else "Нет",
                "Авто Юнит · данные по": economics.get("as_of") if economics else pd.NaT,
                "Ср. чек · пред. 14д": round(avg_check_previous, 2),
                "Ср. чек · посл. 14д": round(avg_check_recent, 2),
                "Тренд ср. чека, %": round(avg_check_trend * 100, 1) if economics else pd.NA,
                "Выручка/раб.день · пред. 14д": round(revenue_previous, 2),
                "Выручка/раб.день · посл. 14д": round(revenue_recent, 2),
                "Тренд выручки/раб.день, %": round(revenue_day_trend * 100, 1) if economics else pd.NA,
                "Медиана цены меню точки": round(price_median, 2),
                "Цена ≥ медианы точки": "Да" if high_value_item else "Нет",
                "Порог тренда, %": round(float(trend_threshold) * 100, 1),
                "Порог большого списания, %": round(float(writeoff_threshold) * 100, 1),
                "Мин. дней SKU для достаточности": int(minimum_sale_days),
                "Цель роста ср. чека, %": round(float(avg_check_growth_target) * 100, 1),
                "Цель роста дохода, %": round(float(revenue_growth_target) * 100, 1),
                "Экономическая рекомендация": economic_recommendation,
                "Причина": "; ".join(reasons),
                "Контроль": "; ".join(checks),
                "Основа при нехватке данных": "Сущность + категория" if sparse_sku else "SKU + свежесть + категория + сущность",
                "Факт по состоянию на": analysis_date,
            }
        )

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    result["_point_sort"] = pd.to_numeric(result["Точка"].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce")
    return result.sort_values(["_point_sort", "Категория", "Сущность", "Название товара"], kind="stable").drop(columns="_point_sort").reset_index(drop=True)


def _ready_plan_check_status_color(status: object) -> tuple[str, str]:
    """Return background/text colors for the system agreement status."""
    value = str(status or "")
    if value.startswith("Красное"):
        return "F4CCCC", "990000"
    if value.startswith("Жёлтое"):
        return "FFF2CC", "7F6000"
    return "D9EAD3", "274E13"


def _ready_plan_check_note(record: dict[str, object]) -> str:
    """Detailed audit trail attached to one ready-plan matrix cell."""
    def num(name: str, digits: int = 1, suffix: str = "") -> str:
        value = pd.to_numeric(pd.Series([record.get(name)]), errors="coerce").iloc[0]
        if pd.isna(value):
            return "—"
        return f"{float(value):.{digits}f}{suffix}"

    def integer(name: str) -> str:
        value = pd.to_numeric(pd.Series([record.get(name)]), errors="coerce").iloc[0]
        if pd.isna(value):
            return "—"
        return str(int(round(float(value))))

    def date_text(value: object) -> str:
        parsed = pd.to_datetime(value, errors="coerce")
        return "—" if pd.isna(parsed) else parsed.strftime("%d.%m.%Y")

    def bullet_lines(text: str) -> list[str]:
        items = [item.strip() for item in str(text or "").split(";") if item.strip()]
        return [f"• {item}" for item in items]

    current_value = pd.to_numeric(pd.Series([record.get("Текущий план")]), errors="coerce").fillna(0).iloc[0]
    recommended_value = pd.to_numeric(pd.Series([record.get("Рекомендованный план")]), errors="coerce").fillna(current_value).iloc[0]
    delta_value = int(round(float(recommended_value - current_value)))
    status = str(record.get("Согласие системы") or "").strip()
    action = str(record.get("Действие") or "Оставить").strip()
    control = str(record.get("Контроль") or "").strip()
    reason = str(record.get("Причина") or "").strip()
    economics = str(record.get("Экономическая рекомендация") or "").strip()
    minimum_text = str(record.get("Минимум действует") or "")
    point_days = pd.to_numeric(pd.Series([record.get("Авто Юнит · дней/нед")]), errors="coerce").iloc[0]
    point_days_text = "—" if pd.isna(point_days) else str(int(float(point_days)))

    target_date = pd.to_datetime(record.get("Дата плана"), errors="coerce")
    fact_date = pd.to_datetime(record.get("Факт по состоянию на"), errors="coerce")
    if pd.notna(target_date):
        sku_start = target_date - pd.DateOffset(months=2)
    else:
        sku_start = pd.NaT
    if pd.notna(fact_date):
        trend_start = fact_date - pd.Timedelta(days=13)
        previous_end = fact_date - pd.Timedelta(days=7)
        recent_start = fact_date - pd.Timedelta(days=6)
    else:
        trend_start = previous_end = recent_start = pd.NaT

    trend_threshold = pd.to_numeric(pd.Series([record.get("Порог тренда, %")]), errors="coerce").iloc[0]
    writeoff_threshold = pd.to_numeric(pd.Series([record.get("Порог большого списания, %")]), errors="coerce").iloc[0]
    min_sale_days = pd.to_numeric(pd.Series([record.get("Мин. дней SKU для достаточности")]), errors="coerce").iloc[0]
    avg_target = pd.to_numeric(pd.Series([record.get("Цель роста ср. чека, %")]), errors="coerce").iloc[0]
    revenue_target = pd.to_numeric(pd.Series([record.get("Цель роста дохода, %")]), errors="coerce").iloc[0]
    trend_threshold_text = "—" if pd.isna(trend_threshold) else f"{float(trend_threshold):.1f}%"
    writeoff_threshold_text = "—" if pd.isna(writeoff_threshold) else f"{float(writeoff_threshold):.1f}%"
    min_sale_days_text = "—" if pd.isna(min_sale_days) else str(int(float(min_sale_days)))
    avg_target_text = "—" if pd.isna(avg_target) else f"{float(avg_target):+.1f}%"
    revenue_target_text = "—" if pd.isna(revenue_target) else f"{float(revenue_target):+.1f}%"

    sale_days_value = pd.to_numeric(pd.Series([record.get("Дней продаж SKU · 2 мес")]), errors="coerce").iloc[0]
    sparse_sku = pd.notna(sale_days_value) and pd.notna(min_sale_days) and float(sale_days_value) < float(min_sale_days)
    sku_trend_value = pd.to_numeric(pd.Series([record.get("Тренд SKU, %")]), errors="coerce").iloc[0]
    entity_trend_value = pd.to_numeric(pd.Series([record.get("Тренд сущности, %")]), errors="coerce").iloc[0]
    category_trend_value = pd.to_numeric(pd.Series([record.get("Тренд категории, %")]), errors="coerce").iloc[0]
    writeoff_value = pd.to_numeric(pd.Series([record.get("Списание, %")]), errors="coerce").iloc[0]
    green_remaining = pd.to_numeric(pd.Series([record.get("Осталось зелёных дней")]), errors="coerce").iloc[0]
    live_green_sales = pd.to_numeric(pd.Series([record.get("Продано в зелёном у живой партии")]), errors="coerce").iloc[0]
    green_avg = pd.to_numeric(pd.Series([record.get("Зелёные продажи / партия")]), errors="coerce").iloc[0]
    grey_avg = pd.to_numeric(pd.Series([record.get("Серые продажи / партия")]), errors="coerce").iloc[0]

    sku_add_rule = pd.notna(green_remaining) and int(green_remaining) == 1 and pd.notna(live_green_sales) and float(live_green_sales) > 0
    entity_growth_rule = (
        pd.notna(entity_trend_value) and pd.notna(trend_threshold)
        and float(entity_trend_value) >= float(trend_threshold)
        and (pd.isna(category_trend_value) or float(category_trend_value) > -float(trend_threshold))
    )
    sku_cut_rule = (
        not sparse_sku
        and pd.notna(sku_trend_value) and pd.notna(trend_threshold)
        and float(sku_trend_value) <= -float(trend_threshold)
        and pd.notna(writeoff_value) and pd.notna(writeoff_threshold)
        and float(writeoff_value) >= float(writeoff_threshold)
    )
    grey_anchor_rule = (
        pd.notna(green_avg) and pd.notna(grey_avg)
        and float(green_avg) > 0 and float(grey_avg) > float(green_avg)
    )
    entity_cut_rule = (
        pd.notna(entity_trend_value) and pd.notna(trend_threshold)
        and float(entity_trend_value) <= -float(trend_threshold)
    )

    if status.startswith("Красное"):
        color_explanation = (
            "КРАСНЫЙ: система существенно не согласна с текущим количеством. "
            "Это ставится, когда итоговая корректировка составляет 2 шт. и более либо одновременно подтверждены "
            "падение SKU, большое списание и падение сущности."
        )
    elif status.startswith("Жёлтое"):
        color_explanation = (
            "ЖЁЛТЫЙ: есть риск или рекомендация к изменению. Это может быть изменение на 1 шт., недостаток истории SKU, "
            "недостижение цели среднего чека/дохода либо дополнительный сигнал Auto Unit."
        )
    else:
        color_explanation = (
            "ЗЕЛЁНЫЙ: система согласна с текущим значением. Существенных сигналов для изменения количества не найдено, "
            "данные не противоречат установленному плану."
        )

    lines = [
        "ПРОВЕРКА ГОТОВОГО МЕНЮ",
        "=" * 34,
        f"Статус: {status}",
        f"Действие: {action}",
        f"Дата плана / отгрузки: {date_text(record.get('Дата плана'))} · {record.get('День недели', '')}",
        f"Факт учтён по: {date_text(record.get('Факт по состоянию на'))}",
        f"Точка: {record.get('Точка', '—')} · магазин {record.get('Номер магазина', '—')}",
        f"Категория: {record.get('Категория', '—')}",
        f"Сущность: {record.get('Сущность', '—')}",
        f"SKU: {record.get('SKU', '—')} · {record.get('Название товара', '—')}",
        "",
        "ИТОГ ПО ЯЧЕЙКЕ",
        f"Категорийщик поставил: {int(round(float(current_value)))} шт.",
        f"Система рекомендует: {int(round(float(recommended_value)))} шт.",
        f"Разница: {delta_value:+d} шт.",
        f"Цена SKU: {num('Цена', 2)}; медиана цены меню этой точки: {num('Медиана цены меню точки', 2)}; цена не ниже медианы: {record.get('Цена ≥ медианы точки', '—')}.",
        "",
        "1. SKU · СРЕДНЕЕ ПО ДНЯМ ФАКТИЧЕСКИХ ПРОДАЖ ЗА 2 МЕСЯЦА",
        f"Период: {date_text(sku_start)}–{date_text(fact_date)}.",
        f"Продажи были в {integer('Дней продаж SKU · 2 мес')} днях; среднее считается только по этим дням и равно {num('СР SKU / день продаж · 2 мес', 2)} шт./день продаж.",
        f"Среднее последних 7 дней: {num('SKU последние 7д', 2)}; предыдущих 7 дней: {num('SKU предыдущие 7д', 2)}; изменение: {num('Тренд SKU, %', 1)}%.",
        f"Порог значимого роста/падения: ±{trend_threshold_text}. Минимум наблюдений для самостоятельной опоры на SKU: {min_sale_days_text} дней.",
        f"Дней покрытия для этой проверки: {integer('Дней покрытия проверки')}; ориентир SKU × цикл: {integer('Ориентир SKU × цикл')} шт.",
        ("Вывод по данным SKU: данных недостаточно — окончательная проверка сильнее опирается на сущность и категорию."
         if sparse_sku else "Вывод по данным SKU: наблюдений достаточно для использования динамики SKU в правилах изменения."),
        "",
        "2. СВЕЖЕСТЬ · ФАКТ ПО ПАРТИЯМ",
        "День отгрузки — День 0 и в свежесть НЕ входит. День 1 начинается на следующий календарный день после даты плана/отгрузки.",
        f"Завершённых партий в анализе: {integer('Завершённых партий')}.",
        f"Средние продажи на одну завершённую партию: зелёный период {num('Зелёные продажи / партия', 2)} шт.; серый период {num('Серые продажи / партия', 2)} шт.",
        f"Списано: {num('Списание, шт.', 2)} шт.; доля списания {num('Списание, %', 1)}%; порог большого списания {writeoff_threshold_text}.",
        f"По последней живой партии продано в зелёном периоде {num('Продано в зелёном у живой партии', 2)} шт.; осталось зелёных дней {integer('Осталось зелёных дней')}.",
        f"Правило +1 «продался в зелёном и остался 1 зелёный день»: {'СРАБОТАЛО' if sku_add_rule else 'не сработало'}.",
        f"Правило ориентира по зелёным продажам при преобладании серого периода: {'СРАБОТАЛО/АКТУАЛЬНО' if grey_anchor_rule else 'не сработало'}.",
        "",
        "3. КАТЕГОРИЯ · 14 ДНЕЙ",
        f"Период сверки: {date_text(trend_start)}–{date_text(fact_date)}.",
        f"Среднее категории за 14 дней: {num('Категория СР/день · 14д', 2)} шт./день.",
        f"Предыдущие 7 дней: {num('Категория предыдущие 7д', 2)} шт./день; последние 7 дней: {num('Категория последние 7д', 2)} шт./день; изменение {num('Тренд категории, %', 1)}%.",
        "Категория используется как контроль: рост сущности не повышает план автоматически, если категория одновременно показывает сильное встречное падение.",
        "",
        "4. СУЩНОСТЬ · ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА",
        f"Предыдущие 7 дней: {num('Сущность предыдущие 7д', 2)} шт./день; последние 7 дней: {num('Сущность последние 7д', 2)} шт./день; изменение {num('Тренд сущности, %', 1)}%.",
        f"Основа расчёта при текущем объёме данных: {record.get('Основа при нехватке данных', '—')}.",
        f"Правило +1 при росте сущности и отсутствии встречного падения категории: {'СРАБОТАЛО' if entity_growth_rule else 'не сработало'}.",
        f"Правило −2 при падении сущности: {'СРАБОТАЛО' if entity_cut_rule else 'не сработало'}.",
        f"Правило −1 при одновременном падении SKU и большом списании: {'СРАБОТАЛО' if sku_cut_rule else 'не сработало'}.",
        "",
        "5. МИНИМАЛЬНОЕ ЗНАЧЕНИЕ И ЦИКЛ",
        f"Номинальный минимум категории: {integer('Минимум номинальный')} шт. (вторые блюда 3; напитки 5; Япония 1; остальные 2).",
        f"Минимум применяется: {minimum_text}.",
        f"График точки по Auto Unit: {point_days_text} дн./нед.; покрытие проверки {integer('Дней покрытия проверки')} дн.",
        ("Особое правило четверга: точка работает 5 дней, поэтому четверг считается загрузкой на 1 день и категорийный минимум отключён."
         if "четверг" in minimum_text.casefold() else "Особое правило четверга для 5-дневной точки к этой ячейке не применяется."),
        "",
        "6. AUTO UNIT · АУДИТОРИЯ ТОЧКИ",
        f"Тип точки: {record.get('Авто Юнит · тип точки', '—') or '—'}.",
        f"Женская доля в ТОП-3 аудитории: {num('Женская доля ТОП-3, %', 1)}%; мужская: {num('Мужская доля ТОП-3, %', 1)}%; женская аудитория преобладает: {record.get('Женская аудитория преобладает', '—')}.",
        f"Позиция распознана как лёгкое овощное блюдо: {record.get('Лёгкое овощное блюдо', '—')}.",
        "Если женская аудитория преобладает, лёгкая овощная позиция получает сигнал +1 только при отсутствии большого списания и сильного падения спроса.",
        "",
        "7. ЭКОНОМИКА ТОЧКИ · СРЕДНИЙ ЧЕК И ДОХОД",
        f"Средний чек: {num('Ср. чек · пред. 14д', 2)} → {num('Ср. чек · посл. 14д', 2)}; изменение {num('Тренд ср. чека, %', 1)}%; цель не ниже {avg_target_text}.",
        f"Выручка на рабочий день: {num('Выручка/раб.день · пред. 14д', 2)} → {num('Выручка/раб.день · посл. 14д', 2)}; изменение {num('Тренд выручки/раб.день, %', 1)}%; цель не ниже {revenue_target_text}.",
        f"Экономическая оценка позиции: {economics or '—'}.",
        "",
        "8. КАКИЕ ПРАВИЛА ФАКТИЧЕСКИ ПОВЛИЯЛИ НА РЕКОМЕНДАЦИЮ",
    ]
    reason_bullets = bullet_lines(reason)
    lines.extend(reason_bullets or ["• Автоматические правила количества не изменили текущее значение."])

    lines.extend(["", "9. ДОПОЛНИТЕЛЬНЫЕ СИГНАЛЫ / ОГРАНИЧЕНИЯ"])
    control_bullets = bullet_lines(control)
    lines.extend(control_bullets or ["• Дополнительных ограничений или сомнений не найдено."])

    lines.extend([
        "",
        "10. ПОЧЕМУ ТАКОЙ ЦВЕТ",
        color_explanation,
        "",
        "ФИНАЛЬНЫЙ ВЫВОД",
        f"Для ячейки {record.get('Точка', '—')} / SKU {record.get('SKU', '—')} система сравнила текущее значение {int(round(float(current_value)))} шт. со всеми доступными уровнями контроля и получила рекомендацию {int(round(float(recommended_value)))} шт. ({delta_value:+d}).",
        "Текущее значение в исходном меню не переписывается автоматически: цвет и эта сноска служат проверкой решения категорийщика.",
    ])
    return "\n".join(lines)

def _ready_plan_check_result_lookup(result: pd.DataFrame) -> dict[tuple[str, str], dict[str, object]]:
    lookup: dict[tuple[str, str], dict[str, object]] = {}
    if result is None or result.empty:
        return lookup
    for _, row in result.iterrows():
        sku = normalize_sku(row.get("SKU"))
        point = normalize_point_label(row.get("Точка"))
        if sku is None or point is None:
            continue
        lookup[(sku, point)] = row.to_dict()
    return lookup


def _ready_plan_check_selected_blocks(file_bytes: bytes, target_date: date):
    """Load only the selected ready-plan day from plan sheets."""
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
    blocks: list[tuple[object, int, int]] = []
    for sheet_name in workbook.sheetnames:
        if not (sheet_name.lower().startswith("план ") and "недел" in sheet_name.lower()):
            continue
        sheet = workbook[sheet_name]
        for start_row, end_row in _find_selected_plan_blocks(sheet, {target_date}):
            # Do not show/export blank tail below the selected menu block.
            while end_row > start_row:
                has_value = any(
                    sheet.cell(end_row, column).value not in (None, "")
                    for column in range(1, min(sheet.max_column, 40) + 1)
                )
                if has_value:
                    break
                end_row -= 1
            blocks.append((sheet, start_row, end_row))
    return workbook, blocks


def ready_plan_check_menu_html(
    file_bytes: bytes,
    result: pd.DataFrame,
    target_date: date,
) -> str:
    """Render the ready plan inside the Check tab in the same grid as the menu."""
    if not file_bytes:
        return ""
    try:
        _, blocks = _ready_plan_check_selected_blocks(file_bytes, target_date)
    except Exception:
        return ""
    if not blocks:
        return ""

    lookup = _ready_plan_check_result_lookup(result)
    parts = [
        "<style>",
        ".plan-check-scroll{overflow-x:auto;max-width:100%;border:1px solid rgba(128,128,128,.35);border-radius:8px;margin:.35rem 0 1rem 0;}",
        ".plan-check-table{border-collapse:collapse;white-space:nowrap;font-size:12px;min-width:max-content;background:transparent;color:inherit;}",
        ".plan-check-table th,.plan-check-table td{border:1px solid rgba(128,128,128,.35);padding:5px 7px;text-align:center;vertical-align:middle;}",
        ".plan-check-table th{font-weight:700;background:rgba(128,128,128,.16);position:sticky;top:0;z-index:2;}",
        ".plan-check-table td:nth-child(5),.plan-check-table th:nth-child(5){text-align:left;min-width:250px;white-space:normal;}",
        ".plan-check-table td:nth-child(4),.plan-check-table th:nth-child(4){text-align:left;min-width:95px;}",
        ".plan-check-sr td{background:#DDEBF7;color:#1F4E78;font-weight:600;}",
        ".plan-check-date{font-weight:700;margin:.75rem 0 .25rem 0;}",
        "</style>",
    ]

    for sheet, start_row, end_row in blocks:
        header_row = start_row + 1
        header_values = [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
        nonempty_header_columns = [i + 1 for i, value in enumerate(header_values) if value not in (None, "")]
        if not nonempty_header_columns:
            continue
        last_column = max(nonempty_header_columns)
        headers = {
            str(sheet.cell(header_row, column).value or "").replace(" ", "").strip(): column
            for column in range(1, last_column + 1)
        }
        code_column = next((column for label, column in headers.items() if label.casefold() in {"код", "код№", "sku", "ску"}), None)
        title_value = " · ".join(
            str(sheet.cell(start_row, column).value).strip()
            for column in range(1, min(last_column, 6) + 1)
            if sheet.cell(start_row, column).value not in (None, "")
        )
        parts.append(f'<div class="plan-check-date">{html.escape(title_value or target_date.strftime("%d.%m.%Y"))}</div>')
        parts.append('<div class="plan-check-scroll"><table class="plan-check-table"><thead><tr>')
        for column in range(1, last_column + 1):
            parts.append(f"<th>{html.escape(str(sheet.cell(header_row, column).value or ''))}</th>")
        parts.append("</tr></thead><tbody>")

        for row_number in range(header_row + 1, end_row + 1):
            row_label = str(sheet.cell(row_number, 1).value or "").strip().upper()
            row_class = ' class="plan-check-sr"' if row_label == "СР" else ""
            sku = normalize_sku(sheet.cell(row_number, code_column).value) if code_column else None
            parts.append(f"<tr{row_class}>")
            for column in range(1, last_column + 1):
                value = sheet.cell(row_number, column).value
                display = "" if value is None else str(value)
                style = ""
                title = ""
                if sku is not None:
                    header = str(sheet.cell(header_row, column).value or "").replace(" ", "").strip()
                    if re.fullmatch(r"Т\d+", header):
                        point = normalize_point_label(header)
                        record = lookup.get((sku, point)) if point else None
                        if record is not None:
                            background, foreground = _ready_plan_check_status_color(record.get("Согласие системы"))
                            style = f' style="background:#{background};color:#{foreground};font-weight:700;cursor:help;"'
                            tooltip = html.escape(_ready_plan_check_note(record), quote=True).replace("\n", "&#10;")
                            title = f' title="{tooltip}"'
                parts.append(f"<td{style}{title}>{html.escape(display)}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table></div>")
    return "".join(parts)


def ready_plan_check_excel(
    matrix_bytes: bytes,
    result: pd.DataFrame,
    target_date: date,
    trend_threshold: float,
    writeoff_threshold: float,
    minimum_sale_days: int,
    avg_check_growth_target: float = 0.05,
    revenue_growth_target: float = 0.05,
    light_vegetable_keywords: tuple[str, ...] | list[str] = PLAN_CHECK_LIGHT_VEGETABLE_DEFAULT,
) -> bytes:
    """Return only the selected menu day, preserving values and annotating T-cells in place."""
    from openpyxl import Workbook

    source_workbook, blocks = _ready_plan_check_selected_blocks(matrix_bytes, target_date)
    export_workbook = Workbook()
    export_workbook.remove(export_workbook.active)
    copied_any = False

    blocks_by_sheet: dict[str, list[tuple[int, int]]] = {}
    for sheet, start_row, end_row in blocks:
        blocks_by_sheet.setdefault(sheet.title, []).append((start_row, end_row))
    for sheet_name, row_ranges in blocks_by_sheet.items():
        source_sheet = source_workbook[sheet_name]
        target_sheet = export_workbook.create_sheet(title=sheet_name)
        _copy_selected_plan_sheet(source_sheet, target_sheet, row_ranges)
        copied_any = True

    if not copied_any:
        raise ValueError("Для выбранной даты не найден блок меню в текущей матрице.")

    lookup = _ready_plan_check_result_lookup(result)
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill("solid", fgColor="F4CCCC")

    for sheet in export_workbook.worksheets:
        header_rows: list[int] = []
        for row_number in range(1, sheet.max_row + 1):
            labels = {
                str(sheet.cell(row_number, column).value or "").replace(" ", "").strip()
                for column in range(1, sheet.max_column + 1)
            }
            if "Код" in labels and any(re.fullmatch(r"Т\d+", label) for label in labels):
                header_rows.append(row_number)
        for header_position, header_row in enumerate(header_rows):
            next_header = header_rows[header_position + 1] if header_position + 1 < len(header_rows) else sheet.max_row + 1
            normalized_headers = {
                str(sheet.cell(header_row, column).value or "").replace(" ", "").strip(): column
                for column in range(1, sheet.max_column + 1)
            }
            code_column = normalized_headers.get("Код", normalized_headers.get("Код№"))
            if code_column is None:
                continue
            for row_number in range(header_row + 1, next_header):
                sku = normalize_sku(sheet.cell(row_number, code_column).value)
                if sku is None:
                    continue
                for point_number in range(1, 30):
                    point_label = f"Т{point_number}"
                    point_column = normalized_headers.get(point_label)
                    if point_column is None:
                        continue
                    record = lookup.get((sku, point_label))
                    if record is None:
                        continue
                    cell = sheet.cell(row_number, point_column)
                    status = str(record.get("Согласие системы") or "")
                    if status.startswith("Красное"):
                        cell.fill = red_fill
                    elif status.startswith("Жёлтое"):
                        cell.fill = yellow_fill
                    else:
                        cell.fill = green_fill
                    note = _ready_plan_check_note(record)
                    if cell.comment is not None and cell.comment.text.strip():
                        note += "\n\nИсходное обоснование значения из плана:\n" + cell.comment.text.strip()
                    detailed_comment = Comment(note, "Проверка системы")
                    detailed_comment.width = 620
                    detailed_comment.height = 720
                    cell.comment = detailed_comment

        # If Apps Script reconstructed a style-free sheet, keep the familiar SR rows visible.
        for row_number in range(1, sheet.max_row + 1):
            if str(sheet.cell(row_number, 1).value or "").strip().upper() == "СР":
                for cell in sheet[row_number]:
                    if cell.fill.fill_type is None:
                        cell.fill = PatternFill("solid", fgColor="DDEBF7")

        sheet.freeze_panes = "F3" if sheet.max_row >= 3 and sheet.max_column >= 6 else None
        sheet.sheet_view.showGridLines = True

    output = io.BytesIO()
    export_workbook.save(output)
    return output.getvalue()


def export_excel(
    sku_point: pd.DataFrame,
    category_profile: pd.DataFrame,
    entity_profile: pd.DataFrame,
    daily_detail: pd.DataFrame,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        category_profile.to_excel(writer, sheet_name="Профиль категорий", index=False)
        entity_profile.to_excel(writer, sheet_name="Профиль сущностей", index=False)
        top3_export = entity_profile[entity_profile["entity_rank_category"] <= 3].copy()
        entity_category_totals = (
            entity_profile.groupby(["point", "category"], as_index=False)["entity_sales"]
            .sum()
            .rename(columns={"entity_sales": "category_entity_sales"})
        )
        top3_export = top3_export.merge(entity_category_totals, on=["point", "category"], how="left")
        top3_export["share_inside_category"] = (
            top3_export["entity_sales"] / top3_export["category_entity_sales"].replace(0, pd.NA)
        )
        top3_export.sort_values(["point", "category", "entity_rank_category"]).to_excel(
            writer, sheet_name="Топ-3 сущности", index=False
        )
        sku_point.to_excel(writer, sheet_name="Продажи SKU", index=False)
        detail_export = daily_detail.copy()
        detail_export.insert(0, "Дата", pd.to_datetime(detail_export["business_date"]).dt.date)
        detail_export.insert(1, "Время продажи", pd.to_datetime(detail_export["sale_datetime"]).dt.time)
        detail_export = detail_export.rename(
            columns={
                "point": "Точка",
                "shop_number": "Номер магазина",
                "sku": "SKU",
                "product_name": "Название товара",
                "category": "Категория",
                "entity": "Сущность",
                "sales": "Продано, шт.",
                "revenue": "Выручка, ₽",
                "plan_quantity": "План партии, шт.",
                "loading_date": "Дата загрузки",
                "freshness_stage": "Статус свежести",
                "batch_expiry_date": "Последний день срока партии",
                "batch_sold_total": "Продано из партии, шт.",
                "batch_live_remaining": "Живой остаток партии, шт.",
                "batch_writeoff_quantity": "Списание партии, шт.",
                "batch_status": "Статус партии",
            }
        ).drop(columns=["business_date", "sale_datetime"])
        preferred_detail_order = [
            "Дата", "Время продажи", "Точка", "Номер магазина", "SKU", "Название товара",
            "Категория", "Сущность", "Продано, шт.", "План партии, шт.", "Выручка, ₽",
            "Дата загрузки", "Продано из партии, шт.", "Живой остаток партии, шт.",
            "Списание партии, шт.", "Последний день срока партии", "Статус свежести", "Статус партии",
        ]
        export_order = [column for column in preferred_detail_order if column in detail_export.columns]
        export_order += [column for column in detail_export.columns if column not in export_order]
        detail_export = detail_export[export_order]
        detail_export.to_excel(writer, sheet_name="Детализация", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column_cells in sheet.columns:
                max_length = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
                sheet.column_dimensions[column_cells[0].column_letter].width = max_length
    return buffer.getvalue()


REPORT_WEEKDAYS_RU = {
    0: "Пн",
    1: "Вт",
    2: "Ср",
    3: "Чт",
    4: "Пт",
    5: "Сб",
    6: "Вс",
}


def _report_date_span(start_date: date, end_date: date) -> list[date]:
    if start_date > end_date:
        return []
    return [start_date + timedelta(days=offset) for offset in range((end_date - start_date).days + 1)]


def build_report_comparison_dates(
    period_1: tuple[date, date],
    period_2: tuple[date, date],
    match_weekdays: bool,
) -> tuple[list[date], list[date], list[tuple[date, date]]]:
    dates_1 = _report_date_span(*period_1)
    dates_2 = _report_date_span(*period_2)

    if not match_weekdays:
        pair_count = min(len(dates_1), len(dates_2))
        pairs = list(zip(dates_1[:pair_count], dates_2[:pair_count]))
        return dates_1, dates_2, pairs

    pairs: list[tuple[date, date]] = []
    for weekday in range(7):
        weekday_1 = [item for item in dates_1 if item.weekday() == weekday]
        weekday_2 = [item for item in dates_2 if item.weekday() == weekday]
        matched_count = min(len(weekday_1), len(weekday_2))
        pairs.extend(zip(weekday_1[:matched_count], weekday_2[:matched_count]))

    pairs = sorted(pairs, key=lambda item: item[0])
    matched_1 = [item[0] for item in pairs]
    matched_2 = [item[1] for item in pairs]
    return matched_1, matched_2, pairs


def prepare_report_sales_frame(
    sales: pd.DataFrame,
    entities: pd.DataFrame,
    period_name: str,
) -> pd.DataFrame:
    """Подготавливает компактный дневной набор для отчёта.

    Данные остаются агрегированными по дню и точке, но SKU и название товара
    сохраняются, чтобы в отчёте можно было раскрыть категорию до конкретных
    товаров без повторного обращения к PostgreSQL.
    """
    compact_columns = [
        "period", "business_date", "point", "shop_number",
        "category", "entity", "sku", "product_name", "sales", "revenue",
    ]
    if sales.empty:
        return pd.DataFrame(columns=compact_columns)

    source_columns = [
        column for column in [
            "business_date", "shop_number", "sku", "product_name", "sold_quantity", "revenue"
        ]
        if column in sales.columns
    ]
    report = sales[source_columns].copy()
    if "product_name" not in report.columns:
        report["product_name"] = ""
    report["business_date"] = pd.to_datetime(report["business_date"], errors="coerce").dt.date
    report["shop_number"] = pd.to_numeric(report["shop_number"], errors="coerce").astype("Int64")
    report = report[report["business_date"].notna() & report["shop_number"].notna()].copy()
    report["shop_number"] = report["shop_number"].astype(int)
    report = report[(report["shop_number"] >= 1) & (report["shop_number"] <= 29)]
    report = report[report["shop_number"] != 11].copy()
    report["point"] = report["shop_number"].map(lambda value: f"Т{int(value)}")
    report["sku"] = report["sku"].map(normalize_sku)

    entity_columns = [
        column for column in ["sku", "category", "entity", "entity_product_name"]
        if column in entities.columns
    ]
    entity_map = entities[entity_columns].copy()
    entity_map["sku"] = entity_map["sku"].map(normalize_sku)
    entity_map = entity_map.drop_duplicates("sku")
    report = report.merge(entity_map, on="sku", how="left", validate="many_to_one")
    report["category"] = report["category"].fillna("Не сопоставлено")
    report["entity"] = report["entity"].fillna("Не сопоставлено")
    report["product_name"] = report["product_name"].fillna("").astype(str).str.strip()
    if "entity_product_name" in report.columns:
        mapped_names = report["entity_product_name"].fillna("").astype(str).str.strip()
        report["product_name"] = mapped_names.where(mapped_names.ne(""), report["product_name"])
        report = report.drop(columns=["entity_product_name"])
    report.loc[report["product_name"].eq(""), "product_name"] = "Без названия"
    report["sales"] = pd.to_numeric(report["sold_quantity"], errors="coerce").fillna(0.0)
    if "revenue" not in report.columns:
        report["revenue"] = 0.0
    report["revenue"] = pd.to_numeric(report["revenue"], errors="coerce").fillna(0.0)

    report = (
        report.groupby(
            [
                "business_date", "point", "shop_number", "category", "entity",
                "sku", "product_name",
            ],
            as_index=False,
            dropna=False,
        )
        .agg(sales=("sales", "sum"), revenue=("revenue", "sum"))
    )
    report["period"] = period_name
    return report[compact_columns]


def _report_group_period_values(
    frame: pd.DataFrame,
    group_columns: list[str],
    value_name: str,
) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=group_columns + [value_name])
    return (
        frame.groupby(group_columns, dropna=False, as_index=False)
        .agg(**{value_name: ("sales", "sum")})
    )


def build_report_category_sku_breakdown(
    frame_1: pd.DataFrame,
    frame_2: pd.DataFrame,
    category: str,
    period_1_dates: list[date],
    period_2_dates: list[date],
) -> pd.DataFrame:
    """Раскрывает выбранную категорию до SKU с количеством продаж и выручкой."""
    days_1 = max(len(period_1_dates), 1)
    days_2 = max(len(period_2_dates), 1)

    def sku_period(
        frame: pd.DataFrame,
        quantity_name: str,
        revenue_name: str,
    ) -> pd.DataFrame:
        columns = ["SKU", "Название товара", "Сущность", quantity_name, revenue_name]
        if frame.empty or "sku" not in frame.columns:
            return pd.DataFrame(columns=columns)
        selected = frame[frame["category"].astype(str).eq(str(category))].copy()
        if selected.empty:
            return pd.DataFrame(columns=columns)
        selected["sku"] = selected["sku"].map(normalize_sku)
        selected["product_name"] = selected.get("product_name", "").fillna("").astype(str).str.strip()
        selected.loc[selected["product_name"].eq(""), "product_name"] = "Без названия"
        selected["entity"] = selected["entity"].fillna("Не сопоставлено").astype(str)
        if "revenue" not in selected.columns:
            selected["revenue"] = 0.0
        selected["revenue"] = pd.to_numeric(selected["revenue"], errors="coerce").fillna(0.0)
        grouped = (
            selected.groupby(["sku", "product_name", "entity"], dropna=False, as_index=False)
            .agg(
                **{
                    quantity_name: ("sales", "sum"),
                    revenue_name: ("revenue", "sum"),
                }
            )
            .rename(columns={
                "sku": "SKU",
                "product_name": "Название товара",
                "entity": "Сущность",
            })
        )
        return grouped[columns]

    left = sku_period(frame_1, "Период 1, шт.", "Выручка П1, ₽")
    right = sku_period(frame_2, "Период 2, шт.", "Выручка П2, ₽")
    result = left.merge(
        right,
        on=["SKU", "Название товара", "Сущность"],
        how="outer",
    ).fillna({
        "Период 1, шт.": 0.0,
        "Период 2, шт.": 0.0,
        "Выручка П1, ₽": 0.0,
        "Выручка П2, ₽": 0.0,
    })
    if result.empty:
        return result

    for column in ["Период 1, шт.", "Период 2, шт.", "Выручка П1, ₽", "Выручка П2, ₽"]:
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0.0)
    result = result[(result["Период 1, шт."] > 0) | (result["Период 2, шт."] > 0)].copy()
    if result.empty:
        return result

    total_1 = float(result["Период 1, шт."].sum())
    total_2 = float(result["Период 2, шт."].sum())
    total_revenue_1 = float(result["Выручка П1, ₽"].sum())
    total_revenue_2 = float(result["Выручка П2, ₽"].sum())

    result["Изменение, шт."] = result["Период 2, шт."] - result["Период 1, шт."]
    result["Изменение, %"] = (
        result["Изменение, шт."].div(result["Период 1, шт."].replace(0, pd.NA)) * 100
    )
    result.loc[result["Период 1, шт."].eq(0), "Изменение, %"] = pd.NA
    result["Изменение выручки, ₽"] = result["Выручка П2, ₽"] - result["Выручка П1, ₽"]
    result["Изменение выручки, %"] = (
        result["Изменение выручки, ₽"].div(result["Выручка П1, ₽"].replace(0, pd.NA)) * 100
    )
    result.loc[result["Выручка П1, ₽"].eq(0), "Изменение выручки, %"] = pd.NA
    result["СР/день П1"] = result["Период 1, шт."] / days_1
    result["СР/день П2"] = result["Период 2, шт."] / days_2
    result = result.sort_values(
        ["Период 2, шт.", "Период 1, шт.", "Название товара"],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)

    total_delta = total_2 - total_1
    total_revenue_delta = total_revenue_2 - total_revenue_1
    total_row = {
        "SKU": "",
        "Название товара": "ВСЕГО КАТЕГОРИИ",
        "Сущность": "",
        "Период 1, шт.": total_1,
        "Период 2, шт.": total_2,
        "Выручка П1, ₽": total_revenue_1,
        "Выручка П2, ₽": total_revenue_2,
        "Изменение, шт.": total_delta,
        "Изменение, %": (total_delta / total_1 * 100) if total_1 else pd.NA,
        "Изменение выручки, ₽": total_revenue_delta,
        "Изменение выручки, %": (
            total_revenue_delta / total_revenue_1 * 100
        ) if total_revenue_1 else pd.NA,
        "СР/день П1": total_1 / days_1,
        "СР/день П2": total_2 / days_2,
    }
    ordered_columns = [
        "SKU", "Название товара", "Сущность",
        "Период 1, шт.", "Период 2, шт.",
        "Выручка П1, ₽", "Выручка П2, ₽",
        "Изменение, шт.", "Изменение, %",
        "Изменение выручки, ₽", "Изменение выручки, %",
        "СР/день П1", "СР/день П2",
    ]
    return pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)[ordered_columns]


def build_report_tables(
    frame_1: pd.DataFrame,
    frame_2: pd.DataFrame,
    period_1_dates: list[date],
    period_2_dates: list[date],
    points: list[str],
) -> dict[str, pd.DataFrame]:
    days_1 = max(len(period_1_dates), 1)
    days_2 = max(len(period_2_dates), 1)

    def comparison(group_columns: list[str], include_revenue: bool = False) -> pd.DataFrame:
        left = _report_group_period_values(frame_1, group_columns, "Период 1, шт.")
        right = _report_group_period_values(frame_2, group_columns, "Период 2, шт.")
        result = left.merge(right, on=group_columns, how="outer").fillna({"Период 1, шт.": 0.0, "Период 2, шт.": 0.0})
        result["Изменение, шт."] = result["Период 2, шт."] - result["Период 1, шт."]
        result["Изменение, %"] = result["Изменение, шт."].div(result["Период 1, шт."].replace(0, pd.NA)) * 100
        zero_base = result["Период 1, шт."].eq(0) & result["Период 2, шт."].gt(0)
        result.loc[zero_base, "Изменение, %"] = pd.NA
        result["СР/день П1"] = result["Период 1, шт."] / days_1
        result["СР/день П2"] = result["Период 2, шт."] / days_2
        result["Изменение СР/день"] = result["СР/день П2"] - result["СР/день П1"]

        if include_revenue:
            def revenue_period(frame: pd.DataFrame, value_name: str) -> pd.DataFrame:
                if frame.empty:
                    return pd.DataFrame(columns=group_columns + [value_name])
                work = frame.copy()
                if "revenue" not in work.columns:
                    work["revenue"] = 0.0
                work["revenue"] = pd.to_numeric(work["revenue"], errors="coerce").fillna(0.0)
                return (
                    work.groupby(group_columns, dropna=False, as_index=False)
                    .agg(**{value_name: ("revenue", "sum")})
                )

            revenue_left = revenue_period(frame_1, "Выручка П1, ₽")
            revenue_right = revenue_period(frame_2, "Выручка П2, ₽")
            result = result.merge(revenue_left, on=group_columns, how="left")
            result = result.merge(revenue_right, on=group_columns, how="left")
            result[["Выручка П1, ₽", "Выручка П2, ₽"]] = result[["Выручка П1, ₽", "Выручка П2, ₽"]].fillna(0.0)
            result["Изменение выручки, ₽"] = result["Выручка П2, ₽"] - result["Выручка П1, ₽"]
            result["Изменение выручки, %"] = (
                result["Изменение выручки, ₽"].div(result["Выручка П1, ₽"].replace(0, pd.NA)) * 100
            )
            result.loc[result["Выручка П1, ₽"].eq(0), "Изменение выручки, %"] = pd.NA
        return result

    category_summary = comparison(["category"], include_revenue=True).rename(columns={"category": "Категория"})
    category_entity = comparison(["category", "entity"]).rename(columns={"category": "Категория", "entity": "Сущность"})
    by_point = comparison(["category", "entity", "point"]).rename(
        columns={"category": "Категория", "entity": "Сущность", "point": "Точка"}
    )

    weekday_order = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    weekday_1 = frame_1.copy()
    weekday_2 = frame_2.copy()
    if not weekday_1.empty:
        weekday_1["День недели"] = weekday_1["business_date"].map(lambda value: REPORT_WEEKDAYS_RU.get(value.weekday(), ""))
    if not weekday_2.empty:
        weekday_2["День недели"] = weekday_2["business_date"].map(lambda value: REPORT_WEEKDAYS_RU.get(value.weekday(), ""))
    left_weekday = _report_group_period_values(weekday_1, ["День недели"], "Период 1, шт.")
    right_weekday = _report_group_period_values(weekday_2, ["День недели"], "Период 2, шт.")
    weekday_summary = left_weekday.merge(right_weekday, on="День недели", how="outer").fillna(0.0)
    weekday_summary["Изменение, шт."] = weekday_summary["Период 2, шт."] - weekday_summary["Период 1, шт."]
    weekday_summary["Изменение, %"] = weekday_summary["Изменение, шт."].div(weekday_summary["Период 1, шт."].replace(0, pd.NA)) * 100
    weekday_summary["_order"] = weekday_summary["День недели"].map({name: idx for idx, name in enumerate(weekday_order)})
    weekday_summary = weekday_summary.sort_values("_order", kind="stable").drop(columns="_order")

    def matrix(frame: pd.DataFrame) -> pd.DataFrame:
        if frame.empty:
            total_row = {"Категория": "ВСЕГО", "Сущность": ""}
            total_row.update({point: 0.0 for point in points})
            total_row["ВСЕГО"] = 0.0
            return pd.DataFrame([total_row], columns=["Категория", "Сущность", *points, "ВСЕГО"])
        pivot = frame.pivot_table(
            index=["category", "entity"],
            columns="point",
            values="sales",
            aggfunc="sum",
            fill_value=0.0,
        )
        pivot = pivot.reindex(columns=points, fill_value=0.0)
        pivot["ВСЕГО"] = pivot.sum(axis=1)
        pivot = pivot.reset_index().rename(columns={"category": "Категория", "entity": "Сущность"})
        total_row = {"Категория": "ВСЕГО", "Сущность": ""}
        for point in points:
            total_row[point] = float(pd.to_numeric(pivot[point], errors="coerce").sum())
        total_row["ВСЕГО"] = float(pd.to_numeric(pivot["ВСЕГО"], errors="coerce").sum())
        return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)


    def category_point_matrix(frame: pd.DataFrame) -> pd.DataFrame:
        """Category totals by point for one selected report period."""
        if frame.empty:
            total_row = {"Категория": "ВСЕГО"}
            total_row.update({point: 0.0 for point in points})
            total_row["ВСЕГО"] = 0.0
            return pd.DataFrame([total_row], columns=["Категория", *points, "ВСЕГО"])
        pivot = frame.pivot_table(
            index=["category"],
            columns="point",
            values="sales",
            aggfunc="sum",
            fill_value=0.0,
        )
        pivot = pivot.reindex(columns=points, fill_value=0.0)
        pivot["ВСЕГО"] = pivot.sum(axis=1)
        pivot = pivot.reset_index().rename(columns={"category": "Категория"})
        total_row = {"Категория": "ВСЕГО"}
        for point in points:
            total_row[point] = float(pd.to_numeric(pivot[point], errors="coerce").sum())
        total_row["ВСЕГО"] = float(pd.to_numeric(pivot["ВСЕГО"], errors="coerce").sum())
        return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    matrix_1 = matrix(frame_1)
    matrix_2 = matrix(frame_2)
    matrix_delta = pd.DataFrame()
    if not matrix_1.empty or not matrix_2.empty:
        key_cols = ["Категория", "Сущность"]
        matrix_delta = matrix_1.merge(
            matrix_2,
            on=key_cols,
            how="outer",
            suffixes=("__1", "__2"),
        ).fillna(0.0)

        # В представлении «Изменение» для каждой точки показываем
        # количественную разницу и рядом процентное изменение.
        result_columns = key_cols.copy()
        for point in points:
            p1 = pd.to_numeric(matrix_delta.get(f"{point}__1", 0.0), errors="coerce").fillna(0.0)
            p2 = pd.to_numeric(matrix_delta.get(f"{point}__2", 0.0), errors="coerce").fillna(0.0)
            delta = p2 - p1
            pct = delta.div(p1.replace(0, pd.NA)) * 100
            pct = pct.mask(p1.eq(0), pd.NA)
            qty_col = f"{point} Δ, шт."
            pct_col = f"{point} Δ, %"
            matrix_delta[qty_col] = delta
            matrix_delta[pct_col] = pct
            result_columns.extend([qty_col, pct_col])

        total_p1 = pd.to_numeric(matrix_delta.get("ВСЕГО__1", 0.0), errors="coerce").fillna(0.0)
        total_p2 = pd.to_numeric(matrix_delta.get("ВСЕГО__2", 0.0), errors="coerce").fillna(0.0)
        total_delta = total_p2 - total_p1
        matrix_delta["ВСЕГО Δ, шт."] = total_delta
        matrix_delta["ВСЕГО Δ, %"] = total_delta.div(total_p1.replace(0, pd.NA)) * 100
        matrix_delta.loc[total_p1.eq(0), "ВСЕГО Δ, %"] = pd.NA
        result_columns.extend(["ВСЕГО Δ, шт.", "ВСЕГО Δ, %"])

        matrix_delta = matrix_delta[result_columns]
        matrix_delta["_is_total"] = matrix_delta["Категория"].eq("ВСЕГО")
        matrix_delta = matrix_delta.sort_values(
            ["_is_total", "Категория", "Сущность"], kind="stable"
        ).drop(columns="_is_total")

    category_matrix_1 = category_point_matrix(frame_1)
    category_matrix_2 = category_point_matrix(frame_2)
    category_matrix_delta = category_matrix_1.merge(
        category_matrix_2,
        on=["Категория"],
        how="outer",
        suffixes=("__1", "__2"),
    ).fillna(0.0)
    category_result_columns = ["Категория"]
    for point in points:
        p1 = pd.to_numeric(category_matrix_delta.get(f"{point}__1", 0.0), errors="coerce").fillna(0.0)
        p2 = pd.to_numeric(category_matrix_delta.get(f"{point}__2", 0.0), errors="coerce").fillna(0.0)
        delta = p2 - p1
        pct = delta.div(p1.replace(0, pd.NA)) * 100
        pct = pct.mask(p1.eq(0), pd.NA)
        qty_col = f"{point} Δ, шт."
        pct_col = f"{point} Δ, %"
        category_matrix_delta[qty_col] = delta
        category_matrix_delta[pct_col] = pct
        category_result_columns.extend([qty_col, pct_col])

    category_total_p1 = pd.to_numeric(
        category_matrix_delta.get("ВСЕГО__1", 0.0), errors="coerce"
    ).fillna(0.0)
    category_total_p2 = pd.to_numeric(
        category_matrix_delta.get("ВСЕГО__2", 0.0), errors="coerce"
    ).fillna(0.0)
    category_total_delta = category_total_p2 - category_total_p1
    category_matrix_delta["ВСЕГО Δ, шт."] = category_total_delta
    category_matrix_delta["ВСЕГО Δ, %"] = (
        category_total_delta.div(category_total_p1.replace(0, pd.NA)) * 100
    )
    category_matrix_delta.loc[category_total_p1.eq(0), "ВСЕГО Δ, %"] = pd.NA
    category_result_columns.extend(["ВСЕГО Δ, шт.", "ВСЕГО Δ, %"])
    category_matrix_delta = category_matrix_delta[category_result_columns]
    category_matrix_delta["_is_total"] = category_matrix_delta["Категория"].eq("ВСЕГО")
    category_matrix_delta = category_matrix_delta.sort_values(
        ["_is_total", "Категория"], kind="stable"
    ).drop(columns="_is_total")

    return {
        "category_summary": category_summary.sort_values("Период 2, шт.", ascending=False, kind="stable"),
        "category_entity": category_entity.sort_values(["Категория", "Период 2, шт."], ascending=[True, False], kind="stable"),
        "by_point": by_point.sort_values(["Точка", "Категория", "Период 2, шт."], ascending=[True, True, False], kind="stable"),
        "weekday_summary": weekday_summary,
        "matrix_1": matrix_1,
        "matrix_2": matrix_2,
        "matrix_delta": matrix_delta,
        "category_matrix_1": category_matrix_1,
        "category_matrix_2": category_matrix_2,
        "category_matrix_delta": category_matrix_delta,
    }


def _append_report_total_row(frame: pd.DataFrame, label_column: str = "Категория") -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    result = frame.copy()
    total_row: dict[str, object] = {}
    for column in result.columns:
        if column == label_column:
            total_row[column] = "ВСЕГО"
        elif column in {"Сущность", "Точка"}:
            total_row[column] = ""
        elif pd.api.types.is_numeric_dtype(result[column]):
            if column == "Изменение, %":
                p1 = pd.to_numeric(result.get("Период 1, шт."), errors="coerce").sum()
                p2 = pd.to_numeric(result.get("Период 2, шт."), errors="coerce").sum()
                total_row[column] = ((p2 - p1) / p1 * 100) if p1 else pd.NA
            elif column == "Изменение выручки, %":
                revenue_1 = pd.to_numeric(result.get("Выручка П1, ₽"), errors="coerce").sum()
                revenue_2 = pd.to_numeric(result.get("Выручка П2, ₽"), errors="coerce").sum()
                total_row[column] = ((revenue_2 - revenue_1) / revenue_1 * 100) if revenue_1 else pd.NA
            elif column == "СР/день П1" or column == "СР/день П2" or column == "Изменение СР/день":
                total_row[column] = pd.NA
            else:
                total_row[column] = pd.to_numeric(result[column], errors="coerce").sum()
        else:
            total_row[column] = ""
    return pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)


def build_report_dynamic_data(
    frame_1: pd.DataFrame,
    frame_2: pd.DataFrame,
    dates_1: list[date],
    dates_2: list[date],
    pairs: list[tuple[date, date]],
    match_weekdays: bool,
    rolling_window: int,
) -> pd.DataFrame:
    qty_1 = frame_1.groupby("business_date")["sales"].sum().to_dict() if not frame_1.empty else {}
    qty_2 = frame_2.groupby("business_date")["sales"].sum().to_dict() if not frame_2.empty else {}

    rows: list[dict[str, object]] = []
    if match_weekdays:
        for index, (date_1, date_2) in enumerate(pairs, start=1):
            rows.append(
                {
                    "Сравнимый день": index,
                    "День недели": REPORT_WEEKDAYS_RU.get(date_1.weekday(), ""),
                    "Дата П1": date_1,
                    "Дата П2": date_2,
                    "Период 1, шт.": float(qty_1.get(date_1, 0.0)),
                    "Период 2, шт.": float(qty_2.get(date_2, 0.0)),
                }
            )
    else:
        max_days = max(len(dates_1), len(dates_2))
        for index in range(max_days):
            date_1 = dates_1[index] if index < len(dates_1) else None
            date_2 = dates_2[index] if index < len(dates_2) else None
            weekday_source = date_2 or date_1
            rows.append(
                {
                    "Сравнимый день": index + 1,
                    "День недели": REPORT_WEEKDAYS_RU.get(weekday_source.weekday(), "") if weekday_source else "",
                    "Дата П1": date_1,
                    "Дата П2": date_2,
                    "Период 1, шт.": float(qty_1.get(date_1, 0.0)) if date_1 else pd.NA,
                    "Период 2, шт.": float(qty_2.get(date_2, 0.0)) if date_2 else pd.NA,
                }
            )

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["СР П1"] = pd.to_numeric(result["Период 1, шт."], errors="coerce").rolling(
        rolling_window, min_periods=1
    ).mean()
    result["СР П2"] = pd.to_numeric(result["Период 2, шт."], errors="coerce").rolling(
        rolling_window, min_periods=1
    ).mean()
    return result


def build_period_comparison_excel(
    tables: dict[str, pd.DataFrame],
    dynamic_data: pd.DataFrame,
    period_1: tuple[date, date],
    period_2: tuple[date, date],
    match_weekdays: bool,
    dates_1: list[date],
    dates_2: list[date],
    rolling_window: int,
    graph_filter_label: str,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        params = pd.DataFrame(
            [
                ["Период 1", f"{period_1[0]:%d.%m.%Y}–{period_1[1]:%d.%m.%Y}"],
                ["Период 2", f"{period_2[0]:%d.%m.%Y}–{period_2[1]:%d.%m.%Y}"],
                ["Сверка одинаковых дней недели", "Да" if match_weekdays else "Нет"],
                ["Дней в сверке П1", len(dates_1)],
                ["Дней в сверке П2", len(dates_2)],
                ["Окно скользящего среднего", rolling_window],
                ["Фильтр графика", graph_filter_label],
            ],
            columns=["Параметр", "Значение"],
        )
        params.to_excel(writer, sheet_name="Параметры", index=False)
        _append_report_total_row(tables["category_summary"]).to_excel(writer, sheet_name="Сводка категорий", index=False)
        _append_report_total_row(tables["category_entity"]).to_excel(writer, sheet_name="Категории-сущности", index=False)
        _append_report_total_row(tables["by_point"]).to_excel(writer, sheet_name="По точкам", index=False)
        tables["weekday_summary"].to_excel(writer, sheet_name="Дни недели", index=False)
        tables["matrix_1"].to_excel(writer, sheet_name="Матрица П1", index=False)
        tables["matrix_2"].to_excel(writer, sheet_name="Матрица П2", index=False)
        tables["matrix_delta"].to_excel(writer, sheet_name="Изменение матрицы", index=False)
        tables["category_matrix_1"].to_excel(writer, sheet_name="Категории-точки П1", index=False)
        tables["category_matrix_2"].to_excel(writer, sheet_name="Категории-точки П2", index=False)
        tables["category_matrix_delta"].to_excel(writer, sheet_name="Категории-точки Δ", index=False)
        dynamic_export = dynamic_data.copy()
        if not dynamic_export.empty:
            for date_column in ["Дата П1", "Дата П2"]:
                dynamic_export[date_column] = pd.to_datetime(dynamic_export[date_column], errors="coerce").dt.date
        dynamic_export.to_excel(writer, sheet_name="Динамика", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill("solid", fgColor="D9EAF7")
        positive_fill = PatternFill("solid", fgColor="E2F0D9")
        negative_fill = PatternFill("solid", fgColor="FCE4D6")

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.auto_filter.ref = sheet.dimensions
            for cells in sheet.columns:
                values = [len(str(cell.value or "")) for cell in cells[: min(len(cells), 3000)]]
                width = min(max(values + [8]) + 2, 34)
                sheet.column_dimensions[cells[0].column_letter].width = width

        for sheet_name in ["Сводка категорий", "Категории-сущности", "По точкам"]:
            sheet = workbook[sheet_name]
            if sheet.max_row > 1:
                for cell in sheet[sheet.max_row]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                headers = {str(cell.value): cell.column for cell in sheet[1]}
                pct_col = headers.get("Изменение, %")
                if pct_col:
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row, pct_col).number_format = '0.0"%"'
                for header in ["Период 1, шт.", "Период 2, шт.", "Изменение, шт.", "СР/день П1", "СР/день П2", "Изменение СР/день"]:
                    col = headers.get(header)
                    if col:
                        for row in range(2, sheet.max_row + 1):
                            sheet.cell(row, col).number_format = '#,##0.0'
                delta_col = headers.get("Изменение, шт.")
                if delta_col:
                    for row in range(2, sheet.max_row):
                        value = sheet.cell(row, delta_col).value
                        if isinstance(value, (int, float)):
                            sheet.cell(row, delta_col).fill = positive_fill if value >= 0 else negative_fill

        for delta_sheet_name in ["Изменение матрицы", "Категории-точки Δ"]:
            matrix_delta_sheet = workbook[delta_sheet_name]
            matrix_delta_headers = {str(cell.value): cell.column for cell in matrix_delta_sheet[1]}
            for header, col in matrix_delta_headers.items():
                if header.endswith("Δ, %"):
                    for row in range(2, matrix_delta_sheet.max_row + 1):
                        matrix_delta_sheet.cell(row, col).number_format = '0.0"%"'
                elif header.endswith("Δ, шт."):
                    for row in range(2, matrix_delta_sheet.max_row + 1):
                        matrix_delta_sheet.cell(row, col).number_format = '+#,##0;-#,##0;0'
                        value = matrix_delta_sheet.cell(row, col).value
                        if isinstance(value, (int, float)):
                            matrix_delta_sheet.cell(row, col).fill = positive_fill if value >= 0 else negative_fill
            if matrix_delta_sheet.max_row > 1:
                for cell in matrix_delta_sheet[matrix_delta_sheet.max_row]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)

        for total_sheet_name in ["Категории-точки П1", "Категории-точки П2"]:
            total_sheet = workbook[total_sheet_name]
            if total_sheet.max_row > 1:
                for cell in total_sheet[total_sheet.max_row]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)

        charts_sheet = workbook.create_sheet("Графики")
        charts_sheet.sheet_view.showGridLines = False
        charts_sheet["A1"] = "Сравнение продаж по категориям"
        charts_sheet["A1"].font = Font(bold=True, size=14)

        category_sheet = workbook["Сводка категорий"]
        category_headers = {str(cell.value): cell.column for cell in category_sheet[1]}
        last_category_row = max(2, category_sheet.max_row - 1)
        if last_category_row >= 2 and category_headers.get("Категория") and category_headers.get("Период 1, шт."):
            bar = BarChart()
            bar.type = "col"
            bar.style = 10
            bar.title = "Категории: Период 1 vs Период 2"
            bar.y_axis.title = "Продано, шт."
            bar.x_axis.title = "Категория"
            data = Reference(
                category_sheet,
                min_col=category_headers["Период 1, шт."],
                max_col=category_headers["Период 2, шт."],
                min_row=1,
                max_row=last_category_row,
            )
            cats = Reference(
                category_sheet,
                min_col=category_headers["Категория"],
                min_row=2,
                max_row=last_category_row,
            )
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 12
            bar.width = 24
            charts_sheet.add_chart(bar, "A3")

        if not dynamic_data.empty:
            dyn_sheet = workbook["Динамика"]
            dyn_headers = {str(cell.value): cell.column for cell in dyn_sheet[1]}
            line = LineChart()
            line.title = f"Динамика продаж · {graph_filter_label}"
            line.y_axis.title = "Продано, шт."
            line.x_axis.title = "Сравнимый день"
            dyn_last_row = dyn_sheet.max_row
            data_cols = [
                dyn_headers.get("Период 1, шт."),
                dyn_headers.get("Период 2, шт."),
                dyn_headers.get("СР П1"),
                dyn_headers.get("СР П2"),
            ]
            data_cols = [col for col in data_cols if col]
            if data_cols:
                for col in data_cols:
                    data = Reference(dyn_sheet, min_col=col, min_row=1, max_row=dyn_last_row)
                    line.add_data(data, titles_from_data=True)
                categories = Reference(
                    dyn_sheet,
                    min_col=dyn_headers.get("Сравнимый день", 1),
                    min_row=2,
                    max_row=dyn_last_row,
                )
                line.set_categories(categories)
                line.height = 12
                line.width = 24
                charts_sheet.add_chart(line, "A28")

    return buffer.getvalue()


def build_period_comparison_html(
    category_chart: go.Figure,
    delta_chart: go.Figure,
    dynamic_chart: go.Figure,
    title: str,
) -> bytes:
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'><title>" + title + "</title></head><body>",
        f"<h1>{title}</h1>",
        pio.to_html(category_chart, full_html=False, include_plotlyjs=True),
        pio.to_html(delta_chart, full_html=False, include_plotlyjs=False),
        pio.to_html(dynamic_chart, full_html=False, include_plotlyjs=False),
        "</body></html>",
    ]
    return "".join(parts).encode("utf-8")


# Основной фирменный заголовок выводится вместе с внешним меню ниже.

try:
    (
        entities,
        entity_reference_source,
        entity_reference_checked_at,
        entity_reference_warning,
        entity_reference_signature,
    ) = get_current_entity_reference()
except Exception as error:
    st.error(f"Не удалось загрузить SKU/категории/сущности из матрицы 2.3: {error}")
    st.stop()

today = date.today()
month_start = today.replace(day=1)
previous_month_end = month_start - timedelta(days=1)
previous_month_start = previous_month_end.replace(day=1)

with st.sidebar:
    st.header("Параметры")
    st.caption("Аналитика спроса · версия 75.11.14 · CATEGORY SKU CALENDAR AVG SEPARATE")
    st.caption("Автозагрузка данных · SEPARATE-MENU")
    st.caption(f"SKU / категории / сущности · {entity_reference_source}")
    if entity_reference_warning and not entity_reference_source.startswith("Apps Script"):
        st.caption(f"Автоисточник справочника временно недоступен: {entity_reference_warning}. Планы загружаются отдельно.")
    with st.expander("Подключение к PostgreSQL", expanded=not bool(os.getenv("PGPASSWORD"))):
        pg_host = st.text_input(
            "Сервер",
            value=os.getenv("PGHOST", "rc1d-7j2b1dkkfbnnu430.mdb.yandexcloud.net"),
        )
        pg_port = st.number_input("Порт", min_value=1, max_value=65535, value=int(os.getenv("PGPORT", "6432")))
        pg_database = st.text_input("База данных", value=os.getenv("PGDATABASE", "analytics_mart"))
        pg_user = st.text_input("Пользователь", value=os.getenv("PGUSER", "sheets_reader"))
        pg_password = st.text_input("Пароль", value=os.getenv("PGPASSWORD", ""), type="password")
        remember_connection = st.checkbox(
            "Запомнить пароль на 30 дней",
            value=bool(remembered_pg.get("PGPASSWORD")),
            help=(
                "Пароль шифруется средствами Windows (DPAPI) и может быть расшифрован "
                "только вашим Windows-пользователем на этом компьютере."
            ),
        )
        if remembered_pg.get("expires_at"):
            try:
                remembered_until = datetime.fromisoformat(remembered_pg["expires_at"])
                st.caption(f"Сохранённый вход действует до {remembered_until:%d.%m.%Y %H:%M}.")
            except ValueError:
                pass
        if st.button("Забыть сохранённый пароль", use_container_width=True, disabled=not REMEMBERED_PG_FILE.exists()):
            forget_remembered_pg_credentials()
            os.environ.pop("PGPASSWORD", None)
            st.session_state.pop("analysis", None)
            st.session_state.pop("analysis_auto_signature_v7590", None)
            st.success("Сохранённый пароль удалён. При следующем подключении введите его снова.")
            st.rerun()
    date_range = st.date_input(
        "Период",
        value=(previous_month_start, previous_month_end),
        max_value=today,
        format="DD.MM.YYYY",
        key="main_period_v7590",
    )
    auto_status = st.empty()
    auto_status.caption("Точки и продажи загружаются автоматически.")

os.environ["PGHOST"] = pg_host.strip()
os.environ["PGPORT"] = str(int(pg_port))
os.environ["PGDATABASE"] = pg_database.strip()
os.environ["PGUSER"] = pg_user.strip()
os.environ["PGPASSWORD"] = pg_password

valid_period = isinstance(date_range, tuple) and len(date_range) == 2
if not valid_period:
    auto_status.warning("Выберите дату начала и дату окончания.")
    st.info("Укажите полный период — после этого точки и продажи загрузятся автоматически.")
    st.stop()

start_date, end_date = date_range
if start_date > end_date:
    auto_status.warning("Дата начала позже даты окончания.")
    st.error("Дата начала периода не может быть позже даты окончания.")
    st.stop()

# Автозагрузка выполняется только при изменении периода/подключения либо раз в 15 минут.
# Обычные переходы по разделам используют уже подготовленный analysis из session_state.
password_signature = (
    hashlib.sha256(pg_password.encode("utf-8")).hexdigest()[:12]
    if pg_password else "no-password"
)
refresh_bucket = int(datetime.now().timestamp() // (15 * 60))
auto_signature = (
    start_date.isoformat(),
    end_date.isoformat(),
    pg_host.strip(),
    int(pg_port),
    pg_database.strip(),
    pg_user.strip(),
    password_signature,
    refresh_bucket,
    entity_reference_signature[:16],
)

analysis_needs_refresh = (
    "analysis" not in st.session_state
    or st.session_state.get("analysis_auto_signature_v7590") != auto_signature
)

if analysis_needs_refresh:
    try:
        with st.spinner("Автоматически загружаю точки и продажи из PostgreSQL…"):
            available_shops = ensure_required_shops(
                load_available_shops(start_date, end_date + timedelta(days=1))
            )
            if available_shops.empty:
                raise RuntimeError("за выбранный период база не вернула магазины")

            # Рабочая сетка: Т1–Т29, Т11 исключена. Т25 добавляется обязательным правилом выше.
            selected = available_shops.copy()
            selected["shop_number"] = pd.to_numeric(selected["shop_number"], errors="coerce")
            selected = selected[
                selected["shop_number"].notna()
                & selected["shop_number"].between(1, 29)
                & selected["shop_number"].ne(11)
            ].copy()
            selected["shop_number"] = selected["shop_number"].astype(int)
            selected = selected.drop_duplicates("shop_number").sort_values("shop_number")

            if selected.empty:
                raise RuntimeError("не найдены рабочие точки Т1–Т29")

            selected_shop_numbers = tuple(selected["shop_number"].tolist())
            point_mapping = {number: f"Т{number}" for number in selected_shop_numbers}
            sales = load_sales(
                start_date,
                end_date + timedelta(days=1),
                selected_shop_numbers,
            )
            if sales.empty:
                raise RuntimeError("за выбранный период продаж не найдено")

            st.session_state["analysis"] = prepare_analysis(sales, entities, point_mapping)
            st.session_state["period"] = (start_date, end_date)
            st.session_state["point_mapping"] = point_mapping
            st.session_state["analysis_auto_signature_v7590"] = auto_signature
            st.session_state["auto_loaded_shops_v7590"] = selected_shop_numbers
            # Старое ручное сопоставление больше не управляет новой навигацией/автозагрузкой.
            st.session_state.pop("shop_mapping", None)

        if remember_connection:
            try:
                remembered_until = save_remembered_pg_credentials(
                    pg_host, int(pg_port), pg_database, pg_user, pg_password
                )
                st.toast(f"Пароль запомнен до {remembered_until:%d.%m.%Y}", icon="🔐")
            except Exception as error:
                st.warning(f"Данные загружены, но пароль не удалось запомнить: {error}")
        elif REMEMBERED_PG_FILE.exists():
            forget_remembered_pg_credentials()

        auto_status.success(
            f"Автозагрузка завершена: {len(selected_shop_numbers)} точек · "
            f"{start_date:%d.%m.%Y}–{end_date:%d.%m.%Y}"
        )
    except Exception as error:
        auto_status.error("Автозагрузка не выполнена.")
        st.error(f"Не удалось автоматически загрузить данные: {error}")
        st.stop()
else:
    cached_points = tuple(st.session_state.get("auto_loaded_shops_v7590", ()))
    auto_status.success(
        f"Данные готовы: {len(cached_points)} точек · "
        f"{start_date:%d.%m.%Y}–{end_date:%d.%m.%Y}"
    )

if "analysis" not in st.session_state:
    st.info("Данные ещё не подготовлены.")
    st.stop()

sku_point, category_profile, entity_profile, daily_detail = st.session_state["analysis"]
period = st.session_state["period"]

MENU_ITEMS = [
    ("Дашборд", ":material/dashboard:"),
    ("Отчет", ":material/description:"),
    ("Топ-3 сущности", ":material/account_tree:"),
    ("Сущности", ":material/storefront:"),
    ("Детализация", ":material/pie_chart:"),
    ("Детализация категории", ":material/sell:"),
    ("ABC продукции", ":material/inventory_2:"),
    ("Анализ категории", ":material/bar_chart:"),
    ("Окно свежести", ":material/calendar_month:"),
    ("Списания категорий", ":material/delete:"),
    ("Прогноз плана", ":material/track_changes:"),
    ("Проверка", ":material/fact_check:"),
]
SECTION_STATE_KEY = "main_section_v759"
MENU_LABELS = [label for label, _ in MENU_ITEMS]


def _open_main_section(label: str) -> None:
    if label in MENU_LABELS:
        st.session_state[SECTION_STATE_KEY] = label
        # Оставляем старый ключ только как совместимость для уже существующих session_state.
        st.session_state["main_tabs_v1"] = label


def _return_to_main_menu() -> None:
    st.session_state[SECTION_STATE_KEY] = None


selected_main_section = st.session_state.get(SECTION_STATE_KEY)
if selected_main_section not in MENU_LABELS:
    selected_main_section = None
    st.session_state[SECTION_STATE_KEY] = None

st.markdown(
    """
<style>
:root {
    --vk-red: #e42f35;
    --vk-ink: #222327;
    --vk-muted: #7b7d84;
    --vk-shell: #f1f1f3;
    --vk-card: #ffffff;
    --vk-line: #e3e3e7;
}
.vk-menu-header {
    position: relative;
    margin: 18px 0 0 0;
    padding: 28px 24px 18px 24px;
    background: var(--vk-shell);
    border-radius: 28px 28px 0 0;
    text-align: center;
}
.vk-menu-brand {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    color: var(--vk-ink);
    font-size: clamp(23px, 2.15vw, 34px);
    line-height: 1.1;
    font-weight: 760;
    letter-spacing: -0.035em;
}
.vk-menu-brand svg { width: 34px; height: 34px; flex: 0 0 auto; }
.vk-menu-subtitle {
    margin-top: 8px;
    color: #8a8c92;
    font-size: 13px;
    font-weight: 450;
    letter-spacing: .01em;
}
.vk-menu-actions {
    position: absolute;
    top: 15px;
    right: 18px;
    display: flex;
    gap: 7px;
}
.vk-menu-actions span {
    width: 25px;
    height: 25px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    border: 1px solid #d6d7db;
    border-radius: 999px;
    color: #666970;
    background: rgba(255,255,255,.55);
    font-size: 11px;
    font-weight: 700;
    user-select: none;
}
.st-key-main_menu_cards_v759 {
    padding: 10px 26px 30px 26px;
    background: var(--vk-shell);
    border-radius: 0 0 28px 28px;
}
.st-key-main_menu_cards_v759 div[data-testid="stButton"] > button {
    width: 100%;
    min-height: 112px;
    margin: 0;
    padding: 16px 12px 14px 12px;
    border: 1px solid var(--vk-line);
    border-radius: 12px;
    background: var(--vk-card);
    box-shadow: 0 5px 13px rgba(25, 27, 34, .10);
    color: var(--vk-ink);
    font-size: 14px;
    font-weight: 650;
    transition: transform .16s ease, box-shadow .16s ease, border-color .16s ease;
    white-space: normal;
}
.st-key-main_menu_cards_v759 div[data-testid="stButton"] > button:hover {
    transform: translateY(-2px);
    border-color: #d0d1d5;
    box-shadow: 0 8px 19px rgba(25, 27, 34, .13);
    color: #b72228;
}
.st-key-section_header_v759 {
    padding: 8px 0 4px 0;
    border-bottom: 1px solid #ececef;
    margin-bottom: 12px;
}
.vk-current-section {
    min-height: 40px;
    display: flex;
    align-items: center;
    font-size: 21px;
    font-weight: 760;
    color: var(--vk-ink);
}
.vk-footer-brand {
    margin: 42px 0 16px 0;
    text-align: center;
    color: var(--vk-red);
    font-weight: 900;
    font-size: clamp(30px, 4vw, 58px);
    line-height: 1;
    letter-spacing: -.04em;
}
@media (max-width: 620px) {
    .vk-menu-header { padding-top: 52px; }
    .vk-menu-actions { left: 50%; right: auto; transform: translateX(-50%); }
    .st-key-main_menu_cards_v759 { padding-left: 16px; padding-right: 16px; }
    .st-key-main_menu_cards_v759 div[data-testid="stButton"] > button { min-height: 90px; }
}
</style>
    """,
    unsafe_allow_html=True,
)

if selected_main_section is None:
    st.markdown(
        """
<div class="vk-menu-header">
    <div class="vk-menu-actions" aria-hidden="true"><span>i</span><span>◐</span><span>A</span></div>
    <div class="vk-menu-brand">
        <svg viewBox="0 0 34 34" fill="none" aria-hidden="true">
            <path d="M4 27V22M9 27V17M14 27V20M19 27V11M24 27V15M29 27V6" stroke="#e42f35" stroke-width="2.7" stroke-linecap="round"/>
            <path d="M4 18L10 14L15 16L21 8L26 11L30 4" stroke="#e42f35" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
        <span>Аналитика спроса</span>
    </div>
    <div class="vk-menu-subtitle">Анализируйте. Принимайте решения. Достигайте результата.</div>
</div>
        """,
        unsafe_allow_html=True,
    )
    with st.container(key="main_menu_cards_v759"):
        first_row = st.columns(4)
        for column, (label, icon) in zip(first_row, MENU_ITEMS[:4]):
            with column:
                st.button(
                    label,
                    key=f"main_menu_card_v759_{MENU_LABELS.index(label)}",
                    icon=icon,
                    use_container_width=True,
                    on_click=_open_main_section,
                    args=(label,),
                )
        second_row = st.columns(4)
        for column, (label, icon) in zip(second_row, MENU_ITEMS[4:8]):
            with column:
                st.button(
                    label,
                    key=f"main_menu_card_v759_{MENU_LABELS.index(label)}",
                    icon=icon,
                    use_container_width=True,
                    on_click=_open_main_section,
                    args=(label,),
                )
        third_row = st.columns(4)
        for column, (label, icon) in zip(third_row, MENU_ITEMS[8:12]):
            with column:
                st.button(
                    label,
                    key=f"main_menu_card_v759_{MENU_LABELS.index(label)}",
                    icon=icon,
                    use_container_width=True,
                    on_click=_open_main_section,
                    args=(label,),
                )
    st.markdown('<div class="vk-footer-brand">ВКУСНО МАРКЕТ</div>', unsafe_allow_html=True)
    st.stop()

with st.container(key="section_header_v759"):
    section_header_columns = st.columns([1.15, 4.85])
    with section_header_columns[0]:
        st.button(
            "В главное меню",
            key="return_to_main_menu_v759",
            icon=":material/arrow_back:",
            use_container_width=True,
            on_click=_return_to_main_menu,
        )
    with section_header_columns[1]:
        st.markdown(
            f'<div class="vk-current-section">{selected_main_section}</div>',
            unsafe_allow_html=True,
        )

# Совместимость с внутренними участками, которые могут читать прежний ключ навигации.
st.session_state["main_tabs_v1"] = selected_main_section

categories = sorted(category_profile["category"].unique())
filter_columns = st.columns(2)
with filter_columns[0]:
    category_filter = st.multiselect("Категории", categories, default=categories)
with filter_columns[1]:
    mapped_point_options = {
        str(label).strip()
        for label in st.session_state.get("point_mapping", {}).values()
        if str(label).strip().startswith("Т") and str(label).strip() != "Т11"
    }
    point_options = sorted(
        mapped_point_options | set(sku_point["point"].dropna().unique()),
        key=lambda value: int(value[1:]),
    )
    point_filter = st.multiselect("Показать точки", point_options, default=point_options)

filtered_category = category_profile[
    category_profile["category"].isin(category_filter) & category_profile["point"].isin(point_filter)
]
filtered_entity = entity_profile[
    entity_profile["category"].isin(category_filter) & entity_profile["point"].isin(point_filter)
]
filtered_sku = sku_point[sku_point["category"].isin(category_filter) & sku_point["point"].isin(point_filter)]
# Не создаём вторую тяжёлую копию всей детализации с FIFO на каждом rerun.
# Партии/даты загрузки рассчитываются ниже только для выбранного пользователем среза.
daily_detail_with_loading = daily_detail

# Необязательный фактический снимок остатков. Источник общий для «Детализации»
# и «Окна свежести», поэтому загрузчик расположен до вкладок.
stock_balances = pd.DataFrame()
stock_snapshot = pd.DataFrame()
selected_stock_date: date | None = None
with st.expander("Фактические остатки по точкам · необязательно", expanded=False):
    st.caption(
        "Загрузите Excel со снимком остатков. Ключ сверки: дата + SKU + точка. "
        "День недели определяется автоматически из даты. Если файл не загружен, приложение работает как раньше."
    )
    stock_controls = st.columns([1.7, 1.0])
    with stock_controls[0]:
        stock_upload = st.file_uploader(
            "Остатки по точкам (.xlsx)",
            type=["xlsx"],
            key="stock_balances_upload_v7546",
            help="Минимальные колонки: Дата, SKU, Остаток и Точка или Магазин.",
        )
    with stock_controls[1]:
        st.download_button(
            "Скачать шаблон остатков",
            data=stock_balance_template_bytes(),
            file_name="шаблон_остатков_по_точкам.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="stock_balance_template_download_v7546",
        )

    if stock_upload is not None:
        try:
            parsed_stock = parse_stock_balances(stock_upload.getvalue())
            stock_balances = resolve_stock_balance_points(
                parsed_stock, st.session_state.get("point_mapping", {})
            )
            if stock_balances.empty:
                st.warning(
                    "Файл прочитан, но после сопоставления магазинов с Т1–Т29 не осталось строк. "
                    "Проверьте колонку «Точка» или «Магазин»."
                )
            else:
                available_stock_dates = sorted(
                    value for value in stock_balances["snapshot_date"].dropna().unique()
                    if value <= today
                )
                if not available_stock_dates:
                    st.warning("В файле нет фактических дат не позднее сегодняшнего дня.")
                else:
                    # По умолчанию берём самый свежий снимок. Пользователь может выбрать любой день/день недели.
                    selected_stock_date = st.selectbox(
                        "Дата снимка остатка",
                        options=available_stock_dates,
                        index=len(available_stock_dates) - 1,
                        format_func=lambda value: f"{value:%d.%m.%Y} · {WEEKDAY_RU.get(value.weekday(), '')}",
                        key="stock_snapshot_date_v7546",
                    )
                    stock_snapshot = stock_balances[
                        stock_balances["snapshot_date"].eq(selected_stock_date)
                    ].copy()
                    stock_snapshot = stock_snapshot[
                        stock_snapshot["point"].isin(point_filter)
                    ].copy()
                    st.session_state["stock_balances_v7546"] = stock_balances
                    st.session_state["stock_snapshot_date_v7546"] = selected_stock_date
                    stock_metrics = st.columns(4)
                    stock_metrics[0].metric("Дата остатка", f"{selected_stock_date:%d.%m.%Y}")
                    stock_metrics[1].metric("День недели", WEEKDAY_RU.get(selected_stock_date.weekday(), "—"))
                    stock_metrics[2].metric("Точек в снимке", stock_snapshot["point"].nunique())
                    stock_metrics[3].metric(
                        "Остаток, шт.",
                        f"{stock_snapshot['actual_stock'].sum():,.0f}".replace(",", " "),
                    )
                    st.caption(
                        "Фактический остаток не распределяется искусственно между партиями. "
                        "Сверка выполняется на уровне SKU + точка, а расчётный остаток складывается только из живых партий на дату снимка."
                    )
        except Exception as error:
            st.error(f"Не удалось прочитать файл остатков: {error}")
    elif isinstance(st.session_state.get("stock_balances_v7546"), pd.DataFrame):
        stock_balances = st.session_state.get("stock_balances_v7546", pd.DataFrame()).copy()
        stored_stock_date = st.session_state.get("stock_snapshot_date_v7546")
        if stored_stock_date is not None and not stock_balances.empty:
            selected_stock_date = stored_stock_date
            stock_snapshot = stock_balances[
                stock_balances["snapshot_date"].eq(selected_stock_date)
                & stock_balances["point"].isin(point_filter)
            ].copy()

filtered_detail = daily_detail[
    daily_detail["category"].isin(category_filter)
    & daily_detail["point"].isin(point_filter)
]

total_sales = filtered_sku["sales"].sum()
total_revenue = filtered_sku["revenue"].sum()
coverage = filtered_sku.loc[filtered_sku["category"] != "Не сопоставлено", "sales"].sum() / total_sales if total_sales else 0
metric_columns = st.columns(5)
metric_columns[0].metric("Продано, шт.", f"{total_sales:,.0f}".replace(",", " "))
metric_columns[1].metric("Выручка, ₽", f"{total_revenue:,.0f}".replace(",", " "))
metric_columns[2].metric("Точек", filtered_sku["point"].nunique())
metric_columns[3].metric("Активных SKU", filtered_sku["sku"].nunique())
metric_columns[4].metric("Покрытие сущностями", f"{coverage:.1%}")


def _load_matrix_context_for_active_tab() -> tuple[bytes, str, str, str]:
    """Load the combo matrix only when an active tab actually needs it."""
    matrix_bytes, source, checked_at, error = get_current_combo_matrix_snapshot()
    st.session_state["combo_matrix_signature_v761"] = _combo_matrix_signature(
        matrix_bytes, source
    )
    _matrix_auto_refresh_watcher()
    return matrix_bytes, source, checked_at, error


def _load_detail_plan_for_active_tab() -> pd.DataFrame:
    """Parse freshness/loading plan lazily for Detail tabs only."""
    matrix_bytes, _, _, _ = _load_matrix_context_for_active_tab()
    if not matrix_bytes:
        return pd.DataFrame()
    try:
        return parse_freshness_plan(matrix_bytes)
    except Exception:
        return pd.DataFrame()


class _MainSection:
    """Совместимый контекст раздела без создания нативной полосы st.tabs."""

    def __init__(self, label: str):
        self.label = label

    @property
    def open(self) -> bool:
        return st.session_state.get(SECTION_STATE_KEY) == self.label

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        return False


tab_dashboard, tab_report, tab_points, tab_entities, tab_detail, tab_category_detail, tab_abc, tab_category_analysis, tab_sales_time, tab_category_writeoffs, tab_forecast, tab_plan_check = [
    _MainSection(label) for label, _ in MENU_ITEMS
]

if tab_report.open:
    with tab_report:
        st.subheader("Сверка продаж по двум периодам")
        st.caption(
            "Отчет сравнивает количество проданной продукции по структуре Категория → Сущность → Точка. "
            "Можно выровнять периоды по одинаковым дням недели: приложение возьмет одинаковое количество "
            "понедельников, вторников и т.д. в обоих периодах."
        )

        report_recent_start = previous_month_start
        report_recent_end = previous_month_end
        report_previous_end = report_recent_start - timedelta(days=1)
        report_previous_start = report_previous_end.replace(day=1)

        period_cols = st.columns(2)
        with period_cols[0]:
            report_period_1_input = st.date_input(
                "Период 1 · база сравнения",
                value=(report_previous_start, report_previous_end),
                max_value=today,
                format="DD.MM.YYYY",
                key="report_period_1_v770",
            )
        with period_cols[1]:
            report_period_2_input = st.date_input(
                "Период 2 · сравниваемый",
                value=(report_recent_start, report_recent_end),
                max_value=today,
                format="DD.MM.YYYY",
                key="report_period_2_v770",
            )

        report_match_weekdays = st.checkbox(
            "Сверять по одинаковым дням недели",
            value=True,
            key="report_same_weekdays_v770",
            help=(
                "Например, если в первом периоде 5 понедельников, а во втором 4, "
                "в сверку войдут по 4 понедельника. То же правило применяется к каждому дню недели."
            ),
        )

        all_report_points = [f"Т{number}" for number in range(1, 30) if number != 11]
        report_points = st.multiselect(
            "Точки отчета",
            all_report_points,
            default=all_report_points,
            key="report_points_v770",
        )

        report_build = st.button(
            "Сформировать отчет",
            type="primary",
            use_container_width=True,
            key="report_build_v770",
        )

        if report_build:
            valid_report_1 = isinstance(report_period_1_input, tuple) and len(report_period_1_input) == 2
            valid_report_2 = isinstance(report_period_2_input, tuple) and len(report_period_2_input) == 2
            if not valid_report_1 or not valid_report_2:
                st.error("Для обоих периодов укажите дату начала и дату окончания.")
            elif not report_points:
                st.error("Выберите хотя бы одну точку.")
            else:
                report_period_1 = tuple(report_period_1_input)
                report_period_2 = tuple(report_period_2_input)
                dates_1, dates_2, date_pairs = build_report_comparison_dates(
                    report_period_1,
                    report_period_2,
                    report_match_weekdays,
                )
                if report_match_weekdays and not date_pairs:
                    st.error("Не удалось подобрать одинаковые дни недели для выбранных периодов.")
                else:
                    point_numbers = tuple(sorted(int(point[1:]) for point in report_points))
                    try:
                        with st.spinner("Загружаю продажи для двух периодов…"):
                            raw_1 = load_sales(
                                report_period_1[0],
                                report_period_1[1] + timedelta(days=1),
                                point_numbers,
                            )
                            raw_2 = load_sales(
                                report_period_2[0],
                                report_period_2[1] + timedelta(days=1),
                                point_numbers,
                            )
                        frame_1 = prepare_report_sales_frame(raw_1, entities, "Период 1")
                        frame_2 = prepare_report_sales_frame(raw_2, entities, "Период 2")
                        if report_match_weekdays:
                            frame_1 = frame_1[frame_1["business_date"].isin(set(dates_1))].copy()
                            frame_2 = frame_2[frame_2["business_date"].isin(set(dates_2))].copy()
                        for stale_key in [
                            "report_category_filter_v770",
                            "report_entity_filter_v770",
                            "report_point_filter_v770",
                            "report_graph_category_v770",
                            "report_graph_entity_v770",
                            "report_graph_point_v770",
                            "report_department_export_v771",
                            "report_category_compare_select_v772",
                        ]:
                            st.session_state.pop(stale_key, None)
                        st.session_state["period_comparison_report_v770"] = {
                            "frame_1": frame_1,
                            "frame_2": frame_2,
                            "period_1": report_period_1,
                            "period_2": report_period_2,
                            "dates_1": dates_1,
                            "dates_2": dates_2,
                            "pairs": date_pairs,
                            "match_weekdays": report_match_weekdays,
                            "points": report_points,
                        }
                    except Exception as error:
                        st.error(f"Не удалось сформировать отчет: {error}")

        report_state = st.session_state.get("period_comparison_report_v770")
        if report_state and st.button(
            "Очистить данные отчета из памяти",
            key="report_clear_memory_v7573",
            help="Удаляет сохранённые данные двух периодов из памяти приложения.",
        ):
            st.session_state.pop("period_comparison_report_v770", None)
            st.rerun()
        if report_state:
            report_frame_1 = report_state["frame_1"]
            report_frame_2 = report_state["frame_2"]
            report_period_1 = report_state["period_1"]
            report_period_2 = report_state["period_2"]
            report_dates_1 = report_state["dates_1"]
            report_dates_2 = report_state["dates_2"]
            report_pairs = report_state["pairs"]
            report_match_weekdays_state = bool(report_state["match_weekdays"])
            report_point_options = list(report_state["points"])

            st.success(
                f"Период 1: {report_period_1[0]:%d.%m.%Y}–{report_period_1[1]:%d.%m.%Y} · "
                f"Период 2: {report_period_2[0]:%d.%m.%Y}–{report_period_2[1]:%d.%m.%Y}"
            )
            if report_match_weekdays_state:
                st.caption(
                    f"Сверка выровнена по дням недели: в каждом периоде используется по {len(report_dates_1)} дней."
                )
            else:
                st.caption(
                    f"Без выравнивания дней недели: П1 — {len(report_dates_1)} дней, П2 — {len(report_dates_2)} дней."
                )

            available_report_categories = sorted(
                set(report_frame_1["category"].dropna().astype(str))
                | set(report_frame_2["category"].dropna().astype(str))
            )
            report_filter_cols = st.columns(3)
            with report_filter_cols[0]:
                selected_report_categories = st.multiselect(
                    "Категории",
                    available_report_categories,
                    default=available_report_categories,
                    key="report_category_filter_v770",
                )

            entity_pool = sorted(
                set(
                    pd.concat(
                        [
                            report_frame_1[report_frame_1["category"].isin(selected_report_categories)]["entity"],
                            report_frame_2[report_frame_2["category"].isin(selected_report_categories)]["entity"],
                        ],
                        ignore_index=True,
                    ).dropna().astype(str)
                )
            )
            with report_filter_cols[1]:
                selected_report_entities = st.multiselect(
                    "Сущности",
                    entity_pool,
                    default=entity_pool,
                    key="report_entity_filter_v770",
                )
            with report_filter_cols[2]:
                selected_report_points = st.multiselect(
                    "Точки в результате",
                    report_point_options,
                    default=report_point_options,
                    key="report_point_filter_v770",
                )

            filtered_report_1 = report_frame_1[
                report_frame_1["category"].isin(selected_report_categories)
                & report_frame_1["entity"].isin(selected_report_entities)
                & report_frame_1["point"].isin(selected_report_points)
            ].copy()
            filtered_report_2 = report_frame_2[
                report_frame_2["category"].isin(selected_report_categories)
                & report_frame_2["entity"].isin(selected_report_entities)
                & report_frame_2["point"].isin(selected_report_points)
            ].copy()

            report_tables = build_report_tables(
                filtered_report_1,
                filtered_report_2,
                report_dates_1,
                report_dates_2,
                selected_report_points,
            )

            total_1 = float(filtered_report_1["sales"].sum()) if not filtered_report_1.empty else 0.0
            total_2 = float(filtered_report_2["sales"].sum()) if not filtered_report_2.empty else 0.0
            total_delta = total_2 - total_1
            total_delta_pct = (total_delta / total_1 * 100) if total_1 else None
            avg_1 = total_1 / max(len(report_dates_1), 1)
            avg_2 = total_2 / max(len(report_dates_2), 1)

            report_metrics = st.columns(6)
            report_metrics[0].metric("Период 1, шт.", f"{total_1:,.0f}".replace(",", " "))
            report_metrics[1].metric("Период 2, шт.", f"{total_2:,.0f}".replace(",", " "))
            report_metrics[2].metric(
                "Изменение, шт.",
                f"{total_delta:+,.0f}".replace(",", " "),
            )
            report_metrics[3].metric(
                "Изменение, %",
                f"{total_delta_pct:+.1f}%" if total_delta_pct is not None else "—",
            )
            report_metrics[4].metric("СР/день П1", f"{avg_1:,.1f}".replace(",", " "))
            report_metrics[5].metric("СР/день П2", f"{avg_2:,.1f}".replace(",", " "))

            st.markdown("#### Сравнение по категориям за выбранные периоды")
            st.caption(
                f"Период 1: {report_period_1[0]:%d.%m.%Y}–{report_period_1[1]:%d.%m.%Y} · "
                f"Период 2: {report_period_2[0]:%d.%m.%Y}–{report_period_2[1]:%d.%m.%Y}. "
                "По каждой категории показаны продажи в штуках и выручка в рублях за оба периода, а также их изменение."
            )
            category_compare_table = report_tables["category_summary"].copy()
            if not category_compare_table.empty:
                category_compare_table = category_compare_table.rename(
                    columns={
                        "Изменение, шт.": "Разница, шт.",
                        "Изменение, %": "Разница, %",
                        "Изменение выручки, ₽": "Разница выручки, ₽",
                        "Изменение выручки, %": "Разница выручки, %",
                    }
                )
                category_total_p1 = float(pd.to_numeric(category_compare_table["Период 1, шт."], errors="coerce").fillna(0).sum())
                category_total_p2 = float(pd.to_numeric(category_compare_table["Период 2, шт."], errors="coerce").fillna(0).sum())
                category_total_delta = category_total_p2 - category_total_p1
                category_total_pct = (category_total_delta / category_total_p1 * 100) if category_total_p1 else None
                category_revenue_p1 = float(pd.to_numeric(category_compare_table.get("Выручка П1, ₽", 0.0), errors="coerce").fillna(0).sum())
                category_revenue_p2 = float(pd.to_numeric(category_compare_table.get("Выручка П2, ₽", 0.0), errors="coerce").fillna(0).sum())
                category_revenue_delta = category_revenue_p2 - category_revenue_p1
                category_revenue_pct = (
                    category_revenue_delta / category_revenue_p1 * 100
                    if category_revenue_p1 else None
                )
                total_row = {column: None for column in category_compare_table.columns}
                total_row["Категория"] = "ВСЕГО"
                total_row["Период 1, шт."] = category_total_p1
                total_row["Период 2, шт."] = category_total_p2
                total_row["Выручка П1, ₽"] = category_revenue_p1
                total_row["Выручка П2, ₽"] = category_revenue_p2
                total_row["Разница, шт."] = category_total_delta
                total_row["Разница, %"] = category_total_pct
                total_row["Разница выручки, ₽"] = category_revenue_delta
                total_row["Разница выручки, %"] = category_revenue_pct
                category_compare_table = pd.concat(
                    [category_compare_table, pd.DataFrame([total_row])],
                    ignore_index=True,
                )
                visible_category_columns = [
                    column for column in [
                        "Категория",
                        "Период 1, шт.", "Период 2, шт.",
                        "Выручка П1, ₽", "Выручка П2, ₽",
                        "Разница, шт.", "Разница, %",
                        "Разница выручки, ₽", "Разница выручки, %",
                    ] if column in category_compare_table.columns
                ]
                category_compare_table = category_compare_table[visible_category_columns]
                category_period_1_label = (
                    f"{report_period_1[0]:%d.%m.%Y}, шт."
                    if report_period_1[0] == report_period_1[1]
                    else f"{report_period_1[0]:%d.%m.%Y}–{report_period_1[1]:%d.%m.%Y}, шт."
                )
                category_period_2_label = (
                    f"{report_period_2[0]:%d.%m.%Y}, шт."
                    if report_period_2[0] == report_period_2[1]
                    else f"{report_period_2[0]:%d.%m.%Y}–{report_period_2[1]:%d.%m.%Y}, шт."
                )
                st.caption("Нажмите на строку категории — ниже откроется список SKU, из которых складываются её продажи.")
                category_selection = st.dataframe(
                    category_compare_table,
                    use_container_width=True,
                    hide_index=True,
                    height=min(620, 38 * len(category_compare_table) + 80),
                    column_config={
                        "Период 1, шт.": st.column_config.NumberColumn(label=category_period_1_label, format="%.0f"),
                        "Период 2, шт.": st.column_config.NumberColumn(label=category_period_2_label, format="%.0f"),
                        "Выручка П1, ₽": st.column_config.NumberColumn(
                            label=f"Выручка П1 · {report_period_1[0]:%d.%m.%Y}–{report_period_1[1]:%d.%m.%Y}, ₽",
                            format="%.2f",
                        ),
                        "Выручка П2, ₽": st.column_config.NumberColumn(
                            label=f"Выручка П2 · {report_period_2[0]:%d.%m.%Y}–{report_period_2[1]:%d.%m.%Y}, ₽",
                            format="%.2f",
                        ),
                        "Разница, шт.": st.column_config.NumberColumn(format="%+.0f"),
                        "Разница, %": st.column_config.NumberColumn(format="%+.1f%%"),
                        "Разница выручки, ₽": st.column_config.NumberColumn(format="%+.2f"),
                        "Разница выручки, %": st.column_config.NumberColumn(format="%+.1f%%"),
                    },
                    on_select="rerun",
                    selection_mode="single-row",
                    key="report_category_compare_select_v772",
                )

                selected_category_rows = list(getattr(category_selection.selection, "rows", []) or [])
                if selected_category_rows:
                    selected_category_index = int(selected_category_rows[0])
                    if 0 <= selected_category_index < len(category_compare_table):
                        selected_drill_category = str(
                            category_compare_table.iloc[selected_category_index]["Категория"]
                        ).strip()
                        if selected_drill_category and selected_drill_category != "ВСЕГО":
                            category_sku_breakdown = build_report_category_sku_breakdown(
                                filtered_report_1,
                                filtered_report_2,
                                selected_drill_category,
                                report_dates_1,
                                report_dates_2,
                            )
                            st.markdown(f"#### SKU внутри категории «{selected_drill_category}»")
                            if category_sku_breakdown.empty:
                                st.info("По выбранной категории нет SKU с продажами в этих двух периодах.")
                            else:
                                sku_count = int(
                                    category_sku_breakdown["SKU"].astype(str).str.strip().ne("").sum()
                                )
                                sku_total_1 = float(
                                    pd.to_numeric(
                                        category_sku_breakdown.iloc[-1]["Период 1, шт."], errors="coerce"
                                    ) or 0.0
                                )
                                sku_total_2 = float(
                                    pd.to_numeric(
                                        category_sku_breakdown.iloc[-1]["Период 2, шт."], errors="coerce"
                                    ) or 0.0
                                )
                                sku_revenue_total_1 = float(
                                    pd.to_numeric(
                                        category_sku_breakdown.iloc[-1]["Выручка П1, ₽"], errors="coerce"
                                    ) or 0.0
                                )
                                sku_revenue_total_2 = float(
                                    pd.to_numeric(
                                        category_sku_breakdown.iloc[-1]["Выручка П2, ₽"], errors="coerce"
                                    ) or 0.0
                                )
                                st.caption(
                                    f"SKU с продажами: {sku_count} · "
                                    f"П1: {sku_total_1:,.0f} шт. / {sku_revenue_total_1:,.0f} ₽ · "
                                    f"П2: {sku_total_2:,.0f} шт. / {sku_revenue_total_2:,.0f} ₽. ".replace(",", " ")
                                    + "Строка «ВСЕГО КАТЕГОРИИ» должна совпадать с выбранной категорией выше и по штукам, и по выручке."
                                )
                                category_sku_display = category_sku_breakdown.rename(
                                    columns={
                                        "Период 1, шт.": "Сумма продаж SKU П1, шт.",
                                        "Период 2, шт.": "Сумма продаж SKU П2, шт.",
                                    }
                                )
                                st.dataframe(
                                    category_sku_display,
                                    use_container_width=True,
                                    hide_index=True,
                                    height=min(720, 38 * len(category_sku_display) + 80),
                                    column_config={
                                        "Сумма продаж SKU П1, шт.": st.column_config.NumberColumn(
                                            label=f"Сумма продаж SKU · {category_period_1_label}", format="%.0f"
                                        ),
                                        "Сумма продаж SKU П2, шт.": st.column_config.NumberColumn(
                                            label=f"Сумма продаж SKU · {category_period_2_label}", format="%.0f"
                                        ),
                                        "Выручка П1, ₽": st.column_config.NumberColumn(
                                            label=f"Выручка SKU П1 · {report_period_1[0]:%d.%m.%Y}–{report_period_1[1]:%d.%m.%Y}, ₽",
                                            format="%.2f",
                                        ),
                                        "Выручка П2, ₽": st.column_config.NumberColumn(
                                            label=f"Выручка SKU П2 · {report_period_2[0]:%d.%m.%Y}–{report_period_2[1]:%d.%m.%Y}, ₽",
                                            format="%.2f",
                                        ),
                                        "Изменение, шт.": st.column_config.NumberColumn(format="%+.0f"),
                                        "Изменение, %": st.column_config.NumberColumn(format="%+.1f%%"),
                                        "Изменение выручки, ₽": st.column_config.NumberColumn(format="%+.2f"),
                                        "Изменение выручки, %": st.column_config.NumberColumn(format="%+.1f%%"),
                                        "СР/день П1": st.column_config.NumberColumn(format="%.2f"),
                                        "СР/день П2": st.column_config.NumberColumn(format="%.2f"),
                                    },
                                )
                        elif selected_drill_category == "ВСЕГО":
                            st.info("Выберите конкретную категорию, а не строку «ВСЕГО», чтобы раскрыть её до SKU.")
            else:
                st.info("Для выбранных фильтров нет категорий для сравнения.")

            st.markdown("#### Сверка по категориям и сущностям")
            report_category_entity_display = _append_report_total_row(report_tables["category_entity"])
            st.dataframe(
                report_category_entity_display,
                use_container_width=True,
                hide_index=True,
                height=min(650, 38 * len(report_category_entity_display) + 80),
                column_config={
                    "Период 1, шт.": st.column_config.NumberColumn(format="%.0f"),
                    "Период 2, шт.": st.column_config.NumberColumn(format="%.0f"),
                    "Изменение, шт.": st.column_config.NumberColumn(format="%+.0f"),
                    "Изменение, %": st.column_config.NumberColumn(format="%+.1f%%"),
                    "СР/день П1": st.column_config.NumberColumn(format="%.1f"),
                    "СР/день П2": st.column_config.NumberColumn(format="%.1f"),
                    "Изменение СР/день": st.column_config.NumberColumn(format="%+.1f"),
                },
            )

            st.markdown("#### Сверка одинаковых дней недели")
            st.dataframe(
                report_tables["weekday_summary"],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Период 1, шт.": st.column_config.NumberColumn(format="%.0f"),
                    "Период 2, шт.": st.column_config.NumberColumn(format="%.0f"),
                    "Изменение, шт.": st.column_config.NumberColumn(format="%+.0f"),
                    "Изменение, %": st.column_config.NumberColumn(format="%+.1f%%"),
                },
            )

            st.markdown("#### Количество продаж по точкам")
            matrix_choice = st.radio(
                "Что показать в матрице",
                ["Период 1", "Период 2", "Изменение"],
                horizontal=True,
                key="report_matrix_choice_v770",
            )
            matrix_key = {
                "Период 1": "matrix_1",
                "Период 2": "matrix_2",
                "Изменение": "matrix_delta",
            }[matrix_choice]
            report_matrix_display = report_tables[matrix_key].copy()
            if matrix_choice == "Изменение":
                st.caption(
                    "По каждой точке: сначала разница в штуках, затем изменение в %. "
                    "В конце — «ВСЕГО Δ, шт.» и «ВСЕГО Δ, %» по всем выбранным точкам."
                )
                matrix_column_config = {}
                for column in report_matrix_display.columns:
                    if column.endswith("Δ, шт."):
                        matrix_column_config[column] = st.column_config.NumberColumn(format="%+.0f")
                    elif column.endswith("Δ, %"):
                        matrix_column_config[column] = st.column_config.NumberColumn(format="%+.1f%%")
            else:
                st.caption("Последний столбец «ВСЕГО» — итог количества продаж по всем выбранным точкам.")
                matrix_column_config = {
                    column: st.column_config.NumberColumn(format="%.0f")
                    for column in report_matrix_display.columns
                    if column not in {"Категория", "Сущность"}
                }
            st.dataframe(
                report_matrix_display,
                use_container_width=True,
                hide_index=True,
                height=min(700, 38 * len(report_matrix_display) + 80),
                column_config=matrix_column_config,
            )

            st.markdown("#### Количество продаж по точкам относительно категории")
            st.caption(
                "Строки — категории, столбцы — выбранные точки. В ячейке показано суммарное количество "
                "продаж категории на конкретной точке за выбранный период. Последний столбец — итог по категории, "
                "строка «ВСЕГО» — итог по каждой точке."
            )
            category_point_choice = st.radio(
                "Период для таблицы по категориям и точкам",
                ["Период 1", "Период 2", "Изменение"],
                horizontal=True,
                key="report_category_point_matrix_choice_v75112",
            )
            category_point_key = {
                "Период 1": "category_matrix_1",
                "Период 2": "category_matrix_2",
                "Изменение": "category_matrix_delta",
            }[category_point_choice]
            category_point_display = report_tables[category_point_key].copy()
            if category_point_choice == "Изменение":
                category_point_column_config = {}
                for column in category_point_display.columns:
                    if column.endswith("Δ, шт."):
                        category_point_column_config[column] = st.column_config.NumberColumn(format="%+.0f")
                    elif column.endswith("Δ, %"):
                        category_point_column_config[column] = st.column_config.NumberColumn(format="%+.1f%%")
            else:
                category_point_column_config = {
                    column: st.column_config.NumberColumn(format="%.0f")
                    for column in category_point_display.columns
                    if column != "Категория"
                }
            st.dataframe(
                category_point_display,
                use_container_width=True,
                hide_index=True,
                height=min(700, 38 * len(category_point_display) + 80),
                column_config=category_point_column_config,
            )

            st.markdown("#### Графически: рост и снижение по категориям")
            category_chart_data = report_tables["category_summary"].copy()
            category_chart = go.Figure()
            category_chart.add_trace(
                go.Bar(
                    x=category_chart_data["Категория"],
                    y=category_chart_data["Период 1, шт."],
                    name=f"П1 · {report_period_1[0]:%d.%m}–{report_period_1[1]:%d.%m}",
                )
            )
            category_chart.add_trace(
                go.Bar(
                    x=category_chart_data["Категория"],
                    y=category_chart_data["Период 2, шт."],
                    name=f"П2 · {report_period_2[0]:%d.%m}–{report_period_2[1]:%d.%m}",
                )
            )
            category_chart.update_layout(
                barmode="group",
                yaxis_title="Продано, шт.",
                xaxis_title="Категория",
                height=480,
                margin=dict(l=20, r=20, t=30, b=80),
            )
            st.plotly_chart(category_chart, use_container_width=True)

            delta_chart_data = category_chart_data.sort_values("Изменение, шт.", ascending=True, kind="stable")
            delta_colors = ["#C0504D" if value < 0 else "#70AD47" for value in delta_chart_data["Изменение, шт."]]
            delta_chart = go.Figure(
                go.Bar(
                    x=delta_chart_data["Изменение, шт."],
                    y=delta_chart_data["Категория"],
                    orientation="h",
                    marker_color=delta_colors,
                    text=delta_chart_data["Изменение, шт."].map(lambda value: f"{value:+,.0f}".replace(",", " ")),
                    textposition="outside",
                    hovertemplate="%{y}<br>Изменение: %{x:+,.0f} шт.<extra></extra>",
                )
            )
            delta_chart.update_layout(
                title="Изменение продаж: Период 2 минус Период 1",
                xaxis_title="Изменение, шт.",
                yaxis_title="",
                height=max(360, 45 * max(len(delta_chart_data), 5)),
                margin=dict(l=20, r=80, t=55, b=30),
            )
            st.plotly_chart(delta_chart, use_container_width=True)

            st.markdown("#### Динамика продаж и средняя линия")
            graph_filter_cols = st.columns(4)
            graph_categories = selected_report_categories or available_report_categories
            with graph_filter_cols[0]:
                graph_category = st.selectbox(
                    "Категория на графике",
                    graph_categories,
                    key="report_graph_category_v770",
                ) if graph_categories else None

            graph_entity_options = ["Все сущности"]
            if graph_category:
                graph_entity_options += sorted(
                    set(
                        pd.concat(
                            [
                                report_frame_1[report_frame_1["category"].eq(graph_category)]["entity"],
                                report_frame_2[report_frame_2["category"].eq(graph_category)]["entity"],
                            ],
                            ignore_index=True,
                        ).dropna().astype(str)
                    )
                )
            with graph_filter_cols[1]:
                graph_entity = st.selectbox(
                    "Сущность",
                    graph_entity_options,
                    key="report_graph_entity_v770",
                )
            with graph_filter_cols[2]:
                graph_point = st.selectbox(
                    "Точка",
                    ["Все точки", *selected_report_points],
                    key="report_graph_point_v770",
                )
            with graph_filter_cols[3]:
                rolling_window = st.slider(
                    "Окно среднего, дней",
                    min_value=2,
                    max_value=14,
                    value=3,
                    step=1,
                    key="report_rolling_window_v770",
                )

            dynamic_frame_1 = filtered_report_1.copy()
            dynamic_frame_2 = filtered_report_2.copy()
            if graph_category:
                dynamic_frame_1 = dynamic_frame_1[dynamic_frame_1["category"].eq(graph_category)]
                dynamic_frame_2 = dynamic_frame_2[dynamic_frame_2["category"].eq(graph_category)]
            if graph_entity != "Все сущности":
                dynamic_frame_1 = dynamic_frame_1[dynamic_frame_1["entity"].eq(graph_entity)]
                dynamic_frame_2 = dynamic_frame_2[dynamic_frame_2["entity"].eq(graph_entity)]
            if graph_point != "Все точки":
                dynamic_frame_1 = dynamic_frame_1[dynamic_frame_1["point"].eq(graph_point)]
                dynamic_frame_2 = dynamic_frame_2[dynamic_frame_2["point"].eq(graph_point)]

            dynamic_data = build_report_dynamic_data(
                dynamic_frame_1,
                dynamic_frame_2,
                report_dates_1,
                report_dates_2,
                report_pairs,
                report_match_weekdays_state,
                rolling_window,
            )
            graph_filter_label = " · ".join(
                [
                    graph_category or "Все категории",
                    graph_entity,
                    graph_point,
                ]
            )

            dynamic_chart = go.Figure()
            if not dynamic_data.empty:
                x_labels = dynamic_data.apply(
                    lambda row: f"{int(row['Сравнимый день'])} · {row['День недели']}", axis=1
                )
                custom_1 = dynamic_data["Дата П1"].map(
                    lambda value: value.strftime("%d.%m.%Y") if isinstance(value, date) else "—"
                )
                custom_2 = dynamic_data["Дата П2"].map(
                    lambda value: value.strftime("%d.%m.%Y") if isinstance(value, date) else "—"
                )
                dynamic_chart.add_trace(
                    go.Bar(
                        x=x_labels,
                        y=dynamic_data["Период 1, шт."],
                        name="Период 1 · факт",
                        customdata=custom_1,
                        hovertemplate="%{x}<br>Дата: %{customdata}<br>Продано: %{y:,.0f} шт.<extra></extra>",
                    )
                )
                dynamic_chart.add_trace(
                    go.Bar(
                        x=x_labels,
                        y=dynamic_data["Период 2, шт."],
                        name="Период 2 · факт",
                        customdata=custom_2,
                        hovertemplate="%{x}<br>Дата: %{customdata}<br>Продано: %{y:,.0f} шт.<extra></extra>",
                    )
                )
                dynamic_chart.add_trace(
                    go.Scatter(
                        x=x_labels,
                        y=dynamic_data["СР П1"],
                        mode="lines+markers",
                        name=f"СР П1 · {rolling_window} дн.",
                        line=dict(width=3, dash="dot"),
                    )
                )
                dynamic_chart.add_trace(
                    go.Scatter(
                        x=x_labels,
                        y=dynamic_data["СР П2"],
                        mode="lines+markers",
                        name=f"СР П2 · {rolling_window} дн.",
                        line=dict(width=3),
                    )
                )
            dynamic_chart.update_layout(
                barmode="group",
                title=f"{graph_filter_label} · факт + скользящее среднее",
                yaxis_title="Продано, шт.",
                xaxis_title="Сравнимый день",
                hovermode="x unified",
                height=520,
                margin=dict(l=20, r=20, t=60, b=80),
            )
            st.plotly_chart(dynamic_chart, use_container_width=True)
            st.caption(
                "Столбцы — фактическое количество продаж за день. Линии СР — скользящее среднее; "
                "его окно можно менять ползунком сверху."
            )

            st.markdown("#### Выгрузка по отделам в Excel")
            st.caption(
                "Выберите отдел (категорию) — приложение сформирует отдельный Excel только по нему. "
                "Выгрузка учитывает текущие периоды, выбранные точки и сущности отчёта."
            )
            department_export_options = sorted(
                set(filtered_report_1["category"].dropna().astype(str))
                | set(filtered_report_2["category"].dropna().astype(str))
            )
            if department_export_options:
                department_export_cols = st.columns([2.2, 1.3])
                with department_export_cols[0]:
                    selected_report_department = st.selectbox(
                        "Отдел / категория для выгрузки",
                        department_export_options,
                        key="report_department_export_v771",
                    )

                department_frame_1 = filtered_report_1[
                    filtered_report_1["category"].eq(selected_report_department)
                ].copy()
                department_frame_2 = filtered_report_2[
                    filtered_report_2["category"].eq(selected_report_department)
                ].copy()
                department_tables = build_report_tables(
                    department_frame_1,
                    department_frame_2,
                    report_dates_1,
                    report_dates_2,
                    selected_report_points,
                )
                department_dynamic = build_report_dynamic_data(
                    department_frame_1,
                    department_frame_2,
                    report_dates_1,
                    report_dates_2,
                    report_pairs,
                    report_match_weekdays_state,
                    rolling_window,
                )
                department_excel = build_period_comparison_excel(
                    department_tables,
                    department_dynamic,
                    report_period_1,
                    report_period_2,
                    report_match_weekdays_state,
                    report_dates_1,
                    report_dates_2,
                    rolling_window,
                    f"Отдел: {selected_report_department}",
                )
                department_filename = re.sub(
                    r"[^0-9A-Za-zА-Яа-яЁё_-]+",
                    "_",
                    selected_report_department,
                ).strip("_") or "department"
                with department_export_cols[1]:
                    st.write("")
                    st.write("")
                    st.download_button(
                        "Скачать отдел Excel",
                        data=department_excel,
                        file_name=(
                            f"report_{department_filename}_"
                            f"{report_period_1[0]:%Y%m%d}_{report_period_1[1]:%Y%m%d}__"
                            f"{report_period_2[0]:%Y%m%d}_{report_period_2[1]:%Y%m%d}.xlsx"
                        ),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="report_department_excel_download_v771",
                    )
                st.caption(
                    f"В Excel попадёт только отдел «{selected_report_department}»: итог, сущности, точки, "
                    "матрицы двух периодов, изменение, дни недели и динамика."
                )
            else:
                st.info("Для текущих фильтров нет отделов для отдельной выгрузки.")

            report_excel = build_period_comparison_excel(
                report_tables,
                dynamic_data,
                report_period_1,
                report_period_2,
                report_match_weekdays_state,
                report_dates_1,
                report_dates_2,
                rolling_window,
                graph_filter_label,
            )
            report_html = build_period_comparison_html(
                category_chart,
                delta_chart,
                dynamic_chart,
                "Отчет сравнения продаж",
            )
            export_cols = st.columns(2)
            with export_cols[0]:
                st.download_button(
                    "Скачать табличный отчет Excel",
                    data=report_excel,
                    file_name=(
                        f"report_sales_{report_period_1[0]:%Y%m%d}_{report_period_1[1]:%Y%m%d}__"
                        f"{report_period_2[0]:%Y%m%d}_{report_period_2[1]:%Y%m%d}.xlsx"
                    ),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="report_excel_download_v770",
                )
            with export_cols[1]:
                st.download_button(
                    "Скачать графический отчет HTML",
                    data=report_html,
                    file_name=(
                        f"report_charts_{report_period_1[0]:%Y%m%d}_{report_period_2[0]:%Y%m%d}.html"
                    ),
                    mime="text/html",
                    use_container_width=True,
                    key="report_html_download_v770",
                )


if tab_dashboard.open:
    with tab_dashboard:
        col1, col2 = st.columns(2)
        point_sales = (
            filtered_sku.groupby("point", as_index=False)
            .agg(sales=("sales", "sum"), revenue=("revenue", "sum"))
        )
        point_sales["point_number"] = point_sales["point"].str[1:].astype(int)
        point_sales = point_sales.sort_values("point_number")

        point_sales_chart = go.Figure()
        point_sales_chart.add_trace(
            go.Bar(
                x=point_sales["point"],
                y=point_sales["sales"],
                name="Продано, шт.",
                text=point_sales["sales"].map(lambda value: f"{value:,.0f}".replace(",", " ")),
                textposition="outside",
                hovertemplate="Точка %{x}<br>Продано: %{y:,.0f} шт.<extra></extra>",
            )
        )
        point_sales_chart.add_trace(
            go.Scatter(
                x=point_sales["point"],
                y=point_sales["revenue"],
                name="Выручка, ₽",
                mode="lines+markers+text",
                yaxis="y2",
                text=point_sales["revenue"].map(lambda value: f"{value / 1000:,.0f} тыс. ₽".replace(",", " ")),
                textposition="top center",
                hovertemplate="Точка %{x}<br>Выручка: %{y:,.0f} ₽<extra></extra>",
            )
        )
        point_sales_chart.update_layout(
            title="Продажи и выручка по точкам",
            xaxis_title="Точка",
            yaxis=dict(title="Продано, шт."),
            yaxis2=dict(title="Выручка, ₽", overlaying="y", side="right", showgrid=False),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
            margin=dict(t=85, r=70, b=45, l=55),
        )
        col1.plotly_chart(point_sales_chart, use_container_width=True)

        category_sales = filtered_sku.groupby("category", as_index=False)["sales"].sum().sort_values("sales", ascending=False)
        col2.plotly_chart(px.bar(category_sales, x="category", y="sales", title="Продажи по категориям"), use_container_width=True)

        if not filtered_detail.empty:
            dashboard_time_sales = filtered_detail.copy()
            dashboard_time_sales["hour"] = pd.to_datetime(
                dashboard_time_sales["sale_datetime"]
            ).dt.hour
            dashboard_time_sales["Время суток"] = dashboard_time_sales["hour"].map(
                lambda hour: "День · 06:00–21:59" if 6 <= hour < 22 else "Ночь · 22:00–05:59"
            )
            dashboard_time_by_point = (
                dashboard_time_sales.groupby(["point", "Время суток"], as_index=False)
                .agg(**{"Продано, шт.": ("sales", "sum"), "Выручка, ₽": ("revenue", "sum")})
                .rename(columns={"point": "Точка"})
            )
            dashboard_time_totals = (
                dashboard_time_by_point.groupby("Точка", as_index=False)
                .agg(
                    **{
                        "Всего по точке, шт.": ("Продано, шт.", "sum"),
                        "Выручка точки, ₽": ("Выручка, ₽", "sum"),
                    }
                )
            )
            dashboard_time_by_point = dashboard_time_by_point.merge(
                dashboard_time_totals, on="Точка", how="left"
            )
            dashboard_time_by_point["Доля времени суток"] = (
                dashboard_time_by_point["Продано, шт."]
                / dashboard_time_by_point["Всего по точке, шт."].replace(0, pd.NA)
            )
            dashboard_point_order = sorted(
                dashboard_time_by_point["Точка"].unique(),
                key=lambda value: int(str(value)[1:]),
            )
            day_total = float(
                dashboard_time_by_point.loc[
                    dashboard_time_by_point["Время суток"].str.startswith("День"),
                    "Продано, шт.",
                ].sum()
            )
            night_total = float(
                dashboard_time_by_point.loc[
                    dashboard_time_by_point["Время суток"].str.startswith("Ночь"),
                    "Продано, шт.",
                ].sum()
            )
            day_revenue = float(
                dashboard_time_by_point.loc[
                    dashboard_time_by_point["Время суток"].str.startswith("День"),
                    "Выручка, ₽",
                ].sum()
            )
            night_revenue = float(
                dashboard_time_by_point.loc[
                    dashboard_time_by_point["Время суток"].str.startswith("Ночь"),
                    "Выручка, ₽",
                ].sum()
            )
            time_total = day_total + night_total
            dashboard_time_metrics = st.columns(3)
            dashboard_time_metrics[0].metric(
                "Дневные продажи, шт.",
                f"{day_total:,.0f}".replace(",", " "),
                delta=f"Выручка: {day_revenue:,.0f} ₽".replace(",", " "),
                delta_color="off",
            )
            dashboard_time_metrics[1].metric(
                "Ночные продажи, шт.",
                f"{night_total:,.0f}".replace(",", " "),
                delta=f"Выручка: {night_revenue:,.0f} ₽".replace(",", " "),
                delta_color="off",
            )
            dashboard_time_metrics[2].metric(
                "Доля ночных продаж",
                f"{night_total / time_total:.1%}" if time_total else "0,0%",
            )
            dashboard_time_chart = px.bar(
                dashboard_time_by_point,
                x="Точка",
                y="Продано, шт.",
                color="Время суток",
                barmode="stack",
                text_auto=".0f",
                title="Дневные и ночные продажи по точкам",
                color_discrete_map={
                    "День · 06:00–21:59": "#F4B183",
                    "Ночь · 22:00–05:59": "#4472C4",
                },
                category_orders={
                    "Точка": dashboard_point_order,
                    "Время суток": ["День · 06:00–21:59", "Ночь · 22:00–05:59"],
                },
                custom_data=[
                    "Доля времени суток", "Всего по точке, шт.",
                    "Выручка, ₽", "Выручка точки, ₽",
                ],
            )
            dashboard_time_chart.update_traces(
                hovertemplate=(
                    "Точка: %{x}<br>%{fullData.name}<br>Продано: %{y:.0f} шт."
                    "<br>Доля в точке: %{customdata[0]:.1%}"
                    "<br>Всего по точке: %{customdata[1]:.0f} шт."
                    "<br>Выручка периода: %{customdata[2]:,.0f} ₽"
                    "<br>Выручка точки: %{customdata[3]:,.0f} ₽<extra></extra>"
                )
            )
            dashboard_time_chart.update_layout(
                xaxis_title="Точка",
                yaxis_title="Продано, шт.",
                legend_title="Время суток",
                height=520,
            )
            st.plotly_chart(dashboard_time_chart, use_container_width=True)

        heat = filtered_category.pivot(index="point", columns="category", values="category_share_point")
        heat_quantity = filtered_category.pivot(index="point", columns="category", values="category_sales")
        point_order = sorted(heat.index, key=lambda value: int(value[1:]))
        heat = heat.reindex(point_order)
        heat_quantity = heat_quantity.reindex(index=point_order, columns=heat.columns)
        heat_labels = heat.copy().astype(object)
        for point in heat.index:
            for category in heat.columns:
                share = heat.loc[point, category]
                quantity = heat_quantity.loc[point, category]
                if pd.isna(share):
                    heat_labels.loc[point, category] = "—"
                else:
                    heat_labels.loc[point, category] = f"{share:.0%}<br>{quantity:,.0f} шт.".replace(",", " ")
        heat_values = heat.stack()
        heat_scale_max = max(float(heat_values.max()), 0.10) if not heat_values.empty else 0.10
        red_limit = 0.05 / heat_scale_max
        green_limit = 0.10 / heat_scale_max
        category_share_colors = [
            [0.0, "#F8696B"],
            [max(0.0, red_limit - 0.000001), "#F8696B"],
            [red_limit, "#FFEB84"],
            [max(red_limit, green_limit - 0.000001), "#FFEB84"],
            [green_limit, "#63BE7B"],
            [1.0, "#008A3B"],
        ]
        heat_figure = px.imshow(
            heat,
            aspect="auto",
            color_continuous_scale=category_share_colors,
            zmin=0,
            zmax=heat_scale_max,
            title="Доля категории и количество продаж по точкам",
            labels={"x": "Категория", "y": "Точка", "color": "Доля категории"},
        )
        heat_figure.update_traces(
            text=heat_labels.to_numpy(),
            texttemplate="%{text}",
            hovertemplate="Точка: %{y}<br>Категория: %{x}<br>%{text}<extra></extra>",
        )
        heat_figure.update_layout(
            height=max(500, 34 * len(heat.index) + 180),
            coloraxis_colorbar=dict(tickformat=".0%", title="Доля<br>10% = зелёная зона"),
        )
        st.plotly_chart(heat_figure, use_container_width=True)

if tab_points.open:
    with tab_points:
        top3 = filtered_entity[filtered_entity["entity_rank_category"] <= 3].copy()
        point_category_totals = filtered_category[["point", "category", "category_sales"]].copy()
        top3 = top3.merge(point_category_totals, on=["point", "category"], how="left")
        top3["entity_share_category"] = top3["entity_sales"] / top3["category_sales"].replace(0, pd.NA)
        top3 = top3.sort_values(["point", "category", "entity_rank_category"])
        top3_display = top3.rename(
            columns={
                "point": "Точка",
                "category": "Категория",
                "entity_rank_category": "Ранг",
                "entity": "Сущность",
                "entity_sales": "Продано, шт.",
                "entity_share_category": "Доля сущности в категории",
                "entity_share_point": "Доля сущности в продажах точки",
            }
        )
        st.caption("Топ-3 сущности внутри каждой категории отдельно по каждой выбранной точке.")
        st.dataframe(
            top3_display[
                [
                    "Категория",
                    "Ранг",
                    "Сущность",
                    "Продано, шт.",
                    "Доля сущности в категории",
                    "Доля сущности в продажах точки",
                ]
            ].style.format(
                {
                    "Продано, шт.": "{:,.0f}",
                    "Доля сущности в категории": "{:.1%}",
                    "Доля сущности в продажах точки": "{:.1%}",
                }
            ),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Категория": st.column_config.TextColumn("Категория", width="medium"),
                "Ранг": st.column_config.NumberColumn("Ранг", width="small", format="%d"),
                "Сущность": st.column_config.TextColumn("Сущность", width="large"),
            },
        )
        chart_top3 = top3.groupby(["category", "entity"], as_index=False)["entity_sales"].sum()
        if not chart_top3.empty:
            category_colors = {
                "Вторые блюда": "#F4B183",
                "Десерты": "#C9A0DC",
                "Завтраки": "#FFD966",
                "Напитки": "#5B9BD5",
                "Салаты": "#70AD47",
                "Супы": "#A5A5A5",
                "Сэндвичи": "#ED7D31",
                "Хлеб": "#C55A11",
                "Япония": "#FF6699",
            }
            category_funnel = (
                filtered_category.groupby("category", as_index=False)["category_sales"]
                .sum()
                .sort_values("category_sales", ascending=False)
            )
            category_total = category_funnel["category_sales"].sum()
            category_funnel["category_share"] = category_funnel["category_sales"] / category_total
            entity_labels = (
                chart_top3.sort_values(["category", "entity_sales"], ascending=[True, False])
                .groupby("category")["entity"]
                .apply(lambda values: " • ".join(values.astype(str).head(3)))
            )
            category_funnel["entities"] = category_funnel["category"].map(entity_labels).fillna("Нет данных")
            category_funnel["label"] = category_funnel.apply(
                lambda row: (
                    f"<b>{row['category']}</b> — {row['category_share']:.1%}<br>"
                    f"{row['entities']}<br>{row['category_sales']:,.0f} шт."
                ).replace(",", " "),
                axis=1,
            )
            funnel = go.Figure(
                go.Funnel(
                    y=category_funnel["category"],
                    x=category_funnel["category_sales"],
                    text=category_funnel["label"],
                    textinfo="text",
                    textposition="inside",
                    marker={
                        "color": [category_colors.get(category, "#7F8C8D") for category in category_funnel["category"]],
                        "line": {"color": "white", "width": 2},
                    },
                    connector={"line": {"color": "#D9D9D9", "width": 1}},
                    hovertemplate="%{text}<extra></extra>",
                )
            )
            funnel.update_layout(
                title="Категории по доле продаж и их топ-3 сущности",
                height=max(600, 90 * len(category_funnel)),
                margin=dict(t=70, l=30, r=30, b=20),
                showlegend=False,
            )
            st.plotly_chart(funnel, use_container_width=True)

if tab_entities.open:
    with tab_entities:
        entity_overall = (
            filtered_sku.groupby(["category", "entity"], as_index=False)
            .agg(
                entity_sales=("sales", "sum"),
                entity_revenue=("revenue", "sum"),
                active_sku=("sku", "nunique"),
            )
        )
        entity_overall["category_sales"] = entity_overall.groupby("category")["entity_sales"].transform("sum")
        entity_overall["entity_share_category"] = (
            entity_overall["entity_sales"] / entity_overall["category_sales"].replace(0, pd.NA)
        )
        entity_overall["entity_rank_category"] = (
            entity_overall.groupby("category")["entity_sales"].rank(method="first", ascending=False).astype(int)
        )
        entity_overall["dominant"] = entity_overall["entity_rank_category"].map(
            lambda rank: "Да" if rank == 1 else "Нет"
        )
        entity_overall = entity_overall.sort_values(["category", "entity_rank_category"])
        entity_display = entity_overall.rename(
            columns={
                "category": "Категория",
                "entity": "Сущность",
                "entity_rank_category": "Место в категории",
                "dominant": "Преобладающая сущность",
                "entity_sales": "Продано, шт.",
                "entity_revenue": "Выручка, ₽",
                "entity_share_category": "Доля сущности в категории",
                "active_sku": "Активных SKU",
            }
        )
        st.caption("Общий анализ по всем выбранным точкам. Доли сущностей рассчитаны внутри соответствующей категории.")
        st.dataframe(
            entity_display[
                [
                    "Категория",
                    "Место в категории",
                    "Преобладающая сущность",
                    "Сущность",
                    "Продано, шт.",
                    "Доля сущности в категории",
                    "Выручка, ₽",
                    "Активных SKU",
                ]
            ].style.format(
                {
                    "Продано, шт.": "{:,.0f}",
                    "Доля сущности в категории": "{:.1%}",
                    "Выручка, ₽": "{:,.0f}",
                }
            ),
            use_container_width=True,
            hide_index=True,
        )

if tab_detail.open:
    with tab_detail:
        detail_loading_plan = _load_detail_plan_for_active_tab()
        st.caption(
            f"Все товары за период {period[0]:%d.%m.%Y}–{period[1]:%d.%m.%Y}: "
            "дата и время продажи, точка, категория, сущность, SKU, количество и выручка."
        )
        if not filtered_detail.empty:
            period_detail = filtered_detail.copy()
            detail_point_options = sorted(
                mapped_point_options | set(period_detail["point"].dropna().unique()),
                key=lambda value: int(value[1:]),
            )
            detail_points = st.multiselect(
                "Точки в детализации",
                detail_point_options,
                default=detail_point_options,
            )
            period_detail = period_detail[period_detail["point"].isin(detail_points)]
            period_detail["Час продажи"] = pd.to_datetime(
                period_detail["sale_datetime"]
            ).dt.hour
            period_detail["Временное окно"] = period_detail["Час продажи"].map(
                lambda hour: "День" if 6 <= hour < 22 else "Ночь"
            )
            detail_filter_columns = st.columns([1.2, 1.8])
            with detail_filter_columns[0]:
                detail_time_window = st.selectbox(
                    "Временное окно",
                    ["Все сутки", "День · 06:00–21:59", "Ночь · 22:00–05:59", "Свой интервал"],
                    key="detail_time_window_v51",
                )
            with detail_filter_columns[1]:
                detail_sku_search = st.text_input(
                    "Поиск SKU или товара в детализации",
                    placeholder="Введите SKU, часть кода или название товара",
                    key="detail_sku_search_v51",
                ).strip()
            custom_start_hour = 0
            custom_end_hour = 23
            if detail_time_window == "Свой интервал":
                custom_time_columns = st.columns(2)
                with custom_time_columns[0]:
                    custom_start_hour = st.selectbox(
                        "Начало интервала",
                        list(range(24)),
                        index=8,
                        format_func=lambda hour: f"{hour:02d}:00",
                        key="detail_custom_start_v51",
                    )
                with custom_time_columns[1]:
                    custom_end_hour = st.selectbox(
                        "Последний включённый час",
                        list(range(24)),
                        index=17,
                        format_func=lambda hour: f"{hour:02d}:59",
                        key="detail_custom_end_v51",
                    )
            if detail_time_window.startswith("День"):
                period_detail = period_detail[period_detail["Час продажи"].between(6, 21)]
            elif detail_time_window.startswith("Ночь"):
                period_detail = period_detail[
                    (period_detail["Час продажи"] >= 22)
                    | (period_detail["Час продажи"] <= 5)
                ]
            elif detail_time_window == "Свой интервал":
                if custom_start_hour <= custom_end_hour:
                    period_detail = period_detail[
                        period_detail["Час продажи"].between(
                            custom_start_hour, custom_end_hour
                        )
                    ]
                else:
                    period_detail = period_detail[
                        (period_detail["Час продажи"] >= custom_start_hour)
                        | (period_detail["Час продажи"] <= custom_end_hour)
                    ]
            if detail_sku_search:
                detail_search_mask = (
                    period_detail["sku"].astype(str).str.contains(
                        detail_sku_search, case=False, regex=False, na=False
                    )
                    | period_detail["product_name"].astype(str).str.contains(
                        detail_sku_search, case=False, regex=False, na=False
                    )
                )
                period_detail = period_detail[detail_search_mask]

            detail_batch_enrichment = st.checkbox(
                "Рассчитать партии, дату загрузки и списания для этого среза",
                value=False,
                key="detail_batch_enrichment_v7573",
                help="Тяжёлый FIFO-расчёт выполняется только по текущему отфильтрованному срезу, чтобы приложение не превышало память Streamlit Cloud.",
            )
            if detail_batch_enrichment and not period_detail.empty and not detail_loading_plan.empty:
                with st.spinner("Сопоставляю продажи с партиями…"):
                    period_detail = attach_loading_dates_to_sales(period_detail, detail_loading_plan)
            else:
                period_detail = period_detail.copy()
                period_detail["loading_date"] = pd.NaT
                period_detail["freshness_stage"] = "Нет партии"
                period_detail["plan_quantity"] = pd.NA
                period_detail["batch_expiry_date"] = pd.NaT
                period_detail["batch_sold_total"] = pd.NA
                period_detail["batch_live_remaining"] = pd.NA
                period_detail["batch_writeoff_quantity"] = pd.NA
                period_detail["batch_status"] = "Не рассчитано"

            detail_metrics = st.columns(4)
            detail_metrics[0].metric(
                "Продано в окне, шт.",
                f"{period_detail['sales'].sum():,.0f}".replace(",", " "),
            )
            detail_metrics[1].metric(
                "Выручка в окне, ₽",
                f"{period_detail['revenue'].sum():,.0f}".replace(",", " "),
            )
            detail_metrics[2].metric("Активных SKU", period_detail["sku"].nunique())
            detail_metrics[3].metric(
                "Дней с продажами", period_detail["business_date"].nunique()
            )

            if not period_detail.empty:
                detail_sku_summary = (
                    period_detail.groupby(
                        ["point", "sku", "product_name", "category"],
                        as_index=False,
                        dropna=False,
                    )
                    .agg(
                        **{
                            "Продано, шт.": ("sales", "sum"),
                            "Выручка, ₽": ("revenue", "sum"),
                            "Дата загрузки": (
                                "loading_date",
                                lambda values: ", ".join(
                                    sorted({pd.Timestamp(value).strftime("%d.%m.%Y") for value in values.dropna()})
                                ) or "—",
                            ),
                            "Дней с продажами": ("business_date", "nunique"),
                        }
                    )
                )
                detail_batch_values = period_detail[
                    period_detail["loading_date"].notna()
                    & pd.to_numeric(period_detail.get("plan_quantity"), errors="coerce").notna()
                ].copy()
                if not detail_batch_values.empty:
                    detail_batch_values = detail_batch_values.drop_duplicates(
                        ["point", "sku", "loading_date"]
                    )
                    detail_batch_summary = (
                        detail_batch_values.groupby(["point", "sku"], as_index=False)
                        .agg(
                            **{
                                "План, шт.": ("plan_quantity", "sum"),
                                "Списание, шт.": ("batch_writeoff_quantity", "sum"),
                            }
                        )
                    )
                    detail_sku_summary = detail_sku_summary.merge(
                        detail_batch_summary, on=["point", "sku"], how="left"
                    )
                else:
                    detail_sku_summary["План, шт."] = 0.0
                    detail_sku_summary["Списание, шт."] = 0.0
                detail_sku_summary[["План, шт.", "Списание, шт."]] = detail_sku_summary[
                    ["План, шт.", "Списание, шт."]
                ].fillna(0)
                if selected_stock_date is not None and not stock_snapshot.empty:
                    detail_stock = (
                        stock_snapshot[stock_snapshot["point"].isin(detail_points)]
                        .groupby(["point", "sku"], as_index=False)["actual_stock"].sum()
                        .rename(columns={"actual_stock": "Факт. остаток, шт."})
                    )
                    detail_sku_summary = detail_sku_summary.merge(
                        detail_stock, on=["point", "sku"], how="left"
                    )
                    detail_sku_summary["Дата остатка"] = selected_stock_date
                else:
                    detail_sku_summary["Факт. остаток, шт."] = pd.NA
                    detail_sku_summary["Дата остатка"] = pd.NaT
                detail_sku_summary = (
                    detail_sku_summary.rename(
                        columns={
                            "point": "Точка",
                            "sku": "SKU",
                            "product_name": "Название товара",
                            "category": "Категория",
                        }
                    )
                    .sort_values("Продано, шт.", ascending=False)
                )
                detail_sku_summary = detail_sku_summary[
                    [
                        "Точка", "SKU", "Название товара", "Категория", "Продано, шт.",
                        "План, шт.", "Списание, шт.", "Факт. остаток, шт.", "Дата остатка",
                        "Выручка, ₽", "Дата загрузки", "Дней с продажами",
                    ]
                ]
                st.markdown("#### Итоги по SKU в выбранном временном окне")
                st.dataframe(
                    detail_sku_summary,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Продано, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "План, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Списание, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Факт. остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Дата остатка": st.column_config.DateColumn(format="DD.MM.YYYY"),
                        "Выручка, ₽": st.column_config.NumberColumn(format="%.0f"),
                        "Дней с продажами": st.column_config.NumberColumn(format="%d"),
                    },
                )
            period_detail["Дата"] = pd.to_datetime(period_detail["business_date"]).dt.strftime("%d.%m.%Y")
            period_detail["Время продажи"] = pd.to_datetime(period_detail["sale_datetime"]).dt.strftime("%H:%M:%S")
            period_detail["Дата загрузки"] = pd.to_datetime(period_detail["loading_date"], errors="coerce")
            period_detail["План партии, шт."] = pd.to_numeric(period_detail.get("plan_quantity"), errors="coerce")
            period_detail["Списание партии, шт."] = pd.to_numeric(
                period_detail.get("batch_writeoff_quantity"), errors="coerce"
            )
            period_detail["Статус свежести"] = period_detail["freshness_stage"].fillna("Нет партии")
            period_detail["Статус партии"] = period_detail.get("batch_status", "Нет партии")
            detail_display = period_detail.rename(
                columns={
                    "point": "Точка",
                    "shop_number": "Номер магазина",
                    "sku": "SKU",
                    "product_name": "Название товара",
                    "category": "Категория",
                    "entity": "Сущность",
                    "sales": "Продано, шт.",
                    "revenue": "Выручка, ₽",
                }
            )
            detail_columns = [
                "Дата",
                "Время продажи",
                "Временное окно",
                "Точка",
                "Категория",
                "Сущность",
                "SKU",
                "Название товара",
                "Продано, шт.",
                "План партии, шт.",
                "Списание партии, шт.",
                "Выручка, ₽",
                "Дата загрузки",
                "Статус свежести",
                "Статус партии",
            ]
            detail_table = detail_display[detail_columns]
            detail_table_screen, display_limit, detail_truncated = styler_safe_preview(
                detail_table, preferred_rows=20000
            )
            if detail_truncated:
                st.info(
                    f"В детализации {len(detail_table):,} строк. На экране показаны первые {display_limit:,}; "
                    "полный объём остаётся в выгрузке. Ограничение применяется только к цветному отображению, "
                    "чтобы большая таблица не падала из-за лимита Pandas Styler.".replace(",", " ")
                )
            st.dataframe(
                detail_table_screen.style.apply(style_loading_date_rows, axis=1),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Продано, шт.": st.column_config.NumberColumn("Продано, шт.", format="%.0f"),
                    "План партии, шт.": st.column_config.NumberColumn("План партии, шт.", format="%.0f"),
                    "Списание партии, шт.": st.column_config.NumberColumn("Списание партии, шт.", format="%.0f"),
                    "Выручка, ₽": st.column_config.NumberColumn("Выручка, ₽", format="%.0f"),
                    "Дата загрузки": st.column_config.DateColumn("Дата загрузки", format="DD.MM.YYYY"),
                },
            )
            st.caption(
                "Дата загрузки и план партии берутся из матрицы 2.3 и привязываются к продаже по FIFO, как в окне свежести. "
                "Списание партии показывается только после полного окончания её жизненного цикла; до этого остаток считается живым. "
                "Светло-зелёная дата — основной период, серая — завершающий период срока. "
                "Пустая/серая дата означает, что продажу нельзя надёжно привязать к плановой партии."
            )
            detail_csv = detail_table.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
            st.download_button(
                "Скачать отфильтрованную детализацию (CSV)",
                data=detail_csv,
                file_name="детализация_по_времени.csv",
                mime="text/csv",
                key="download_time_detail_v51",
            )
        else:
            st.info("Для выбранных фильтров детализация отсутствует.")

if tab_category_detail.open:
    with tab_category_detail:
        detail_loading_plan = _load_detail_plan_for_active_tab()
        st.subheader("Детализация продаж внутри категории")
        st.caption(
            f"Период анализа: {period[0]:%d.%m.%Y}–{period[1]:%d.%m.%Y}. "
            "Сначала выберите одну категорию. Таблица ниже учитывает только её товары. "
            "Можно выбрать несколько SKU одновременно: они появятся на одном графике, а покупки будут объединены в общей таблице."
        )
        category_detail_source = daily_detail.copy()
        category_detail_source["sale_datetime"] = pd.to_datetime(
            category_detail_source["sale_datetime"], errors="coerce"
        )
        category_detail_source["business_date"] = pd.to_datetime(
            category_detail_source["business_date"], errors="coerce"
        ).dt.date
        category_detail_categories = sorted(
            category_detail_source["category"].dropna().astype(str).unique()
        )
        if not category_detail_categories:
            st.info("В выбранном периоде нет категорий для детализации.")
        else:
            category_detail_controls = st.columns([1.15, 1.85])
            with category_detail_controls[0]:
                selected_detail_category = st.selectbox(
                    "Категория",
                    category_detail_categories,
                    key="category_detail_category_v72",
                )
            category_detail_mapping = st.session_state.get("point_mapping", {})
            category_detail_point_options = sorted(
                {
                    str(label).strip()
                    for label in category_detail_mapping.values()
                    if str(label).strip().startswith("Т") and str(label).strip() != "Т11"
                },
                key=lambda value: int(value[1:]),
            )
            with category_detail_controls[1]:
                selected_detail_points = st.multiselect(
                    "Точки",
                    category_detail_point_options,
                    default=category_detail_point_options,
                    key="category_detail_points_v72",
                    help="При выборе нескольких точек их продажи суммируются по SKU.",
                )

            category_only = category_detail_source[
                category_detail_source["category"].astype(str).eq(selected_detail_category)
                & category_detail_source["point"].isin(selected_detail_points)
            ].copy()
            if not category_only.empty and not detail_loading_plan.empty:
                # Здесь FIFO считается только по одной выбранной категории, а не по всей базе.
                category_only = attach_loading_dates_to_sales(category_only, detail_loading_plan)
            else:
                category_only["loading_date"] = pd.NaT
                category_only["freshness_stage"] = "Нет партии"
                category_only["plan_quantity"] = pd.NA
                category_only["batch_expiry_date"] = pd.NaT
                category_only["batch_sold_total"] = pd.NA
                category_only["batch_live_remaining"] = pd.NA
                category_only["batch_writeoff_quantity"] = pd.NA
                category_only["batch_status"] = "Нет партии"
            if category_only.empty:
                st.info("По выбранной категории и точкам продаж за период нет.")
            else:
                category_sales_summary = (
                    category_only.groupby(["sku"], as_index=False, dropna=False)
                    .agg(
                        product_name=("product_name", "last"),
                        sold_total=("sales", "sum"),
                        revenue_total=("revenue", "sum"),
                        sales_loading_dates=(
                            "loading_date",
                            lambda values: ", ".join(
                                sorted({pd.Timestamp(value).strftime("%d.%m.%Y") for value in values.dropna()})
                            ) or "—",
                        ),
                        first_purchase=("sale_datetime", "min"),
                        last_purchase=("sale_datetime", "max"),
                        sale_days=("business_date", "nunique"),
                    )
                )
                category_batch_summary = build_loading_batch_summary(
                    detail_loading_plan,
                    category_only,
                    point_labels=selected_detail_points,
                    category=selected_detail_category,
                    display_start=period[0],
                    display_end=period[1],
                )
                if not category_batch_summary.empty:
                    category_plan_summary = (
                        category_batch_summary.groupby("sku", as_index=False, dropna=False)
                        .agg(
                            plan_product_name=("product_name", "last"),
                            planned_total=("analyst_plan", "sum"),
                            live_remaining_total=("live_remaining", "sum"),
                            writeoff_total=("writeoff_quantity", "sum"),
                            plan_loading_dates=(
                                "plan_date",
                                lambda values: ", ".join(
                                    sorted({pd.Timestamp(value).strftime("%d.%m.%Y") for value in values.dropna()})
                                ) or "—",
                            ),
                        )
                    )
                    category_sku_summary = category_sales_summary.merge(
                        category_plan_summary, on="sku", how="outer"
                    )
                else:
                    category_sku_summary = category_sales_summary.copy()
                    category_sku_summary["plan_product_name"] = pd.NA
                    category_sku_summary["planned_total"] = 0.0
                    category_sku_summary["live_remaining_total"] = 0.0
                    category_sku_summary["writeoff_total"] = 0.0
                    category_sku_summary["plan_loading_dates"] = "—"

                category_sku_summary["product_name"] = category_sku_summary["product_name"].fillna(
                    category_sku_summary["plan_product_name"]
                )
                for numeric_column in [
                    "sold_total", "revenue_total", "sale_days", "planned_total",
                    "live_remaining_total", "writeoff_total",
                ]:
                    category_sku_summary[numeric_column] = pd.to_numeric(
                        category_sku_summary.get(numeric_column), errors="coerce"
                    ).fillna(0)
                category_sku_summary["loading_dates"] = category_sku_summary.apply(
                    lambda row: (
                        row["plan_loading_dates"]
                        if str(row.get("plan_loading_dates", "—")) != "—"
                        else str(row.get("sales_loading_dates", "—"))
                    ),
                    axis=1,
                )
                if selected_stock_date is not None and not stock_snapshot.empty:
                    category_stock_summary = (
                        stock_snapshot[stock_snapshot["point"].isin(selected_detail_points)]
                        .groupby("sku", as_index=False)["actual_stock"].sum()
                        .rename(columns={"actual_stock": "actual_stock_total"})
                    )
                    category_sku_summary = category_sku_summary.merge(
                        category_stock_summary, on="sku", how="left"
                    )
                    category_sku_summary["stock_snapshot_date"] = selected_stock_date
                else:
                    category_sku_summary["actual_stock_total"] = pd.NA
                    category_sku_summary["stock_snapshot_date"] = pd.NaT
                category_sku_summary = category_sku_summary.sort_values(
                    ["writeoff_total", "planned_total", "sold_total", "last_purchase"],
                    ascending=[False, False, False, False],
                    na_position="last",
                ).reset_index(drop=True)

                category_sku_display = category_sku_summary.rename(
                    columns={
                        "sku": "SKU",
                        "product_name": "Название товара",
                        "sold_total": "Продано суммарно, шт.",
                        "planned_total": "План, шт.",
                        "live_remaining_total": "Живой остаток, шт.",
                        "writeoff_total": "Списание, шт.",
                        "revenue_total": "Выручка, ₽",
                        "loading_dates": "Дата загрузки",
                        "actual_stock_total": "Факт. остаток, шт.",
                        "stock_snapshot_date": "Дата остатка",
                        "first_purchase": "Первая покупка",
                        "last_purchase": "Последняя покупка",
                        "sale_days": "Дней с продажами",
                    }
                )[
                    [
                        "SKU", "Название товара", "Продано суммарно, шт.", "План, шт.",
                        "Живой остаток, шт.", "Списание, шт.", "Факт. остаток, шт.", "Дата остатка",
                        "Выручка, ₽", "Дата загрузки", "Первая покупка", "Последняя покупка", "Дней с продажами",
                    ]
                ]
                category_metrics = st.columns(5)
                category_metrics[0].metric(
                    "Продано в категории, шт.",
                    f"{category_only['sales'].sum():,.0f}".replace(",", " "),
                )
                category_metrics[1].metric(
                    "План загрузки, шт.",
                    f"{category_batch_summary['analyst_plan'].sum():,.0f}".replace(",", " ")
                    if not category_batch_summary.empty else "0",
                )
                category_metrics[2].metric(
                    "Списание, шт.",
                    f"{category_batch_summary['writeoff_quantity'].sum():,.0f}".replace(",", " ")
                    if not category_batch_summary.empty else "0",
                )
                category_metrics[3].metric(
                    "Выручка категории, ₽",
                    f"{category_only['revenue'].sum():,.0f}".replace(",", " "),
                )
                category_metrics[4].metric("SKU", category_sku_summary["sku"].nunique())

                st.markdown("#### SKU выбранной категории")
                st.caption("Выберите одну или несколько строк SKU. Все выбранные позиции будут показаны вместе на графике и в таблице покупок.")
                category_selection = st.dataframe(
                    category_sku_display,
                    use_container_width=True,
                    hide_index=True,
                    on_select="rerun",
                    selection_mode="multi-row",
                    key="category_detail_sku_table_v75_multi",
                    column_config={
                        "Продано суммарно, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "План, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Живой остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Списание, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Факт. остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Дата остатка": st.column_config.DateColumn(format="DD.MM.YYYY"),
                        "Выручка, ₽": st.column_config.NumberColumn(format="%.0f"),
                        "Первая покупка": st.column_config.DatetimeColumn(format="DD.MM.YYYY HH:mm"),
                        "Последняя покупка": st.column_config.DatetimeColumn(format="DD.MM.YYYY HH:mm"),
                        "Дней с продажами": st.column_config.NumberColumn(format="%d"),
                    },
                )
                selected_rows = sorted({int(row) for row in category_selection.selection.rows})
                if selected_rows:
                    selected_summary = category_sku_summary.iloc[selected_rows].copy()
                    selected_skus = selected_summary["sku"].astype(str).tolist()
                    selected_names = dict(
                        zip(selected_summary["sku"].astype(str), selected_summary["product_name"].astype(str))
                    )
                    selected_sku_history = category_only[
                        category_only["sku"].astype(str).isin(selected_skus)
                    ].copy().sort_values(["sale_datetime", "sku", "point"])
                    selected_sku_batches = (
                        category_batch_summary[
                            category_batch_summary["sku"].astype(str).isin(selected_skus)
                        ].copy()
                        if not category_batch_summary.empty else pd.DataFrame()
                    )

                    if len(selected_skus) == 1:
                        selected_sku = selected_skus[0]
                        st.markdown(f"#### {selected_sku} — {selected_names.get(selected_sku, '')}")
                    else:
                        st.markdown(f"#### Сравнение выбранных SKU: {len(selected_skus)}")
                        selected_labels = [
                            f"{sku} — {selected_names.get(sku, '')}" for sku in selected_skus
                        ]
                        st.caption("Выбрано: " + " • ".join(selected_labels))

                    plan_total = (
                        float(selected_sku_batches["analyst_plan"].sum())
                        if not selected_sku_batches.empty else 0.0
                    )
                    sold_total = (
                        float(selected_sku_history["sales"].sum())
                        if not selected_sku_history.empty else 0.0
                    )
                    live_total = (
                        float(selected_sku_batches["live_remaining"].sum())
                        if not selected_sku_batches.empty else 0.0
                    )
                    writeoff_total = (
                        float(selected_sku_batches["writeoff_quantity"].sum())
                        if not selected_sku_batches.empty else 0.0
                    )
                    actual_stock_series = (
                        pd.to_numeric(selected_summary.get("actual_stock_total"), errors="coerce")
                        if "actual_stock_total" in selected_summary.columns
                        else pd.Series(dtype="float64")
                    )
                    actual_stock_available = bool(actual_stock_series.notna().any())
                    actual_stock_total = actual_stock_series.fillna(0).sum() if actual_stock_available else 0.0
                    sku_metrics = st.columns(6)
                    sku_metrics[0].metric("Выбрано SKU", len(selected_skus))
                    sku_metrics[1].metric("План, шт.", f"{plan_total:,.0f}".replace(",", " "))
                    sku_metrics[2].metric("Продано, шт.", f"{sold_total:,.0f}".replace(",", " "))
                    sku_metrics[3].metric("Живой остаток, шт.", f"{live_total:,.0f}".replace(",", " "))
                    sku_metrics[4].metric("Списание, шт.", f"{writeoff_total:,.0f}".replace(",", " "))
                    sku_metrics[5].metric(
                        "Факт. остаток, шт.",
                        f"{actual_stock_total:,.0f}".replace(",", " ") if actual_stock_available else "—",
                        help=(
                            f"Снимок на {selected_stock_date:%d.%m.%Y} · {WEEKDAY_RU.get(selected_stock_date.weekday(), '')}"
                            if selected_stock_date is not None else "Загрузите файл остатков выше вкладок."
                        ),
                    )

                    purchase_metrics = st.columns(3)
                    purchase_metrics[0].metric(
                        "Выручка, ₽",
                        f"{selected_sku_history['revenue'].sum():,.0f}".replace(",", " ")
                        if not selected_sku_history.empty else "0",
                    )
                    purchase_metrics[1].metric(
                        "Первая покупка",
                        selected_sku_history["sale_datetime"].min().strftime("%d.%m.%Y %H:%M")
                        if not selected_sku_history.empty else "—",
                    )
                    purchase_metrics[2].metric(
                        "Последняя покупка",
                        selected_sku_history["sale_datetime"].max().strftime("%d.%m.%Y %H:%M")
                        if not selected_sku_history.empty else "—",
                    )

                    if not selected_sku_batches.empty:
                        batch_display = selected_sku_batches.rename(
                            columns={
                                "sku": "SKU",
                                "product_name": "Название товара",
                                "plan_date": "Дата загрузки",
                                "point": "Точка",
                                "analyst_plan": "План, шт.",
                                "sold_from_batch": "Продано из партии, шт.",
                                "live_remaining": "Живой остаток, шт.",
                                "writeoff_quantity": "Списание, шт.",
                                "expiry_date": "Последний день срока",
                                "batch_status": "Статус партии",
                            }
                        )[
                            [
                                "SKU", "Название товара", "Дата загрузки", "Точка", "План, шт.",
                                "Продано из партии, шт.", "Живой остаток, шт.", "Списание, шт.",
                                "Последний день срока", "Статус партии",
                            ]
                        ].sort_values(["Дата загрузки", "SKU", "Точка"])
                        st.markdown("#### Партии выбранных SKU по плану меню")
                        st.dataframe(
                            batch_display,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Дата загрузки": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                "План, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Продано из партии, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Живой остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Списание, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Последний день срока": st.column_config.DateColumn(format="DD.MM.YYYY"),
                            },
                        )
                        st.caption(
                            "План берётся непосредственно из 2.3 Матрицы КОМБО. Остаток становится списанием "
                            "только со следующего дня после последнего дня жизненного цикла партии."
                        )

                    depletion_chart = go.Figure()
                    if not selected_sku_history.empty:
                        hourly_total = selected_sku_history.copy()
                        hourly_total["Час"] = hourly_total["sale_datetime"].dt.floor("h")
                        hourly_total = (
                            hourly_total.groupby("Час", as_index=False)
                            .agg(**{"Продано за час, шт.": ("sales", "sum")})
                            .sort_values("Час")
                        )
                        depletion_chart.add_bar(
                            x=hourly_total["Час"],
                            y=hourly_total["Продано за час, шт."],
                            name="Продано за час · все выбранные",
                            marker_color="#4CAF50",
                            opacity=0.35,
                            hovertemplate="%{x|%d.%m.%Y %H:%M}<br>Всего продано: %{y:.0f} шт.<extra></extra>",
                        )

                        writeoff_legend_shown = False
                        for sku_value in selected_skus:
                            sku_history = selected_sku_history[
                                selected_sku_history["sku"].astype(str).eq(sku_value)
                            ].copy()
                            sku_hourly = sku_history.copy()
                            if not sku_hourly.empty:
                                sku_hourly["Час"] = sku_hourly["sale_datetime"].dt.floor("h")
                                sku_hourly = (
                                    sku_hourly.groupby("Час", as_index=False)
                                    .agg(**{"Продано за час, шт.": ("sales", "sum")})
                                    .sort_values("Час")
                                )
                            else:
                                sku_hourly = pd.DataFrame(columns=["Час", "Продано за час, шт."])

                            sku_writeoffs = pd.DataFrame(columns=["Час", "Списание, шт."])
                            if not selected_sku_batches.empty:
                                sku_writeoffs = selected_sku_batches[
                                    selected_sku_batches["sku"].astype(str).eq(sku_value)
                                    & (
                                        pd.to_numeric(
                                            selected_sku_batches["writeoff_quantity"], errors="coerce"
                                        ).fillna(0) > 0
                                    )
                                ].copy()
                                if not sku_writeoffs.empty:
                                    sku_writeoffs["Час"] = (
                                        pd.to_datetime(sku_writeoffs["expiry_date"], errors="coerce")
                                        + pd.Timedelta(days=1)
                                    )
                                    sku_writeoffs["Списание, шт."] = pd.to_numeric(
                                        sku_writeoffs["writeoff_quantity"], errors="coerce"
                                    ).fillna(0.0)
                                    sku_writeoffs = (
                                        sku_writeoffs.groupby("Час", as_index=False)["Списание, шт."]
                                        .sum()
                                        .sort_values("Час")
                                    )

                            event_frames = []
                            if not sku_hourly.empty:
                                sales_events = sku_hourly[["Час", "Продано за час, шт."]].copy()
                                sales_events["Списание, шт."] = 0.0
                                event_frames.append(sales_events)
                            if not sku_writeoffs.empty:
                                writeoff_events_sku = sku_writeoffs.copy()
                                writeoff_events_sku["Продано за час, шт."] = 0.0
                                event_frames.append(writeoff_events_sku)
                            if not event_frames:
                                continue

                            sku_events = pd.concat(event_frames, ignore_index=True)
                            sku_events = (
                                sku_events.groupby("Час", as_index=False)[
                                    ["Продано за час, шт.", "Списание, шт."]
                                ]
                                .sum()
                                .sort_values("Час")
                            )
                            sku_events["Выбыло за событие, шт."] = (
                                sku_events["Продано за час, шт."] + sku_events["Списание, шт."]
                            )
                            sku_events["Накопительное выбытие, шт."] = sku_events[
                                "Выбыло за событие, шт."
                            ].cumsum()

                            short_name = selected_names.get(sku_value, "")
                            trace_name = f"{sku_value} · {short_name}" if short_name else sku_value
                            depletion_chart.add_scatter(
                                x=sku_events["Час"],
                                y=sku_events["Накопительное выбытие, шт."],
                                name=trace_name,
                                mode="lines+markers",
                                line=dict(width=3),
                                hovertemplate=(
                                    f"SKU {sku_value}<br>%{{x|%d.%m.%Y %H:%M}}"
                                    "<br>Накопительно выбыло: %{y:.0f} шт.<extra></extra>"
                                ),
                            )

                            if not sku_writeoffs.empty:
                                for _, writeoff_row in sku_writeoffs.iterrows():
                                    event_time = pd.Timestamp(writeoff_row["Час"])
                                    writeoff_qty = float(writeoff_row["Списание, шт."])
                                    before_rows = sku_events[sku_events["Час"] < event_time]
                                    cumulative_before = (
                                        float(before_rows["Накопительное выбытие, шт."].iloc[-1])
                                        if not before_rows.empty else 0.0
                                    )
                                    cumulative_after_writeoff = cumulative_before + writeoff_qty
                                    depletion_chart.add_scatter(
                                        x=[event_time, event_time],
                                        y=[cumulative_before, cumulative_after_writeoff],
                                        mode="lines",
                                        line=dict(color="#D32F2F", width=5),
                                        name="Списание · выбытие",
                                        showlegend=not writeoff_legend_shown,
                                        hoverinfo="skip",
                                    )
                                    depletion_chart.add_scatter(
                                        x=[event_time],
                                        y=[cumulative_after_writeoff],
                                        mode="markers+text",
                                        marker=dict(color="#D32F2F", size=12, symbol="diamond"),
                                        text=[f"−{writeoff_qty:.0f}"],
                                        textposition="top center",
                                        textfont=dict(color="#B71C1C", size=11),
                                        name="Списание · выбытие",
                                        showlegend=False,
                                        hovertemplate=(
                                            f"SKU {sku_value}<br>%{{x|%d.%m.%Y %H:%M}}"
                                            f"<br>Списано: {writeoff_qty:.0f} шт."
                                            "<br>Накопительно выбыло: %{y:.0f} шт.<extra></extra>"
                                        ),
                                    )
                                    writeoff_legend_shown = True

                    if not selected_sku_batches.empty:
                        for sku_value in selected_skus:
                            sku_batches = selected_sku_batches[
                                selected_sku_batches["sku"].astype(str).eq(sku_value)
                            ].copy()
                            if sku_batches.empty:
                                continue
                            plan_events = (
                                sku_batches.groupby("plan_date", as_index=False)["analyst_plan"].sum()
                                .sort_values("plan_date")
                            )
                            plan_events["Дата события"] = pd.to_datetime(plan_events["plan_date"])
                            plan_events["Накопленный план, шт."] = plan_events["analyst_plan"].cumsum()
                            depletion_chart.add_scatter(
                                x=plan_events["Дата события"],
                                y=plan_events["Накопленный план, шт."],
                                name=f"План · {sku_value}",
                                mode="lines+markers",
                                line=dict(width=2, dash="dot", shape="hv"),
                                hovertemplate=(
                                    f"План SKU {sku_value}<br>%{{x|%d.%m.%Y}}"
                                    "<br>План накопительно: %{y:.0f} шт.<extra></extra>"
                                ),
                            )

                        writeoff_events = selected_sku_batches[
                            pd.to_numeric(
                                selected_sku_batches["writeoff_quantity"], errors="coerce"
                            ).fillna(0) > 0
                        ].copy()
                        if not writeoff_events.empty:
                            writeoff_events["Дата списания"] = (
                                pd.to_datetime(writeoff_events["expiry_date"], errors="coerce")
                                + pd.Timedelta(days=1)
                            )
                            writeoff_events = (
                                writeoff_events.groupby("Дата списания", as_index=False)["writeoff_quantity"]
                                .sum()
                                .sort_values("Дата списания")
                            )

                    sku_loading_dates = (
                        [
                            pd.Timestamp(value).date()
                            for value in selected_sku_batches["plan_date"].dropna().unique()
                        ]
                        if not selected_sku_batches.empty else []
                    )
                    add_freshness_bands_to_depletion_chart(
                        depletion_chart, sku_loading_dates, selected_detail_category
                    )
                    if len(depletion_chart.data) > 0:
                        chart_dates: list[pd.Timestamp] = []
                        if not selected_sku_history.empty:
                            chart_dates.extend(
                                pd.to_datetime(selected_sku_history["sale_datetime"], errors="coerce")
                                .dropna().tolist()
                            )
                        if not selected_sku_batches.empty:
                            chart_dates.extend(
                                pd.to_datetime(selected_sku_batches["plan_date"], errors="coerce")
                                .dropna().tolist()
                            )
                        if "writeoff_events" in locals() and not writeoff_events.empty:
                            chart_dates.extend(
                                pd.to_datetime(writeoff_events["Дата списания"], errors="coerce")
                                .dropna().tolist()
                            )
                        chart_dates = [pd.Timestamp(value) for value in chart_dates if pd.notna(value)]
                        xaxis_range = None
                        if chart_dates:
                            xaxis_range = [
                                min(chart_dates) - pd.Timedelta(hours=12),
                                max(chart_dates) + pd.Timedelta(hours=12),
                            ]
                        depletion_chart.update_layout(
                            title=(
                                "Время выбытия выбранных SKU по часам"
                                if len(selected_skus) > 1 else "Время выбытия товара по часам"
                            ),
                            xaxis_title="Дата и час продажи / события партии",
                            yaxis_title="Количество, шт.",
                            hovermode="x unified",
                            barmode="overlay",
                            legend=dict(orientation="h", y=1.08, x=0),
                            margin=dict(t=90, l=20, r=20, b=30),
                            height=470,
                        )
                        if xaxis_range is not None:
                            depletion_chart.update_xaxes(range=xaxis_range)
                        st.plotly_chart(depletion_chart, use_container_width=True)
                        st.caption(
                            "Линия SKU теперь показывает полное накопительное выбытие: продажи + списание. "
                            "Красный вертикальный скачок и ромб появляются в 00:00 следующего дня после последнего "
                            "дня срока и показывают фактический объём списания. Пунктирная линия — накопленный план; "
                            "зелёные столбцы — продажи за час."
                        )

                    if not selected_sku_history.empty:
                        sku_interval_display = selected_sku_history[
                            [
                                "sale_datetime", "business_date", "shop_number", "point", "sku", "product_name",
                                "sales", "plan_quantity", "revenue", "loading_date", "freshness_stage",
                                "batch_writeoff_quantity", "batch_status",
                            ]
                        ].rename(
                            columns={
                                "sale_datetime": "Дата и время покупки",
                                "business_date": "Операционный день",
                                "shop_number": "Номер магазина",
                                "point": "Точка",
                                "sku": "SKU",
                                "product_name": "Название товара",
                                "sales": "Продано, шт.",
                                "plan_quantity": "План партии, шт.",
                                "revenue": "Выручка, ₽",
                                "loading_date": "Дата загрузки",
                                "freshness_stage": "Статус свежести",
                                "batch_writeoff_quantity": "Списание партии, шт.",
                                "batch_status": "Статус партии",
                            }
                        )
                        st.markdown("#### Все покупки выбранных SKU за выбранный период")
                        sku_interval_display = sku_interval_display[
                            [
                                "Дата и время покупки", "Операционный день", "Номер магазина", "Точка",
                                "SKU", "Название товара", "Продано, шт.", "План партии, шт.", "Выручка, ₽",
                                "Дата загрузки", "Списание партии, шт.", "Статус свежести", "Статус партии",
                            ]
                        ].sort_values(["Дата и время покупки", "Точка", "SKU"])
                        sku_interval_screen, sku_interval_limit, sku_interval_truncated = styler_safe_preview(
                            sku_interval_display, preferred_rows=20000
                        )
                        if sku_interval_truncated:
                            st.info(
                                f"У выбранных SKU {len(sku_interval_display):,} строк покупок. Для цветного просмотра "
                                f"показаны первые {sku_interval_limit:,}; расчёты и данные остаются полными.".replace(",", " ")
                            )
                        st.dataframe(
                            sku_interval_screen.style.apply(style_loading_date_rows, axis=1),
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Дата и время покупки": st.column_config.DatetimeColumn(format="DD.MM.YYYY HH:mm:ss"),
                                "Операционный день": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                "Номер магазина": st.column_config.NumberColumn(format="%d"),
                                "Продано, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "План партии, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Выручка, ₽": st.column_config.NumberColumn(format="%.0f"),
                                "Дата загрузки": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                "Списание партии, шт.": st.column_config.NumberColumn(format="%.0f"),
                            },
                        )
                    elif not selected_sku_batches.empty:
                        st.info("По выбранным SKU есть плановые партии, но в выбранном периоде фактических покупок нет.")
                else:
                    st.info("Выберите одну или несколько строк SKU в таблице, чтобы открыть общий график и все покупки.")

if tab_abc.open:
    with tab_abc:
        st.subheader("ABC-анализ всей продукции")
        st.caption(
            "ABC рассчитывается отдельно внутри каждой категории: A — первые 80% накопительного "
            "результата, B — следующие до 95%, C — оставшаяся часть. Количество и доход анализируются "
            "независимо, затем объединяются в матрицу AA, AB, BC и т. д."
        )
        abc_controls = st.columns([1.25, 1.75])
        with abc_controls[0]:
            abc_period = st.date_input(
                "Период ABC-анализа",
                value=period,
                key="abc_product_period_v39",
            )
        abc_mapping = st.session_state.get("point_mapping", {})
        abc_point_to_shop = {
            str(label).strip(): int(shop_number)
            for shop_number, label in abc_mapping.items()
            if str(label).strip().startswith("Т") and str(label).strip() != "Т11"
        }
        abc_point_options = sorted(abc_point_to_shop, key=lambda label: int(label[1:]))
        with abc_controls[1]:
            abc_selected_points = st.multiselect(
                "Точки ABC-анализа",
                abc_point_options,
                default=abc_point_options,
                key="abc_product_points_v39",
                help="Продажи выбранных точек суммируются перед ABC-классификацией.",
            )
        calculate_abc_button = st.button(
            "Рассчитать ABC продукции",
            type="primary",
            key="calculate_product_abc_v39",
        )
        if calculate_abc_button:
            valid_abc_period = isinstance(abc_period, (tuple, list)) and len(abc_period) == 2
            if not valid_abc_period:
                st.error("Выберите дату начала и дату окончания ABC-анализа.")
            elif not abc_selected_points:
                st.error("Выберите хотя бы одну точку.")
            else:
                abc_start_date, abc_end_date = abc_period
                abc_shop_numbers = tuple(
                    abc_point_to_shop[label] for label in abc_selected_points
                )
                try:
                    abc_sales_source = load_forecast_history(
                        abc_start_date,
                        abc_end_date + timedelta(days=1),
                        abc_shop_numbers,
                    )
                    abc_result = calculate_product_abc(abc_sales_source, entities)
                except Exception as error:
                    st.error(f"Не удалось рассчитать ABC продукции: {error}")
                    abc_result = pd.DataFrame()
                st.session_state["product_abc_result_v39"] = abc_result
                st.session_state["product_abc_context_v39"] = {
                    "start": abc_start_date,
                    "end": abc_end_date,
                    "points": list(abc_selected_points),
                }

        product_abc_result = st.session_state.get("product_abc_result_v39", pd.DataFrame())
        product_abc_context = st.session_state.get("product_abc_context_v39")
        if product_abc_result.empty:
            st.info("Выберите период и точки, затем нажмите «Рассчитать ABC продукции».")
        else:
            if product_abc_context:
                st.caption(
                    f"Расчёт: {product_abc_context['start']:%d.%m.%Y}–"
                    f"{product_abc_context['end']:%d.%m.%Y}; точки: "
                    f"{', '.join(product_abc_context['points'])}."
                )
            abc_filter_columns = st.columns([1.6, 1.4])
            abc_categories = sorted(product_abc_result["category"].astype(str).unique())
            with abc_filter_columns[0]:
                abc_selected_categories = st.multiselect(
                    "Категории продукции",
                    abc_categories,
                    default=abc_categories,
                    key="abc_product_categories_v39",
                )
            with abc_filter_columns[1]:
                abc_sku_search = st.text_input(
                    "Поиск SKU или товара",
                    placeholder="SKU, часть кода или название",
                    key="abc_product_search_v39",
                ).strip()
            abc_settings_columns = st.columns([1.25, 1.15, 1.55, 1.05])
            with abc_settings_columns[0]:
                abc_class_basis = st.selectbox(
                    "Класс ABC по",
                    ["Количеству", "Доходу"],
                    key="abc_class_basis_v55",
                )
            with abc_settings_columns[1]:
                abc_selected_classes = st.multiselect(
                    "Классы ABC",
                    ["A", "B", "C"],
                    default=["A", "B", "C"],
                    key="abc_selected_classes_v55",
                )
            abc_sort_options = {
                "Количество продаж": "sold_quantity",
                "Сумма дохода": "revenue",
                "Доля количества": "quantity_share_category",
                "Доля дохода": "revenue_share_category",
                "Ранг по количеству": "quantity_rank",
                "Ранг по доходу": "revenue_rank",
            }
            with abc_settings_columns[2]:
                abc_sort_label = st.selectbox(
                    "Сортировать по",
                    list(abc_sort_options),
                    key="abc_sort_metric_v55",
                )
            with abc_settings_columns[3]:
                abc_sort_direction = st.selectbox(
                    "Порядок",
                    ["По убыванию", "По возрастанию"],
                    key="abc_sort_direction_v55",
                )
            abc_sort_inside_category = st.checkbox(
                "Сортировать отдельно внутри каждой категории",
                value=True,
                key="abc_sort_inside_category_v55",
            )
            abc_filtered = product_abc_result[
                product_abc_result["category"].isin(abc_selected_categories)
            ].copy()
            if abc_sku_search:
                abc_filtered = abc_filtered[
                    abc_filtered["sku"].astype(str).str.contains(
                        abc_sku_search, case=False, regex=False, na=False
                    )
                    | abc_filtered["product_name"].astype(str).str.contains(
                        abc_sku_search, case=False, regex=False, na=False
                    )
                ].copy()
            abc_class_column = (
                "abc_quantity" if abc_class_basis == "Количеству" else "abc_revenue"
            )
            abc_filtered = abc_filtered[
                abc_filtered[abc_class_column].isin(abc_selected_classes)
            ].copy()
            abc_sort_column = abc_sort_options[abc_sort_label]
            abc_sort_ascending = abc_sort_direction == "По возрастанию"
            if abc_sort_inside_category:
                abc_filtered = abc_filtered.sort_values(
                    ["category", abc_sort_column],
                    ascending=[True, abc_sort_ascending],
                    kind="stable",
                )
            else:
                abc_filtered = abc_filtered.sort_values(
                    abc_sort_column, ascending=abc_sort_ascending, kind="stable"
                )

            if abc_filtered.empty:
                st.warning("По выбранным фильтрам продукция не найдена.")
            else:
                abc_metric_columns = st.columns(4)
                abc_metric_columns[0].metric(
                    "Продано, шт.",
                    f"{abc_filtered['sold_quantity'].sum():,.0f}".replace(",", " "),
                )
                abc_metric_columns[1].metric(
                    "Сумма дохода, ₽",
                    f"{abc_filtered['revenue'].sum():,.0f}".replace(",", " "),
                )
                abc_metric_columns[2].metric("SKU", abc_filtered["sku"].nunique())
                abc_metric_columns[3].metric("Категорий", abc_filtered["category"].nunique())

                abc_card_colors = {
                    "A": ("#E2F0D9", "#548235"),
                    "B": ("#FFF2CC", "#BF9000"),
                    "C": ("#FCE4D6", "#C00000"),
                }
                abc_card_columns = st.columns(3)
                for card_column, abc_class in zip(abc_card_columns, ["A", "B", "C"]):
                    class_rows = abc_filtered[abc_filtered[abc_class_column] == abc_class]
                    background, accent = abc_card_colors[abc_class]
                    with card_column:
                        st.markdown(
                            f"""
                            <div style="background:{background};border-left:6px solid {accent};
                            border-radius:10px;padding:12px 16px;margin:4px 0 14px 0;">
                              <div style="font-size:22px;font-weight:700;color:{accent};">Класс {abc_class}</div>
                              <div style="font-size:14px;margin-top:5px;">SKU: <b>{class_rows['sku'].nunique()}</b></div>
                              <div style="font-size:14px;">Продано: <b>{class_rows['sold_quantity'].sum():,.0f} шт.</b></div>
                              <div style="font-size:14px;">Доход: <b>{class_rows['revenue'].sum():,.0f} ₽</b></div>
                            </div>
                            """.replace(",", " "),
                            unsafe_allow_html=True,
                        )

                abc_category_totals = (
                    abc_filtered.groupby("category", as_index=False)
                    .agg(sold_quantity=("sold_quantity", "sum"), revenue=("revenue", "sum"))
                    .sort_values("sold_quantity", ascending=False)
                )
                abc_chart_columns = st.columns(2)
                abc_quantity_chart = px.bar(
                    abc_category_totals,
                    x="category",
                    y="sold_quantity",
                    title="Количество продаж по категориям",
                    labels={"category": "Категория", "sold_quantity": "Продано, шт."},
                    text_auto=".0f",
                )
                abc_quantity_chart.update_traces(marker_color="#5B9BD5")
                abc_chart_columns[0].plotly_chart(abc_quantity_chart, use_container_width=True)
                abc_revenue_chart = px.bar(
                    abc_category_totals.sort_values("revenue", ascending=False),
                    x="category",
                    y="revenue",
                    title="Сумма дохода по категориям",
                    labels={"category": "Категория", "revenue": "Доход, ₽"},
                    text_auto=".0f",
                )
                abc_revenue_chart.update_traces(marker_color="#70AD47")
                abc_chart_columns[1].plotly_chart(abc_revenue_chart, use_container_width=True)

                st.markdown("### ABC-анализ категорий")
                st.caption(
                    "Нажмите на поле ниже и выберите одну или несколько категорий. "
                    "График и таблица пересчитаются только по выбранным категориям."
                )
                category_abc_available = sorted(
                    product_abc_result["category"].dropna().astype(str).unique().tolist()
                )
                stored_category_selection = st.session_state.get("category_abc_selected_v57")
                if stored_category_selection is not None:
                    valid_stored_categories = [
                        category for category in stored_category_selection
                        if category in category_abc_available
                    ]
                    if valid_stored_categories != list(stored_category_selection):
                        st.session_state["category_abc_selected_v57"] = (
                            valid_stored_categories or category_abc_available
                        )
                category_abc_selected = st.multiselect(
                    "Категории для отдельного анализа",
                    options=category_abc_available,
                    default=category_abc_available,
                    key="category_abc_selected_v57",
                    placeholder="Выберите категории",
                )
                if not category_abc_selected:
                    st.info("Выберите хотя бы одну категорию в плашке выше.")

                category_abc_source = product_abc_result[
                    product_abc_result["category"].isin(category_abc_selected)
                ].copy()
                category_abc_base = (
                    category_abc_source.groupby("category", as_index=False)
                    .agg(
                        sold_quantity=("sold_quantity", "sum"),
                        revenue=("revenue", "sum"),
                        active_sku=("sku", "nunique"),
                    )
                )
                category_quantity_abc = category_abc_base.sort_values(
                    ["sold_quantity", "category"], ascending=[False, True]
                ).copy()
                category_quantity_total = category_quantity_abc["sold_quantity"].sum()
                category_quantity_abc["quantity_share"] = (
                    category_quantity_abc["sold_quantity"] / category_quantity_total
                    if category_quantity_total else 0.0
                )
                category_quantity_abc["quantity_cumulative_share"] = (
                    category_quantity_abc["quantity_share"].cumsum()
                )
                category_quantity_abc["abc_quantity"] = category_quantity_abc[
                    "quantity_cumulative_share"
                ].map(lambda value: "A" if value <= 0.80 else ("B" if value <= 0.95 else "C"))
                category_quantity_abc["quantity_rank"] = range(
                    1, len(category_quantity_abc) + 1
                )

                category_revenue_abc = category_abc_base.sort_values(
                    ["revenue", "category"], ascending=[False, True]
                ).copy()
                category_revenue_total = category_revenue_abc["revenue"].sum()
                category_revenue_abc["revenue_share"] = (
                    category_revenue_abc["revenue"] / category_revenue_total
                    if category_revenue_total else 0.0
                )
                category_revenue_abc["revenue_cumulative_share"] = (
                    category_revenue_abc["revenue_share"].cumsum()
                )
                category_revenue_abc["abc_revenue"] = category_revenue_abc[
                    "revenue_cumulative_share"
                ].map(lambda value: "A" if value <= 0.80 else ("B" if value <= 0.95 else "C"))
                category_revenue_abc["revenue_rank"] = range(
                    1, len(category_revenue_abc) + 1
                )
                category_abc = category_quantity_abc.merge(
                    category_revenue_abc[
                        [
                            "category", "revenue_share", "revenue_cumulative_share",
                            "abc_revenue", "revenue_rank",
                        ]
                    ],
                    on="category",
                    how="left",
                    validate="one_to_one",
                )
                category_abc["abc_matrix"] = (
                    category_abc["abc_quantity"] + category_abc["abc_revenue"]
                )
                category_abc_class_column = (
                    "abc_quantity" if abc_class_basis == "Количеству" else "abc_revenue"
                )
                category_abc = category_abc[
                    category_abc[category_abc_class_column].isin(abc_selected_classes)
                ].copy()
                category_abc_sort_column = (
                    "sold_quantity" if abc_class_basis == "Количеству" else "revenue"
                )
                category_abc = category_abc.sort_values(
                    category_abc_sort_column,
                    ascending=abc_sort_ascending,
                    kind="stable",
                )

                if category_abc.empty:
                    st.info("Для выбранных классов категории не найдены.")
                else:
                    category_abc_chart = px.bar(
                        category_abc,
                        x="category",
                        y=category_abc_sort_column,
                        color=category_abc_class_column,
                        text_auto=".0f",
                        title=(
                            "Категории по количеству продаж и классу ABC"
                            if abc_class_basis == "Количеству"
                            else "Категории по доходу и классу ABC"
                        ),
                        labels={
                            "category": "Категория",
                            "sold_quantity": "Продано, шт.",
                            "revenue": "Доход, ₽",
                            category_abc_class_column: "Класс ABC",
                        },
                        color_discrete_map={"A": "#70AD47", "B": "#FFD966", "C": "#F8696B"},
                        category_orders={"category": category_abc["category"].tolist()},
                    )
                    category_abc_chart.update_layout(height=470)
                    st.plotly_chart(category_abc_chart, use_container_width=True)

                    category_abc_display = category_abc.rename(
                        columns={
                            "category": "Категория",
                            "sold_quantity": "Количество, шт.",
                            "revenue": "Сумма, ₽",
                            "active_sku": "Активных SKU",
                            "quantity_share": "Доля количества",
                            "quantity_cumulative_share": "Накопительная доля количества",
                            "abc_quantity": "ABC количество",
                            "quantity_rank": "Ранг по количеству",
                            "revenue_share": "Доля дохода",
                            "revenue_cumulative_share": "Накопительная доля дохода",
                            "abc_revenue": "ABC доход",
                            "revenue_rank": "Ранг по доходу",
                            "abc_matrix": "Матрица ABC",
                        }
                    )
                    for category_percent_column in [
                        "Доля количества", "Накопительная доля количества",
                        "Доля дохода", "Накопительная доля дохода",
                    ]:
                        category_abc_display[category_percent_column] = (
                            category_abc_display[category_percent_column] * 100
                        )
                    category_abc_columns = [
                        "Категория", "Количество, шт.", "Сумма, ₽",
                        "ABC количество", "ABC доход", "Матрица ABC",
                    ]
                    def category_abc_cell_style(value: object) -> str:
                        text = str(value).strip().upper()
                        if text == "A" or text == "AA":
                            return "background-color: #E2F0D9; color: #375623; font-weight: 700"
                        if "C" in text:
                            return "background-color: #FCE4D6; color: #9C0006; font-weight: 700"
                        if text:
                            return "background-color: #FFF2CC; color: #7F6000; font-weight: 700"
                        return ""
                    category_abc_styled = category_abc_display[category_abc_columns].style.map(
                        category_abc_cell_style,
                        subset=["ABC количество", "ABC доход", "Матрица ABC"],
                    )
                    st.dataframe(
                        category_abc_styled,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Количество, шт.": st.column_config.NumberColumn(format="%.0f"),
                            "Сумма, ₽": st.column_config.NumberColumn(format="%.0f"),
                        },
                    )

                st.markdown("#### Распределение SKU выбранных категорий по классам")
                abc_distribution_source = product_abc_result[
                    product_abc_result["category"].isin(category_abc_selected)
                ].copy()
                if abc_distribution_source.empty:
                    st.info("Выберите категории, чтобы увидеть распределение SKU по классам ABC.")
                else:
                    abc_class_order = ["A", "B", "C"]
                    abc_class_colors = {"A": "#70AD47", "B": "#FFD966", "C": "#F8696B"}

                    def build_abc_sku_count(source: pd.DataFrame, class_column: str) -> pd.DataFrame:
                        complete_index = pd.MultiIndex.from_product(
                            [category_abc_selected, abc_class_order],
                            names=["category", class_column],
                        )
                        return (
                            source.groupby(["category", class_column])["sku"].nunique()
                            .reindex(complete_index, fill_value=0)
                            .rename("sku_count")
                            .reset_index()
                        )

                    abc_quantity_sku_count = build_abc_sku_count(
                        abc_distribution_source, "abc_quantity"
                    )
                    abc_revenue_sku_count = build_abc_sku_count(
                        abc_distribution_source, "abc_revenue"
                    )
                    abc_sku_chart_columns = st.columns(2)
                    abc_sku_quantity_chart = px.bar(
                        abc_quantity_sku_count,
                        x="category",
                        y="sku_count",
                        color="abc_quantity",
                        barmode="group",
                        text_auto=".0f",
                        title="Количество SKU в A/B/C по количеству продаж",
                        labels={
                            "category": "Категория",
                            "sku_count": "Количество SKU",
                            "abc_quantity": "ABC количество",
                        },
                        color_discrete_map=abc_class_colors,
                        category_orders={
                            "category": category_abc_selected,
                            "abc_quantity": abc_class_order,
                        },
                    )
                    abc_sku_quantity_chart.update_layout(height=430)
                    abc_sku_chart_columns[0].plotly_chart(
                        abc_sku_quantity_chart, use_container_width=True
                    )

                    abc_sku_revenue_chart = px.bar(
                        abc_revenue_sku_count,
                        x="category",
                        y="sku_count",
                        color="abc_revenue",
                        barmode="group",
                        text_auto=".0f",
                        title="Количество SKU в A/B/C по доходу",
                        labels={
                            "category": "Категория",
                            "sku_count": "Количество SKU",
                            "abc_revenue": "ABC доход",
                        },
                        color_discrete_map=abc_class_colors,
                        category_orders={
                            "category": category_abc_selected,
                            "abc_revenue": abc_class_order,
                        },
                    )
                    abc_sku_revenue_chart.update_layout(height=430)
                    abc_sku_chart_columns[1].plotly_chart(
                        abc_sku_revenue_chart, use_container_width=True
                    )

                abc_linked_filtered = abc_distribution_source.copy()
                if abc_sku_search and not abc_linked_filtered.empty:
                    abc_linked_filtered = abc_linked_filtered[
                        abc_linked_filtered["sku"].astype(str).str.contains(
                            abc_sku_search, case=False, regex=False, na=False
                        )
                        | abc_linked_filtered["product_name"].astype(str).str.contains(
                            abc_sku_search, case=False, regex=False, na=False
                        )
                    ].copy()
                abc_linked_filtered = abc_linked_filtered[
                    abc_linked_filtered[abc_class_column].isin(abc_selected_classes)
                ].copy()
                if abc_sort_inside_category:
                    abc_linked_filtered = abc_linked_filtered.sort_values(
                        ["category", abc_sort_column],
                        ascending=[True, abc_sort_ascending],
                        kind="stable",
                    )
                else:
                    abc_linked_filtered = abc_linked_filtered.sort_values(
                        abc_sort_column, ascending=abc_sort_ascending, kind="stable"
                    )

                abc_display = abc_linked_filtered.rename(
                    columns={
                        "category": "Категория",
                        "sku": "SKU",
                        "product_name": "Наименование продукции",
                        "sold_quantity": "Количество продаж, шт.",
                        "revenue": "Сумма дохода, ₽",
                        "quantity_rank": "Ранг по количеству",
                        "quantity_share_category": "Доля количества в категории",
                        "quantity_cumulative_share": "Накопительная доля количества",
                        "abc_quantity": "ABC количество",
                        "revenue_rank": "Ранг по доходу",
                        "revenue_share_category": "Доля дохода в категории",
                        "revenue_cumulative_share": "Накопительная доля дохода",
                        "abc_revenue": "ABC доход",
                        "abc_matrix": "Матрица ABC",
                    }
                )
                for abc_percent_column in [
                    "Доля количества в категории", "Накопительная доля количества",
                    "Доля дохода в категории", "Накопительная доля дохода",
                ]:
                    abc_display[abc_percent_column] = abc_display[abc_percent_column] * 100
                abc_display_columns = [
                    "Категория", "SKU", "Наименование продукции", "Количество продаж, шт.",
                    "Сумма дохода, ₽", "Ранг по количеству", "Доля количества в категории",
                    "Накопительная доля количества", "ABC количество", "Ранг по доходу",
                    "Доля дохода в категории", "Накопительная доля дохода", "ABC доход",
                    "Матрица ABC",
                ]
                st.markdown("#### ABC продукции внутри категорий")
                st.caption(
                    "Таблица связана с плашкой «Категории для отдельного анализа». "
                    "Она показывает конкретные SKU только из выбранных категорий."
                )
                def abc_cell_style(value: object) -> str:
                    text = str(value).strip().upper()
                    if text == "A" or text == "AA":
                        return "background-color: #E2F0D9; color: #375623; font-weight: 700"
                    if "C" in text:
                        return "background-color: #FCE4D6; color: #9C0006; font-weight: 700"
                    if text:
                        return "background-color: #FFF2CC; color: #7F6000; font-weight: 700"
                    return ""
                abc_styled_table = abc_display[abc_display_columns].style.map(
                    abc_cell_style,
                    subset=["ABC количество", "ABC доход", "Матрица ABC"],
                )
                st.dataframe(
                    abc_styled_table,
                    use_container_width=True,
                    hide_index=True,
                    height=min(850, 35 * len(abc_display) + 80),
                    column_config={
                        "Количество продаж, шт.": st.column_config.NumberColumn(format="%.0f"),
                        "Сумма дохода, ₽": st.column_config.NumberColumn(format="%.2f"),
                        "Ранг по количеству": st.column_config.NumberColumn(format="%d"),
                        "Ранг по доходу": st.column_config.NumberColumn(format="%d"),
                        "Доля количества в категории": st.column_config.NumberColumn(format="%.1f%%"),
                        "Накопительная доля количества": st.column_config.NumberColumn(format="%.1f%%"),
                        "Доля дохода в категории": st.column_config.NumberColumn(format="%.1f%%"),
                        "Накопительная доля дохода": st.column_config.NumberColumn(format="%.1f%%"),
                    },
                )
                abc_csv = abc_display[abc_display_columns].to_csv(
                    index=False, sep=";", decimal=","
                ).encode("utf-8-sig")
                st.download_button(
                    "Скачать ABC продукции (CSV)",
                    data=abc_csv,
                    file_name="ABC_продукции.csv",
                    mime="text/csv",
                    key="download_product_abc_v39",
                )

if tab_category_analysis.open:
    with tab_category_analysis:
        st.subheader("Анализ категории по дням недели")
        st.caption(
            "Выберите категорию, один или несколько дней недели и до четырёх периодов для сравнения. "
            "Среднее рассчитывается по всем подходящим календарным датам; дни без продаж входят в "
            "расчёт как нулевые. При выборе нескольких точек показываются отдельные значения и общий итог."
        )
        weekday_names = {
            0: "Понедельник", 1: "Вторник", 2: "Среда", 3: "Четверг",
            4: "Пятница", 5: "Суббота", 6: "Воскресенье",
        }
        category_analysis_controls = st.columns([1.2, 1.8, 1.0])
        with category_analysis_controls[0]:
            selected_analysis_category = st.selectbox(
                "Категория",
                sorted(entities["category"].dropna().astype(str).unique()),
                key="category_analysis_category_v37",
            )
        with category_analysis_controls[1]:
            selected_weekday_names = st.multiselect(
                "Дни недели",
                list(weekday_names.values()),
                default=[weekday_names[int(period[1].weekday())]],
                key="category_analysis_weekdays_v40",
                help="Можно выбрать сразу несколько дней, например понедельник, четверг и воскресенье.",
            )
        with category_analysis_controls[2]:
            category_period_count = st.slider(
                "Периодов для сравнения",
                min_value=1,
                max_value=4,
                value=1,
                key="category_analysis_period_count_v40",
            )

        default_period_days = 56
        category_period_columns = st.columns(category_period_count)
        category_periods: list[tuple[str, date, date]] = []
        for period_index, period_column in enumerate(category_period_columns):
            default_end = period[1] - timedelta(days=default_period_days * period_index)
            default_start = default_end - timedelta(days=default_period_days - 1)
            with period_column:
                selected_range = st.date_input(
                    f"Период {period_index + 1}",
                    value=(default_start, default_end),
                    key=f"category_analysis_period_{period_index + 1}_v40",
                )
            if isinstance(selected_range, (tuple, list)) and len(selected_range) == 2:
                range_start, range_end = selected_range
                if range_start > range_end:
                    range_start, range_end = range_end, range_start
                category_periods.append((f"Период {period_index + 1}", range_start, range_end))
        category_mapping = st.session_state.get("point_mapping", {})
        category_point_to_shop = {
            str(label).strip(): int(shop_number)
            for shop_number, label in category_mapping.items()
            if str(label).strip().startswith("Т") and str(label).strip() != "Т11"
        }
        category_point_options = sorted(
            category_point_to_shop,
            key=lambda label: int(label[1:]),
        )
        selected_category_points = st.multiselect(
            "Точки анализа категории",
            category_point_options,
            default=category_point_options[:1],
            key="category_analysis_points_v37",
            help="При выборе нескольких точек приложение покажет каждую отдельно и их общую сумму.",
        )
        category_sku_search = st.text_input(
            "Поиск по SKU или названию товара",
            placeholder="Введите полный SKU, часть кода или название товара",
            key="category_analysis_sku_search_v38",
            help="Поиск фильтрует показатели, графики и таблицу SKU внутри выбранной категории.",
        ).strip()

        if not selected_weekday_names and not category_sku_search:
            st.info("Выберите хотя бы один день недели.")
        elif len(category_periods) != category_period_count:
            st.info("Для каждого периода укажите начальную и конечную дату.")
        elif not selected_category_points:
            st.info("Выберите хотя бы одну точку.")
        else:
            selected_weekdays = {
                number for number, name in weekday_names.items() if name in selected_weekday_names
            }
            target_rows = []
            for period_name, range_start, range_end in category_periods:
                if category_sku_search:
                    target_rows.append({"Период": period_name, "business_date": range_start})
                else:
                    for target_date in pd.date_range(range_start, range_end, freq="D").date:
                        if target_date.weekday() in selected_weekdays:
                            target_rows.append({"Период": period_name, "business_date": target_date})
            category_target_frame = pd.DataFrame(
                target_rows, columns=["Период", "business_date"]
            )
            category_target_dates = sorted(category_target_frame["business_date"].unique())
            category_shop_numbers = tuple(
                category_point_to_shop[label] for label in selected_category_points
            )
            if not category_target_dates:
                st.info("В выбранных периодах нет указанных дней недели для анализа категории.")
                category_sales_history = pd.DataFrame()
            else:
                try:
                    category_sales_history = load_forecast_history(
                        min(range_start for _, range_start, _ in category_periods),
                        max(range_end for _, _, range_end in category_periods) + timedelta(days=1),
                        category_shop_numbers,
                    )
                except Exception as error:
                    st.error(f"Не удалось загрузить продажи для анализа категории: {error}")
                    category_sales_history = pd.DataFrame()

            if category_sales_history.empty:
                if category_sku_search:
                    st.info("За выбранный период продажи SKU не найдены.")
                else:
                    st.info("За выбранные одинаковые дни недели продажи не найдены.")
            else:
                category_sales_history["point"] = (
                    pd.to_numeric(category_sales_history["shop_number"], errors="coerce")
                    .astype("Int64")
                    .map({shop: point for point, shop in category_point_to_shop.items()})
                )
                category_sales_history = category_sales_history.merge(
                    entities[["sku", "category"]], on="sku", how="left", validate="many_to_one"
                )
                category_sales_history["category"] = category_sales_history["category"].fillna(
                    "Не сопоставлено"
                )
                category_lifecycle_history = category_sales_history[
                    category_sales_history["category"] == selected_analysis_category
                ].copy()
                category_full_period_frames = []
                for period_name, range_start, range_end in category_periods:
                    period_frame = category_lifecycle_history[
                        category_lifecycle_history["business_date"].between(
                            range_start, range_end, inclusive="both"
                        )
                    ].copy()
                    period_frame["Период"] = period_name
                    category_full_period_frames.append(period_frame)
                category_full_period_history = pd.concat(
                    category_full_period_frames, ignore_index=True
                )
                category_sales_history = category_lifecycle_history.merge(
                    category_target_frame, on="business_date", how="inner"
                )
                if category_sku_search:
                    lifecycle_sku_mask = (
                        category_lifecycle_history["sku"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                        | category_lifecycle_history["product_name"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                    )
                    category_lifecycle_history = category_lifecycle_history[
                        lifecycle_sku_mask
                    ].copy()
                    full_period_sku_mask = (
                        category_full_period_history["sku"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                        | category_full_period_history["product_name"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                    )
                    category_full_period_history = category_full_period_history[
                        full_period_sku_mask
                    ].copy()
                    category_sku_mask = (
                        category_sales_history["sku"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                        | category_sales_history["product_name"].astype(str).str.contains(
                            category_sku_search, case=False, regex=False, na=False
                        )
                    )
                    category_sales_history = category_sales_history[category_sku_mask].copy()
                    lifecycle_candidates = (
                        category_lifecycle_history[["sku", "product_name"]]
                        .drop_duplicates("sku")
                        .sort_values(["product_name", "sku"])
                    )
                    if not lifecycle_candidates.empty:
                        lifecycle_candidate_labels = {
                            f"{row.sku} — {row.product_name}": row.sku
                            for row in lifecycle_candidates.itertuples(index=False)
                        }
                        selected_lifecycle_label = st.selectbox(
                            "Выберите конкретный SKU",
                            list(lifecycle_candidate_labels),
                            key="category_analysis_selected_sku_v49",
                        )
                        selected_lifecycle_sku = lifecycle_candidate_labels[
                            selected_lifecycle_label
                        ]
                        category_lifecycle_history = category_lifecycle_history[
                            category_lifecycle_history["sku"] == selected_lifecycle_sku
                        ].copy()
                        category_sales_history = category_sales_history[
                            category_sales_history["sku"] == selected_lifecycle_sku
                        ].copy()
                        category_full_period_history = category_full_period_history[
                            category_full_period_history["sku"] == selected_lifecycle_sku
                        ].copy()

                category_table_history = category_full_period_history

                if False and category_sku_search and not category_lifecycle_history.empty:
                    lifecycle_days = product_lifecycle_days(selected_analysis_category)
                    lifecycle_products = category_lifecycle_history[
                        ["sku", "product_name"]
                    ].drop_duplicates()
                    lifecycle_rows = []
                    lifecycle_sale_day_rows = []
                    lifecycle_day_grid_rows = []
                    for target_row in category_target_frame.itertuples(index=False):
                        cycle_start = target_row.business_date
                        cycle_end = min(
                            cycle_start + timedelta(days=lifecycle_days - 1),
                            category_period_end_map[target_row.Период],
                        )
                        for point_name in selected_category_points:
                            point_cycle = category_lifecycle_history[
                                (category_lifecycle_history["point"] == point_name)
                                & (category_lifecycle_history["business_date"] >= cycle_start)
                                & (category_lifecycle_history["business_date"] <= cycle_end)
                            ]
                            cycle_totals = point_cycle.groupby("sku")["sold_quantity"].sum()
                            for product_row in lifecycle_products.itertuples(index=False):
                                product_cycle = point_cycle[point_cycle["sku"] == product_row.sku]
                                product_daily_sales = (
                                    product_cycle.groupby("business_date")["sold_quantity"].sum()
                                )
                                cycle_length = (cycle_end - cycle_start).days + 1
                                for lifecycle_day in range(cycle_length):
                                    lifecycle_date = cycle_start + timedelta(days=lifecycle_day)
                                    lifecycle_quantity = float(
                                        product_daily_sales.get(lifecycle_date, 0.0)
                                    )
                                    lifecycle_day_grid_rows.append(
                                        {
                                            "Период": target_row.Период,
                                            "Точка": point_name,
                                            "SKU": product_row.sku,
                                            "Название товара": product_row.product_name,
                                            "Начало цикла": cycle_start,
                                            "Дата": lifecycle_date,
                                            "День цикла": f"День {lifecycle_day + 1}",
                                            "Продано, шт.": max(lifecycle_quantity, 0.0),
                                        }
                                    )
                                product_daily_sales = product_daily_sales[product_daily_sales > 0]
                                for sale_date, sale_quantity in product_daily_sales.items():
                                    lifecycle_sale_day_rows.append(
                                        {
                                            "Период": target_row.Период,
                                            "Точка": point_name,
                                            "SKU": product_row.sku,
                                            "Название товара": product_row.product_name,
                                            "Дата продажи": sale_date,
                                            "Продано в день, шт.": float(sale_quantity),
                                        }
                                    )
                                lifecycle_rows.append(
                                    {
                                        "Период": target_row.Период,
                                        "Точка": point_name,
                                        "SKU": product_row.sku,
                                        "Название товара": product_row.product_name,
                                        "Начало цикла": cycle_start,
                                        "Конец цикла": cycle_end,
                                        "Продано за цикл, шт.": float(
                                            cycle_totals.get(product_row.sku, 0.0)
                                        ),
                                    }
                                )
                    lifecycle_detail = pd.DataFrame(lifecycle_rows)
                    lifecycle_sale_days = pd.DataFrame(
                        lifecycle_sale_day_rows,
                        columns=[
                            "Период", "Точка", "SKU", "Название товара",
                            "Дата продажи", "Продано в день, шт.",
                        ],
                    ).drop_duplicates(
                        ["Период", "Точка", "SKU", "Дата продажи"], keep="last"
                    )
                    lifecycle_summary_source = lifecycle_detail.copy()
                    lifecycle_sale_days_source = lifecycle_sale_days.copy()
                    if len(selected_category_points) > 1:
                        lifecycle_combined = (
                            lifecycle_detail.groupby(
                                ["Период", "SKU", "Название товара", "Начало цикла", "Конец цикла"],
                                as_index=False,
                            )["Продано за цикл, шт."].sum()
                        )
                        lifecycle_combined.insert(1, "Точка", "Все выбранные")
                        lifecycle_summary_source = pd.concat(
                            [lifecycle_combined, lifecycle_detail], ignore_index=True
                        )
                        lifecycle_combined_days = (
                            lifecycle_sale_days.groupby(
                                ["Период", "SKU", "Название товара", "Дата продажи"],
                                as_index=False,
                            )["Продано в день, шт."].sum()
                        )
                        lifecycle_combined_days.insert(1, "Точка", "Все выбранные")
                        lifecycle_sale_days_source = pd.concat(
                            [lifecycle_combined_days, lifecycle_sale_days], ignore_index=True
                        )
                    lifecycle_summary = (
                        lifecycle_summary_source.groupby(
                            ["Период", "Точка", "SKU", "Название товара"], as_index=False
                        )
                        .agg(
                            **{
                                "Жизненных циклов": ("Начало цикла", "nunique"),
                                "Продано всего, шт.": ("Продано за цикл, шт.", "sum"),
                                "Среднее за жизненный цикл, шт.": (
                                    "Продано за цикл, шт.", "mean"
                                ),
                                "Медиана за жизненный цикл, шт.": (
                                    "Продано за цикл, шт.", "median"
                                ),
                            }
                        )
                        .sort_values(
                            ["Период", "Точка", "Среднее за жизненный цикл, шт."],
                            ascending=[True, True, False],
                        )
                    )
                    if lifecycle_sale_days_source.empty:
                        lifecycle_days_summary = pd.DataFrame(
                            columns=[
                                "Период", "Точка", "SKU", "Название товара",
                                "Дней продажи", "Продано в дни продажи, шт.",
                                "Среднее за день продажи, шт.",
                            ]
                        )
                    else:
                        lifecycle_days_summary = (
                            lifecycle_sale_days_source.groupby(
                                ["Период", "Точка", "SKU", "Название товара"], as_index=False
                            )
                            .agg(
                                **{
                                    "Дней продажи": ("Дата продажи", "nunique"),
                                    "Продано в дни продажи, шт.": ("Продано в день, шт.", "sum"),
                                }
                            )
                        )
                        lifecycle_days_summary["Среднее за день продажи, шт."] = (
                            lifecycle_days_summary["Продано в дни продажи, шт."]
                            / lifecycle_days_summary["Дней продажи"]
                        )
                    lifecycle_summary = lifecycle_summary.merge(
                        lifecycle_days_summary,
                        on=["Период", "Точка", "SKU", "Название товара"],
                        how="left",
                    )
                    lifecycle_summary[
                        ["Дней продажи", "Продано в дни продажи, шт.", "Среднее за день продажи, шт."]
                    ] = lifecycle_summary[
                        ["Дней продажи", "Продано в дни продажи, шт.", "Среднее за день продажи, шт."]
                    ].fillna(0.0)
                    metric_source = (
                        lifecycle_combined
                        if len(selected_category_points) > 1
                        else lifecycle_detail
                    )
                    metric_day_source = (
                        lifecycle_combined_days
                        if len(selected_category_points) > 1
                        else lifecycle_sale_days
                    )
                    sold_days_count = len(metric_day_source)
                    sales_on_sold_days = (
                        float(metric_day_source["Продано в день, шт."].sum())
                        if not metric_day_source.empty else 0.0
                    )
                    average_per_sold_day = (
                        sales_on_sold_days / sold_days_count if sold_days_count else 0.0
                    )
                    st.markdown("#### Среднее выбранного SKU по жизненному циклу")
                    lifecycle_metric_columns = st.columns(4)
                    lifecycle_metric_columns[0].metric(
                        "Срок анализа SKU", f"{lifecycle_days} дней"
                    )
                    lifecycle_metric_columns[1].metric(
                        "Дней продажи", sold_days_count
                    )
                    lifecycle_metric_columns[2].metric(
                        "Продано в эти дни, шт.",
                        f"{sales_on_sold_days:,.0f}".replace(",", " "),
                    )
                    lifecycle_metric_columns[3].metric(
                        "Среднее за день продажи, шт.",
                        f"{average_per_sold_day:,.2f}".replace(",", " "),
                    )
                    st.caption(
                        f"Жизненный цикл: {lifecycle_days} дней. «Дней продажи» — число уникальных "
                        "дат внутри выбранных циклов, когда продажи SKU были больше нуля. Среднее "
                        "равно проданному количеству, делённому только на эти дни."
                    )
                    st.metric(
                        "Проанализировано циклов",
                        len(metric_source[["Период", "Начало цикла"]].drop_duplicates()),
                    )
                    st.dataframe(
                        lifecycle_summary,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Жизненных циклов": st.column_config.NumberColumn(format="%d"),
                            "Продано всего, шт.": st.column_config.NumberColumn(format="%.0f"),
                            "Дней продажи": st.column_config.NumberColumn(format="%d"),
                            "Продано в дни продажи, шт.": st.column_config.NumberColumn(format="%.0f"),
                            "Среднее за день продажи, шт.": st.column_config.NumberColumn(format="%.2f"),
                            "Среднее за жизненный цикл, шт.": st.column_config.NumberColumn(
                                format="%.2f"
                            ),
                            "Медиана за жизненный цикл, шт.": st.column_config.NumberColumn(
                                format="%.2f"
                            ),
                        },
                    )
                    lifecycle_day_grid = pd.DataFrame(lifecycle_day_grid_rows)
                    if not lifecycle_day_grid.empty:
                        lifecycle_day_grid["Строка"] = (
                            lifecycle_day_grid["Период"].astype(str)
                            + " · " + lifecycle_day_grid["Точка"].astype(str)
                            + " · цикл "
                            + pd.to_datetime(lifecycle_day_grid["Начало цикла"]).dt.strftime("%d.%m.%Y")
                        )
                        lifecycle_heat = lifecycle_day_grid.pivot_table(
                            index="Строка",
                            columns="День цикла",
                            values="Продано, шт.",
                            aggfunc="sum",
                            fill_value=0.0,
                        )
                        lifecycle_heat = lifecycle_heat.reindex(
                            columns=[f"День {number}" for number in range(1, lifecycle_days + 1)]
                        )
                        lifecycle_date_heat = lifecycle_day_grid.pivot_table(
                            index="Строка",
                            columns="День цикла",
                            values="Дата",
                            aggfunc="first",
                        ).reindex(index=lifecycle_heat.index, columns=lifecycle_heat.columns)
                        lifecycle_date_text = lifecycle_date_heat.apply(
                            lambda column: pd.to_datetime(column).dt.strftime("%d.%m.%Y")
                        )
                        lifecycle_heat_text = lifecycle_heat.apply(
                            lambda column: column.map(
                                lambda value: "" if pd.isna(value) else str(int(round(value)))
                            )
                        )
                        heat_max = max(float(lifecycle_heat.stack().max()), 1.0)
                        lifecycle_heatmap = go.Figure(
                            data=go.Heatmap(
                                z=lifecycle_heat.to_numpy(),
                                x=lifecycle_heat.columns.tolist(),
                                y=lifecycle_heat.index.tolist(),
                                text=lifecycle_heat_text.to_numpy(),
                                customdata=lifecycle_date_text.to_numpy(),
                                texttemplate="%{text}",
                                hovertemplate=(
                                    "%{y}<br>%{x}<br>Дата: %{customdata}<br>Продано: %{z:.0f} шт."
                                    "<extra></extra>"
                                ),
                                colorscale=[
                                    [0.0, "#D9D9D9"],
                                    [0.000001, "#D9D9D9"],
                                    [0.000002, "#A9D18E"],
                                    [1.0, "#008A3B"],
                                ],
                                zmin=0,
                                zmax=heat_max,
                                colorbar_title="Продано, шт.",
                                xgap=4,
                                ygap=4,
                            )
                        )
                        lifecycle_heatmap.update_layout(
                            title=f"Жизненные циклы SKU {selected_lifecycle_sku}",
                            xaxis_title="День жизненного цикла",
                            yaxis_title="Период · точка · дата начала цикла",
                            height=max(360, min(1000, 70 + 34 * len(lifecycle_heat))),
                            margin=dict(l=20, r=20, t=60, b=30),
                        )
                        st.plotly_chart(lifecycle_heatmap, use_container_width=True)
                        st.caption(
                            "Зелёная ячейка — SKU продавался в этот день жизненного цикла. "
                            "Серая ячейка — продаж выбранного SKU не было. Число внутри — продано, шт."
                        )
                elif False and category_sku_search:
                    st.warning(
                        f"Для запроса «{category_sku_search}» продажи в жизненном цикле не найдены."
                    )

                if category_table_history.empty:
                    if category_sku_search:
                        st.warning(
                            f"В категории «{selected_analysis_category}» по запросу "
                            f"«{category_sku_search}» продажи не найдены."
                        )
                    else:
                        st.info("В выбранной категории за эти дни продаж не найдено.")
                else:
                    if category_sku_search:
                        st.caption(
                            f"Показаны результаты поиска «{category_sku_search}» внутри категории "
                            f"«{selected_analysis_category}»."
                        )
                    period_day_counts = (
                        category_target_frame.groupby("Период")["business_date"].nunique().to_dict()
                    )
                    analysis_day_count = int(sum(period_day_counts.values()))
                    combined_daily = (
                        category_sales_history.groupby(["Период", "business_date"], as_index=False)[
                            "sold_quantity"
                        ].sum()
                        .rename(columns={"business_date": "Дата", "sold_quantity": "Продано, шт."})
                    )
                    complete_dates = category_target_frame.rename(columns={"business_date": "Дата"})
                    combined_daily = complete_dates.merge(
                        combined_daily, on=["Период", "Дата"], how="left"
                    ).fillna({"Продано, шт.": 0.0})
                    combined_daily["Дата"] = pd.to_datetime(combined_daily["Дата"])
                    point_daily = (
                        category_sales_history.groupby(["Период", "point", "business_date"], as_index=False)[
                            "sold_quantity"
                        ].sum()
                    )
                    point_totals = (
                        point_daily.groupby(["Период", "point"], as_index=False)["sold_quantity"].sum()
                        .rename(columns={"point": "Точка", "sold_quantity": "Продано всего, шт."})
                    )
                    point_grid = pd.MultiIndex.from_product(
                        [[name for name, _, _ in category_periods], selected_category_points],
                        names=["Период", "Точка"],
                    ).to_frame(index=False)
                    point_totals = point_grid.merge(
                        point_totals, on=["Период", "Точка"], how="left"
                    )
                    point_totals["Продано всего, шт."] = point_totals["Продано всего, шт."].fillna(0.0)
                    point_totals["Среднее за выбранный день, шт."] = (
                        point_totals["Продано всего, шт."]
                        / point_totals["Период"].map(period_day_counts)
                    )
                    combined_total = float(combined_daily["Продано, шт."].sum())
                    combined_average = combined_total / analysis_day_count

                    if not category_sku_search:
                        category_metric_columns = st.columns(4)
                        category_metric_columns[0].metric(
                            "Среднее категории, шт.", f"{combined_average:,.1f}".replace(",", " ")
                        )
                        category_metric_columns[1].metric(
                            "Продано за выбранные дни, шт.",
                            f"{combined_total:,.0f}".replace(",", " "),
                        )
                        category_metric_columns[2].metric("Дней в расчёте", analysis_day_count)
                        category_metric_columns[3].metric("Точек", len(selected_category_points))
                        st.caption(
                            "Выбранные дни: "
                            + ", ".join(name.lower() for name in selected_weekday_names)
                            + ". Средние и итоги ниже разделены по выбранным периодам."
                        )

                        category_chart_columns = st.columns(2)
                        category_daily_chart = px.bar(
                            combined_daily,
                            x="Дата",
                            y="Продано, шт.",
                            title=f"Продажи категории «{selected_analysis_category}» по выбранным датам",
                            text_auto=".0f",
                            color="Период",
                            barmode="group",
                        )
                        category_chart_columns[0].plotly_chart(
                            category_daily_chart, use_container_width=True
                        )

                        point_chart_data = point_totals[
                            ["Период", "Точка", "Среднее за выбранный день, шт."]
                        ].copy()
                        if len(selected_category_points) > 1:
                            combined_by_period = combined_daily.groupby("Период", as_index=False)[
                                "Продано, шт."
                            ].sum()
                            combined_by_period["Среднее за выбранный день, шт."] = (
                                combined_by_period["Продано, шт."]
                                / combined_by_period["Период"].map(period_day_counts)
                            )
                            combined_by_period["Точка"] = "Все выбранные"
                            point_chart_data = pd.concat(
                                [
                                    point_chart_data,
                                    combined_by_period[
                                        ["Период", "Точка", "Среднее за выбранный день, шт."]
                                    ],
                                ],
                                ignore_index=True,
                            )
                        point_average_chart = px.bar(
                            point_chart_data,
                            x="Точка",
                            y="Среднее за выбранный день, шт.",
                            title="Средние продажи категории по точкам",
                            text_auto=".1f",
                            color="Период",
                            barmode="group",
                        )
                        category_chart_columns[1].plotly_chart(
                            point_average_chart, use_container_width=True
                        )

                    sku_daily_by_point = (
                        category_table_history.groupby(
                            ["Период", "point", "sku", "product_name", "business_date"],
                            as_index=False,
                            dropna=False,
                        )["sold_quantity"].sum()
                    )
                    sku_daily_by_point = sku_daily_by_point[
                        sku_daily_by_point["sold_quantity"] > 0
                    ]
                    sku_by_point = (
                        sku_daily_by_point.groupby(
                            ["Период", "point", "sku", "product_name"],
                            as_index=False,
                            dropna=False,
                        )
                        .agg(
                            total_sales=("sold_quantity", "sum"),
                            active_days=("business_date", "nunique"),
                        )
                    )
                    sku_by_point["average_sales"] = (
                        sku_by_point["total_sales"] / sku_by_point["active_days"]
                    )
                    sku_by_point = sku_by_point.rename(
                        columns={
                            "point": "Точка", "sku": "SKU", "product_name": "Название товара",
                            "total_sales": "Продано всего, шт.", "active_days": "Дней с продажами",
                            "average_sales": "Среднее за день продажи, шт.",
                        }
                    )
                    sku_tables = [sku_by_point]
                    if len(selected_category_points) > 1:
                        sku_daily_combined = (
                            category_table_history.groupby(
                                ["Период", "sku", "product_name", "business_date"],
                                as_index=False,
                                dropna=False,
                            )["sold_quantity"].sum()
                        )
                        sku_daily_combined = sku_daily_combined[
                            sku_daily_combined["sold_quantity"] > 0
                        ]
                        sku_combined = (
                            sku_daily_combined.groupby(
                                ["Период", "sku", "product_name"],
                                as_index=False,
                                dropna=False,
                            )
                            .agg(
                                total_sales=("sold_quantity", "sum"),
                                active_days=("business_date", "nunique"),
                            )
                        )
                        sku_combined["average_sales"] = (
                            sku_combined["total_sales"] / sku_combined["active_days"]
                        )
                        sku_combined.insert(0, "Точка", "Все выбранные")
                        sku_combined = sku_combined.rename(
                            columns={
                                "sku": "SKU", "product_name": "Название товара",
                                "total_sales": "Продано всего, шт.", "active_days": "Дней с продажами",
                                "average_sales": "Среднее за день продажи, шт.",
                            }
                        )
                        sku_tables.insert(0, sku_combined)
                    category_sku_table = pd.concat(sku_tables, ignore_index=True).sort_values(
                        ["Период", "Точка", "Среднее за день продажи, шт."],
                        ascending=[True, True, False]
                    ).reset_index(drop=True)
                    selected_category_table_sku = None
                    if not category_sku_search:
                        st.markdown("#### Средние продажи позиций SKU внутри категории")
                        st.caption(
                            "Таблица SKU всегда рассчитывается по всем календарным датам выбранного "
                            "периода. Выберите строку по SKU или названию, чтобы сразу открыть факт "
                            "продаж этого SKU по дням."
                        )
                        category_sku_selection = st.dataframe(
                            category_sku_table,
                            use_container_width=True,
                            hide_index=True,
                            on_select="rerun",
                            selection_mode="single-row",
                            key="category_analysis_sku_table_v74",
                            column_config={
                                "Продано всего, шт.": st.column_config.NumberColumn(format="%.0f"),
                                "Среднее за день продажи, шт.": st.column_config.NumberColumn(format="%.2f"),
                                "Дней с продажами": st.column_config.NumberColumn(format="%d"),
                            },
                        )
                        category_selected_rows = list(category_sku_selection.selection.rows)
                        if category_selected_rows:
                            selected_category_table_sku = str(
                                category_sku_table.iloc[int(category_selected_rows[0])]["SKU"]
                            )
                            selected_lifecycle_sku = selected_category_table_sku
                            sku_daily_by_point = sku_daily_by_point[
                                sku_daily_by_point["sku"].astype(str).eq(
                                    selected_category_table_sku
                                )
                            ].copy()
                            if len(selected_category_points) > 1:
                                sku_daily_combined = sku_daily_combined[
                                    sku_daily_combined["sku"].astype(str).eq(
                                        selected_category_table_sku
                                    )
                                ].copy()
                    if category_sku_search or selected_category_table_sku:
                        sku_daily_display = sku_daily_by_point.rename(
                            columns={
                                "point": "Точка",
                                "sku": "SKU",
                                "product_name": "Название товара",
                                "business_date": "Дата продажи",
                                "sold_quantity": "Продано за день, шт.",
                            }
                        )
                        daily_tables = [sku_daily_display]
                        if len(selected_category_points) > 1:
                            sku_daily_combined_display = sku_daily_combined.copy()
                            sku_daily_combined_display.insert(1, "point", "Все выбранные")
                            sku_daily_combined_display = sku_daily_combined_display.rename(
                                columns={
                                    "point": "Точка",
                                    "sku": "SKU",
                                    "product_name": "Название товара",
                                    "business_date": "Дата продажи",
                                    "sold_quantity": "Продано за день, шт.",
                                }
                            )
                            daily_tables.insert(0, sku_daily_combined_display)
                        sku_daily_table = pd.concat(daily_tables, ignore_index=True)
                        sku_daily_table["Дата продажи"] = pd.to_datetime(
                            sku_daily_table["Дата продажи"]
                        )
                        sku_daily_table["День недели"] = (
                            sku_daily_table["Дата продажи"].dt.weekday.map(weekday_names)
                        )
                        sku_daily_table = sku_daily_table.sort_values(
                            ["Период", "Точка", "Дата продажи"]
                        )
                        st.markdown("#### Факт продаж выбранного SKU по дням")
                        st.caption(
                            "Показаны только даты внутри выбранного периода, когда продажи SKU были "
                            "больше нуля. Под каждой датой указан соответствующий день недели."
                        )
                        st.dataframe(
                            sku_daily_table[
                                [
                                    "Период", "Точка", "Дата продажи", "День недели", "SKU",
                                    "Название товара", "Продано за день, шт.",
                                ]
                            ],
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Дата продажи": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                "Продано за день, шт.": st.column_config.NumberColumn(format="%.0f")
                            },
                        )
                        calendar_rows = []
                        for period_name, range_start, range_end in category_periods:
                            for point_name in selected_category_points:
                                point_sales_lookup = (
                                    sku_daily_by_point[
                                        (sku_daily_by_point["Период"] == period_name)
                                        & (sku_daily_by_point["point"] == point_name)
                                    ]
                                    .set_index("business_date")["sold_quantity"]
                                    .to_dict()
                                )
                                for calendar_date in pd.date_range(
                                    range_start, range_end, freq="D"
                                ).date:
                                    calendar_rows.append(
                                        {
                                            "Период": period_name,
                                            "Точка": point_name,
                                            "Дата": calendar_date,
                                            "Продано, шт.": float(
                                                point_sales_lookup.get(calendar_date, 0.0)
                                            ),
                                        }
                                    )
                            if len(selected_category_points) > 1:
                                combined_sales_lookup = (
                                    sku_daily_combined[
                                        sku_daily_combined["Период"] == period_name
                                    ]
                                    .set_index("business_date")["sold_quantity"]
                                    .to_dict()
                                )
                                for calendar_date in pd.date_range(
                                    range_start, range_end, freq="D"
                                ).date:
                                    calendar_rows.append(
                                        {
                                            "Период": period_name,
                                            "Точка": "Все выбранные",
                                            "Дата": calendar_date,
                                            "Продано, шт.": float(
                                                combined_sales_lookup.get(calendar_date, 0.0)
                                            ),
                                        }
                                    )
                        sku_calendar = pd.DataFrame(calendar_rows)
                        sku_calendar["Строка"] = (
                            sku_calendar["Период"].astype(str)
                            + " · " + sku_calendar["Точка"].astype(str)
                        )
                        calendar_heat = sku_calendar.pivot_table(
                            index="Строка",
                            columns="Дата",
                            values="Продано, шт.",
                            aggfunc="sum",
                        ).sort_index(axis=1)

                        # Среднее за выбранный период показываем отдельным узким блоком
                        # справа от последней календарной даты — визуально отдельно от дат.
                        # Бизнес-логика: СР = всё проданное количество / число ФАКТИЧЕСКИХ
                        # дней продаж SKU. Календарные дни с нулём в знаменатель не входят.
                        calendar_sales_sum = calendar_heat.fillna(0.0).sum(axis=1)
                        calendar_active_sale_days = calendar_heat.fillna(0.0).gt(0).sum(axis=1)
                        calendar_average = calendar_sales_sum.div(
                            calendar_active_sale_days.where(calendar_active_sale_days > 0)
                        ).fillna(0.0)
                        calendar_dates = list(calendar_heat.columns)
                        calendar_x_values = [
                            calendar_date.strftime("%Y-%m-%d") for calendar_date in calendar_dates
                        ]
                        calendar_tick_labels = [
                            f"{calendar_date:%d.%m}<br>{weekday_names[calendar_date.weekday()]}"
                            for calendar_date in calendar_dates
                        ]
                        calendar_hover_labels = [
                            f"Дата: {calendar_date:%d.%m.%Y}" for calendar_date in calendar_dates
                        ]
                        calendar_customdata = [
                            calendar_hover_labels for _ in range(len(calendar_heat.index))
                        ]
                        calendar_text = calendar_heat.applymap(
                            lambda value: "" if pd.isna(value) else str(int(round(value)))
                        )
                        average_text = calendar_average.map(
                            lambda value: "" if pd.isna(value) else f"{float(value):.1f}"
                        )

                        calendar_heat_max = max(float(calendar_heat.stack().max()), 1.0)
                        calendar_rows_count = max(len(calendar_heat.index), 1)
                        # Блок среднего остаётся компактным даже при длинном выбранном периоде.
                        average_width = min(0.10, max(0.055, 2.2 / max(len(calendar_dates), 1)))
                        date_width = 1.0 - average_width
                        sku_calendar_chart = make_subplots(
                            rows=1,
                            cols=2,
                            shared_yaxes=True,
                            column_widths=[date_width, average_width],
                            horizontal_spacing=0.012,
                        )
                        sku_calendar_chart.add_trace(
                            go.Heatmap(
                                z=calendar_heat.to_numpy(),
                                x=calendar_x_values,
                                y=calendar_heat.index.tolist(),
                                text=calendar_text.to_numpy(),
                                customdata=calendar_customdata,
                                texttemplate="%{text}",
                                hovertemplate=(
                                    "%{y}<br>%{customdata}<br>Продано: %{z:.0f} шт."
                                    "<extra></extra>"
                                ),
                                colorscale=[
                                    [0.0, "#D9D9D9"],
                                    [0.000001, "#D9D9D9"],
                                    [0.000002, "#A9D18E"],
                                    [1.0, "#008A3B"],
                                ],
                                zmin=0,
                                zmax=calendar_heat_max,
                                showscale=True,
                                colorbar=dict(
                                    title="Продано, шт.",
                                    x=1.075,
                                    len=min(0.86, max(0.36, 0.18 + 0.12 * calendar_rows_count)),
                                ),
                                xgap=3,
                                ygap=4,
                            ),
                            row=1,
                            col=1,
                        )
                        sku_calendar_chart.add_trace(
                            go.Heatmap(
                                z=calendar_average.to_numpy().reshape(-1, 1),
                                x=["СР за период"],
                                y=calendar_heat.index.tolist(),
                                text=average_text.to_numpy().reshape(-1, 1),
                                texttemplate="%{text}",
                                hovertemplate=(
                                    "%{y}<br>СР за фактический день продаж: %{z:.2f} шт."
                                    "<extra></extra>"
                                ),
                                colorscale=[
                                    [0.0, "#D9D9D9"],
                                    [0.000001, "#D9D9D9"],
                                    [0.000002, "#A9D18E"],
                                    [1.0, "#008A3B"],
                                ],
                                zmin=0,
                                zmax=calendar_heat_max,
                                showscale=False,
                                xgap=3,
                                ygap=4,
                            ),
                            row=1,
                            col=2,
                        )
                        sku_calendar_chart.update_layout(
                            title=f"Календарь продаж SKU {selected_lifecycle_sku}",
                            height=max(330, 100 + 55 * len(calendar_heat)),
                            margin=dict(l=20, r=95, t=60, b=55),
                        )
                        sku_calendar_chart.update_xaxes(
                            title_text="Дата выбранного периода",
                            tickmode="array",
                            tickvals=calendar_x_values,
                            ticktext=calendar_tick_labels,
                            tickangle=-45,
                            row=1,
                            col=1,
                        )
                        sku_calendar_chart.update_xaxes(
                            title_text="",
                            tickmode="array",
                            tickvals=["СР за период"],
                            ticktext=["СР<br>за период"],
                            tickangle=0,
                            showgrid=False,
                            row=1,
                            col=2,
                        )
                        sku_calendar_chart.update_yaxes(
                            title_text="Период · точка",
                            row=1,
                            col=1,
                        )
                        sku_calendar_chart.update_yaxes(
                            showticklabels=False,
                            title_text="",
                            row=1,
                            col=2,
                        )
                        st.plotly_chart(sku_calendar_chart, use_container_width=True)
                        st.caption(
                            "Зелёный — SKU продавался, число внутри — продано за день. "
                            "Серый — в эту дату продаж SKU не было. Справа от последней даты "
                            "отдельным блоком показано «СР за период» по каждой точке. Расчёт: "
                            "всё проданное количество SKU / количество фактических дней продаж; "
                            "дни с нулевыми продажами в знаменатель не входят."
                        )



if tab_sales_time.open:
    with tab_sales_time:
        matrix_snapshot_bytes, matrix_snapshot_source, matrix_snapshot_checked_at, matrix_snapshot_error = (
            _load_matrix_context_for_active_tab()
        )
        st.session_state["freshness_category_source_v62"] = pd.DataFrame()
        st.session_state["freshness_category_source_by_point_v63"] = pd.DataFrame()
        st.session_state["freshness_category_context_v62"] = None
        st.subheader("Окно свежести партии SKU")
        st.caption(
            "Дата плана считается датой отгрузки. Отсчёт срока начинается на следующий день. "
            "Вторые блюда: 3 зелёных + 2 серых дня; салаты, супы, завтраки, сэндвичи, "
            "десерты и хлеб: 2 зелёных + 1 серый день; Япония: 1 зелёный + 1 серый день; "
            "напитки: 4 зелёных + 3 серых дня. "
            "Скидка 40% применяется только в последний день срока."
        )
        # Основной источник — Apps Script, связанный с Google Sheet «2.3 Матрица КОМБО».
        # Снимок обновляется автоматически каждые 15 минут; локальный XLSX — только резерв.
        matrix_status_columns = st.columns([4.0, 1.0])
        with matrix_status_columns[0]:
            if matrix_snapshot_source.startswith("Apps Script"):
                st.success(
                    f"Матрица: {matrix_snapshot_source}. Автообновление каждые 15 минут · "
                    f"проверено {matrix_snapshot_checked_at.replace('T', ' ')}."
                )
            else:
                st.warning(
                    f"Матрица сейчас загружена из резерва: {matrix_snapshot_source}. "
                    "Приложение продолжит пытаться получить актуальную матрицу через Apps Script каждые 15 минут."
                )
                if matrix_snapshot_error:
                    st.caption(f"Причина: {matrix_snapshot_error}")
        with matrix_status_columns[1]:
            if st.button(
                "Обновить матрицу сейчас",
                use_container_width=True,
                key="refresh_google_matrix_now_v761",
            ):
                _fetch_apps_script_matrix_snapshot.clear()
                _fetch_apps_script_entity_reference.clear()
                st.rerun()

        if not matrix_snapshot_bytes:
            st.error("Матрица 2.3 недоступна ни через Apps Script, ни в резервной копии.")
            sales_time_plans = pd.DataFrame()
        else:
            try:
                sales_time_plans = parse_freshness_plan(matrix_snapshot_bytes)
            except Exception as error:
                st.error(f"Не удалось прочитать матрицу 2.3: {error}")
                sales_time_plans = pd.DataFrame()

        if sales_time_plans.empty:
            st.warning(
                "В матрице 2.3 не найдены основные блоки «План на день кухня» "
                "с датами, SKU и колонками Т1–Т29 на листах 1–4 недели."
            )
        else:
            available_plan_dates = sorted(sales_time_plans["plan_date"].dropna().unique())
            matrix_start = available_plan_dates[0]
            matrix_end = available_plan_dates[-1]
            st.success(
                f"Матрица 2.3 загружена: планы 1–4 недель; "
                f"дат — {len(available_plan_dates)}, SKU — {sales_time_plans['sku'].nunique()}; "
                f"период {matrix_start:%d.%m.%Y}–{matrix_end:%d.%m.%Y}."
            )

            start_key = "freshness_period_start_v75"
            end_key = "freshness_period_end_v75"
            anchor_key = "freshness_period_anchor_v75"
            if (
                st.session_state.get(start_key) not in available_plan_dates
                or st.session_state.get(end_key) not in available_plan_dates
            ):
                st.session_state[start_key] = matrix_start
                st.session_state[end_key] = matrix_end
                st.session_state[anchor_key] = None

            st.markdown("#### Период плана")
            st.caption(
                "По умолчанию выбраны все даты из планов 1–4 недель. "
                "Чтобы выбрать свой период, нажмите первую дату, затем последнюю."
            )
            if st.button(
                "Все даты 1–4 недель",
                key="freshness_all_dates_v75",
                type="primary",
                use_container_width=False,
            ):
                st.session_state[start_key] = matrix_start
                st.session_state[end_key] = matrix_end
                st.session_state[anchor_key] = None
                st.rerun()

            weekday_short = {
                0: "Пн", 1: "Вт", 2: "Ср", 3: "Чт", 4: "Пт", 5: "Сб", 6: "Вс",
            }
            selected_start = st.session_state[start_key]
            selected_end = st.session_state[end_key]
            for row_start in range(0, len(available_plan_dates), 7):
                row_dates = available_plan_dates[row_start : row_start + 7]
                date_columns = st.columns(7)
                for offset, plan_date in enumerate(row_dates):
                    is_selected = selected_start <= plan_date <= selected_end
                    if date_columns[offset].button(
                        f"{plan_date:%d.%m} · {weekday_short[plan_date.weekday()]}",
                        key=f"freshness_date_button_v75_{plan_date.isoformat()}",
                        type="primary" if is_selected else "secondary",
                        use_container_width=True,
                    ):
                        anchor_date = st.session_state.get(anchor_key)
                        if anchor_date is None:
                            st.session_state[start_key] = plan_date
                            st.session_state[end_key] = plan_date
                            st.session_state[anchor_key] = plan_date
                        else:
                            st.session_state[start_key] = min(anchor_date, plan_date)
                            st.session_state[end_key] = max(anchor_date, plan_date)
                            st.session_state[anchor_key] = None
                        st.rerun()

            shipment_start = st.session_state[start_key]
            shipment_end = st.session_state[end_key]
            if shipment_start > shipment_end:
                shipment_start, shipment_end = shipment_end, shipment_start
            selected_plan_dates = [
                plan_date for plan_date in available_plan_dates
                if shipment_start <= plan_date <= shipment_end
            ]
            st.info(
                f"Выбрано дат плана: {len(selected_plan_dates)} · "
                f"{shipment_start:%d.%m.%Y}–{shipment_end:%d.%m.%Y}"
            )
            period_plans = sales_time_plans[
                sales_time_plans["plan_date"].isin(selected_plan_dates)
            ].copy()
            current_mapping = st.session_state.get("point_mapping", {})
            point_to_shop = {
                label: int(shop_number)
                for shop_number, label in current_mapping.items()
                if str(label).startswith("Т") and str(label) != "Т11"
            }
            available_point_numbers = sorted(
                period_plans["point_number"].dropna().astype(int).unique()
            )
            available_point_labels = [
                f"Т{number}" for number in available_point_numbers
                if number != 11 and f"Т{number}" in point_to_shop
            ]
            if not available_point_labels:
                st.warning(
                    "Для выбранного периода нет сопоставленных точек. Сначала найдите магазины и задайте им названия Т1–Т29."
                )
            else:
                selected_time_points = st.multiselect(
                    "Точки",
                    available_point_labels,
                    default=[available_point_labels[0]],
                    key="sales_time_points_v33",
                    help="При выборе нескольких точек план, продажи, выручка и списания суммируются.",
                )
                selected_point_numbers = [int(label[1:]) for label in selected_time_points]
                selected_shop_numbers = tuple(
                    point_to_shop[label] for label in selected_time_points
                )
                selected_plan_rows = period_plans[
                    period_plans["point_number"].isin(selected_point_numbers)
                ].copy()

                if selected_stock_date is not None and not stock_snapshot.empty:
                    selected_stock_rows = stock_snapshot[
                        stock_snapshot["point"].isin(selected_time_points)
                    ]
                    st.success(
                        f"Фактические остатки подключены: {selected_stock_date:%d.%m.%Y} · "
                        f"{WEEKDAY_RU.get(selected_stock_date.weekday(), '')}; "
                        f"точек в выбранном снимке — {selected_stock_rows['point'].nunique()}, "
                        f"SKU — {selected_stock_rows['sku'].nunique()}."
                    )

                # Верхний блок — фактический вид общего меню за реальные даты
                # плана. При смене точки меняются только количества.
                period_menu = selected_plan_rows.merge(
                    entities[["sku", "category", "entity"]], on="sku", how="left"
                )
                period_menu["category"] = period_menu["category"].fillna(
                    period_menu["matrix_category"]
                ).map(normalize_matrix_category)
                period_menu["entity"] = period_menu["entity"].fillna("Не сопоставлено")
                menu_pivot = period_menu.pivot_table(
                    index=["category", "entity", "sku", "product_name"],
                    columns="plan_date",
                    values="analyst_plan",
                    aggfunc="sum",
                    fill_value=0,
                ).reset_index()
                menu_pivot = menu_pivot.rename(
                    columns={
                        "category": "Категория",
                        "entity": "Сущность",
                        "sku": "SKU",
                        "product_name": "Название товара",
                        **{
                            plan_date: plan_date.strftime("%d.%m.%Y")
                            for plan_date in available_plan_dates
                            if plan_date in menu_pivot.columns
                        },
                    }
                )
                category_order = [
                    "Завтраки", "Салаты", "Супы", "Вторые блюда", "Сэндвичи",
                    "Япония", "Десерты", "Напитки", "Хлеб",
                ]
                category_rank = {category: index for index, category in enumerate(category_order)}
                menu_pivot["_category_order"] = menu_pivot["Категория"].map(category_rank).fillna(99)
                menu_pivot = (
                    menu_pivot.sort_values(
                        ["_category_order", "Категория", "Сущность", "Название товара", "SKU"],
                        kind="stable",
                    )
                    .drop(columns="_category_order")
                    .reset_index(drop=True)
                )
                category_colors = {
                    "Завтраки": "#FFF2CC",
                    "Салаты": "#E2F0D9",
                    "Супы": "#DDEBF7",
                    "Вторые блюда": "#FCE4D6",
                    "Сэндвичи": "#E4DFEC",
                    "Япония": "#D9EAD3",
                    "Десерты": "#F4CCCC",
                    "Напитки": "#D9EAF7",
                    "Хлеб": "#EDEDED",
                }

                def color_menu_category(row: pd.Series) -> list[str]:
                    fill = category_colors.get(str(row.get("Категория", "")), "#F7F7F7")
                    return [f"background-color: {fill}; color: #202124;" for _ in row.index]

                date_menu_columns = [
                    plan_date.strftime("%d.%m.%Y")
                    for plan_date in selected_plan_dates
                    if plan_date.strftime("%d.%m.%Y") in menu_pivot.columns
                ]
                menu_screen, menu_screen_limit, menu_screen_truncated = styler_safe_preview(menu_pivot)
                menu_styler = menu_screen.style.apply(color_menu_category, axis=1)
                if date_menu_columns:
                    menu_styler = menu_styler.format(
                        {column: lambda value: "—" if pd.isna(value) else f"{float(value):.0f}" for column in date_menu_columns}
                    )

                st.markdown(
                    f"#### Меню за {shipment_start:%d.%m.%Y}–{shipment_end:%d.%m.%Y} · "
                    f"{', '.join(selected_time_points) if selected_time_points else 'точки не выбраны'}"
                )
                st.caption(
                    "Меню отсортировано по категориям; каждая категория выделена своим цветом. "
                    "В колонках дат показан план выбранных точек."
                )
                if menu_screen_truncated:
                    st.info(
                        f"Меню содержит {len(menu_pivot):,} строк. В цветном просмотре показаны первые "
                        f"{menu_screen_limit:,}, чтобы не превышать безопасный лимит отрисовки.".replace(",", " ")
                    )
                menu_selection = st.dataframe(
                    menu_styler,
                    use_container_width=True,
                    hide_index=True,
                    height=min(700, 38 * len(menu_pivot) + 80),
                    on_select="rerun",
                    selection_mode="single-row",
                    key="freshness_menu_sku_selection_v757",
                    column_config={
                        "Категория": st.column_config.TextColumn(width="medium"),
                        "Сущность": st.column_config.TextColumn(width="medium"),
                        "SKU": st.column_config.TextColumn(width="small"),
                        "Название товара": st.column_config.TextColumn(width="large"),
                    },
                )
                st.caption(
                    "Нажмите на строку нужного SKU в меню — ниже откроется история его продаж "
                    "по выбранной точке за период, заданный слева в «Параметрах»."
                )

                selected_menu_rows = list(menu_selection.selection.rows)
                if selected_menu_rows:
                    selected_menu_index = int(selected_menu_rows[0])
                    if 0 <= selected_menu_index < len(menu_screen):
                        selected_menu_item = menu_screen.iloc[selected_menu_index]
                        selected_history_sku = normalize_sku(selected_menu_item.get("SKU"))
                        selected_history_name = str(selected_menu_item.get("Название товара", "") or "")

                        history_start, history_end = period
                        sku_history = daily_detail.copy()
                        sku_history["sku"] = sku_history["sku"].map(normalize_sku)
                        sku_history["business_date"] = pd.to_datetime(
                            sku_history["business_date"], errors="coerce"
                        ).dt.date
                        sku_history = sku_history[
                            sku_history["sku"].eq(selected_history_sku)
                            & sku_history["point"].isin(selected_time_points)
                            & sku_history["business_date"].between(
                                history_start, history_end, inclusive="both"
                            )
                        ].copy()

                        st.markdown(
                            f"#### История продаж SKU {selected_history_sku or '—'} · {selected_history_name}"
                        )
                        st.caption(
                            f"Период из «Параметров»: {history_start:%d.%m.%Y}–{history_end:%d.%m.%Y} · "
                            f"точки: {', '.join(selected_time_points) if selected_time_points else 'не выбраны'}."
                        )

                        if sku_history.empty:
                            st.warning(
                                "По выбранному SKU и точке за период из «Параметров» продаж не найдено."
                            )
                        else:
                            sku_history["sales"] = pd.to_numeric(
                                sku_history["sales"], errors="coerce"
                            ).fillna(0.0)
                            sku_history["revenue"] = pd.to_numeric(
                                sku_history["revenue"], errors="coerce"
                            ).fillna(0.0)
                            sold_history = sku_history[sku_history["sales"] > 0].copy()
                            sold_days_count = int(sold_history["business_date"].nunique())
                            avg_sales_per_day = (
                                float(sold_history["sales"].sum()) / sold_days_count
                                if sold_days_count > 0
                                else 0.0
                            )
                            history_metrics = st.columns(5)
                            history_metrics[0].metric(
                                "Продано, шт.",
                                f"{sku_history['sales'].sum():,.0f}".replace(",", " "),
                            )
                            history_metrics[1].metric(
                                "СР за день, шт.",
                                f"{avg_sales_per_day:,.1f}".replace(",", " "),
                            )
                            history_metrics[2].metric(
                                "Выручка, ₽",
                                f"{sku_history['revenue'].sum():,.0f}".replace(",", " "),
                            )
                            history_metrics[3].metric(
                                "Дней с продажами",
                                sold_days_count,
                            )
                            history_metrics[4].metric(
                                "Продаж / строк",
                                int(len(sold_history)),
                            )

                            daily_history = (
                                sku_history.groupby(
                                    ["business_date", "point"], as_index=False, dropna=False
                                )
                                .agg(
                                    **{
                                        "Продано, шт.": ("sales", "sum"),
                                        "Выручка, ₽": ("revenue", "sum"),
                                    }
                                )
                                .rename(columns={"business_date": "Дата", "point": "Точка"})
                                .sort_values(["Дата", "Точка"], kind="stable")
                            )
                            daily_history["Дата"] = pd.to_datetime(daily_history["Дата"], errors="coerce")
                            history_chart = px.line(
                                daily_history,
                                x="Дата",
                                y="Продано, шт.",
                                color="Точка" if len(selected_time_points) > 1 else None,
                                markers=True,
                                title=(
                                    f"Продажи SKU {selected_history_sku} по дням · "
                                    f"{history_start:%d.%m.%Y}–{history_end:%d.%m.%Y}"
                                ),
                            )
                            history_chart.update_layout(
                                xaxis_title="Дата",
                                yaxis_title="Продано, шт.",
                                hovermode="x unified",
                                margin=dict(l=20, r=20, t=60, b=20),
                                height=390,
                            )
                            st.plotly_chart(history_chart, use_container_width=True)

                            st.markdown("##### Календарь продаж за выбранный период")
                            sales_by_date = (
                                daily_history.groupby("Дата", as_index=False)["Продано, шт."].sum()
                                .sort_values("Дата", kind="stable")
                            )
                            sales_by_date["Дата"] = pd.to_datetime(sales_by_date["Дата"], errors="coerce")
                            sales_by_date = sales_by_date.dropna(subset=["Дата"]).copy()
                            sales_map = {
                                ts.date(): float(qty)
                                for ts, qty in zip(sales_by_date["Дата"], sales_by_date["Продано, шт."])
                            }

                            calendar_months = []
                            month_cursor = date(history_start.year, history_start.month, 1)
                            last_month = date(history_end.year, history_end.month, 1)
                            while month_cursor <= last_month:
                                calendar_months.append(month_cursor)
                                if month_cursor.month == 12:
                                    month_cursor = date(month_cursor.year + 1, 1, 1)
                                else:
                                    month_cursor = date(month_cursor.year, month_cursor.month + 1, 1)

                            month_tabs = st.tabs([month_date.strftime("%m.%Y") for month_date in calendar_months])
                            weekday_labels = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
                            for month_tab, month_date in zip(month_tabs, calendar_months):
                                with month_tab:
                                    month_weeks = pycalendar.Calendar(firstweekday=0).monthdatescalendar(
                                        month_date.year, month_date.month
                                    )
                                    z_values = []
                                    text_values = []
                                    for week in month_weeks:
                                        week_qty = []
                                        week_text = []
                                        for day_item in week:
                                            if (
                                                day_item.month != month_date.month
                                                or day_item < history_start
                                                or day_item > history_end
                                            ):
                                                week_qty.append(None)
                                                week_text.append("")
                                            else:
                                                qty_value = float(sales_map.get(day_item, 0.0))
                                                week_qty.append(qty_value)
                                                week_text.append(f"{day_item.day}<br>{qty_value:,.0f} шт.".replace(',', ' '))
                                        z_values.append(week_qty)
                                        text_values.append(week_text)

                                    month_title = f"{MONTH_NAMES_RU.get(month_date.month, month_date.strftime('%m'))} {month_date.year}"
                                    month_calendar_fig = go.Figure(
                                        data=go.Heatmap(
                                            z=z_values,
                                            x=weekday_labels,
                                            y=[f"Неделя {idx + 1}" for idx in range(len(z_values))],
                                            text=text_values,
                                            texttemplate="%{text}",
                                            textfont={"size": 12},
                                            colorscale=[
                                                [0.0, '#f3f6fb'],
                                                [0.2, '#dcefe3'],
                                                [0.5, '#a9d3b0'],
                                                [1.0, '#4f8a61'],
                                            ],
                                            hovertemplate="%{x}<br>%{text}<extra></extra>",
                                            colorbar_title="Продано, шт.",
                                            zmin=0,
                                            xgap=3,
                                            ygap=3,
                                        )
                                    )
                                    month_calendar_fig.update_layout(
                                        title=f"Календарь продаж · {month_title}",
                                        margin=dict(l=20, r=20, t=55, b=20),
                                        height=max(260, 95 + len(z_values) * 70),
                                    )
                                    month_calendar_fig.update_yaxes(autorange='reversed', title='')
                                    month_calendar_fig.update_xaxes(title='')
                                    st.plotly_chart(month_calendar_fig, use_container_width=True)
                                    st.caption(
                                        'В ячейке показаны день месяца и суммарное количество продаж SKU за этот день.'
                                    )

                            history_table = sku_history.copy()
                            history_table["Дата"] = pd.to_datetime(
                                history_table["business_date"], errors="coerce"
                            ).dt.strftime("%d.%m.%Y")
                            history_table["Время"] = pd.to_datetime(
                                history_table["sale_datetime"], errors="coerce"
                            ).dt.strftime("%H:%M")
                            history_table = history_table.rename(
                                columns={
                                    "point": "Точка",
                                    "sales": "Продано, шт.",
                                    "revenue": "Выручка, ₽",
                                }
                            )
                            history_table = history_table[
                                ["Дата", "Время", "Точка", "Продано, шт.", "Выручка, ₽"]
                            ].sort_values(["Дата", "Время", "Точка"], kind="stable")
                            st.dataframe(
                                history_table,
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Дата": st.column_config.TextColumn(width="small"),
                                    "Время": st.column_config.TextColumn(width="small"),
                                    "Точка": st.column_config.TextColumn(width="small"),
                                    "Продано, шт.": st.column_config.NumberColumn(format="%.0f"),
                                    "Выручка, ₽": st.column_config.NumberColumn(format="%.2f"),
                                },
                            )

                if selected_shop_numbers:
                    try:
                        history_start_for_freshness = shipment_start - timedelta(days=8)
                        history_end_for_freshness = shipment_end + timedelta(days=8)
                        if selected_stock_date is not None:
                            history_start_for_freshness = min(
                                history_start_for_freshness, selected_stock_date - timedelta(days=8)
                            )
                            history_end_for_freshness = max(
                                history_end_for_freshness, selected_stock_date + timedelta(days=1)
                            )
                        time_sales = load_forecast_history(
                            history_start_for_freshness,
                            history_end_for_freshness,
                            selected_shop_numbers,
                        )
                    except Exception as error:
                        st.error(f"Не удалось загрузить продажи партии: {error}")
                        time_sales = pd.DataFrame()
                else:
                    time_sales = pd.DataFrame()

                point_results: list[pd.DataFrame] = []
                for point_label, point_number, shop_number in zip(
                    selected_time_points, selected_point_numbers, selected_shop_numbers
                ):
                    point_plan_rows = sales_time_plans[
                        (sales_time_plans["point_number"] == point_number)
                        & sales_time_plans["plan_date"].between(
                            shipment_start - timedelta(days=7), shipment_end, inclusive="both"
                        )
                    ].copy()
                    point_sales_history = time_sales[
                        pd.to_numeric(time_sales.get("shop_number"), errors="coerce") == shop_number
                    ].copy() if not time_sales.empty else pd.DataFrame()
                    point_result = build_sales_time_period(
                        point_plan_rows,
                        point_sales_history,
                        entities,
                        shipment_start,
                        shipment_end,
                    )
                    if not point_result.empty:
                        point_result.insert(0, "Точка", point_label)
                        point_results.append(point_result)

                sales_time_menu = (
                    pd.concat(point_results, ignore_index=True)
                    if point_results else pd.DataFrame()
                )
                # Сохраняем исходную детализацию по каждой точке до суммирования.
                # Верхняя таблица партий по-прежнему показывает общий итог по выбранным точкам,
                # а по клику на строку ниже можно раскрыть реализацию именно этой партии
                # отдельно по Т1–Т29.
                sales_time_menu_by_point = sales_time_menu.copy()

                # 75.11.15: прямой табличный вид «матрица меню -> факт по дням свежести».
                # Каждая точка показывается отдельно; строки = SKU из её плана с количеством > 0.
                point_menu_view = prepare_freshness_point_menu_view(
                    sales_time_menu_by_point,
                    as_of_date=date.today(),
                )
                st.markdown("#### Меню по точкам · продажи по дням свежести")
                st.caption(
                    "Таблица строится от выбранных дат формирования меню в матрице. "
                    "Для каждого SKU показан план точки и фактическое количество продаж в День 1, День 2 и далее. "
                    "День 1 начинается на следующий календарный день после даты меню. "
                    "Зелёные ячейки — основной период свежести; серые — завершающие дни срока; «—» — день вне срока или ещё не наступил."
                )
                if point_menu_view.empty:
                    st.info(
                        "Для выбранного периода и точек в матрице нет SKU с планом больше нуля, "
                        "поэтому таблица окна свежести не сформирована."
                    )
                else:
                    export_col, summary_col = st.columns([1.0, 2.2])
                    with export_col:
                        try:
                            freshness_point_excel = build_freshness_point_menu_excel(
                                sales_time_menu_by_point,
                                shipment_start,
                                shipment_end,
                            )
                        except Exception as export_error:
                            freshness_point_excel = b""
                            st.caption(f"Excel-выгрузка временно недоступна: {export_error}")
                        if freshness_point_excel:
                            st.download_button(
                                "Скачать меню по точкам · Excel",
                                data=freshness_point_excel,
                                file_name=(
                                    f"окно_свежести_по_точкам_"
                                    f"{shipment_start:%Y-%m-%d}_{shipment_end:%Y-%m-%d}.xlsx"
                                ),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="freshness_point_menu_excel_v75115",
                            )
                    with summary_col:
                        st.info(
                            f"Точек: {point_menu_view['Точка'].nunique()} · "
                            f"строк меню: {len(point_menu_view):,} · "
                            f"SKU: {point_menu_view['SKU'].nunique()}".replace(",", " ")
                        )

                    def style_freshness_point_menu(row: pd.Series) -> list[str]:
                        styles = [""] * len(row.index)
                        category = row.get("Категория", "")
                        green_days = product_green_days(category)
                        shelf_days = product_lifecycle_days(category)
                        for day_number, column_name in enumerate(FRESHNESS_POINT_DAY_COLUMNS, start=1):
                            if column_name not in row.index:
                                continue
                            position = row.index.get_loc(column_name)
                            value = row.get(column_name)
                            numeric_value = pd.to_numeric(
                                str(value).replace(" ", "") if value != "—" else None,
                                errors="coerce",
                            )
                            if day_number <= green_days:
                                if pd.notna(numeric_value) and float(numeric_value) > 0:
                                    styles[position] = (
                                        "background-color: #B6D7A8; color: #274E13; font-weight: 700"
                                    )
                                else:
                                    styles[position] = (
                                        "background-color: #D9EAD3; color: #274E13; font-weight: 600"
                                    )
                            elif day_number <= shelf_days:
                                styles[position] = (
                                    "background-color: #E7E6E6; color: #333333; font-weight: 600"
                                )
                            else:
                                styles[position] = "background-color: #F2F2F2; color: #B7B7B7"
                        return styles

                    point_order_for_view = [
                        point for point in selected_time_points
                        if point in set(point_menu_view["Точка"].dropna().astype(str))
                    ]
                    for point_index, point_label in enumerate(point_order_for_view):
                        point_table = point_menu_view[
                            point_menu_view["Точка"].astype(str).eq(point_label)
                        ].drop(columns="Точка").reset_index(drop=True)
                        point_shop = point_to_shop.get(point_label)
                        point_plan_total = pd.to_numeric(
                            point_table.get("План, шт."), errors="coerce"
                        ).fillna(0).sum()
                        point_green_sales = pd.to_numeric(
                            point_table.get("Продано в свежесть, шт."), errors="coerce"
                        ).fillna(0).sum()
                        point_title = (
                            f"{point_label} · магазин {point_shop} · "
                            f"SKU {point_table['SKU'].nunique()} · план {point_plan_total:,.0f} шт. · "
                            f"продано в свежесть {point_green_sales:,.0f} шт."
                        ).replace(",", " ")
                        with st.expander(point_title, expanded=(point_index == 0)):
                            point_screen, point_screen_limit, point_screen_truncated = styler_safe_preview(
                                point_table
                            )
                            if point_screen_truncated:
                                st.info(
                                    f"Показаны первые {point_screen_limit:,} строк из {len(point_table):,}. "
                                    "Excel-выгрузка содержит все строки.".replace(",", " ")
                                )
                            st.dataframe(
                                point_screen.style.apply(style_freshness_point_menu, axis=1),
                                use_container_width=True,
                                hide_index=True,
                                height=min(620, 38 * len(point_screen) + 80),
                                column_config={
                                    "Дата меню": st.column_config.DateColumn("Дата меню", format="DD.MM.YYYY"),
                                    "План, шт.": st.column_config.NumberColumn("План, шт.", format="%.0f"),
                                    "Окно, дней": st.column_config.NumberColumn("Окно, дней", format="%d"),
                                    "Продано в свежесть, шт.": st.column_config.NumberColumn(format="%.0f"),
                                    "Продано за срок, шт.": st.column_config.NumberColumn(format="%.0f"),
                                    "Остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                                },
                            )

                if len(selected_time_points) > 1 and not sales_time_menu.empty:
                    timeline_sum_columns = [
                        "Отгружено по плану", "День 1", "День 2", "День 3", "День 4", "День 5",
                        "День 6", "День 7",
                        "Продано в зелёный период", "Продано в серый период", "Продано за срок",
                        "Расчётный остаток", "Списания", "Выручка SKU, ₽", "Убыток от списания, ₽",
                        "Вес часов выбытия", "Количество для скорости",
                    ]
                    timeline_group_columns = [
                        "Дата отгрузки", "Категория", "Сущность", "SKU", "Название товара",
                        "Срок годности, дней", "Зелёный период, дней", "Последний день срока",
                        "Скидка в последний день", "Статус партии",
                    ]
                    sales_time_menu = (
                        sales_time_menu.groupby(timeline_group_columns, as_index=False, dropna=False)[
                            timeline_sum_columns
                        ].sum()
                    )
                    sales_time_menu["Скорость выбытия, ч"] = safe_ratio(
                        sales_time_menu["Вес часов выбытия"],
                        sales_time_menu["Количество для скорости"],
                    ).round(1)
                    sales_time_menu.insert(0, "Точка", ", ".join(selected_time_points))
                # Для вкладки «Списания» сохраняем также несуммированную детализацию по каждой точке.
                # Это позволяет раскрыть категорию до SKU, а SKU — до конкретных Т1–Т29.
                st.session_state["freshness_category_source_by_point_v63"] = sales_time_menu_by_point.copy()
                st.session_state["freshness_category_source_v62"] = sales_time_menu.copy()
                st.session_state["freshness_category_context_v62"] = {
                    "start": shipment_start,
                    "end": shipment_end,
                    "points": list(selected_time_points),
                }
                st.markdown("#### Продажи партий в окне свежести")
                st.caption(
                    "Можно одновременно отфильтровать данные по категории и быстро найти конкретный SKU. "
                    "Фильтры применяются к показателям, таблице партий, сравнительной таблице блюд и графику. "
                    "Нажмите строку партии в таблице — ниже откроется её реализация отдельно по каждой выбранной точке."
                )

                category_filter_options = [
                    category for category in category_order
                    if category in set(sales_time_menu.get("Категория", pd.Series(dtype=str)).dropna().astype(str))
                ]
                extra_category_options = sorted(
                    set(sales_time_menu.get("Категория", pd.Series(dtype=str)).dropna().astype(str))
                    - set(category_filter_options)
                )
                category_filter_options.extend(extra_category_options)
                category_filter_key = "sales_time_categories_v752"
                stored_categories = st.session_state.get(category_filter_key, [])
                if not isinstance(stored_categories, list):
                    stored_categories = list(stored_categories) if stored_categories else []
                valid_stored_categories = [
                    category for category in stored_categories if category in category_filter_options
                ]
                if valid_stored_categories != stored_categories:
                    st.session_state[category_filter_key] = valid_stored_categories

                filter_category_col, filter_sku_col = st.columns([1.0, 1.15])
                with filter_category_col:
                    selected_time_categories = st.multiselect(
                        "Категории",
                        options=category_filter_options,
                        default=st.session_state.get(category_filter_key, []),
                        key=category_filter_key,
                        placeholder="Все категории",
                        help=(
                            "Оставьте поле пустым, чтобы видеть все категории. Можно выбрать одну или несколько — "
                            "тогда нижняя таблица и график сравнят блюда только внутри выбранных категорий."
                        ),
                    )
                with filter_sku_col:
                    sku_time_search = st.text_input(
                        "Быстрый поиск SKU",
                        placeholder="SKU, часть кода или название блюда",
                        key="sales_time_sku_search_v752",
                        help="Поиск работает внутри выбранных категорий и применяется ко всем нижним блокам.",
                    ).strip()

                if selected_time_categories and not sales_time_menu.empty:
                    sales_time_menu = sales_time_menu[
                        sales_time_menu["Категория"].astype(str).isin(selected_time_categories)
                    ].copy()
                if sku_time_search and not sales_time_menu.empty:
                    sku_search_mask = (
                        sales_time_menu["SKU"].astype(str).str.contains(
                            sku_time_search, case=False, regex=False, na=False
                        )
                        | sales_time_menu["Название товара"].astype(str).str.contains(
                            sku_time_search, case=False, regex=False, na=False
                        )
                    )
                    sales_time_menu = sales_time_menu[sku_search_mask].copy()

                if sales_time_menu.empty:
                    if sku_time_search or selected_time_categories:
                        active_filter_parts = []
                        if selected_time_categories:
                            active_filter_parts.append(
                                "категориям: " + ", ".join(selected_time_categories)
                            )
                        if sku_time_search:
                            active_filter_parts.append(f"поиску: {sku_time_search}")
                        st.warning(
                            "По выбранным фильтрам ничего не найдено (" + "; ".join(active_filter_parts) + ")."
                        )
                    else:
                        st.info("В выбранном периоде и точке отгрузки в плане отсутствуют.")
                else:
                    time_metrics = st.columns(7)
                    time_metrics[0].metric("SKU в отгрузке", sales_time_menu["SKU"].nunique())
                    time_metrics[1].metric(
                        "Отгружено, шт.",
                        f"{sales_time_menu['Отгружено по плану'].sum():,.0f}".replace(",", " "),
                    )
                    time_metrics[2].metric(
                        "Продано в основной период, шт.",
                        f"{sales_time_menu['Продано в зелёный период'].sum():,.0f}".replace(",", " "),
                    )
                    time_metrics[3].metric(
                        "Списания, шт.",
                        f"{sales_time_menu['Списания'].sum():,.0f}".replace(",", " "),
                    )
                    time_metrics[4].metric(
                        "Выручка SKU, ₽",
                        f"{sales_time_menu['Выручка SKU, ₽'].sum():,.0f}".replace(",", " "),
                    )
                    time_metrics[5].metric(
                        "Убыток списаний, ₽",
                        f"{sales_time_menu['Убыток от списания, ₽'].sum():,.0f}".replace(",", " "),
                    )
                    total_speed_quantity = sales_time_menu["Количество для скорости"].sum()
                    average_depletion_hours = (
                        sales_time_menu["Вес часов выбытия"].sum() / total_speed_quantity
                        if total_speed_quantity > 0 else None
                    )
                    time_metrics[6].metric(
                        "Скорость выбытия",
                        f"{average_depletion_hours:.1f} ч" if average_depletion_hours is not None else "—",
                        help=(
                            "Среднее число часов от начала следующего дня после отгрузки "
                            "до фактической продажи единицы товара. Меньше — быстрее."
                        ),
                    )

                    display_time_menu = sales_time_menu.copy()
                    timeline_columns = [f"День {day_number}" for day_number in range(1, 8)]
                    # Визуальная временная шкала текстовая: будущие дни и дни после срока показываем «—».
                    for column_name in timeline_columns:
                        display_time_menu[column_name] = display_time_menu[column_name].map(
                            lambda value: f"{float(value):,.0f}".replace(",", " ")
                        ).astype(object)
                    today_for_freshness = date.today()
                    for row_index, category in display_time_menu["Категория"].items():
                        normalized_category = normalize_matrix_category(category)
                        shelf_days = product_lifecycle_days(category)
                        shipment_date_row = pd.to_datetime(
                            display_time_menu.at[row_index, "Дата отгрузки"], errors="coerce"
                        )

                        # Для категории «Япония» жизненный цикл строго двухдневный:
                        # День 1 — основной (зелёный), День 2 — завершающий (серый),
                        # в 00:00 Дня 3 остаток уже переходит в списание (красный).
                        if normalized_category == "Япония":
                            display_time_menu.loc[row_index, timeline_columns[3:]] = "—"
                            if pd.notna(shipment_date_row):
                                shipment_day = shipment_date_row.date()
                                writeoff_day = shipment_day + timedelta(days=3)
                                for day_number in (1, 2):
                                    if shipment_day + timedelta(days=day_number) > today_for_freshness:
                                        display_time_menu.at[row_index, f"День {day_number}"] = "—"
                                if writeoff_day <= today_for_freshness:
                                    writeoff_qty = float(
                                        pd.to_numeric(
                                            pd.Series([display_time_menu.at[row_index, "Списания"]]),
                                            errors="coerce",
                                        ).fillna(0).iloc[0]
                                    )
                                    display_time_menu.at[row_index, "День 3"] = (
                                        f"Списание {writeoff_qty:,.0f}".replace(",", " ")
                                        if writeoff_qty > 0
                                        else "Срок завершён"
                                    )
                                else:
                                    display_time_menu.at[row_index, "День 3"] = "—"
                        else:
                            display_time_menu.loc[row_index, timeline_columns[shelf_days:]] = "—"
                            if pd.notna(shipment_date_row):
                                shipment_day = shipment_date_row.date()
                                for day_number in range(1, shelf_days + 1):
                                    if shipment_day + timedelta(days=day_number) > today_for_freshness:
                                        display_time_menu.at[row_index, f"День {day_number}"] = "—"

                    def style_sales_timeline(row: pd.Series) -> list[str]:
                        styles = [""] * len(row)
                        normalized_category = normalize_matrix_category(row["Категория"])
                        green_days = product_green_days(row["Категория"])
                        shelf_days = product_lifecycle_days(row["Категория"])
                        for day_number, column_name in enumerate(timeline_columns, start=1):
                            position = row.index.get_loc(column_name)
                            if day_number <= green_days:
                                styles[position] = "background-color: #D9EAD3; color: #274E13; font-weight: 600"
                            elif day_number <= shelf_days:
                                styles[position] = "background-color: #D9D9D9; color: #333333; font-weight: 600"
                            elif normalized_category == "Япония" and day_number == 3:
                                styles[position] = "background-color: #F4CCCC; color: #990000; font-weight: 700"
                            else:
                                styles[position] = "color: #B7B7B7"
                        writeoff_position = row.index.get_loc("Списания")
                        if float(row["Списания"]) > 0:
                            styles[writeoff_position] = (
                                "background-color: #F4CCCC; color: #990000; font-weight: 700"
                            )
                        else:
                            styles[writeoff_position] = (
                                "background-color: #F2F2F2; color: #666666; font-weight: 600"
                            )
                        revenue_position = row.index.get_loc("Выручка SKU, ₽")
                        styles[revenue_position] = (
                            "background-color: #E2F0D9; color: #274E13; font-weight: 700"
                        )
                        loss_position = row.index.get_loc("Убыток от списания, ₽")
                        if float(row["Убыток от списания, ₽"]) > 0:
                            styles[loss_position] = (
                                "background-color: #F4CCCC; color: #990000; font-weight: 700"
                            )
                        else:
                            styles[loss_position] = "color: #666666"
                        return styles

                    display_columns = [
                        "Точка", "Дата отгрузки", "Категория", "Сущность", "SKU", "Название товара",
                        "Отгружено по плану",
                        *timeline_columns, "Расчётный остаток", "Списания", "Статус партии",
                        "Продано в зелёный период", "Продано в серый период",
                        "Продано за срок", "Выручка SKU, ₽", "Убыток от списания, ₽", "Последний день срока",
                        "Скорость выбытия, ч", "Скидка в последний день",
                    ]
                    freshness_table = display_time_menu[display_columns]
                    freshness_screen, freshness_screen_limit, freshness_screen_truncated = styler_safe_preview(
                        freshness_table
                    )
                    if freshness_screen_truncated:
                        st.info(
                            f"В окне свежести {len(freshness_table):,} строк. Цветная таблица показывает первые "
                            f"{freshness_screen_limit:,}; сами расчёты выполняются по всему объёму данных.".replace(",", " ")
                        )
                    freshness_selection = st.dataframe(
                        freshness_screen.style.apply(style_sales_timeline, axis=1),
                        use_container_width=True,
                        hide_index=True,
                        height=min(760, 38 * len(display_time_menu) + 80),
                        on_select="rerun",
                        selection_mode="single-row",
                        key="freshness_batches_table_v7545",
                        column_config={
                            "Дата отгрузки": st.column_config.DateColumn(
                                "Дата отгрузки", format="DD.MM.YYYY"
                            ),
                            "Последний день срока": st.column_config.DateColumn(
                                "Последний день срока", format="DD.MM.YYYY"
                            ),
                            "Выручка SKU, ₽": st.column_config.NumberColumn(
                                "Выручка SKU, ₽", format="%.2f"
                            ),
                            "Убыток от списания, ₽": st.column_config.NumberColumn(
                                "Убыток от списания, ₽", format="%.2f"
                            ),
                            "Скорость выбытия, ч": st.column_config.NumberColumn(
                                "Скорость выбытия, ч", format="%.1f",
                                help="Среднее время до продажи единицы; меньшее значение означает более быстрое выбытие.",
                            ),
                        },
                    )
                    st.caption(
                        "Зелёный — основной период оценки продаж. Серый — завершающий период срока годности. "
                        "Списание появляется только после окончания последнего дня жизненного цикла партии. "
                        "Пока срок не завершён — остаток остаётся живым, а будущие дни показаны знаком «—». "
                        "Скидка 40% действует только в последний серый день. Нажмите строку, чтобы раскрыть точки."
                    )

                    # Детализация выбранной суммарной партии по точкам.
                    selected_freshness_rows = list(freshness_selection.selection.rows)
                    if selected_freshness_rows and not sales_time_menu_by_point.empty:
                        selected_preview_index = int(selected_freshness_rows[0])
                        if 0 <= selected_preview_index < len(freshness_screen):
                            selected_batch = freshness_screen.iloc[selected_preview_index]
                            selected_batch_sku = str(selected_batch.get("SKU", ""))
                            selected_batch_date = pd.to_datetime(
                                selected_batch.get("Дата отгрузки"), errors="coerce"
                            )
                            selected_batch_category = str(selected_batch.get("Категория", ""))
                            selected_batch_entity = str(selected_batch.get("Сущность", ""))
                            selected_batch_name = str(selected_batch.get("Название товара", ""))

                            point_batch_detail = sales_time_menu_by_point.copy()
                            point_batch_detail = point_batch_detail[
                                point_batch_detail["SKU"].astype(str).eq(selected_batch_sku)
                            ].copy()
                            if pd.notna(selected_batch_date):
                                point_batch_detail = point_batch_detail[
                                    pd.to_datetime(
                                        point_batch_detail["Дата отгрузки"], errors="coerce"
                                    ).dt.normalize().eq(selected_batch_date.normalize())
                                ].copy()
                            if selected_batch_category:
                                point_batch_detail = point_batch_detail[
                                    point_batch_detail["Категория"].astype(str).eq(selected_batch_category)
                                ].copy()
                            if selected_batch_entity:
                                point_batch_detail = point_batch_detail[
                                    point_batch_detail["Сущность"].astype(str).eq(selected_batch_entity)
                                ].copy()

                            if point_batch_detail.empty:
                                st.info("Для выбранной партии нет детализации по точкам.")
                            else:
                                point_batch_detail["Магазин"] = point_batch_detail["Точка"].map(
                                    lambda label: point_to_shop.get(str(label), pd.NA)
                                )
                                for numeric_column in [
                                    "Отгружено по плану", "Продано в зелёный период",
                                    "Продано в серый период", "Продано за срок",
                                    "Расчётный остаток", "Списания", "Выручка SKU, ₽",
                                    "Убыток от списания, ₽", "Вес часов выбытия",
                                    "Количество для скорости",
                                ] + timeline_columns:
                                    if numeric_column in point_batch_detail.columns:
                                        point_batch_detail[numeric_column] = numeric_series(
                                            point_batch_detail[numeric_column]
                                        ).fillna(0.0)
                                point_batch_detail["Живой остаток, шт."] = (
                                    point_batch_detail["Расчётный остаток"]
                                    - point_batch_detail["Списания"]
                                ).clip(lower=0)
                                point_batch_detail["Реализация, %"] = (
                                    safe_ratio(
                                        point_batch_detail["Продано за срок"],
                                        point_batch_detail["Отгружено по плану"],
                                    ) * 100
                                ).fillna(0.0).clip(0, 100).round(1)
                                point_batch_detail["Списание, %"] = (
                                    safe_ratio(
                                        point_batch_detail["Списания"],
                                        point_batch_detail["Отгружено по плану"],
                                    ) * 100
                                ).fillna(0.0).clip(0, 100).round(1)
                                point_batch_detail["Скорость выбытия, ч"] = safe_ratio(
                                    point_batch_detail["Вес часов выбытия"],
                                    point_batch_detail["Количество для скорости"],
                                ).round(1)
                                point_batch_detail = point_batch_detail.sort_values(
                                    "Точка", key=lambda series: pd.to_numeric(
                                        series.astype(str).str.extract(r"(\d+)", expand=False),
                                        errors="coerce",
                                    )
                                ).reset_index(drop=True)

                                point_detail_display = point_batch_detail.copy()
                                for column_name in timeline_columns:
                                    point_detail_display[column_name] = point_detail_display[column_name].map(
                                        lambda value: f"{float(value):,.0f}".replace(",", " ")
                                    ).astype(object)
                                for row_index, category in point_detail_display["Категория"].items():
                                    shelf_days = product_lifecycle_days(category)
                                    point_detail_display.loc[row_index, timeline_columns[shelf_days:]] = "—"
                                    shipment_date_row = pd.to_datetime(
                                        point_detail_display.at[row_index, "Дата отгрузки"], errors="coerce"
                                    )
                                    if pd.notna(shipment_date_row):
                                        shipment_day = shipment_date_row.date()
                                        for day_number in range(1, shelf_days + 1):
                                            if shipment_day + timedelta(days=day_number) > today_for_freshness:
                                                point_detail_display.at[row_index, f"День {day_number}"] = "—"

                                point_detail_columns = [
                                    "Точка", "Магазин", "Категория", "Отгружено по плану",
                                    "Продано в зелёный период", "Продано в серый период",
                                    "Продано за срок", "Реализация, %",
                                    "Живой остаток, шт.", "Списания", "Списание, %",
                                    *timeline_columns, "Выручка SKU, ₽",
                                    "Убыток от списания, ₽", "Скорость выбытия, ч",
                                    "Последний день срока", "Статус партии",
                                ]
                                point_detail_table = point_detail_display[point_detail_columns]
                                point_detail_screen, point_detail_limit, point_detail_truncated = styler_safe_preview(
                                    point_detail_table
                                )
                                st.markdown(
                                    f"#### Реализация по точкам · SKU {selected_batch_sku} — {selected_batch_name}"
                                )
                                if pd.notna(selected_batch_date):
                                    st.caption(
                                        f"Партия от {selected_batch_date:%d.%m.%Y}. "
                                        "Сумма по строкам ниже соответствует выбранной суммарной партии сверху."
                                    )
                                else:
                                    st.caption(
                                        "Сумма по строкам ниже соответствует выбранной суммарной партии сверху."
                                    )
                                if point_detail_truncated:
                                    st.info(
                                        f"Показаны первые {point_detail_limit:,} строк детализации.".replace(",", " ")
                                    )
                                st.dataframe(
                                    point_detail_screen.style.apply(style_sales_timeline, axis=1),
                                    use_container_width=True,
                                    hide_index=True,
                                    height=min(520, 38 * len(point_detail_screen) + 80),
                                    column_config={
                                        "Магазин": st.column_config.NumberColumn("Магазин", format="%d"),
                                        "Отгружено по плану": st.column_config.NumberColumn(format="%.0f"),
                                        "Продано в зелёный период": st.column_config.NumberColumn(format="%.0f"),
                                        "Продано в серый период": st.column_config.NumberColumn(format="%.0f"),
                                        "Продано за срок": st.column_config.NumberColumn(format="%.0f"),
                                        "Реализация, %": st.column_config.NumberColumn(format="%.1f%%"),
                                        "Живой остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                                        "Списания": st.column_config.NumberColumn(format="%.0f"),
                                        "Списание, %": st.column_config.NumberColumn(format="%.1f%%"),
                                        "Выручка SKU, ₽": st.column_config.NumberColumn(format="%.2f"),
                                        "Убыток от списания, ₽": st.column_config.NumberColumn(format="%.2f"),
                                        "Скорость выбытия, ч": st.column_config.NumberColumn(format="%.1f"),
                                        "Последний день срока": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                    },
                                )
                                detail_metrics = st.columns(5)
                                detail_metrics[0].metric(
                                    "План, шт.",
                                    f"{point_batch_detail['Отгружено по плану'].sum():,.0f}".replace(",", " "),
                                )
                                detail_metrics[1].metric(
                                    "Продано, шт.",
                                    f"{point_batch_detail['Продано за срок'].sum():,.0f}".replace(",", " "),
                                )
                                detail_metrics[2].metric(
                                    "Живой остаток, шт.",
                                    f"{point_batch_detail['Живой остаток, шт.'].sum():,.0f}".replace(",", " "),
                                )
                                detail_metrics[3].metric(
                                    "Списано, шт.",
                                    f"{point_batch_detail['Списания'].sum():,.0f}".replace(",", " "),
                                )
                                detail_metrics[4].metric(
                                    "Выручка, ₽",
                                    f"{point_batch_detail['Выручка SKU, ₽'].sum():,.0f}".replace(",", " "),
                                )

                                # Фактический остаток относится к SKU + точке на дату снимка. Он не
                                # приписывается отдельной партии без фактического batch-id, поэтому
                                # показываем отдельную честную сверку со всеми живыми партиями SKU.
                                if selected_stock_date is not None and not stock_snapshot.empty:
                                    actual_selected_stock = stock_snapshot[
                                        stock_snapshot["sku"].astype(str).eq(selected_batch_sku)
                                        & stock_snapshot["point"].isin(selected_time_points)
                                    ].copy()
                                    calculated_stock = build_calculated_stock_snapshot(
                                        sales_time_plans,
                                        time_sales,
                                        entities,
                                        point_to_shop,
                                        selected_time_points,
                                        selected_stock_date,
                                    )
                                    stock_comparison = merge_actual_and_calculated_stock(
                                        actual_selected_stock,
                                        calculated_stock,
                                        selected_time_points,
                                    )
                                    if not stock_comparison.empty:
                                        stock_comparison = stock_comparison[
                                            stock_comparison["SKU"].astype(str).eq(selected_batch_sku)
                                        ].copy()
                                    if not stock_comparison.empty:
                                        stock_comparison.insert(0, "Дата остатка", selected_stock_date)
                                        stock_comparison.insert(1, "День недели", WEEKDAY_RU.get(selected_stock_date.weekday(), ""))
                                        stock_columns = [
                                            "Дата остатка", "День недели", "Точка", "Магазин", "SKU",
                                            "Название товара", "Категория", "Расчётный живой остаток, шт.",
                                            "Факт. остаток, шт.", "Отклонение, шт.", "Статус сверки",
                                        ]
                                        stock_comparison_display = stock_comparison[stock_columns].copy()
                                        st.markdown(
                                            f"#### Фактический остаток по точкам · {selected_stock_date:%d.%m.%Y} · "
                                            f"{WEEKDAY_RU.get(selected_stock_date.weekday(), '')}"
                                        )
                                        st.caption(
                                            "Факт сравнивается с суммой расчётных живых остатков всех действующих партий "
                                            "этого SKU в точке на дату снимка. Поэтому остаток не приписывается выбранной партии искусственно."
                                        )
                                        st.dataframe(
                                            stock_comparison_display.style.apply(style_stock_comparison, axis=1),
                                            use_container_width=True,
                                            hide_index=True,
                                            column_config={
                                                "Дата остатка": st.column_config.DateColumn(format="DD.MM.YYYY"),
                                                "Магазин": st.column_config.NumberColumn(format="%d"),
                                                "Расчётный живой остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                                                "Факт. остаток, шт.": st.column_config.NumberColumn(format="%.0f"),
                                                "Отклонение, шт.": st.column_config.NumberColumn(format="%+.0f"),
                                            },
                                        )
                                        actual_total = pd.to_numeric(
                                            stock_comparison["Факт. остаток, шт."], errors="coerce"
                                        ).fillna(0).sum()
                                        calculated_total = pd.to_numeric(
                                            stock_comparison["Расчётный живой остаток, шт."], errors="coerce"
                                        ).fillna(0).sum()
                                        stock_metrics = st.columns(3)
                                        stock_metrics[0].metric(
                                            "Расчётный живой остаток",
                                            f"{calculated_total:,.0f}".replace(",", " "),
                                        )
                                        stock_metrics[1].metric(
                                            "Фактический остаток",
                                            f"{actual_total:,.0f}".replace(",", " "),
                                        )
                                        stock_metrics[2].metric(
                                            "Отклонение",
                                            f"{actual_total - calculated_total:+,.0f}".replace(",", " "),
                                        )
                                        stock_chart = stock_comparison[[
                                            "Точка", "Расчётный живой остаток, шт.", "Факт. остаток, шт."
                                        ]].melt(
                                            id_vars="Точка", var_name="Показатель", value_name="Количество, шт."
                                        )
                                        st.plotly_chart(
                                            px.bar(
                                                stock_chart,
                                                x="Точка",
                                                y="Количество, шт.",
                                                color="Показатель",
                                                barmode="group",
                                                title="Расчётный и фактический остаток по точкам",
                                            ),
                                            use_container_width=True,
                                        )
                                    else:
                                        st.info(
                                            f"В снимке остатков на {selected_stock_date:%d.%m.%Y} нет SKU "
                                            f"{selected_batch_sku} по выбранным точкам."
                                        )

                    # Сравнение нескольких блюд/SKU. Основная таблица и график работают как по одной,
                    # так и по нескольким выбранным категориям.
                    st.markdown("#### Сравнение блюд в выбранных категориях")
                    sku_lifecycle = (
                        sales_time_menu.groupby(
                            ["Категория", "SKU", "Название товара"], as_index=False
                        )
                        .agg(
                            Отгружено=("Отгружено по плану", "sum"),
                            Зелёные_продажи=("Продано в зелёный период", "sum"),
                            Серые_продажи=("Продано в серый период", "sum"),
                            Списания=("Списания", "sum"),
                            Остаток=("Расчётный остаток", "sum"),
                            Выручка=("Выручка SKU, ₽", "sum"),
                            Убыток=("Убыток от списания, ₽", "sum"),
                            Вес_часов=("Вес часов выбытия", "sum"),
                            Количество_для_скорости=("Количество для скорости", "sum"),
                        )
                    )
                    sku_lifecycle = sku_lifecycle[sku_lifecycle["Отгружено"] > 0].copy()

                    completed_rows = sales_time_menu[
                        sales_time_menu["Статус партии"].astype(str).str.startswith("Срок завершён")
                    ].copy()
                    if not completed_rows.empty:
                        completed_eval = (
                            completed_rows.groupby(
                                ["Категория", "SKU", "Название товара"], as_index=False
                            )
                            .agg(
                                Завершено_отгружено=("Отгружено по плану", "sum"),
                                Завершено_зелёные=("Продано в зелёный период", "sum"),
                                Завершено_серые=("Продано в серый период", "sum"),
                                Завершено_списания=("Списания", "sum"),
                            )
                        )
                        sku_lifecycle = sku_lifecycle.merge(
                            completed_eval,
                            on=["Категория", "SKU", "Название товара"],
                            how="left",
                        )
                    else:
                        for column_name in [
                            "Завершено_отгружено", "Завершено_зелёные",
                            "Завершено_серые", "Завершено_списания",
                        ]:
                            sku_lifecycle[column_name] = 0.0

                    for column_name in [
                        "Завершено_отгружено", "Завершено_зелёные",
                        "Завершено_серые", "Завершено_списания",
                    ]:
                        sku_lifecycle[column_name] = pd.to_numeric(
                            sku_lifecycle[column_name], errors="coerce"
                        ).fillna(0.0)

                    # GroupBy can return object/nullable columns when some rows contain pd.NA.
                    # Normalize every freshness metric before arithmetic so pandas 2.3+/Python 3.14
                    # never leaves the result as dtype=object (which breaks Series.round).
                    for column_name in [
                        "Отгружено", "Зелёные_продажи", "Серые_продажи", "Списания",
                        "Остаток", "Выручка", "Убыток", "Вес_часов", "Количество_для_скорости",
                    ]:
                        sku_lifecycle[column_name] = numeric_series(sku_lifecycle[column_name]).fillna(0.0)

                    sku_lifecycle["Живой остаток"] = (
                        sku_lifecycle["Остаток"] - sku_lifecycle["Списания"]
                    ).clip(lower=0)
                    sku_lifecycle["Скорость выбытия, ч"] = safe_ratio(
                        sku_lifecycle["Вес_часов"],
                        sku_lifecycle["Количество_для_скорости"],
                    ).round(1)
                    sku_lifecycle["Продано вовремя, %"] = (
                        safe_ratio(sku_lifecycle["Зелёные_продажи"], sku_lifecycle["Отгружено"]) * 100
                    ).fillna(0.0).clip(0, 100)
                    sku_lifecycle["Продано в конце срока, %"] = (
                        safe_ratio(sku_lifecycle["Серые_продажи"], sku_lifecycle["Отгружено"]) * 100
                    ).fillna(0.0).clip(0, 100)
                    sku_lifecycle["Списано, %"] = (
                        safe_ratio(sku_lifecycle["Списания"], sku_lifecycle["Отгружено"]) * 100
                    ).fillna(0.0).clip(0, 100)
                    sku_lifecycle["Живой остаток, %"] = (
                        safe_ratio(sku_lifecycle["Живой остаток"], sku_lifecycle["Отгружено"]) * 100
                    ).fillna(0.0).clip(0, 100)

                    def lifecycle_status(row: pd.Series) -> str:
                        # Оценка делается только по партиям с уже завершившимся жизненным циклом.
                        # Будущие и активные партии не могут ухудшить оценку и не считаются списанием.
                        if float(row["Завершено_отгружено"]) <= 0:
                            return "Наблюдение · цикл не завершён"
                        completed_shipped = max(float(row["Завершено_отгружено"]), 1.0)
                        completed_shares = {
                            "Топ товар": float(row["Завершено_зелёные"]) / completed_shipped,
                            "Товар хуже": float(row["Завершено_серые"]) / completed_shipped,
                            "Наихудший": float(row["Завершено_списания"]) / completed_shipped,
                        }
                        return max(completed_shares, key=completed_shares.get)

                    sku_lifecycle["Оценка"] = sku_lifecycle.apply(lifecycle_status, axis=1)
                    sku_lifecycle["Рейтинг с учётом плана"] = (
                        sku_lifecycle["Зелёные_продажи"]
                        - 0.5 * sku_lifecycle["Серые_продажи"]
                        - 2.0 * sku_lifecycle["Списания"]
                    )

                    if sku_lifecycle.empty:
                        st.info("Недостаточно данных для сравнения блюд.")
                    else:
                        comparison_table = sku_lifecycle[
                            [
                                "Категория", "SKU", "Название товара", "Отгружено",
                                "Зелёные_продажи", "Серые_продажи", "Живой остаток",
                                "Списания", "Списано, %", "Скорость выбытия, ч",
                                "Выручка", "Убыток", "Оценка",
                            ]
                        ].copy()
                        comparison_table = comparison_table.rename(
                            columns={
                                "Отгружено": "Отгружено, шт.",
                                "Зелёные_продажи": "Продано вовремя, шт.",
                                "Серые_продажи": "Продано в конце срока, шт.",
                                "Живой остаток": "Живой остаток, шт.",
                                "Списания": "Списано, шт.",
                                "Выручка": "Выручка, ₽",
                                "Убыток": "Убыток списания, ₽",
                            }
                        )
                        comparison_table = comparison_table.sort_values(
                            ["Категория", "Оценка", "Продано вовремя, шт."],
                            ascending=[True, True, False],
                            kind="stable",
                        )
                        st.dataframe(
                            comparison_table,
                            use_container_width=True,
                            hide_index=True,
                            height=min(650, 38 * len(comparison_table) + 80),
                            column_config={
                                "Списано, %": st.column_config.NumberColumn(format="%.1f%%"),
                                "Скорость выбытия, ч": st.column_config.NumberColumn(format="%.1f"),
                                "Выручка, ₽": st.column_config.NumberColumn(format="%.2f"),
                                "Убыток списания, ₽": st.column_config.NumberColumn(format="%.2f"),
                            },
                        )
                        st.caption(
                            "Оценка «Топ товар / Товар хуже / Наихудший» строится только по партиям, "
                            "у которых жизненный цикл уже завершён. Для блюд только с будущими или активными "
                            "партиями показывается «Наблюдение · цикл не завершён»."
                        )

                        st.markdown("#### Графическая оценка блюд")
                        lifecycle_sort = st.selectbox(
                            "Сортировка графической оценки",
                            [
                                "Комплексный рейтинг с учётом плана",
                                "Лучшие продажи в зелёные дни",
                                "Наибольшая доля списаний",
                                "Наибольшая доля продаж в серые дни",
                                "Наибольший убыток от списания",
                                "Наибольший план отгрузки",
                            ],
                            key="sales_time_sku_sort_v752",
                        )
                        if lifecycle_sort == "Комплексный рейтинг с учётом плана":
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Рейтинг с учётом плана", "Отгружено"], ascending=[False, False]
                            )
                        elif lifecycle_sort == "Наибольшая доля списаний":
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Списано, %", "Убыток"], ascending=[False, False]
                            )
                        elif lifecycle_sort == "Наибольшая доля продаж в серые дни":
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Продано в конце срока, %", "Списано, %"], ascending=[False, False]
                            )
                        elif lifecycle_sort == "Наибольший убыток от списания":
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Убыток", "Списано, %"], ascending=[False, False]
                            )
                        elif lifecycle_sort == "Наибольший план отгрузки":
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Отгружено", "Списано, %"], ascending=[False, False]
                            )
                        else:
                            sku_lifecycle = sku_lifecycle.sort_values(
                                ["Продано вовремя, %", "Продано в конце срока, %", "Списано, %"],
                                ascending=[False, True, True],
                            )

                        chart_data = sku_lifecycle.copy()
                        chart_data["SKU"] = chart_data["SKU"].astype(str)
                        chart_data["Подпись"] = (
                            chart_data["Категория"].astype(str)
                            + " · " + chart_data["SKU"]
                            + " · " + chart_data["Оценка"]
                        )
                        selected_category_caption = (
                            ", ".join(selected_time_categories)
                            if selected_time_categories else "все категории"
                        )
                        st.caption(
                            f"Сравниваются блюда: {selected_category_caption}. "
                            f"На графике SKU: {len(chart_data):,}.".replace(",", " ")
                        )

                        lifecycle_chart = go.Figure()
                        lifecycle_chart.add_bar(
                            y=chart_data["Подпись"],
                            x=chart_data["Зелёные_продажи"],
                            name="Продано вовремя",
                            orientation="h",
                            marker_color="#70AD47",
                            customdata=chart_data[[
                                "Название товара", "Отгружено", "Продано вовремя, %",
                                "Выручка", "Рейтинг с учётом плана",
                            ]],
                            hovertemplate=(
                                "<b>%{y}</b><br>%{customdata[0]}<br>"
                                "Зелёный период: %{x:.0f} шт. (%{customdata[2]:.1f}%)<br>"
                                "Отгружено: %{customdata[1]:.0f} шт.<br>"
                                "Выручка: %{customdata[3]:,.0f} ₽<br>"
                                "Текущий рейтинг: %{customdata[4]:.1f}<extra></extra>"
                            ),
                        )
                        lifecycle_chart.add_bar(
                            y=chart_data["Подпись"],
                            x=chart_data["Серые_продажи"],
                            name="Продано в конце срока",
                            orientation="h",
                            marker_color="#A6A6A6",
                            customdata=chart_data[["Название товара", "Продано в конце срока, %"]],
                            hovertemplate=(
                                "<b>%{y}</b><br>%{customdata[0]}<br>"
                                "Серый период: %{x:.0f} шт. (%{customdata[1]:.1f}%)<extra></extra>"
                            ),
                        )
                        lifecycle_chart.add_bar(
                            y=chart_data["Подпись"],
                            x=chart_data["Живой остаток"],
                            name="Живой остаток",
                            orientation="h",
                            marker_color="#5B9BD5",
                            customdata=chart_data[["Название товара", "Живой остаток, %"]],
                            hovertemplate=(
                                "<b>%{y}</b><br>%{customdata[0]}<br>"
                                "Живой остаток: %{x:.0f} шт. (%{customdata[1]:.1f}%)<br>"
                                "Партия ещё не обязана быть списана.<extra></extra>"
                            ),
                        )
                        lifecycle_chart.add_bar(
                            y=chart_data["Подпись"],
                            x=chart_data["Списания"],
                            name="Списано",
                            orientation="h",
                            marker_color="#C00000",
                            customdata=chart_data[["Название товара", "Списано, %", "Убыток"]],
                            hovertemplate=(
                                "<b>%{y}</b><br>%{customdata[0]}<br>"
                                "Списано: %{x:.0f} шт. (%{customdata[1]:.1f}%)<br>"
                                "Убыток: %{customdata[2]:,.0f} ₽<extra></extra>"
                            ),
                        )
                        lifecycle_chart.update_layout(
                            barmode="stack",
                            height=max(520, 31 * len(chart_data) + 180),
                            margin=dict(l=20, r=20, t=60, b=20),
                            xaxis=dict(title="Количество от плановой отгрузки, шт.", rangemode="tozero"),
                            yaxis=dict(title="", autorange="reversed"),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                            hovermode="y unified",
                        )
                        st.plotly_chart(lifecycle_chart, use_container_width=True)
                        st.caption(
                            "График сравнивает несколько блюд выбранной категории: зелёный — продажи вовремя, "
                            "серый — продажи в завершающий срок, синий — живой остаток активных/будущих партий, "
                            "красный — только фактическое списание после завершения жизненного цикла."
                        )
if tab_category_writeoffs.open:
    with tab_category_writeoffs:
        st.subheader("Списания по категориям и SKU")
        st.caption(
            "Расчёт использует период и точки из вкладки «Окно свежести». "
            "Нажмите категорию, чтобы увидеть суммарные списания по SKU; затем нажмите SKU, "
            "чтобы раскрыть его в разрезе точек."
        )
        category_writeoff_source = st.session_state.get(
            "freshness_category_source_v62", pd.DataFrame()
        )
        category_writeoff_source_by_point = st.session_state.get(
            "freshness_category_source_by_point_v63", pd.DataFrame()
        )
        category_writeoff_context = st.session_state.get("freshness_category_context_v62")
        if category_writeoff_source.empty:
            st.info(
                "Сначала выберите период и точки во вкладке «Окно свежести». "
                "После расчёта здесь появятся списания."
            )
        else:
            if category_writeoff_context:
                st.caption(
                    f"Период: {category_writeoff_context['start']:%d.%m.%Y}–"
                    f"{category_writeoff_context['end']:%d.%m.%Y}; точки: "
                    f"{', '.join(category_writeoff_context['points'])}."
                )

            # В старой сессии отдельной детализации по точкам может ещё не быть.
            # Для одной точки агрегированный источник эквивалентен; для нескольких точек
            # пользователь получит полноценный разрез после повторного расчёта «Окна свежести».
            if category_writeoff_source_by_point.empty:
                category_writeoff_source_by_point = category_writeoff_source.copy()

            writeoff_category_options = sorted(
                category_writeoff_source["Категория"].dropna().astype(str).unique().tolist()
            )
            stored_writeoff_categories = st.session_state.get("writeoff_categories_v62")
            if stored_writeoff_categories is not None:
                valid_writeoff_categories = [
                    category for category in stored_writeoff_categories
                    if category in writeoff_category_options
                ]
                if valid_writeoff_categories != list(stored_writeoff_categories):
                    st.session_state["writeoff_categories_v62"] = (
                        valid_writeoff_categories or writeoff_category_options
                    )
            selected_writeoff_categories = st.multiselect(
                "Категории",
                options=writeoff_category_options,
                default=writeoff_category_options,
                key="writeoff_categories_v62",
                placeholder="Выберите категории",
            )

            category_writeoff_detail = category_writeoff_source[
                category_writeoff_source["Категория"].isin(selected_writeoff_categories)
            ].copy()
            category_writeoff_by_point = category_writeoff_source_by_point[
                category_writeoff_source_by_point["Категория"].isin(selected_writeoff_categories)
            ].copy()

            if category_writeoff_detail.empty:
                st.warning("Выберите хотя бы одну категорию.")
            else:
                category_writeoff_summary = (
                    category_writeoff_detail.groupby("Категория", as_index=False)
                    .agg(
                        **{
                            "SKU": ("SKU", "nunique"),
                            "Отгружено, шт.": ("Отгружено по плану", "sum"),
                            "Продано за срок, шт.": ("Продано за срок", "sum"),
                            "Списано, шт.": ("Списания", "sum"),
                            "Выручка, ₽": ("Выручка SKU, ₽", "sum"),
                            "Сумма списания, ₽": ("Убыток от списания, ₽", "sum"),
                        }
                    )
                    .sort_values(["Списано, шт.", "Категория"], ascending=[False, True])
                    .reset_index(drop=True)
                )
                category_writeoff_summary["Доля списания, %"] = (
                    safe_ratio(
                        category_writeoff_summary["Списано, шт."],
                        category_writeoff_summary["Отгружено, шт."],
                    ) * 100
                ).fillna(0.0).round(1)

                # Суммарный SKU-уровень строим из исходной детализации по точкам,
                # чтобы значения не зависели от экранного суммирования в «Окне свежести».
                writeoff_sku_summary_all = (
                    category_writeoff_by_point.groupby(
                        ["Категория", "SKU", "Название товара"], as_index=False, dropna=False
                    )
                    .agg(
                        **{
                            "Отгружено, шт.": ("Отгружено по плану", "sum"),
                            "Продано за срок, шт.": ("Продано за срок", "sum"),
                            "Списано, шт.": ("Списания", "sum"),
                            "Выручка, ₽": ("Выручка SKU, ₽", "sum"),
                            "Сумма списания, ₽": ("Убыток от списания, ₽", "sum"),
                        }
                    )
                )
                writeoff_sku_summary_all["Доля списания, %"] = (
                    safe_ratio(
                        writeoff_sku_summary_all["Списано, шт."],
                        writeoff_sku_summary_all["Отгружено, шт."],
                    ) * 100
                ).fillna(0.0).round(1)
                writeoff_sku_summary = writeoff_sku_summary_all[
                    pd.to_numeric(writeoff_sku_summary_all["Списано, шт."], errors="coerce").fillna(0) > 0
                ].copy()
                writeoff_sku_summary = writeoff_sku_summary.sort_values(
                    ["Категория", "Списано, шт.", "Сумма списания, ₽"],
                    ascending=[True, False, False],
                ).reset_index(drop=True)

                sku_with_writeoff_counts = (
                    writeoff_sku_summary.groupby("Категория")["SKU"].nunique()
                    if not writeoff_sku_summary.empty else pd.Series(dtype="int64")
                )
                category_writeoff_summary["SKU со списанием"] = (
                    category_writeoff_summary["Категория"].map(sku_with_writeoff_counts).fillna(0).astype(int)
                )
                # Ставим показатель рядом с общим числом SKU.
                category_cols = list(category_writeoff_summary.columns)
                if "SKU со списанием" in category_cols:
                    category_cols.remove("SKU со списанием")
                    sku_pos = category_cols.index("SKU") + 1 if "SKU" in category_cols else 1
                    category_cols.insert(sku_pos, "SKU со списанием")
                    category_writeoff_summary = category_writeoff_summary[category_cols]

                total_plan_quantity = category_writeoff_summary["Отгружено, шт."].sum()
                total_writeoff_quantity = category_writeoff_summary["Списано, шт."].sum()
                total_writeoff_amount = category_writeoff_summary["Сумма списания, ₽"].sum()
                writeoff_metrics = st.columns(5)
                writeoff_metrics[0].metric("Категорий", len(category_writeoff_summary))
                writeoff_metrics[1].metric(
                    "Отгружено, шт.", f"{total_plan_quantity:,.0f}".replace(",", " ")
                )
                writeoff_metrics[2].metric(
                    "Списано, шт.", f"{total_writeoff_quantity:,.0f}".replace(",", " ")
                )
                writeoff_metrics[3].metric(
                    "Сумма списания, ₽", f"{total_writeoff_amount:,.0f}".replace(",", " ")
                )
                writeoff_metrics[4].metric(
                    "Доля списания",
                    f"{(total_writeoff_quantity / total_plan_quantity * 100) if total_plan_quantity else 0:.1f}%",
                )

                writeoff_chart = px.bar(
                    category_writeoff_summary,
                    x="Категория",
                    y="Списано, шт.",
                    color="Доля списания, %",
                    color_continuous_scale=["#70AD47", "#FFD966", "#F8696B", "#C00000"],
                    text_auto=".0f",
                    title="Количество и доля списаний по категориям",
                    hover_data=["Отгружено, шт.", "Сумма списания, ₽", "SKU со списанием"],
                )
                writeoff_chart.update_layout(height=440)
                st.plotly_chart(writeoff_chart, use_container_width=True)

                st.markdown("#### Итог по категориям")
                st.caption(
                    "Нажмите строку категории — ниже откроется суммарный список SKU, по которым было списание."
                )
                category_writeoff_selection = st.dataframe(
                    category_writeoff_summary,
                    use_container_width=True,
                    hide_index=True,
                    height=min(620, 38 * len(category_writeoff_summary) + 80),
                    on_select="rerun",
                    selection_mode="single-row",
                    key="writeoff_category_select_v63",
                    column_config={
                        "SKU": st.column_config.NumberColumn(format="%d"),
                        "SKU со списанием": st.column_config.NumberColumn(format="%d"),
                        "Доля списания, %": st.column_config.NumberColumn(format="%.1f%%"),
                        "Выручка, ₽": st.column_config.NumberColumn(format="%.2f"),
                        "Сумма списания, ₽": st.column_config.NumberColumn(format="%.2f"),
                    },
                )

                selected_category_rows = list(
                    getattr(category_writeoff_selection.selection, "rows", []) or []
                )
                selected_writeoff_category = None
                selected_writeoff_sku = None
                if selected_category_rows:
                    selected_category_index = int(selected_category_rows[0])
                    if 0 <= selected_category_index < len(category_writeoff_summary):
                        selected_writeoff_category = str(
                            category_writeoff_summary.iloc[selected_category_index]["Категория"]
                        )

                if selected_writeoff_category:
                    selected_category_sku = writeoff_sku_summary[
                        writeoff_sku_summary["Категория"].astype(str).eq(selected_writeoff_category)
                    ].copy().reset_index(drop=True)
                    st.markdown(
                        f"#### SKU списания категории «{selected_writeoff_category}» · суммарно"
                    )
                    if selected_category_sku.empty:
                        st.info("В выбранной категории за этот период фактических списаний нет.")
                    else:
                        st.caption(
                            "Значения SKU суммированы по всем выбранным точкам. Нажмите SKU — ниже появится разрез по точкам."
                        )
                        sku_writeoff_selection = st.dataframe(
                            selected_category_sku,
                            use_container_width=True,
                            hide_index=True,
                            height=min(650, 38 * len(selected_category_sku) + 80),
                            on_select="rerun",
                            selection_mode="single-row",
                            key="writeoff_sku_select_v63",
                            column_config={
                                "Доля списания, %": st.column_config.NumberColumn(format="%.1f%%"),
                                "Выручка, ₽": st.column_config.NumberColumn(format="%.2f"),
                                "Сумма списания, ₽": st.column_config.NumberColumn(format="%.2f"),
                            },
                        )
                        selected_sku_rows = list(
                            getattr(sku_writeoff_selection.selection, "rows", []) or []
                        )
                        if selected_sku_rows:
                            selected_sku_index = int(selected_sku_rows[0])
                            if 0 <= selected_sku_index < len(selected_category_sku):
                                selected_writeoff_sku = str(
                                    selected_category_sku.iloc[selected_sku_index]["SKU"]
                                )
                                selected_writeoff_name = str(
                                    selected_category_sku.iloc[selected_sku_index]["Название товара"]
                                )

                                sku_point_detail = category_writeoff_by_point[
                                    category_writeoff_by_point["Категория"].astype(str).eq(selected_writeoff_category)
                                    & category_writeoff_by_point["SKU"].astype(str).eq(selected_writeoff_sku)
                                ].copy()
                                if not sku_point_detail.empty:
                                    point_group_columns = ["Точка"]
                                    sku_point_summary = (
                                        sku_point_detail.groupby(
                                            point_group_columns, as_index=False, dropna=False
                                        )
                                        .agg(
                                            **{
                                                "Отгружено, шт.": ("Отгружено по плану", "sum"),
                                                "Продано за срок, шт.": ("Продано за срок", "sum"),
                                                "Списано, шт.": ("Списания", "sum"),
                                                "Выручка, ₽": ("Выручка SKU, ₽", "sum"),
                                                "Сумма списания, ₽": ("Убыток от списания, ₽", "sum"),
                                            }
                                        )
                                    )
                                    sku_point_summary["Доля списания, %"] = (
                                        safe_ratio(
                                            sku_point_summary["Списано, шт."],
                                            sku_point_summary["Отгружено, шт."],
                                        ) * 100
                                    ).fillna(0.0).round(1)
                                    sku_point_summary = sku_point_summary.sort_values(
                                        ["Списано, шт.", "Сумма списания, ₽", "Точка"],
                                        ascending=[False, False, True],
                                    ).reset_index(drop=True)

                                    st.markdown(
                                        f"#### SKU {selected_writeoff_sku} · {selected_writeoff_name} · по точкам"
                                    )
                                    st.caption(
                                        "Показаны все выбранные точки, где этот SKU был в плане. "
                                        "Нулевое списание оставлено для сравнения между точками."
                                    )
                                    st.dataframe(
                                        sku_point_summary,
                                        use_container_width=True,
                                        hide_index=True,
                                        height=min(560, 38 * len(sku_point_summary) + 80),
                                        column_config={
                                            "Доля списания, %": st.column_config.NumberColumn(format="%.1f%%"),
                                            "Выручка, ₽": st.column_config.NumberColumn(format="%.2f"),
                                            "Сумма списания, ₽": st.column_config.NumberColumn(format="%.2f"),
                                        },
                                    )
                                else:
                                    st.info("Для выбранного SKU нет детализации по точкам.")

                # Полная выгрузка сохраняет иерархию: категории -> SKU со списанием -> точки.
                writeoff_point_export = (
                    category_writeoff_by_point.groupby(
                        ["Категория", "SKU", "Название товара", "Точка"],
                        as_index=False,
                        dropna=False,
                    )
                    .agg(
                        **{
                            "Отгружено, шт.": ("Отгружено по плану", "sum"),
                            "Продано за срок, шт.": ("Продано за срок", "sum"),
                            "Списано, шт.": ("Списания", "sum"),
                            "Выручка, ₽": ("Выручка SKU, ₽", "sum"),
                            "Сумма списания, ₽": ("Убыток от списания, ₽", "sum"),
                        }
                    )
                )
                writeoff_point_export["Доля списания, %"] = (
                    safe_ratio(
                        writeoff_point_export["Списано, шт."],
                        writeoff_point_export["Отгружено, шт."],
                    ) * 100
                ).fillna(0.0).round(1)
                if not writeoff_sku_summary.empty and not writeoff_point_export.empty:
                    affected_keys = set(
                        zip(
                            writeoff_sku_summary["Категория"].astype(str),
                            writeoff_sku_summary["SKU"].astype(str),
                        )
                    )
                    writeoff_point_export = writeoff_point_export[
                        [
                            (str(category), str(sku)) in affected_keys
                            for category, sku in zip(
                                writeoff_point_export["Категория"],
                                writeoff_point_export["SKU"],
                            )
                        ]
                    ].copy()
                writeoff_point_export = writeoff_point_export.sort_values(
                    ["Категория", "SKU", "Списано, шт.", "Точка"],
                    ascending=[True, True, False, True],
                ).reset_index(drop=True)

                writeoff_export_buffer = io.BytesIO()
                with pd.ExcelWriter(writeoff_export_buffer, engine="openpyxl") as writer:
                    category_writeoff_summary.to_excel(
                        writer, sheet_name="Итог по категориям", index=False
                    )
                    writeoff_sku_summary.to_excel(
                        writer, sheet_name="SKU списания суммарно", index=False
                    )
                    writeoff_point_export.to_excel(
                        writer, sheet_name="SKU по точкам", index=False
                    )
                    for worksheet in writer.book.worksheets:
                        worksheet.freeze_panes = "A2"
                        worksheet.auto_filter.ref = worksheet.dimensions
                        for cells in worksheet.columns:
                            width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 42)
                            worksheet.column_dimensions[cells[0].column_letter].width = width
                st.download_button(
                    "Скачать списания: категории, SKU и точки (Excel)",
                    data=writeoff_export_buffer.getvalue(),
                    file_name=(
                        f"списания_категории_SKU_точки_"
                        f"{category_writeoff_context['start']:%Y-%m-%d}_"
                        f"{category_writeoff_context['end']:%Y-%m-%d}.xlsx"
                        if category_writeoff_context else "списания_категории_SKU_точки.xlsx"
                    ),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_category_writeoffs_v63",
                )


if tab_plan_check.open:
    with tab_plan_check:
        st.subheader("Проверка готового плана")
        st.caption(
            "Проверка НЕ строит новый план. Базой всегда остаётся уже заполненное значение категорийщика "
            "из текущей матрицы. Система проверяет его по SKU, фактической свежести, категории, сущности "
            "и параметрам точки из «Авто Юнит точки ВМ», после чего предлагает только корректировку к текущему значению."
        )

        check_matrix_bytes, check_matrix_source, check_matrix_checked_at, check_matrix_error = _load_matrix_context_for_active_tab()
        if not check_matrix_bytes:
            st.error("Текущая матрица недоступна. Проверьте Apps Script или резервный combo_matrix.xlsx.")
            if check_matrix_error:
                st.caption(f"Причина: {check_matrix_error}")
        else:
            source_text = check_matrix_source or "текущая матрица"
            checked_text = str(check_matrix_checked_at or "").replace("T", " ")
            st.success(f"Готовый план подключён: {source_text}" + (f" · проверено {checked_text}" if checked_text else ""))
            if check_matrix_error:
                st.caption(f"Использован резервный источник. Основной источник сообщил: {check_matrix_error}")

            try:
                ready_matrix_plans = parse_analyst_plan_history(check_matrix_bytes)
            except Exception as error:
                st.error(f"Не удалось прочитать заполненный план из матрицы: {error}")
                ready_matrix_plans = pd.DataFrame()

            auto_unit_bytes = b""
            auto_unit_source = ""
            with st.expander("Авто Юнит точки ВМ · параметры точек и экономика", expanded=False):
                st.caption(
                    "Используются график работы, ТОП-3 аудитории, дневная выручка, чеки и средний чек. "
                    "Если файл auto_unit_points_vm.xlsx / «Авто Юнит точки ВМ.xlsx» лежит рядом с app.py, он подключается автоматически."
                )
                auto_unit_upload = st.file_uploader(
                    "Заменить источник Авто Юнит (.xlsx)",
                    type=["xlsx"],
                    key="ready_plan_check_auto_unit_upload_v2",
                )
                if auto_unit_upload is not None:
                    auto_unit_bytes = auto_unit_upload.getvalue()
                    auto_unit_source = auto_unit_upload.name
                elif AUTO_UNIT_FILE.exists():
                    try:
                        auto_unit_bytes = AUTO_UNIT_FILE.read_bytes()
                        auto_unit_source = AUTO_UNIT_FILE.name
                    except OSError:
                        auto_unit_bytes = b""

            if auto_unit_bytes:
                try:
                    auto_unit_profiles, auto_unit_daily = parse_auto_unit_points(auto_unit_bytes)
                except Exception as error:
                    st.warning(f"Авто Юнит не удалось прочитать: {error}")
                    auto_unit_profiles, auto_unit_daily = pd.DataFrame(), pd.DataFrame()
            else:
                auto_unit_profiles, auto_unit_daily = pd.DataFrame(), pd.DataFrame()

            if not auto_unit_profiles.empty:
                active_profiles = auto_unit_profiles[auto_unit_profiles["active"].fillna(False)].copy()
                auto_metrics = st.columns(4)
                auto_metrics[0].metric("Авто Юнит · точек", active_profiles["point_number"].nunique())
                auto_metrics[1].metric("5-дневных", int(pd.to_numeric(active_profiles["work_days_per_week"], errors="coerce").eq(5).sum()))
                auto_metrics[2].metric("Женская аудитория", int(active_profiles["female_dominant"].fillna(False).sum()))
                latest_auto_date = pd.to_datetime(auto_unit_daily.get("business_date"), errors="coerce").max() if not auto_unit_daily.empty else pd.NaT
                auto_metrics[3].metric("Данные по", latest_auto_date.strftime("%d.%m.%Y") if pd.notna(latest_auto_date) else "—")
                st.caption(f"Источник Авто Юнит: {auto_unit_source}")
            else:
                st.warning("Авто Юнит не подключён. Базовая проверка SKU/свежести/категории/сущности продолжит работать, но правила графика, аудитории и среднего чека применены не будут.")

            if ready_matrix_plans.empty:
                st.warning("В текущей матрице не найдены заполненные основные блоки «План на день кухня».")
            else:
                ready_matrix_plans["plan_date"] = pd.to_datetime(ready_matrix_plans["plan_date"], errors="coerce").dt.date
                check_dates = sorted(ready_matrix_plans["plan_date"].dropna().unique().tolist())
                future_dates = [value for value in check_dates if value >= today]
                default_check_date = future_dates[0] if future_dates else check_dates[-1]
                default_index = check_dates.index(default_check_date)

                check_control_columns = st.columns([1.5, 1.0, 1.0, 1.0, 1.0, 1.0])
                with check_control_columns[0]:
                    check_target_date = st.selectbox(
                        "Дата готового плана",
                        check_dates,
                        index=default_index,
                        format_func=lambda value: f"{value:%d.%m.%Y} · {WEEKDAY_RU.get(value.weekday(), '')}",
                        key="ready_plan_check_date_v2",
                    )
                with check_control_columns[1]:
                    check_trend_percent = st.number_input("Рост / падение, %", min_value=1.0, max_value=100.0, value=10.0, step=1.0, key="ready_plan_check_trend_v2")
                with check_control_columns[2]:
                    check_writeoff_percent = st.number_input("Большое списание, %", min_value=1.0, max_value=100.0, value=20.0, step=1.0, key="ready_plan_check_writeoff_v2")
                with check_control_columns[3]:
                    check_min_sale_days = st.number_input("Мин. дней SKU", min_value=1, max_value=30, value=3, step=1, key="ready_plan_check_min_days_v2")
                with check_control_columns[4]:
                    check_avg_target = st.number_input("Цель ср. чека, %", min_value=-50.0, max_value=100.0, value=5.0, step=1.0, key="ready_plan_check_avg_target_v2")
                with check_control_columns[5]:
                    check_revenue_target = st.number_input("Цель дохода, %", min_value=-50.0, max_value=100.0, value=5.0, step=1.0, key="ready_plan_check_revenue_target_v2")

                light_keywords_text = st.text_input(
                    "Ключи лёгких овощных блюд",
                    value=", ".join(PLAN_CHECK_LIGHT_VEGETABLE_DEFAULT),
                    key="ready_plan_check_light_keywords_v2",
                    help="По этим фрагментам в категории/сущности/названии определяется лёгкое овощное блюдо для правила женской аудитории.",
                )
                light_keywords = tuple(
                    value.strip() for value in light_keywords_text.split(",") if value.strip()
                ) or PLAN_CHECK_LIGHT_VEGETABLE_DEFAULT

                target_plan_preview = ready_matrix_plans[ready_matrix_plans["plan_date"].eq(check_target_date)].copy()
                plan_preview_metrics = st.columns(4)
                plan_preview_metrics[0].metric("Строк плана", f"{len(target_plan_preview):,}".replace(",", " "))
                plan_preview_metrics[1].metric("Точек", target_plan_preview["point_number"].nunique())
                plan_preview_metrics[2].metric("SKU", target_plan_preview["sku"].nunique())
                plan_preview_metrics[3].metric("План, шт.", f"{pd.to_numeric(target_plan_preview['analyst_plan'], errors='coerce').fillna(0).sum():,.0f}".replace(",", " "))
                st.caption("Проверка выполняется по всем точкам и категориям выбранной даты. Рекомендация всегда считается от текущего значения в матрице.")

                with st.expander("Логика проверки", expanded=False):
                    st.markdown(
                        f"""
**Основа**  
1. SKU — среднее только по дням фактических продаж за 2 месяца.  
2. Цикл — фактическая свежесть FIFO: зелёные/серые продажи и списания; День 1 начинается на следующий день после даты отгрузки/плана.  
3. Категория — 14 дней: последние 7 против предыдущих 7.  
4. Сущность — fallback при нехватке данных SKU и отдельный сигнал роста/падения.

**Добавить к уже проставленному значению**  
- +1: товар продаётся в зелёном периоде, у живой партии остался 1 зелёный день.  
- +1: сущность растёт минимум на **{check_trend_percent:.0f}%**, категория не противоречит.  
- +1: по Авто Юнит женская аудитория преобладает, а SKU определён как лёгкое овощное блюдо и нет сильного риска списаний/падения.

**Убавить от уже проставленного значения**  
- −1: SKU падает минимум на **{check_trend_percent:.0f}%** и списание ≥ **{check_writeoff_percent:.0f}%**.  
- если серые продажи преобладают — ориентир на зелёные продажи.  
- −2: сущность падает минимум на **{check_trend_percent:.0f}%**.

**Минимумы**: вторые блюда 3 · остальные 2 · напитки 5 · Япония 1.  
**Исключение Авто Юнит**: на 5-дневной точке в **четверг** проверяется загрузка только **1 дня**, поэтому нижний минимум категории в этот день не применяется.

**Экономика точки**: последние 14 дней среднего чека и выручки/рабочий день сравниваются с предыдущими 14. Цели: ср. чек ≥ **{check_avg_target:+.1f}%**, доход ≥ **{check_revenue_target:+.1f}%**.

**Цвет согласия**  
🟢 зелёный — система согласна; 🟡 жёлтый — сомнение/рекомендация; 🔴 красный — полное несогласие.
                        """
                    )

                check_signature = (
                    _combo_matrix_signature(check_matrix_bytes, check_matrix_source),
                    check_target_date,
                    round(float(check_trend_percent), 3),
                    round(float(check_writeoff_percent), 3),
                    int(check_min_sale_days),
                    round(float(check_avg_target), 3),
                    round(float(check_revenue_target), 3),
                    light_keywords,
                    hashlib.sha256(auto_unit_bytes).hexdigest()[:16] if auto_unit_bytes else "NO_AUTO_UNIT",
                    "ALL_POINTS",
                    "ALL_CATEGORIES",
                )
                if st.session_state.get("ready_plan_check_signature_v2") != check_signature:
                    st.session_state.pop("ready_plan_check_result_v2", None)

                if st.button("Проверить готовый план", type="primary", use_container_width=True, key="ready_plan_check_button_v2"):
                    analysis_date = min(today, check_target_date - timedelta(days=1))
                    history_from = (pd.Timestamp(check_target_date) - pd.DateOffset(months=2) - pd.Timedelta(days=7)).date()
                    check_point_mapping = {
                        int(shop): str(label)
                        for shop, label in st.session_state.get("point_mapping", {}).items()
                        if str(label).startswith("Т") and str(label) != "Т11"
                    }
                    if not check_point_mapping:
                        check_point_mapping = {number: f"Т{number}" for number in range(1, 30) if number != 11}
                    try:
                        with st.spinner("Проверяю текущие значения плана: SKU, свежесть, категория, сущность и Авто Юнит…"):
                            check_sales = load_forecast_history(history_from, analysis_date + timedelta(days=1), tuple(sorted(check_point_mapping)))
                            check_result = build_ready_plan_check(
                                ready_matrix_plans,
                                check_sales,
                                entities,
                                check_target_date,
                                check_point_mapping,
                                trend_threshold=float(check_trend_percent) / 100.0,
                                writeoff_threshold=float(check_writeoff_percent) / 100.0,
                                minimum_sale_days=int(check_min_sale_days),
                                selected_points=None,
                                selected_categories=None,
                                auto_unit_profiles=auto_unit_profiles,
                                auto_unit_daily=auto_unit_daily,
                                avg_check_growth_target=float(check_avg_target) / 100.0,
                                revenue_growth_target=float(check_revenue_target) / 100.0,
                                light_vegetable_keywords=light_keywords,
                            )
                    except Exception as error:
                        st.error(f"Не удалось выполнить проверку готового плана: {error}")
                        check_result = pd.DataFrame()
                    if check_result.empty:
                        st.warning("По выбранной дате результат проверки пуст.")
                    else:
                        st.session_state["ready_plan_check_result_v2"] = check_result
                        st.session_state["ready_plan_check_signature_v2"] = check_signature

                check_result = st.session_state.get("ready_plan_check_result_v2", pd.DataFrame())
                if isinstance(check_result, pd.DataFrame) and not check_result.empty:
                    delta_values = pd.to_numeric(check_result["Изменение"], errors="coerce").fillna(0)
                    agreement_text = check_result["Согласие системы"].astype(str)
                    check_metrics = st.columns(7)
                    check_metrics[0].metric("Проверено", f"{len(check_result):,}".replace(",", " "))
                    check_metrics[1].metric("Добавить", int((delta_values > 0).sum()))
                    check_metrics[2].metric("Убавить", int((delta_values < 0).sum()))
                    check_metrics[3].metric("Оставить", int((delta_values == 0).sum()))
                    check_metrics[4].metric("🟢 Согласие", int(agreement_text.str.startswith("Зелёное").sum()))
                    check_metrics[5].metric("🟡 Сомнение", int(agreement_text.str.startswith("Жёлтое").sum()))
                    check_metrics[6].metric("🔴 Несогласие", int(agreement_text.str.startswith("Красное").sum()))

                    st.markdown("#### Проверка прямо в меню")
                    st.caption(
                        "Показан только выбранный день готового меню. Значения категорийщика не заменяются. "
                        "Цвет наносится только на ячейки Т1–Т29: зелёный — согласие, жёлтый — сомнение/рекомендация, "
                        "красный — полное несогласие. Наведите курсор на ячейку, чтобы увидеть объяснение системы."
                    )
                    matrix_html = ready_plan_check_menu_html(
                        check_matrix_bytes,
                        check_result,
                        check_target_date,
                    )
                    if matrix_html:
                        st.markdown(matrix_html, unsafe_allow_html=True)
                    else:
                        st.warning("Не удалось построить вид меню для выбранной даты. Сам расчёт проверки сохранён.")

                    check_excel = ready_plan_check_excel(
                        check_matrix_bytes,
                        check_result,
                        check_target_date,
                        float(check_trend_percent) / 100.0,
                        float(check_writeoff_percent) / 100.0,
                        int(check_min_sale_days),
                        avg_check_growth_target=float(check_avg_target) / 100.0,
                        revenue_growth_target=float(check_revenue_target) / 100.0,
                        light_vegetable_keywords=light_keywords,
                    )
                    st.download_button(
                        "Скачать проверенное меню (Excel)",
                        data=check_excel,
                        file_name=f"проверенное_меню_{check_target_date:%Y-%m-%d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="ready_plan_check_download_v3",
                    )
                    st.caption(
                        "В Excel сохраняется только выбранный блок меню. Текущие значения остаются как у категорийщика; "
                        "цвет показывает оценку системы, а в примечании каждой проверенной ячейки Т находится причина анализа. "
                        "Отсчёт свежести начинается на следующий день после даты отгрузки/плана."
                    )


if tab_forecast.open:
    with tab_forecast:
        st.subheader("Прогноз плана по меню из текущей матрицы")
        st.caption(
            "Меню и даты плана загружаются автоматически из текущей матрицы. "
            "Сначала выберите нужные даты из тех, которые найдены в матрице; затем приложение "
            "рассчитает только выбранные дни и заполнит Т1–Т29 только в соответствующих блоках меню. "
            "Период для расчёта среднего выбирается вручную ползунком ниже — от 1 до 26 недель (примерно 6 месяцев). "
            "Для каждого SKU и каждой точки среднее считается только по дням, когда SKU действительно продавался, "
            "и только внутри выбранного исторического периода. "
            "Множитель загрузки зависит от категории: Япония — ×1, вторые блюда — ×3, "
            "напитки — ×4, остальные категории — ×2. Т11 остаётся пустой."
        )

        matrix_bytes, matrix_source, matrix_checked_at, matrix_error = _load_matrix_context_for_active_tab()
        if not matrix_bytes:
            st.error("Текущая матрица недоступна. Проверьте подключение Apps Script или резервный combo_matrix.xlsx.")
            if matrix_error:
                st.caption(f"Причина: {matrix_error}")
        else:
            forecast_matrix_signature = _combo_matrix_signature(matrix_bytes, matrix_source)
            forecast_signature_key = "forecast_matrix_signature_v760"
            if st.session_state.get(forecast_signature_key) != forecast_matrix_signature:
                for stale_key in [
                    "forecast_result",
                    "forecast_menu",
                    "forecast_history_from",
                    "forecast_target_date",
                    "forecast_first_target_date",
                    "forecast_matrix_bytes",
                    "forecast_selected_matrix_dates_v760",
                    "forecast_preview_block_v760",
                    "forecast_calculated_dates_v760",
                    "forecast_calculated_lookback_weeks_v75111",
                ]:
                    st.session_state.pop(stale_key, None)
                st.session_state[forecast_signature_key] = forecast_matrix_signature

            source_text = matrix_source or "текущая матрица"
            checked_text = str(matrix_checked_at or "").replace("T", " ")
            st.success(
                f"Матрица подключена: {source_text}"
                + (f" · проверено {checked_text}" if checked_text else "")
            )
            if matrix_error:
                st.caption(f"Использован резервный источник. Основной источник сообщил: {matrix_error}")

            try:
                matrix_menu, matrix_capacity = parse_menu_matrix(matrix_bytes)
            except Exception as error:
                st.error(f"Не удалось прочитать меню из текущей матрицы: {error}")
                matrix_menu = pd.DataFrame()
                matrix_capacity = pd.DataFrame()

            if matrix_menu.empty:
                st.warning("В текущей матрице не найдено меню с датой, SKU и колонками точек Т.")
            else:
                all_target_dates = sorted(matrix_menu["target_date"].dropna().unique().tolist())
                date_key = "forecast_selected_matrix_dates_v760"
                if date_key in st.session_state:
                    clean_dates = [
                        value for value in st.session_state[date_key]
                        if value in all_target_dates
                    ]
                    if clean_dates != st.session_state[date_key]:
                        st.session_state[date_key] = clean_dates
                else:
                    st.session_state[date_key] = list(all_target_dates)

                selected_target_dates = st.multiselect(
                    "Дни плана из матрицы — выбрать для расчёта и заполнения",
                    all_target_dates,
                    format_func=lambda value: (
                        f"{value:%d.%m.%Y} · {WEEKDAY_RU.get(value.weekday(), '')}"
                    ),
                    key=date_key,
                    help=(
                        "В расчёт и в готовую выгрузку попадут только выбранные даты. "
                        "Меню каждой даты берётся непосредственно из соответствующего блока текущей матрицы."
                    ),
                )

                if not selected_target_dates:
                    st.info("Выберите хотя бы одну дату плана из матрицы.")
                else:
                    selected_target_dates = sorted(selected_target_dates)
                    selected_matrix_menu = matrix_menu[
                        matrix_menu["target_date"].isin(selected_target_dates)
                    ].copy()

                    selected_blocks = (
                        selected_matrix_menu[["target_date", "sheet", "day_label"]]
                        .drop_duplicates()
                        .sort_values(["target_date", "sheet"])
                    )
                    block_records = selected_blocks.to_dict("records")
                    preview_key = "forecast_preview_block_v760"
                    if (
                        preview_key in st.session_state
                        and st.session_state[preview_key] not in block_records
                    ):
                        st.session_state.pop(preview_key, None)
                    selected_block = st.selectbox(
                        "Предпросмотр меню выбранного дня",
                        block_records,
                        format_func=lambda item: (
                            f"{item['target_date']:%d.%m.%Y} · "
                            f"{WEEKDAY_RU.get(item['target_date'].weekday(), '')}"
                        ),
                        key=preview_key,
                    )
                    selected_menu = selected_matrix_menu[
                        (selected_matrix_menu["target_date"] == selected_block["target_date"])
                        & (selected_matrix_menu["sheet"] == selected_block["sheet"])
                    ].copy()
                    preview_menu = selected_menu.merge(
                        entities[["sku", "category", "entity"]], on="sku", how="left"
                    )
                    preview_menu["category"] = preview_menu["category"].fillna(
                        preview_menu["matrix_category"]
                    )
                    preview_menu["entity"] = preview_menu["entity"].fillna("Не сопоставлено")

                    forecast_metrics = st.columns(4)
                    forecast_metrics[0].metric("Выбрано дат", len(selected_target_dates))
                    forecast_metrics[1].metric("SKU в дне", preview_menu["sku"].nunique())
                    forecast_metrics[2].metric("Категорий", preview_menu["category"].nunique())
                    forecast_metrics[3].metric("Сущностей", preview_menu["entity"].nunique())

                    with st.expander("Показать меню выбранного дня"):
                        st.dataframe(
                            preview_menu[["sku", "product_name", "category", "entity", "price"]].rename(
                                columns={
                                    "sku": "SKU",
                                    "product_name": "Название товара",
                                    "category": "Категория",
                                    "entity": "Сущность",
                                    "price": "Цена",
                                }
                            ),
                            use_container_width=True,
                            hide_index=True,
                        )

                    first_target_date = selected_target_dates[0]
                    last_target_date = selected_target_dates[-1]
                    lookback_weeks = st.slider(
                        "Период для расчёта среднего, недель",
                        min_value=1,
                        max_value=26,
                        value=8,
                        step=1,
                        key="forecast_average_period_weeks_v75111",
                        help=(
                            "Ползунок задаёт глубину истории перед каждой датой плана. "
                            "Например, 4 недели = предыдущие 28 календарных дней, а 26 недель ≈ 6 месяцев истории. "
                            "Среднее SKU всё равно считается только по фактическим дням продаж внутри выбранного окна."
                        ),
                    )
                    history_from = first_target_date - timedelta(weeks=int(lookback_weeks))
                    st.info(
                        f"К расчёту выбрано {len(selected_target_dates)} дат: "
                        f"{', '.join(value.strftime('%d.%m.%Y') for value in selected_target_dates)}. "
                        f"Период среднего: {int(lookback_weeks)} нед. ({int(lookback_weeks) * 7} дн.). "
                        "Для каждой даты используется меню именно её блока и отдельное историческое окно "
                        "непосредственно перед этой датой."
                    )

                    forecast_button = st.button(
                        "Рассчитать план для выбранных дат",
                        type="primary",
                        key="calculate_forecast",
                        use_container_width=True,
                    )
                    if forecast_button:
                        forecast_points = {
                            int(number): label
                            for number, label in st.session_state.get("point_mapping", {}).items()
                            if str(label).startswith("Т")
                            and 1 <= int(str(label)[1:]) <= 29
                            and int(str(label)[1:]) != 11
                        }
                        if not forecast_points:
                            forecast_points = {
                                number: f"Т{number}"
                                for number in range(1, 30)
                                if number != 11
                            }
                        try:
                            forecast_history = load_forecast_history(
                                history_from,
                                last_target_date,
                                tuple(sorted(forecast_points)),
                            )
                        except Exception as error:
                            st.error(f"Не удалось загрузить историю продаж: {error}")
                            forecast_history = pd.DataFrame()

                        if forecast_history.empty:
                            st.warning("За исторический период продажи не найдены.")
                        else:
                            forecast_parts: list[pd.DataFrame] = []
                            for (target_date, sheet_name), menu_block in selected_matrix_menu.groupby(
                                ["target_date", "sheet"], sort=True
                            ):
                                block_history_from = target_date - timedelta(weeks=int(lookback_weeks))
                                block_history = forecast_history[
                                    pd.to_datetime(
                                        forecast_history["business_date"], errors="coerce"
                                    ).dt.date.between(
                                        block_history_from,
                                        target_date - timedelta(days=1),
                                        inclusive="both",
                                    )
                                ].copy()
                                block_result = calculate_sku_daily_forecast(
                                    menu_block,
                                    block_history,
                                    entities,
                                    target_date,
                                    forecast_points,
                                    lookback_weeks=int(lookback_weeks),
                                )
                                if not block_result.empty:
                                    forecast_parts.append(block_result)

                            forecast_result = (
                                pd.concat(forecast_parts, ignore_index=True)
                                if forecast_parts else pd.DataFrame()
                            )
                            if forecast_result.empty:
                                st.warning("Не удалось сформировать расчёт для выбранных дат меню.")
                                st.stop()

                            st.session_state["forecast_result"] = forecast_result
                            st.session_state["forecast_menu"] = selected_matrix_menu
                            st.session_state["forecast_history_from"] = history_from
                            st.session_state["forecast_target_date"] = last_target_date
                            st.session_state["forecast_first_target_date"] = first_target_date
                            st.session_state["forecast_matrix_bytes"] = matrix_bytes
                            st.session_state["forecast_calculated_dates_v760"] = list(
                                selected_target_dates
                            )
                            st.session_state["forecast_calculated_lookback_weeks_v75111"] = int(
                                lookback_weeks
                            )

                    if "forecast_result" in st.session_state:
                        forecast_result = st.session_state["forecast_result"]
                        calculated_dates = st.session_state.get(
                            "forecast_calculated_dates_v760", []
                        )
                        calculated_lookback_weeks = int(
                            st.session_state.get("forecast_calculated_lookback_weeks_v75111", 0) or 0
                        )
                        calculation_current = (
                            list(selected_target_dates) == list(calculated_dates)
                            and int(lookback_weeks) == calculated_lookback_weeks
                        )
                        if not calculation_current:
                            changed_parts = []
                            if list(selected_target_dates) != list(calculated_dates):
                                changed_parts.append("даты плана")
                            if int(lookback_weeks) != calculated_lookback_weeks:
                                changed_parts.append("период среднего")
                            st.warning(
                                f"После последнего расчёта изменены: {', '.join(changed_parts)}. "
                                "Нажмите «Рассчитать план для выбранных дат», чтобы пересчитать готовый план."
                            )

                        st.markdown("#### Рекомендованный план")
                        result_metrics = st.columns(5)
                        result_metrics[0].metric("Дат в готовом плане", forecast_result["Дата плана"].nunique())
                        result_metrics[1].metric("Точек", forecast_result["Точка"].nunique())
                        result_metrics[2].metric("SKU", forecast_result["SKU"].nunique())
                        result_metrics[3].metric(
                            "План, шт.",
                            f"{forecast_result['Рекомендованный план'].sum(skipna=True):,.0f}".replace(",", " "),
                        )
                        result_metrics[4].metric(
                            "Пустых ячеек без данных",
                            int(forecast_result["Рекомендованный план"].isna().sum()),
                        )

                        result_point_options = sorted(
                            forecast_result["Точка"].unique(),
                            key=lambda value: int(str(value)[1:]),
                        )
                        result_points = st.multiselect(
                            "Показать точки прогноза",
                            result_point_options,
                            default=result_point_options[:1],
                            key="forecast_result_points",
                        )
                        visible_forecast = forecast_result[
                            forecast_result["Точка"].isin(result_points)
                        ]
                        st.dataframe(
                            visible_forecast,
                            use_container_width=True,
                            hide_index=True,
                        )

                        hide_average_values = st.checkbox(
                            "Убрать числовые значения СР из выгрузки",
                            value=False,
                            key="forecast_hide_average_values_export",
                            help=(
                                "Синяя строка СР, подпись и цикл загрузки останутся. "
                                "Будут очищены только числовые средние по точкам и итог СР в колонке ПЛАН."
                            ),
                        )
                        export_period_label = _forecast_export_date_label(forecast_result)
                        st.success(f"Выгрузка плана меню: {export_period_label}")
                        st.caption(
                            "В скачанный Excel попадут только выбранные даты меню. Остальные дни и недели "
                            "исходной матрицы полностью исключаются из выгрузки. Столбцы Ф удаляются только "
                            "из скачанной копии; исходная матрица не изменяется."
                        )

                        if calculation_current:
                            filled_matrix = fill_forecast_into_matrix(
                                st.session_state["forecast_matrix_bytes"],
                                forecast_result,
                                st.session_state["forecast_history_from"],
                                st.session_state["forecast_target_date"],
                                hide_average_values=hide_average_values,
                            )
                            st.download_button(
                                "Скачать готовый план по выбранным датам",
                                data=filled_matrix,
                                file_name=(
                                    f"план_меню_"
                                    f"{st.session_state['forecast_first_target_date']:%Y-%m-%d}_"
                                    f"{st.session_state['forecast_target_date']:%Y-%m-%d}.xlsx"
                                ),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="forecast_download_selected_dates_v760",
                                use_container_width=True,
                            )

    excel_bytes = export_excel(filtered_sku, filtered_category, filtered_entity, filtered_detail)
    st.download_button(
        "Скачать Excel",
        data=excel_bytes,
        file_name=f"структура_спроса_{period[0]:%Y-%m-%d}_{period[1]:%Y-%m-%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown('<div class="vk-footer-brand">ВКУСНО МАРКЕТ</div>', unsafe_allow_html=True)
